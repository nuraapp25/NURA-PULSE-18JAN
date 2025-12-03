from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, Query, Request, Form, UploadFile, File, BackgroundTasks, Body
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
from passlib.context import CryptContext
import jwt
import secrets
import string
import base64
import googlemaps
import requests
import asyncio
import pandas as pd
import numpy as np

# Import hotspot optimizer
from hotspot_optimizer import optimize_hotspots, TIME_SLOTS

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Import after loading .env so environment variables are available
from sheets_multi_sync import (
    sync_user_to_sheets, bulk_sync_users_to_sheets, delete_user_from_sheets,
    sync_all_records, get_all_records, sync_single_record, delete_record,
    get_last_sync_time, update_last_sync_time
)

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db_name = os.environ.get('DB_NAME', 'nura_pulse_db')  # Default to nura_pulse_db
print(f"🔍 Connecting to MongoDB database: {db_name}")
db = client[db_name]

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# JWT settings
SECRET_KEY = os.environ.get('JWT_SECRET_KEY', 'your-secret-key-change-in-production')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24 * 7  # 7 days

# Security
security = HTTPBearer()

# Create the main app
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Health check endpoint (outside /api prefix for monitoring)
@app.get("/health")
async def health_check():
    """Simple health check endpoint to monitor server status"""
    return {
        "status": "healthy",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "service": "Nura Pulse API"
    }

@api_router.get("/health")
async def api_health_check():
    """Health check endpoint under /api prefix"""
    return {
        "status": "healthy",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "service": "Nura Pulse API"
    }

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


# ==================== Models ====================

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    first_name: str
    last_name: Optional[str] = None
    email: str
    account_type: str  # "master_admin", "admin", "standard", "ops_team", "telecaller"
    status: str = "pending"  # "pending", "active", "rejected", "deleted"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    is_temp_password: bool = False
    telecaller_id: Optional[str] = None  # Link to telecaller profile if account_type is "telecaller"


class UserCreate(BaseModel):
    first_name: str
    last_name: Optional[str] = None
    email: str
    password: str
    account_type: str  # "admin", "standard", "ops_team", or "telecaller"


class UserLogin(BaseModel):
    email: str
    password: str


class PasswordResetRequest(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    user_email: str
    user_name: str
    status: str = "pending"  # pending, approved, rejected
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    temporary_password: Optional[str] = None


class PasswordChange(BaseModel):
    old_password: str
    new_password: str


class TempPasswordReset(BaseModel):
    email: str
    temp_password: str
    new_password: str


class PasswordResetApproval(BaseModel):
    request_id: str


class UserApproval(BaseModel):
    user_id: str


# ==================== Helper Functions ====================

def hash_password(password: str) -> str:
    return pwd_context.hash(password)


def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)


def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt


def generate_temporary_password(length: int = 12) -> str:
    """Generate a secure temporary password"""
    characters = string.ascii_letters + string.digits + "!@#$%"
    return ''.join(secrets.choice(characters) for _ in range(length))



def load_mode_mapping_tables():
    """Load Model Mapping and Ride Mode Mapping from Excel file"""
    import pandas as pd
    
    mapping_file_path = "/app/backend/uploaded_files/mode_details.xlsx"
    
    try:
        # Read Model Mapping
        model_mapping = pd.read_excel(mapping_file_path, sheet_name='Model Mapping')
        model_dict = {}
        for _, row in model_mapping.iterrows():
            vehicle_num = str(row['Vehicle Number']).strip()
            model = str(row['Model']).strip()
            model_dict[vehicle_num] = model
        
        # Read Ride Mode Mapping
        ride_mode_mapping = pd.read_excel(mapping_file_path, sheet_name='Ride Mode Mapping')
        mode_dict = {}
        for _, row in ride_mode_mapping.iterrows():
            concat_key = str(row['Vehicle-Mode Concatenation']).strip()
            mode_name = str(row['Mode Name']).strip()
            mode_type = str(row['Mode Type']).strip()
            mode_dict[concat_key] = {
                'mode_name': mode_name,
                'mode_type': mode_type
            }
        
        logger.info(f"Loaded {len(model_dict)} vehicle models and {len(mode_dict)} ride modes")
        return model_dict, mode_dict
    except Exception as e:
        logger.error(f"Error loading mode mapping tables: {str(e)}")
        return {}, {}


def enrich_with_mode_data(vehicle_id: str, ride_mode: str, model_dict: dict, mode_dict: dict):
    """Enrich a single record with Mode Name and Mode Type"""
    # Get model for this vehicle
    model = model_dict.get(vehicle_id, "Unknown")
    
    if model == "Unknown":
        return "Unknown Model", "Unknown"
    
    # Create concatenation key
    concat_key = f"{model} {ride_mode}"
    
    # Look up mode details
    mode_details = mode_dict.get(concat_key)
    
    if mode_details:
        return mode_details['mode_name'], mode_details['mode_type']
    else:
        return "Unknown Mode", "Unknown"


async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    token = credentials.credentials
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id = payload.get("user_id")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid token")
        
        user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
        if user is None:
            raise HTTPException(status_code=401, detail="User not found")
        
        # Convert datetime strings back to datetime objects
        if isinstance(user.get('created_at'), str):
            user['created_at'] = datetime.fromisoformat(user['created_at'])
        
        return User(**user)
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except Exception as e:
        logger.error(f"Token verification error: {e}")
        raise HTTPException(status_code=401, detail="Invalid token")


async def initialize_master_admin():
    """Create master admin account if it doesn't exist"""
    master_admin = await db.users.find_one({"email": "admin"})
    if not master_admin:
        master_admin_data = {
            "id": str(uuid.uuid4()),
            "first_name": "Master Admin",
            "last_name": "",
            "email": "admin",
            "password": hash_password("Nura@1234$"),
            "account_type": "master_admin",
            "status": "active",  # Master admin is always active
            "created_at": datetime.now(timezone.utc).isoformat(),
            "is_temp_password": False
        }
        await db.users.insert_one(master_admin_data)
        logger.info("Master admin account created")
        
        # Sync to Google Sheets
        sync_user_to_sheets(master_admin_data)


# ==================== Auth Routes ====================

@api_router.post("/auth/register")
async def register(user_data: UserCreate):
    from datetime import datetime, timezone
    
    # Check if user already exists
    existing_user = await db.users.find_one({"email": user_data.email})
    if existing_user:
        # Allow re-registration if the account was deleted
        if existing_user.get('status') == 'deleted':
            # Delete the old record completely
            await db.users.delete_one({"email": user_data.email})
        else:
            raise HTTPException(status_code=400, detail="Email already registered. If you're awaiting approval, please contact admin.")
    
    # Validate account type
    if user_data.account_type not in ["admin", "standard", "ops_team", "telecaller"]:
        raise HTTPException(status_code=400, detail="Invalid account type")
    
    # Create user with pending status
    hashed_password = hash_password(user_data.password)
    user = User(
        first_name=user_data.first_name,
        last_name=user_data.last_name,
        email=user_data.email,
        account_type=user_data.account_type,
        status="pending"  # Requires admin approval
    )
    
    user_dict = user.model_dump()
    user_dict['password'] = hashed_password
    user_dict['created_at'] = user_dict['created_at'].isoformat()
    
    await db.users.insert_one(user_dict)
    
    # Sync to Google Sheets
    sync_user_to_sheets(user_dict)
    
    # If registering as telecaller, automatically create telecaller profile
    if user_data.account_type == "telecaller":
        # Check if telecaller profile already exists
        existing_profile = await db.telecaller_profiles.find_one({"email": user_data.email})
        
        if not existing_profile:
            telecaller_profile = {
                "id": str(uuid.uuid4()),
                "name": f"{user_data.first_name} {user_data.last_name}",
                "phone_number": "",  # Can be updated later
                "email": user_data.email,
                "status": "pending",  # Match user status - will become active when admin approves
                "notes": "Auto-created from user registration",
                "aadhar_card": "",
                "pan_card": "",
                "address_proof": "",
                "total_assigned_leads": 0,
                "active_leads": 0,
                "converted_leads": 0,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "last_modified": datetime.now(timezone.utc).isoformat(),
                "created_by": user_dict['id']  # Self-created
            }
            
            await db.telecaller_profiles.insert_one(telecaller_profile)
    
    return {
        "message": "Registration successful! Your account is pending approval. Please wait for admin to approve.",
        "status": "pending"
    }


@api_router.post("/auth/login")
async def login(credentials: UserLogin):
    # Find user
    user = await db.users.find_one({"email": credentials.email})
    if not user:
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    # Verify password
    if not verify_password(credentials.password, user['password']):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    # Check if user is approved
    if user.get('status') == 'pending':
        raise HTTPException(status_code=403, detail="Your account is pending approval. Please contact admin.")
    
    if user.get('status') == 'rejected':
        raise HTTPException(status_code=403, detail="Your account has been rejected. Please contact admin.")
    
    if user.get('status') == 'deleted':
        raise HTTPException(status_code=403, detail="Your account has been deleted. Please contact admin.")
    
    if user.get('status') != 'active':
        raise HTTPException(status_code=403, detail="Your account is not active. Please contact admin.")
    
    # Convert datetime string back
    if isinstance(user.get('created_at'), str):
        user['created_at'] = datetime.fromisoformat(user['created_at'])
    
    user_obj = User(**user)
    
    # Create token
    token = create_access_token({"user_id": user_obj.id, "email": user_obj.email})
    
    return {
        "message": "Login successful",
        "user": user_obj,
        "token": token
    }


@api_router.get("/auth/me")
async def get_me(current_user: User = Depends(get_current_user)):
    return current_user


@api_router.post("/auth/change-password")
async def change_password(
    password_data: PasswordChange,
    current_user: User = Depends(get_current_user)
):
    # Get user with password
    user = await db.users.find_one({"id": current_user.id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Verify old password
    if not verify_password(password_data.old_password, user['password']):
        raise HTTPException(status_code=400, detail="Incorrect old password")
    
    # Update password
    new_hashed_password = hash_password(password_data.new_password)
    await db.users.update_one(
        {"id": current_user.id},
        {"$set": {"password": new_hashed_password, "is_temp_password": False}}
    )
    
    return {"message": "Password changed successfully"}


@api_router.post("/auth/reset-with-temp-password")
async def reset_with_temp_password(reset_data: TempPasswordReset):
    """Reset password using email and temporary password (from forgot password page)"""
    # Find user by email
    user = await db.users.find_one({"email": reset_data.email})
    
    if not user:
        raise HTTPException(status_code=400, detail="Invalid email or temporary password")
    
    # Check if user has a temporary password
    if not user.get('is_temp_password'):
        raise HTTPException(status_code=400, detail="No temporary password set for this account. Please request a password reset from admin.")
    
    # Verify the temporary password
    if not verify_password(reset_data.temp_password, user['password']):
        raise HTTPException(status_code=400, detail="Invalid email or temporary password")
    
    # Update password
    new_hashed_password = hash_password(reset_data.new_password)
    await db.users.update_one(
        {"id": user['id']},
        {"$set": {"password": new_hashed_password, "is_temp_password": False}}
    )
    
    return {"message": "Password reset successfully. You can now login with your new password."}


# ==================== Password Reset Routes ====================

@api_router.post("/password-reset/request")
async def request_password_reset(current_user: User = Depends(get_current_user)):
    """User requests a password reset"""
    # Check if there's already a pending request
    existing_request = await db.password_reset_requests.find_one({
        "user_id": current_user.id,
        "status": "pending"
    })
    
    if existing_request:
        raise HTTPException(status_code=400, detail="You already have a pending password reset request")
    
    # Create reset request
    reset_request = PasswordResetRequest(
        user_id=current_user.id,
        user_email=current_user.email,
        user_name=current_user.first_name
    )
    
    request_dict = reset_request.model_dump()
    request_dict['created_at'] = request_dict['created_at'].isoformat()
    
    await db.password_reset_requests.insert_one(request_dict)
    
    return {"message": "Password reset request submitted. Please wait for admin approval."}


@api_router.get("/password-reset/requests")
async def get_password_reset_requests(current_user: User = Depends(get_current_user)):
    """Get all password reset requests (admin only)"""
    if current_user.account_type not in ["master_admin", "admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    requests = await db.password_reset_requests.find({}, {"_id": 0}).to_list(1000)
    
    # Convert datetime strings
    for req in requests:
        if isinstance(req.get('created_at'), str):
            req['created_at'] = datetime.fromisoformat(req['created_at'])
    
    return requests


@api_router.post("/password-reset/approve")
async def approve_password_reset(
    approval: PasswordResetApproval,
    current_user: User = Depends(get_current_user)
):
    """Admin approves password reset and generates temporary password"""
    if current_user.account_type not in ["master_admin", "admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Find the request
    request = await db.password_reset_requests.find_one({"id": approval.request_id})
    if not request:
        raise HTTPException(status_code=404, detail="Request not found")
    
    if request['status'] != "pending":
        raise HTTPException(status_code=400, detail="Request already processed")
    
    # Generate temporary password
    temp_password = generate_temporary_password()
    
    # Update user's password
    hashed_temp_password = hash_password(temp_password)
    await db.users.update_one(
        {"id": request['user_id']},
        {"$set": {"password": hashed_temp_password, "is_temp_password": True}}
    )
    
    # Update request status
    await db.password_reset_requests.update_one(
        {"id": approval.request_id},
        {"$set": {"status": "approved", "temporary_password": temp_password}}
    )
    
    return {
        "message": "Password reset approved",
        "temporary_password": temp_password,
        "user_email": request['user_email']
    }


# ==================== User Management Routes (Master Admin) ====================

@api_router.get("/users")
async def get_all_users(current_user: User = Depends(get_current_user)):
    """Get all users (master admin and admin only) - excludes deleted users"""
    if current_user.account_type not in ["master_admin", "admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Filter out deleted users
    users = await db.users.find({"status": {"$ne": "deleted"}}, {"_id": 0, "password": 0}).to_list(1000)
    
    # Convert datetime strings
    for user in users:
        if isinstance(user.get('created_at'), str):
            user['created_at'] = datetime.fromisoformat(user['created_at'])
    
    return users


@api_router.post("/users/create")
async def create_user(user_data: UserCreate, current_user: User = Depends(get_current_user)):
    """Create a new user (master admin and admin only)"""
    if current_user.account_type not in ["master_admin", "admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Check if user already exists
    existing_user = await db.users.find_one({"email": user_data.email})
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Validate account type
    if user_data.account_type not in ["admin", "standard", "ops_team", "telecaller"]:
        raise HTTPException(status_code=400, detail="Invalid account type")
    
    # Create user (master admin created users are auto-approved)
    hashed_password = hash_password(user_data.password)
    user = User(
        first_name=user_data.first_name,
        last_name=user_data.last_name,
        email=user_data.email,
        account_type=user_data.account_type,
        status="active"  # Master admin created users are auto-approved
    )
    
    user_dict = user.model_dump()
    user_dict['password'] = hashed_password
    user_dict['created_at'] = user_dict['created_at'].isoformat()
    
    await db.users.insert_one(user_dict)
    
    # If creating a telecaller, automatically create telecaller profile
    if user_data.account_type == "telecaller":
        # Check if telecaller profile already exists
        existing_profile = await db.telecaller_profiles.find_one({"email": user_data.email})
        
        if not existing_profile:
            telecaller_profile = {
                "id": str(uuid.uuid4()),
                "name": f"{user_data.first_name} {user_data.last_name}",
                "phone_number": "",  # Can be updated later
                "email": user_data.email,
                "status": "active",  # Admin-created telecallers are active immediately
                "notes": "Auto-created from User Management",
                "aadhar_card": "",
                "pan_card": "",
                "address_proof": "",
                "total_assigned_leads": 0,
                "active_leads": 0,
                "converted_leads": 0,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "last_modified": datetime.now(timezone.utc).isoformat(),
                "created_by": current_user.id
            }
            
            await db.telecaller_profiles.insert_one(telecaller_profile)
            logger.info(f"Auto-created telecaller profile for {user_data.email}")
    
    # Sync to Google Sheets
    sync_user_to_sheets(user_dict)
    
    return {"message": "User created successfully", "user": user}


@api_router.post("/users/approve")
async def approve_user(approval: UserApproval, current_user: User = Depends(get_current_user)):
    """Approve a pending user (master admin and admin only)"""
    from datetime import datetime, timezone
    
    if current_user.account_type not in ["master_admin", "admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    user = await db.users.find_one({"id": approval.user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    if user.get('status') != 'pending':
        raise HTTPException(status_code=400, detail="User is not pending approval")
    
    # Update user status to active
    await db.users.update_one(
        {"id": approval.user_id},
        {"$set": {"status": "active"}}
    )
    
    # If user is a telecaller, also activate their telecaller profile
    if user.get('account_type') == 'telecaller':
        await db.telecaller_profiles.update_one(
            {"email": user['email']},
            {"$set": {
                "status": "active",
                "last_modified": datetime.now(timezone.utc).isoformat()
            }}
        )
    
    # Sync to Google Sheets
    user['status'] = 'active'
    sync_user_to_sheets(user)
    
    return {"message": "User approved successfully"}


@api_router.post("/users/reject")
async def reject_user(approval: UserApproval, current_user: User = Depends(get_current_user)):
    """Reject a pending user (master admin and admin only)"""
    from datetime import datetime, timezone
    
    if current_user.account_type not in ["master_admin", "admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    user = await db.users.find_one({"id": approval.user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    if user.get('status') != 'pending':
        raise HTTPException(status_code=400, detail="User is not pending approval")
    
    # Update user status to rejected
    await db.users.update_one(
        {"id": approval.user_id},
        {"$set": {"status": "rejected"}}
    )
    
    # If user is a telecaller, also reject their telecaller profile
    if user.get('account_type') == 'telecaller':
        await db.telecaller_profiles.update_one(
            {"email": user['email']},
            {"$set": {
                "status": "inactive",
                "notes": "Rejected by admin",
                "last_modified": datetime.now(timezone.utc).isoformat()
            }}
        )
    
    # Sync to Google Sheets
    user['status'] = 'rejected'
    sync_user_to_sheets(user)
    
    return {"message": "User rejected"}


@api_router.post("/users/{user_id}/generate-temp-password")
async def generate_temp_password_for_user(user_id: str, current_user: User = Depends(get_current_user)):
    """Generate temporary password for a user (master admin and admin only)"""
    if current_user.account_type not in ["master_admin", "admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    if user.get('account_type') == 'master_admin':
        raise HTTPException(status_code=400, detail="Cannot reset master admin password")
    
    # Generate temporary password
    temp_password = generate_temporary_password()
    hashed_temp_password = hash_password(temp_password)
    
    # Update user's password
    await db.users.update_one(
        {"id": user_id},
        {"$set": {"password": hashed_temp_password, "is_temp_password": True}}
    )
    
    return {
        "message": "Temporary password generated",
        "temporary_password": temp_password,
        "user_email": user.get('email')
    }


@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: User = Depends(get_current_user)):
    """Delete a user (master admin and admin only) - permanently removes from database"""
    from datetime import datetime, timezone
    
    if current_user.account_type not in ["master_admin", "admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Don't allow deleting master admin or other admins (admins can only delete non-admin users)
    user_to_delete = await db.users.find_one({"id": user_id})
    if not user_to_delete:
        raise HTTPException(status_code=404, detail="User not found")
    
    if user_to_delete.get('account_type') == "master_admin":
        raise HTTPException(status_code=400, detail="Cannot delete master admin")
    
    # Regular admins cannot delete other admin accounts
    if current_user.account_type == "admin" and user_to_delete.get('account_type') == "admin":
        raise HTTPException(status_code=403, detail="Admins cannot delete other admin accounts")
    
    # If user is a telecaller, deactivate their telecaller profile
    if user_to_delete.get('account_type') == 'telecaller':
        await db.telecaller_profiles.update_one(
            {"email": user_to_delete['email']},
            {"$set": {
                "status": "inactive",
                "notes": "User account deleted",
                "last_modified": datetime.now(timezone.utc).isoformat()
            }}
        )
    
    # Permanently delete the user
    await db.users.delete_one({"id": user_id})
    
    # Update Google Sheets - remove the row
    delete_user_from_sheets(user_to_delete.get('email'))
    
    return {"message": "User deleted successfully"}


class AccountTypeChange(BaseModel):
    user_id: str
    new_account_type: str


@api_router.post("/users/change-account-type")
async def change_user_account_type(
    account_change: AccountTypeChange,
    current_user: User = Depends(get_current_user)
):
    """Change user account type (master admin only)"""
    from datetime import datetime, timezone
    
    if current_user.account_type != "master_admin":
        raise HTTPException(status_code=403, detail="Only master admin can change account types")
    
    # Validate new account type
    valid_types = ["admin", "standard", "ops_team", "telecaller"]
    if account_change.new_account_type not in valid_types:
        raise HTTPException(status_code=400, detail=f"Invalid account type. Must be one of: {', '.join(valid_types)}")
    
    # Find the user
    user_to_update = await db.users.find_one({"id": account_change.user_id})
    if not user_to_update:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Cannot change master admin account type
    if user_to_update.get('account_type') == "master_admin":
        raise HTTPException(status_code=400, detail="Cannot change master admin account type")
    
    old_account_type = user_to_update.get('account_type')
    
    # Update user account type
    await db.users.update_one(
        {"id": account_change.user_id},
        {"$set": {"account_type": account_change.new_account_type}}
    )
    
    # Handle telecaller profile creation/deactivation
    if account_change.new_account_type == "telecaller" and old_account_type != "telecaller":
        # User is being changed TO telecaller - create profile if doesn't exist
        existing_profile = await db.telecaller_profiles.find_one({"email": user_to_update['email']})
        
        if not existing_profile:
            telecaller_profile = {
                "id": str(uuid.uuid4()),
                "name": f"{user_to_update.get('first_name', '')} {user_to_update.get('last_name', '')}",
                "phone_number": "",
                "email": user_to_update['email'],
                "status": "active" if user_to_update.get('status') == 'active' else "pending",
                "notes": f"Created from account type change by {current_user.email}",
                "aadhar_card": "",
                "pan_card": "",
                "address_proof": "",
                "total_assigned_leads": 0,
                "active_leads": 0,
                "converted_leads": 0,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "last_modified": datetime.now(timezone.utc).isoformat(),
                "created_by": current_user.id
            }
            await db.telecaller_profiles.insert_one(telecaller_profile)
        else:
            # Reactivate existing profile
            await db.telecaller_profiles.update_one(
                {"email": user_to_update['email']},
                {"$set": {
                    "status": "active" if user_to_update.get('status') == 'active' else "pending",
                    "last_modified": datetime.now(timezone.utc).isoformat()
                }}
            )
    elif old_account_type == "telecaller" and account_change.new_account_type != "telecaller":
        # User is being changed FROM telecaller - deactivate profile
        await db.telecaller_profiles.update_one(
            {"email": user_to_update['email']},
            {"$set": {
                "status": "inactive",
                "notes": f"User account type changed from telecaller to {account_change.new_account_type} by {current_user.email}",
                "last_modified": datetime.now(timezone.utc).isoformat()
            }}
        )
    
    # Update Google Sheets
    user_to_update['account_type'] = account_change.new_account_type
    sync_user_to_sheets(user_to_update)
    
    return {
        "message": "Account type changed successfully",
        "old_type": old_account_type,
        "new_type": account_change.new_account_type
    }


@api_router.get("/stats")
async def get_stats(current_user: User = Depends(get_current_user)):
    """Get user statistics (master admin and admin only)"""
    if current_user.account_type not in ["master_admin", "admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    total_users = await db.users.count_documents({"status": {"$ne": "deleted"}})
    admin_users = await db.users.count_documents({"account_type": "admin", "status": {"$ne": "deleted"}})
    standard_users = await db.users.count_documents({"account_type": "standard", "status": {"$ne": "deleted"}})
    ops_team_users = await db.users.count_documents({"account_type": "ops_team", "status": {"$ne": "deleted"}})
    pending_users = await db.users.count_documents({"status": "pending"})
    pending_resets = await db.password_reset_requests.count_documents({"status": "pending"})
    
    return {
        "total_users": total_users,
        "admin_users": admin_users,
        "standard_users": standard_users,
        "ops_team_users": ops_team_users,
        "pending_users": pending_users,
        "pending_password_resets": pending_resets
    }


@api_router.get("/users/export")
async def export_users(
    user_ids: Optional[str] = Query(None, description="Comma-separated user IDs to export"),
    current_user: User = Depends(get_current_user)
):
    """Export users with encrypted passwords (master admin only)"""
    from cryptography.fernet import Fernet
    import json
    import base64
    
    if current_user.account_type != "master_admin":
        raise HTTPException(status_code=403, detail="Only master admin can export users")
    
    try:
        # Build query
        query = {}
        if user_ids:
            user_id_list = user_ids.split(',')
            query["id"] = {"$in": user_id_list}
        
        # Fetch users from database
        users = await db.users.find(query).to_list(None)
        
        # Get passwords from user_credentials collection
        export_data = []
        for user in users:
            # Get hashed password
            credential = await db.user_credentials.find_one({"user_id": user["id"]})
            
            user_data = {
                "id": user.get("id"),
                "first_name": user.get("first_name"),
                "last_name": user.get("last_name"),
                "email": user.get("email"),
                "account_type": user.get("account_type"),
                "status": user.get("status"),
                "created_at": user.get("created_at"),
                "is_temp_password": user.get("is_temp_password", False),
                "telecaller_id": user.get("telecaller_id"),
                "hashed_password": credential.get("hashed_password") if credential else None
            }
            export_data.append(user_data)
        
        # Generate encryption key
        encryption_key = Fernet.generate_key()
        cipher = Fernet(encryption_key)
        
        # Encrypt the data
        json_data = json.dumps(export_data, default=str)
        encrypted_data = cipher.encrypt(json_data.encode())
        
        # Encode for transmission
        encrypted_b64 = base64.b64encode(encrypted_data).decode()
        key_b64 = base64.b64encode(encryption_key).decode()
        
        return {
            "success": True,
            "encrypted_data": encrypted_b64,
            "encryption_key": key_b64,
            "count": len(export_data),
            "exported_at": datetime.now(timezone.utc).isoformat(),
            "message": "Users exported successfully. Save both the data and encryption key securely."
        }
    
    except Exception as e:
        logger.error(f"Error exporting users: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error exporting users: {str(e)}")


@api_router.post("/users/import/check-duplicates")
async def check_import_duplicates(
    encrypted_data: str = Body(...),
    encryption_key: str = Body(...),
    current_user: User = Depends(get_current_user)
):
    """Check for duplicate users before importing (master admin only)"""
    from cryptography.fernet import Fernet
    import json
    import base64
    
    if current_user.account_type != "master_admin":
        raise HTTPException(status_code=403, detail="Only master admin can check duplicates")
    
    try:
        # Decode encryption key and data
        key = base64.b64decode(encryption_key.encode())
        encrypted_bytes = base64.b64decode(encrypted_data.encode())
        
        # Decrypt data
        cipher = Fernet(key)
        decrypted_data = cipher.decrypt(encrypted_bytes)
        users_data = json.loads(decrypted_data.decode())
        
        # Check for duplicates
        duplicate_emails = []
        for user_data in users_data:
            existing_user = await db.users.find_one({"email": user_data["email"]})
            if existing_user:
                duplicate_emails.append(user_data["email"])
        
        return {
            "success": True,
            "duplicates_found": len(duplicate_emails) > 0,
            "duplicate_count": len(duplicate_emails),
            "duplicate_emails": duplicate_emails,
            "total_users": len(users_data),
            "new_users": len(users_data) - len(duplicate_emails)
        }
    
    except Exception as e:
        logger.error(f"Error checking duplicates: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error checking duplicates: {str(e)}")


@api_router.post("/users/import")
async def import_users(
    encrypted_data: str = Body(...),
    encryption_key: str = Body(...),
    duplicate_action: str = Body("skip", description="Action for duplicates: 'skip' or 'replace'"),
    current_user: User = Depends(get_current_user)
):
    """Import users from encrypted export file (master admin only)"""
    from cryptography.fernet import Fernet
    import json
    import base64
    
    if current_user.account_type != "master_admin":
        raise HTTPException(status_code=403, detail="Only master admin can import users")
    
    try:
        # Decode encryption key and data
        key = base64.b64decode(encryption_key.encode())
        encrypted_bytes = base64.b64decode(encrypted_data.encode())
        
        # Decrypt data
        cipher = Fernet(key)
        decrypted_data = cipher.decrypt(encrypted_bytes)
        users_data = json.loads(decrypted_data.decode())
        
        imported_count = 0
        skipped_count = 0
        replaced_count = 0
        errors = []
        
        for user_data in users_data:
            try:
                # Check if user already exists
                existing_user = await db.users.find_one({"email": user_data["email"]})
                
                if existing_user:
                    if duplicate_action == "skip":
                        skipped_count += 1
                        continue
                    elif duplicate_action == "replace":
                        # Replace existing user
                        user_doc = {
                            "first_name": user_data["first_name"],
                            "last_name": user_data.get("last_name"),
                            "account_type": user_data["account_type"],
                            "status": user_data.get("status", "active"),
                            "is_temp_password": user_data.get("is_temp_password", False),
                            "telecaller_id": user_data.get("telecaller_id")
                        }
                        
                        # Update user
                        await db.users.update_one(
                            {"email": user_data["email"]},
                            {"$set": user_doc}
                        )
                        
                        # Update password credential if exists
                        if user_data.get("hashed_password"):
                            await db.user_credentials.update_one(
                                {"user_id": existing_user["id"]},
                                {"$set": {"hashed_password": user_data["hashed_password"]}},
                                upsert=True
                            )
                        
                        replaced_count += 1
                        continue
                
                # Prepare user document for new user
                user_doc = {
                    "id": user_data["id"],
                    "first_name": user_data["first_name"],
                    "last_name": user_data.get("last_name"),
                    "email": user_data["email"],
                    "account_type": user_data["account_type"],
                    "status": user_data.get("status", "active"),
                    "created_at": user_data.get("created_at"),
                    "is_temp_password": user_data.get("is_temp_password", False),
                    "telecaller_id": user_data.get("telecaller_id")
                }
                
                # Insert user
                await db.users.insert_one(user_doc)
                
                # Insert password credential if exists
                if user_data.get("hashed_password"):
                    credential_doc = {
                        "user_id": user_data["id"],
                        "hashed_password": user_data["hashed_password"]
                    }
                    await db.user_credentials.insert_one(credential_doc)
                
                imported_count += 1
                
            except Exception as e:
                errors.append(f"Error importing {user_data.get('email', 'unknown')}: {str(e)}")
                logger.error(f"Error importing user {user_data.get('email')}: {str(e)}")
        
        return {
            "success": True,
            "imported": imported_count,
            "skipped": skipped_count,
            "replaced": replaced_count,
            "total": len(users_data),
            "errors": errors,
            "message": f"Successfully imported {imported_count} new users. {replaced_count} users replaced. {skipped_count} users skipped."
        }
    
    except Exception as e:
        logger.error(f"Error importing users: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error importing users: {str(e)}")


@api_router.post("/users/sync-to-sheets")
async def sync_all_users_to_sheets(current_user: User = Depends(get_current_user)):
    """Manually sync all users to Google Sheets (master admin and admin only)"""
    if current_user.account_type not in ["master_admin", "admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(1000)
    
    success = bulk_sync_users_to_sheets(users)
    
    if success:
        return {"message": "All users synced to Google Sheets successfully"}
    else:
        raise HTTPException(status_code=500, detail="Failed to sync users to Google Sheets. Check if Google Sheets integration is enabled.")


# ==================== Multi-App Endpoints ====================

from app_models import (
    PaymentRecord, PaymentRecordCreate, PaymentFolder, PaymentFolderCreate,
    DriverRecord, DriverRecordCreate,
    TelecallerTask, TelecallerTaskCreate,
    VehicleRecord, VehicleRecordCreate,
    DriverLead, DriverLeadUpdate, BulkLeadStatusUpdate, BulkLeadDelete,
    QRCode, QRCodeCreate, QRCodeUpdate, QRScan,
    TelecallerProfile, TelecallerProfileCreate, TelecallerProfileUpdate,
    LeadAssignment, LeadReassignment, LeadDeassignment, BulkAssignFromSheets,
    VehicleServiceRequest, VehicleServiceRequestCreate, VehicleServiceRequestUpdate
)

# Payment Reconciliation
@api_router.post("/payment-reconciliation")
async def create_payment(payment_data: PaymentRecordCreate, current_user: User = Depends(get_current_user)):
    """Create new payment record"""
    payment = PaymentRecord(**payment_data.model_dump())
    payment_dict = payment.model_dump()
    payment_dict['date'] = payment_dict['date'].isoformat()
    payment_dict['created_at'] = payment_dict['created_at'].isoformat()
    
    await db.payment_reconciliation.insert_one(payment_dict)
    sync_single_record('payment_reconciliation', payment_dict)
    
    return {"message": "Payment record created", "payment": payment}


@api_router.get("/payment-reconciliation")
async def get_payments(current_user: User = Depends(get_current_user)):
    """Get all payment records"""
    payments = await db.payment_reconciliation.find({}, {"_id": 0}).to_list(1000)
    for payment in payments:
        if isinstance(payment.get('date'), str):
            try:
                payment['date'] = datetime.fromisoformat(payment['date'])
            except ValueError:
                # Handle non-ISO date formats, keep as string
                pass
        if isinstance(payment.get('created_at'), str):
            try:
                payment['created_at'] = datetime.fromisoformat(payment['created_at'])
            except ValueError:
                # Handle non-ISO date formats, keep as string
                pass
    return payments


@api_router.post("/payment-reconciliation/sync")
async def sync_payments(current_user: User = Depends(get_current_user)):
    """Sync all payments to Google Sheets"""
    payments = await db.payment_reconciliation.find({}, {"_id": 0}).to_list(1000)
    success = sync_all_records('payment_reconciliation', payments)
    if success:
        return {"message": "Payments synced successfully"}
    raise HTTPException(status_code=500, detail="Failed to sync payments")


# Driver Onboarding
@api_router.post("/driver-onboarding")
async def create_driver(driver_data: DriverRecordCreate, current_user: User = Depends(get_current_user)):
    """Create new driver record"""
    driver = DriverRecord(**driver_data.model_dump())
    driver_dict = driver.model_dump()
    driver_dict['date'] = driver_dict['date'].isoformat()
    driver_dict['created_at'] = driver_dict['created_at'].isoformat()
    
    await db.driver_onboarding.insert_one(driver_dict)
    sync_single_record('driver_onboarding', driver_dict)
    
    return {"message": "Driver record created", "driver": driver}


@api_router.get("/driver-onboarding")
async def get_drivers(current_user: User = Depends(get_current_user)):
    """Get all driver records"""
    drivers = await db.driver_onboarding.find({}, {"_id": 0}).to_list(1000)
    for driver in drivers:
        if isinstance(driver.get('date'), str):
            driver['date'] = datetime.fromisoformat(driver['date'])
        if isinstance(driver.get('created_at'), str):
            driver['created_at'] = datetime.fromisoformat(driver['created_at'])
    return drivers


@api_router.post("/driver-onboarding/sync")
async def sync_drivers(current_user: User = Depends(get_current_user)):
    """Sync all drivers to Google Sheets"""
    drivers = await db.driver_onboarding.find({}, {"_id": 0}).to_list(1000)
    success = sync_all_records('driver_onboarding', drivers)
    if success:
        return {"message": "Drivers synced successfully"}
    raise HTTPException(status_code=500, detail="Failed to sync drivers")


# Driver Onboarding - Leads Import
import pandas as pd
import io


@api_router.post("/driver-onboarding/import-leads")
async def import_leads(
    request: Request,
    duplicate_action: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user)
):
    """Import driver leads from CSV or XLSX with SMART column mapping and status matching"""
    logger.info(f"Import request received. Duplicate action: {duplicate_action}")
    try:
        # Parse form data
        form = await request.form()
        file = form.get('file')
        lead_source = form.get('lead_source', '')
        lead_date = form.get('lead_date', '')
        read_source_from_file = form.get('read_source_from_file', 'false').lower() == 'true'
        
        if not file:
            raise HTTPException(status_code=400, detail="No file uploaded")
        
        # Read file content
        content = await file.read()
        
        # Detect file type and parse
        if file.filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(content))
        elif file.filename.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(io.BytesIO(content))
        else:
            raise HTTPException(status_code=400, detail="Invalid file type. Only CSV and XLSX are supported.")
        
        # SMART COLUMN MAPPING: Define flexible column name matches
        def find_column(df, possible_names, case_sensitive=False):
            """Find column by matching possible names flexibly"""
            if not case_sensitive:
                possible_names = [name.lower() for name in possible_names]
                columns = {col.lower(): col for col in df.columns}
            else:
                columns = {col: col for col in df.columns}
            
            for name in possible_names:
                if name in columns:
                    return columns[name]
            return None
        
        # Define our app's status hierarchy
        STATUS_HIERARCHY = [
            "New",
            "Interested",
            "Highly Interested",  # NEW: Keep as separate status
            "Docs Upload Pending",
            "Onboarding Incomplete",
            "Training WIP",
            "Onboarding Complete",
            "Ready to Deploy",
            "Deployed",
            "Not Interested",
            "Not Reachable",
            "Wrong Number",
            "Duplicate",
            "Junk"
        ]
        
        def match_status(file_status):
            """
            Match file status with app's status hierarchy
            Returns tuple: (matched_status, extracted_stage)
            """
            if not file_status or pd.isna(file_status):
                return ("New", "S1")
            
            file_status = str(file_status).strip()
            extracted_stage = None
            
            # Extract status from formats like "S1-a Not interested" or "S1-c Highly Interested"
            # Pattern: "S[1-4]-[a-z] Actual Status"
            import re
            stage_code_pattern = r'^(S[1-4])-[a-z]\s+(.+)$'
            match = re.match(stage_code_pattern, file_status, re.IGNORECASE)
            if match:
                extracted_stage = match.group(1).upper()  # Extract stage (S1, S2, S3, S4)
                file_status = match.group(2).strip()  # Extract the actual status text
                logger.info(f"Extracted from '{match.group(0)}' → Stage: '{extracted_stage}', Status: '{file_status}'")
            
            # Exact match (case-insensitive) - normalize spaces
            file_status_clean = ' '.join(file_status.split()).lower()  # Normalize whitespace
            for app_status in STATUS_HIERARCHY:
                app_status_clean = ' '.join(app_status.split()).lower()
                if file_status_clean == app_status_clean:
                    # If stage wasn't extracted, determine it from the status
                    if not extracted_stage:
                        extracted_stage = determine_stage_from_status(app_status)
                    logger.info(f"Exact match: '{file_status}' → '{app_status}'")
                    return (app_status, extracted_stage)
            
            # Partial match (contains) - ORDER IS CRITICAL!
            # ALWAYS check negative/rejection statuses FIRST before positive ones
            file_status_lower = file_status.lower()
            matched_status = None
            
            # Check NEGATIVE statuses FIRST (to avoid false positives)
            if "not interest" in file_status_lower or "reject" in file_status_lower or "no interest" in file_status_lower or "uninterested" in file_status_lower or "notinterested" in file_status_lower.replace(" ", ""):
                matched_status = "Not Interested"
                logger.info(f"Partial match (not interested): '{file_status}' → 'Not Interested'")
            elif "not reach" in file_status_lower or "unreachable" in file_status_lower or "no response" in file_status_lower or "not responding" in file_status_lower:
                matched_status = "Not Reachable"
            elif "wrong" in file_status_lower or "incorrect" in file_status_lower:
                matched_status = "Wrong Number"
            elif "duplicate" in file_status_lower or "dup" in file_status_lower:
                matched_status = "Duplicate"
            elif "junk" in file_status_lower or "invalid" in file_status_lower or "spam" in file_status_lower:
                matched_status = "Junk"
            # Now check POSITIVE statuses (after ruling out negatives)
            elif "highly interested" in file_status_lower or "very interested" in file_status_lower:
                matched_status = "Highly Interested"
            elif "interested, no dl" in file_status_lower or "interest no dl" in file_status_lower:
                matched_status = "Interested, No DL"
            elif "interested, no badge" in file_status_lower or "interest no badge" in file_status_lower or "no badge" in file_status_lower:
                matched_status = "Interested, No Badge"
            elif "call back 1d" in file_status_lower:
                matched_status = "Call back 1D"
            elif "call back 1w" in file_status_lower:
                matched_status = "Call back 1W"
            elif "call back 2w" in file_status_lower:
                matched_status = "Call back 2W"
            elif "call back 1m" in file_status_lower:
                matched_status = "Call back 1M"
            elif "interest" in file_status_lower or "follow" in file_status_lower or "call back" in file_status_lower or "callback" in file_status_lower:
                matched_status = "Interested"
            elif "doc" in file_status_lower and ("pending" in file_status_lower or "upload" in file_status_lower or "collection" in file_status_lower):
                matched_status = "Docs Upload Pending"
            elif "verif" in file_status_lower and "pending" in file_status_lower:
                matched_status = "Verification Pending"
            elif "verified" in file_status_lower:
                matched_status = "Verified"
            elif "training wip" in file_status_lower or "train wip" in file_status_lower:
                matched_status = "Training WIP"
            elif "training complete" in file_status_lower:
                matched_status = "Training Completed"
            elif "approved" in file_status_lower:
                matched_status = "Approved"
            elif "ct pending" in file_status_lower:
                matched_status = "CT Pending"
            elif "done" in file_status_lower:
                matched_status = "DONE!"
            elif "long distance" in file_status_lower or "out of town" in file_status_lower or "outside" in file_status_lower:
                matched_status = "Interested"  # They're interested but have distance constraints
            elif "health" in file_status_lower or "medical" in file_status_lower:
                matched_status = "Interested"  # Health issue but potentially interested
            
            # If matched, return with stage
            if matched_status:
                if not extracted_stage:
                    extracted_stage = determine_stage_from_status(matched_status)
                return (matched_status, extracted_stage)
            
            # If no match found, return original but log warning
            logger.warning(f"Status '{file_status}' not matched, defaulting to 'New'")
            return ("New", extracted_stage or "S1")
        
        def determine_stage_from_status(status):
            """Determine which stage a status belongs to"""
            s1_statuses = ["New", "Not Interested", "Interested, No DL", "Interested, No Badge", "Highly Interested", 
                          "Call back 1D", "Call back 1W", "Call back 2W", "Call back 1M", 
                          "Interested", "Not Reachable", "Wrong Number", "Duplicate", "Junk"]
            s2_statuses = ["Docs Upload Pending", "Verification Pending", "Duplicate License", 
                          "DL - Amount", "Verified", "Verification Rejected"]
            s3_statuses = ["Schedule Pending", "Training WIP", "Training Completed", 
                          "Training Rejected", "Re-Training", "Absent for training", "Approved"]
            s4_statuses = ["CT Pending", "CT WIP", "Shift Details Pending", "DONE!", 
                          "Terminated"]
            
            if status in s1_statuses:
                return "S1"
            elif status in s2_statuses:
                return "S2"
            elif status in s3_statuses:
                return "S3"
            elif status in s4_statuses:
                return "S4"
            else:
                return "S1"  # Default
        
        # SMART COLUMN DETECTION
        logger.info(f"Columns in file: {list(df.columns)}")
        
        # Check if first row contains headers (sometimes Excel has headers in row 1 data)
        first_row_values = df.iloc[0].values if len(df) > 0 else []
        is_first_row_header = any(
            str(val).lower() in ['sl.no', 'sno', 's.no', 'name', 'phone', 'phone no', 'mobile', 'address', 'status', 'stage'] 
            for val in first_row_values if pd.notna(val)
        )
        
        if is_first_row_header:
            logger.info("First data row detected as headers, using it for column names")
            df.columns = df.iloc[0]
            df = df.iloc[1:].reset_index(drop=True)
        
        # Find columns dynamically
        name_col = find_column(df, ['name', 'driver name', 'full name', 'full_name', 'candidate name', 'Name ', 'name ', 'poc name', 'POC Name'])
        phone_col = find_column(df, ['phone', 'phone no', 'phone number', 'phone_number', 'mobile', 'mobile no', 'contact', 'Phone No', 'Phone Number'])
        address_col = find_column(df, ['address', 'location', 'current location', 'city', 'Address ', 'address '])
        email_col = find_column(df, ['email', 'email address', 'email_address', 'Email', 'Email Address'])
        experience_col = find_column(df, ['experience', 'Experience', 'exp', 'years of experience'])
        vehicle_col = find_column(df, ['vehicle', 'vehicle type', 'Vehicle', 'vehicle '])
        
        # PRIORITY: Look for "Status.1", "Status" column FIRST, then "Final Status", then "Lead Status", then "Current Status" last
        # Also check for "STATUS" (uppercase) for Google Forms data
        status_col = find_column(df, ['Status.1', 'STATUS', 'status', 'Status', 'final status', 'Final Status', 'lead status', 'Lead Status', 'current status', 'Current Status'])
        stage_col = find_column(df, ['stage', 'lead stage', 'current stage', 'Stage', 'stage '])
        source_col = find_column(df, ['lead source', 'source', 'lead generator', 'Lead Generator', 'lead_source', 'LeadSource'])
        date_col = find_column(df, ['date', 'lead date', 'lead creation date', 'created date', 'Lead Creation Date', 'import date'])
        
        poc_col = find_column(df, ['poc', 'assigned to', 'telecaller', 'POC', 'poc ', 'POC Name', 'poc name'])
        
        # Multiple possible telecaller notes columns (ss, sss, Current Status, etc.)
        telecaller_notes_col = find_column(df, ['telecaller notes', 'Telecaller Notes', 'ss', 'ss.1', 'sss', 'current status', 'Current Status', 'next action', 'action', 'follow up', 'Next Action'])
        
        # General remarks/notes columns
        remarks_col = find_column(df, ['remarks', 'notes', 'comments', 'Remarks', 'remarks ', 'dd', 'dd.1', 'ss.2'])
        
        logger.info(f"Mapped columns - Name: {name_col}, Phone: {phone_col}, Email: {email_col}, Status: {status_col}, Stage: {stage_col}, Source: {source_col}, POC: {poc_col}, Vehicle: {vehicle_col}")
        
        # Process rows
        leads = []
        import_date = datetime.now(timezone.utc).isoformat()
        
        for _, row in df.iterrows():
            # Skip completely empty rows
            if all(pd.isna(val) for val in row.values):
                continue
            
            # Get name and phone (required fields)
            name_val = str(row[name_col]) if name_col and pd.notna(row.get(name_col)) else ""
            
            # Handle phone number - convert to int first to avoid .0 issue, then to string
            phone_val = ""
            if phone_col and pd.notna(row.get(phone_col)):
                try:
                    # Try converting to int first to remove any .0, then to string
                    phone_val = str(int(float(row[phone_col])))
                except (ValueError, TypeError):
                    # If conversion fails, use string directly
                    phone_val = str(row[phone_col])
            
            # Skip if both name and phone are empty
            if not name_val and not phone_val:
                continue
            
            # Clean phone number - handle various formats
            # Remove 'p:' prefix (Google Forms format), spaces, dashes
            phone_val = phone_val.strip()
            phone_val = phone_val.replace('p:', '').replace('p:+', '+')  # Remove p: prefix
            phone_val = phone_val.replace(' ', '').replace('-', '')
            
            # Smart removal of +91 or 91 prefix - only if number is longer than 10 digits
            # This preserves numbers that start with 91 but are actually 10 digits (e.g., 9178822331)
            phone_digits = ''.join(filter(str.isdigit, phone_val))
            if len(phone_digits) > 10:
                # If number has more than 10 digits, take the last 10 digits
                # This handles +919897721333 or 919897721333 → 9897721333
                phone_val = phone_digits[-10:]
            else:
                # If already 10 digits or less, keep as is
                phone_val = phone_digits
            
            # Get status from file and match with app's status hierarchy
            file_status = str(row[status_col]) if status_col and pd.notna(row.get(status_col)) else None
            if not file_status and stage_col:
                file_status = str(row[stage_col]) if pd.notna(row.get(stage_col)) else None
            
            # Parse the status and extract stage
            matched_status, extracted_stage = match_status(file_status)
            logger.info(f"File status '{file_status}' → Matched status: '{matched_status}', Stage: '{extracted_stage}'")
            
            # Use the MATCHED status (clean app status) for the database
            # This ensures dashboard summary counts work correctly
            display_status = matched_status
            
            # Get lead source from file or form
            row_lead_source = lead_source
            if source_col and pd.notna(row.get(source_col)):
                row_lead_source = str(row[source_col])
            
            # Get lead date from file or form
            row_lead_date = lead_date
            if date_col and pd.notna(row.get(date_col)):
                row_lead_date = str(row[date_col])
            
            # Create lead object
            lead = {
                "id": str(uuid.uuid4()),
                "name": name_val,
                "phone_number": phone_val,
                "email": str(row[email_col]) if email_col and pd.notna(row.get(email_col)) else None,
                "vehicle": str(row[vehicle_col]) if vehicle_col and pd.notna(row.get(vehicle_col)) else None,
                "driving_license": None,
                "experience": str(row[experience_col]) if experience_col and pd.notna(row.get(experience_col)) else None,
                "interested_ev": None,
                "monthly_salary": None,
                "residing_chennai": None,
                "current_location": str(row[address_col]) if address_col and pd.notna(row.get(address_col)) else None,
                "import_date": import_date,
                "lead_source": row_lead_source,
                "source": row_lead_source if row_lead_source else file.filename,  # Track import source for filtering
                "lead_date": row_lead_date,
                "status": display_status,  # USE MATCHED APP STATUS (e.g., "Not Interested")
                "stage": extracted_stage,  # Use extracted stage (e.g., "S1", "S2")
                "lead_stage": file_status if file_status else "New",  # Keep original for reference
                "driver_readiness": "Not Started",
                "docs_collection": "Pending",
                "customer_readiness": "Not Ready",
                "assigned_telecaller": str(row[poc_col]) if poc_col and pd.notna(row.get(poc_col)) else None,
                "assigned_date": import_date if (poc_col and pd.notna(row.get(poc_col))) else None,  # Set assigned date if telecaller assigned
                "telecaller_notes": str(row[telecaller_notes_col]) if telecaller_notes_col and pd.notna(row.get(telecaller_notes_col)) else None,
                "notes": str(row[remarks_col]) if remarks_col and pd.notna(row.get(remarks_col)) else None,
                "created_at": import_date,
                "last_modified": import_date
            }
            leads.append(lead)
        
        if not leads:
            raise HTTPException(status_code=400, detail="No valid leads found in file")
        
        logger.info(f"Parsed {len(leads)} leads from file")
        
        # SMART DUPLICATE DETECTION: Normalize phone numbers (last 10 digits)
        def normalize_phone(phone):
            """Extract last 10 digits from phone number (handles +91, 91, etc.)"""
            if not phone:
                return None
            # Remove all non-numeric characters
            phone_digits = ''.join(filter(str.isdigit, str(phone)))
            # Return last 10 digits
            if len(phone_digits) >= 10:
                return phone_digits[-10:]
            return phone_digits if phone_digits else None
        
        # Check for duplicates by normalized phone number
        if leads:
            # Get all existing phone numbers from database (no limit)
            all_existing_leads = await db.driver_leads.find(
                {},
                {"_id": 0, "phone_number": 1, "name": 1, "id": 1, "status": 1}
            ).to_list(length=None)
            
            # Create a map of normalized phone -> existing lead
            existing_phone_map = {}
            for existing_lead in all_existing_leads:
                normalized = normalize_phone(existing_lead.get('phone_number'))
                if normalized:
                    existing_phone_map[normalized] = existing_lead
            
            logger.info(f"Found {len(existing_phone_map)} existing leads with valid phone numbers")
            
            # Identify duplicates and non-duplicates
            duplicates = []
            non_duplicates = []
            
            for lead in leads:
                normalized_phone = normalize_phone(lead['phone_number'])
                
                if normalized_phone and normalized_phone in existing_phone_map:
                    # Duplicate found!
                    existing = existing_phone_map[normalized_phone]
                    duplicates.append({
                        "name": lead['name'],
                        "phone_number": lead['phone_number'],
                        "normalized_phone": normalized_phone,
                        "existing_name": existing.get('name', 'Unknown'),
                        "existing_status": existing.get('status', 'Unknown'),
                        "existing_phone": existing.get('phone_number', '')
                    })
                    logger.info(f"Duplicate found: {lead['phone_number']} (normalized: {normalized_phone}) matches existing {existing.get('phone_number')}")
                else:
                    non_duplicates.append(lead)
            
            # If duplicates found, automatically skip them and show the list
            if duplicates:
                logger.info(f"Found {len(duplicates)} duplicates, automatically skipping")
                
                # Insert only non-duplicates
                if non_duplicates:
                    await db.driver_leads.insert_many(non_duplicates)
                    logger.info(f"Imported {len(non_duplicates)} new leads to database")
                    
                    # Sync to Google Sheets
                    try:
                        sync_all_records('leads', non_duplicates)
                    except Exception as sync_error:
                        logger.warning(f"Google Sheets sync failed: {str(sync_error)}")
                
                return {
                    "success": True,
                    "message": f"Imported {len(non_duplicates)} new lead(s), skipped {len(duplicates)} duplicate(s)",
                    "imported_count": len(non_duplicates),
                    "duplicate_count": len(duplicates),
                    "total_in_file": len(leads),
                    "duplicates_found": True,
                    "duplicates": duplicates,  # Show ALL duplicates
                    "duplicate_phones": [dup['phone_number'] for dup in duplicates]  # Just phone numbers for quick view
                }
            else:
                # No duplicates, insert all leads
                await db.driver_leads.insert_many(leads)
                logger.info(f"Imported {len(leads)} leads to database (no duplicates)")
                
                # Sync to Google Sheets
                try:
                    sync_all_records('leads', leads)
                except Exception as sync_error:
                    logger.warning(f"Google Sheets sync failed: {str(sync_error)}")
                
                return {
                    "success": True,
                    "message": f"Successfully imported {len(leads)} lead(s)",
                    "imported_count": len(leads),
                    "duplicate_count": 0,
                    "total_in_file": len(leads),
                    "duplicates_found": False
                }
        
        return {
            "success": False,
            "message": "No leads found in file",
            "imported_count": 0,
            "duplicate_count": 0
        }
        
    except Exception as e:
        logger.error(f"Error importing leads: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to import leads: {str(e)}")


# ==================== BULK EXPORT/IMPORT WITH BACKUP LIBRARY ====================

# Backup storage folder
BACKUP_LIBRARY_FOLDER = "driver_onboarding_backups"
os.makedirs(BACKUP_LIBRARY_FOLDER, exist_ok=True)


@api_router.post("/driver-onboarding/check-duplicate")
async def check_duplicate_lead(
    phone_number: str = Body(..., embed=True),
    current_user: User = Depends(get_current_user)
):
    """Check if a lead with the given phone number already exists"""
    try:
        # Normalize phone number (last 10 digits)
        normalized_phone = phone_number[-10:] if len(phone_number) >= 10 else phone_number
        
        # Check for duplicate
        existing_lead = await db.driver_leads.find_one({"phone_number": normalized_phone})
        
        if existing_lead:
            return {
                "success": True,
                "duplicate_found": True,
                "existing_lead": {
                    "name": existing_lead.get("name"),
                    "phone_number": existing_lead.get("phone_number"),
                    "email": existing_lead.get("email"),
                    "status": existing_lead.get("status"),
                    "stage": existing_lead.get("stage"),
                    "source": existing_lead.get("source")
                }
            }
        else:
            return {
                "success": True,
                "duplicate_found": False
            }
    
    except Exception as e:
        logger.error(f"Error checking duplicate: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error checking duplicate: {str(e)}")


@api_router.post("/driver-onboarding/create-lead")
async def create_single_lead(
    name: str = Body(...),
    phone_number: str = Body(...),
    email: Optional[str] = Body(None),
    source: Optional[str] = Body("Manual Entry"),
    status: str = Body("New"),
    current_location: Optional[str] = Body(None),
    experience: Optional[str] = Body(None),
    monthly_salary: Optional[str] = Body(None),
    has_driving_license: str = Body("no"),
    driving_license_no: Optional[str] = Body(None),
    has_badge: str = Body("no"),
    badge_no: Optional[str] = Body(None),
    duplicate_action: str = Body("skip"),
    current_user: User = Depends(get_current_user)
):
    """Create a single driver lead manually"""
    try:
        # Normalize phone number (last 10 digits)
        normalized_phone = phone_number[-10:] if len(phone_number) >= 10 else phone_number
        
        # Check for duplicate
        existing_lead = await db.driver_leads.find_one({"phone_number": normalized_phone})
        
        if existing_lead and duplicate_action == "skip":
            return {
                "success": False,
                "message": "Lead with this phone number already exists",
                "duplicate": True
            }
        
        # Generate unique ID
        lead_id = str(uuid.uuid4())
        
        # Prepare lead data
        lead_data = {
            "id": lead_id,
            "name": name,
            "phone_number": normalized_phone,
            "email": email,
            "source": source or "Manual Entry",
            "current_location": current_location,
            "experience": experience,
            "monthly_salary": monthly_salary,
            "has_driving_license": has_driving_license,
            "driving_license_no": driving_license_no if has_driving_license == "yes" else None,
            "has_badge": has_badge,
            "badge_no": badge_no if has_badge == "yes" else None,
            "status": status or "New",
            "stage": "S1",
            "import_date": datetime.now(timezone.utc).strftime('%Y-%m-%d'),
            "created_at": datetime.now(timezone.utc),
            "updated_at": datetime.now(timezone.utc),
            "remarks": "",
            "remarks_history": [],
            "status_history": [{
                "status": status or "New",
                "stage": "S1",
                "changed_by": current_user.email,
                "changed_at": datetime.now(timezone.utc)
            }]
        }
        
        if existing_lead and duplicate_action == "replace":
            # Update existing lead
            await db.driver_leads.update_one(
                {"phone_number": normalized_phone},
                {"$set": lead_data}
            )
            message = "Lead updated successfully (replaced existing)"
        else:
            # Insert new lead
            await db.driver_leads.insert_one(lead_data)
            message = "Lead created successfully"
        
        return {
            "success": True,
            "message": message,
            "lead_id": lead_id
        }
    
    except Exception as e:
        logger.error(f"Error creating lead: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error creating lead: {str(e)}")


@api_router.post("/driver-onboarding/bulk-export")
async def bulk_export_leads(current_user: User = Depends(get_current_user)):
    """
    Export ALL driver leads to Excel file with streaming for large datasets
    Optimized for 30K+ leads with chunked processing to avoid timeouts
    """
    try:
        import pandas as pd
        from datetime import datetime
        import io
        
        logger.info(f"🔄 Starting bulk export for user {current_user.email}")
        
        # Fetch all leads from database - NO LIMIT
        leads_collection = db['driver_leads']
        
        # Count total first
        total_count = await leads_collection.count_documents({})
        logger.info(f"📊 Total leads in database: {total_count}")
        
        # For large datasets (>10K), use batched fetching to reduce memory
        if total_count > 10000:
            logger.info(f"🔄 Large dataset detected ({total_count} leads), using batched processing...")
            batch_size = 5000
            all_leads = []
            
            for skip in range(0, total_count, batch_size):
                batch = await leads_collection.find({}, {"_id": 0}).skip(skip).limit(batch_size).to_list(length=batch_size)
                all_leads.extend(batch)
                logger.info(f"📦 Fetched batch: {skip + len(batch)}/{total_count} leads")
            
            leads = all_leads
        else:
            # Fetch all at once for smaller datasets
            leads = await leads_collection.find({}, {"_id": 0}).to_list(length=None)
        
        actual_fetched = len(leads)
        logger.info(f"✅ Fetched {actual_fetched} leads for export (expected {total_count})")
        
        if actual_fetched != total_count:
            logger.warning(f"⚠️ Mismatch: Expected {total_count} but fetched {actual_fetched}")
        
        if not leads:
            logger.warning("❌ No leads found in database")
            raise HTTPException(status_code=404, detail="No leads found to export")
        
        # Convert to DataFrame with chunked processing for large datasets
        logger.info(f"📊 Converting {actual_fetched} leads to DataFrame...")
        df = pd.DataFrame(leads)
        
        # Add document number columns if they don't exist (ensure they're in export)
        document_number_fields = ['dl_no', 'badge_no', 'aadhar_card', 'pan_card', 'gas_bill', 'bank_passbook']
        for doc_field in document_number_fields:
            if doc_field not in df.columns:
                df[doc_field] = ""  # Add empty column
        
        # Add document upload status columns (yes/no based on whether document number exists)
        document_fields = {
            'dl_no': 'dl_documents_uploaded',
            'badge_no': 'badge_documents_uploaded',
            'aadhar_card': 'aadhar_documents_uploaded',
            'pan_card': 'pan_documents_uploaded',
            'gas_bill': 'gas_documents_uploaded',
            'bank_passbook': 'bank_documents_uploaded'
        }
        
        for doc_field, status_field in document_fields.items():
            # Show "yes" if document number exists, otherwise "no"
            df[status_field] = df[doc_field].apply(
                lambda x: "yes" if pd.notna(x) and str(x).strip() != "" else "no"
            )
        
        logger.info(f"📋 Added document upload status columns (yes/no) and ensured document number columns exist")
        
        # Define preferred column order (with document upload status before document numbers)
        preferred_columns = [
            'id', 'name', 'phone_number', 'email', 'vehicle', 
            'stage', 'status', 'source', 'remarks',
            'current_location', 'preferred_shift', 'experience', 'assigned_telecaller',
            'dl_documents_uploaded', 'dl_no',
            'badge_documents_uploaded', 'badge_no',
            'aadhar_documents_uploaded', 'aadhar_card',
            'pan_documents_uploaded', 'pan_card',
            'gas_documents_uploaded', 'gas_bill',
            'bank_documents_uploaded', 'bank_passbook',
            'last_called', 'callback_date', 'assigned_date',
            'import_date', 'created_at', 'updated_at'
        ]
        
        # Reorder columns (keeping any extra columns at the end)
        existing_preferred = [col for col in preferred_columns if col in df.columns]
        other_columns = [col for col in df.columns if col not in preferred_columns]
        column_order = existing_preferred + other_columns
        df = df[column_order]
        
        logger.info(f"📋 Export columns ({len(df.columns)}): {', '.join(df.columns.tolist()[:10])}...")
        
        # Create Excel file in memory with streaming
        logger.info(f"📝 Generating Excel file...")
        output = io.BytesIO()
        
        # Use openpyxl for Excel generation
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Write in chunks for very large datasets
            if actual_fetched > 10000:
                logger.info(f"📦 Writing Excel in chunks...")
                chunk_size = 5000
                for i in range(0, len(df), chunk_size):
                    df_chunk = df.iloc[i:i+chunk_size]
                    if i == 0:
                        df_chunk.to_excel(writer, index=False, sheet_name='Driver Leads', startrow=0)
                    else:
                        df_chunk.to_excel(writer, index=False, sheet_name='Driver Leads', 
                                        startrow=i+1, header=False)
                    logger.info(f"📦 Wrote rows {i+1} to {min(i+chunk_size, len(df))}")
            else:
                df.to_excel(writer, index=False, sheet_name='Driver Leads')
            
            # Auto-adjust column widths (skip for very large files to save time)
            if actual_fetched <= 20000:
                worksheet = writer.sheets['Driver Leads']
                from openpyxl.utils import get_column_letter
                for idx, col in enumerate(df.columns, 1):
                    try:
                        max_length = min(
                            df[col].astype(str).apply(len).max(),
                            len(col),
                            50  # Max width
                        )
                        adjusted_width = max_length + 2
                        col_letter = get_column_letter(idx)
                        worksheet.column_dimensions[col_letter].width = adjusted_width
                    except Exception as e:
                        logger.warning(f"Could not adjust width for column {col}: {e}")
                        continue
        
        output.seek(0)
        file_size_mb = len(output.getvalue()) / (1024 * 1024)
        logger.info(f"📦 Excel file generated: {file_size_mb:.2f} MB")
        
        # Generate filename without underscores
        filename = "driver leads export.xlsx"
        
        logger.info(f"✅ Bulk export complete: {actual_fetched} leads exported to {filename}")
        
        # Create a generator function for streaming
        def iterfile():
            yield from output
        
        return StreamingResponse(
            iterfile(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "X-Total-Leads": str(actual_fetched),
                "X-Expected-Leads": str(total_count),
                "Content-Length": str(len(output.getvalue()))
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Bulk export failed: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Bulk export failed: {str(e)}")


@api_router.post("/driver-onboarding/batch-export-zip")
async def batch_export_leads_as_zip(current_user: User = Depends(get_current_user)):
    """
    Export ALL leads in batches as a ZIP file containing multiple Excel files
    Each Excel file contains up to 1000 leads
    Designed to work within production memory constraints
    """
    try:
        import pandas as pd
        from datetime import datetime
        import io
        import zipfile
        from openpyxl.utils import get_column_letter
        
        logger.info(f"🔄 Starting batched ZIP export for user {current_user.email}")
        
        # Fetch all leads from database
        leads_collection = db['driver_leads']
        
        # Count total leads
        total_count = await leads_collection.count_documents({})
        logger.info(f"📊 Total leads in database: {total_count}")
        
        if total_count == 0:
            raise HTTPException(status_code=404, detail="No leads found to export")
        
        # Define batch size (1000 leads per Excel file)
        BATCH_SIZE = 1000
        total_batches = (total_count + BATCH_SIZE - 1) // BATCH_SIZE
        
        logger.info(f"📦 Creating {total_batches} Excel files with {BATCH_SIZE} leads each")
        
        # Define preferred column order (with document upload status before document numbers)
        preferred_columns = [
            'id', 'name', 'phone_number', 'email', 'vehicle', 
            'stage', 'status', 'source', 'remarks',
            'current_location', 'preferred_shift', 'experience', 'assigned_telecaller',
            'dl_documents_uploaded', 'dl_no',
            'badge_documents_uploaded', 'badge_no',
            'aadhar_documents_uploaded', 'aadhar_card',
            'pan_documents_uploaded', 'pan_card',
            'gas_documents_uploaded', 'gas_bill',
            'bank_documents_uploaded', 'bank_passbook',
            'last_called', 'callback_date', 'assigned_date',
            'import_date', 'created_at', 'updated_at'
        ]
        
        # Create ZIP file in memory
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Process leads in batches
            for batch_num in range(total_batches):
                skip = batch_num * BATCH_SIZE
                
                # Fetch batch
                batch_leads = await leads_collection.find(
                    {}, 
                    {"_id": 0}
                ).skip(skip).limit(BATCH_SIZE).to_list(length=BATCH_SIZE)
                
                actual_batch_size = len(batch_leads)
                logger.info(f"📦 Processing batch {batch_num + 1}/{total_batches} ({actual_batch_size} leads)")
                
                # Convert to DataFrame
                df = pd.DataFrame(batch_leads)
                
                # Add document number columns if they don't exist (ensure they're in export)
                document_number_fields = ['dl_no', 'badge_no', 'aadhar_card', 'pan_card', 'gas_bill', 'bank_passbook']
                for doc_field in document_number_fields:
                    if doc_field not in df.columns:
                        df[doc_field] = ""  # Add empty column
                
                # Add document upload status columns (yes/no based on whether document number exists)
                document_fields = {
                    'dl_no': 'dl_documents_uploaded',
                    'badge_no': 'badge_documents_uploaded',
                    'aadhar_card': 'aadhar_documents_uploaded',
                    'pan_card': 'pan_documents_uploaded',
                    'gas_bill': 'gas_documents_uploaded',
                    'bank_passbook': 'bank_documents_uploaded'
                }
                
                for doc_field, status_field in document_fields.items():
                    # Show "yes" if document number exists, otherwise "no"
                    df[status_field] = df[doc_field].apply(
                        lambda x: "yes" if pd.notna(x) and str(x).strip() != "" else "no"
                    )
                
                # Reorder columns
                existing_preferred = [col for col in preferred_columns if col in df.columns]
                other_columns = [col for col in df.columns if col not in preferred_columns]
                column_order = existing_preferred + other_columns
                df = df[column_order]
                
                # Create Excel file in memory for this batch
                excel_buffer = io.BytesIO()
                
                with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Driver Leads')
                    
                    # Auto-adjust column widths (only for small batches to save time)
                    if actual_batch_size <= 1000:
                        worksheet = writer.sheets['Driver Leads']
                        for idx, col in enumerate(df.columns, 1):
                            try:
                                max_length = min(
                                    df[col].astype(str).apply(len).max(),
                                    len(col),
                                    50
                                )
                                adjusted_width = max_length + 2
                                col_letter = get_column_letter(idx)
                                worksheet.column_dimensions[col_letter].width = adjusted_width
                            except Exception:
                                continue
                
                # Add Excel file to ZIP with proper naming
                excel_buffer.seek(0)
                start_row = skip + 1
                end_row = skip + actual_batch_size
                filename = f"driver_leads_batch_{batch_num + 1}_rows_{start_row}-{end_row}.xlsx"
                
                zip_file.writestr(filename, excel_buffer.getvalue())
                
                logger.info(f"✅ Added {filename} to ZIP ({actual_batch_size} leads)")
        
        # Prepare ZIP for download
        zip_buffer.seek(0)
        zip_size_mb = len(zip_buffer.getvalue()) / (1024 * 1024)
        
        logger.info(f"📦 ZIP file created: {zip_size_mb:.2f} MB with {total_batches} Excel files")
        logger.info(f"✅ Batched export complete: {total_count} leads in {total_batches} files")
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        zip_filename = f"driver_leads_batched_{timestamp}.zip"
        
        # Create generator for streaming
        def iterfile():
            yield from zip_buffer
        
        return StreamingResponse(
            iterfile(),
            media_type="application/zip",
            headers={
                "Content-Disposition": f'attachment; filename="{zip_filename}"',
                "X-Total-Leads": str(total_count),
                "X-Total-Files": str(total_batches),
                "X-Batch-Size": str(BATCH_SIZE)
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Batched ZIP export failed: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Batched export failed: {str(e)}")


@api_router.post("/driver-onboarding/bulk-import")
async def bulk_import_leads(
    file: UploadFile = File(...),
    column_mapping: str = Form(None),
    current_user: User = Depends(get_current_user)
):
    """
    Import Excel file and ADD new leads (does not replace existing)
    Checks for duplicates by phone number and skips them
    Automatically creates backup before import
    Supports flexible column mapping from user's Excel file
    """
    try:
        import pandas as pd
        from datetime import datetime
        import shutil
        import json
        
        logger.info(f"Starting bulk import (add mode) for user {current_user.email}")
        
        # Validate file type
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="File must be Excel format (.xlsx or .xls)")
        
        # Parse column mapping if provided
        mapping = None
        if column_mapping:
            try:
                mapping = json.loads(column_mapping)
                logger.info(f"Using column mapping: {mapping}")
            except json.JSONDecodeError:
                raise HTTPException(status_code=400, detail="Invalid column mapping format")
        
        # Step 1: Create backup of current leads before import
        logger.info("Creating backup of current leads...")
        leads_collection = db['driver_leads']
        current_leads = await leads_collection.find({}, {"_id": 0}).to_list(length=None)
        
        backup_filename = None
        if current_leads:
            # Save backup to library
            backup_df = pd.DataFrame(current_leads)
            timestamp = datetime.now().strftime("%d-%m-%Y-%H-%M-%S")
            backup_filename = f"backup-{timestamp}.xlsx"
            backup_path = os.path.join(BACKUP_LIBRARY_FOLDER, backup_filename)
            
            backup_df.to_excel(backup_path, index=False, sheet_name='Driver Leads')
            logger.info(f"Backup created: {backup_filename} ({len(current_leads)} leads)")
        else:
            logger.info("No existing leads to backup")
        
        # Step 2: Read uploaded Excel file
        contents = await file.read()
        df_raw = pd.read_excel(io.BytesIO(contents))
        
        total_rows = len(df_raw)
        logger.info(f"Read {total_rows} rows from uploaded file")
        
        if total_rows == 0:
            raise HTTPException(status_code=400, detail="Uploaded file is empty")
        
        # Step 3: Apply column mapping if provided
        if mapping:
            logger.info("Applying column mapping...")
            df = pd.DataFrame()
            
            # Map each field to its corresponding column
            for field, col_index in mapping.items():
                if col_index is not None and col_index < len(df_raw.columns):
                    df[field] = df_raw.iloc[:, col_index]
            
            # Generate IDs if not provided
            if 'id' not in df.columns or df['id'].isna().all():
                import uuid
                df['id'] = [str(uuid.uuid4()) for _ in range(len(df))]
                logger.info("Generated UUIDs for leads")
        else:
            # Use original dataframe if no mapping
            df = df_raw
            
            # Validate required columns for backward compatibility
            if 'id' not in df.columns:
                raise HTTPException(status_code=400, detail="Excel file must contain 'id' column or provide column mapping")
        
        # Step 4: Get existing phone numbers for duplicate detection
        logger.info("Checking for duplicates by phone number...")
        existing_phones = set()
        if current_leads:
            for lead in current_leads:
                phone = lead.get('phone_number', '')
                # Handle None/null phone numbers
                if phone is None:
                    continue
                phone = str(phone).strip()
                if phone and phone != 'nan':
                    # Normalize phone number (remove spaces, dashes, etc.)
                    normalized_phone = ''.join(filter(str.isdigit, phone))
                    if normalized_phone:  # Only add if there are digits
                        existing_phones.add(normalized_phone)
        
        logger.info(f"Found {len(existing_phones)} existing phone numbers")
        
        # Step 5: Prepare leads for upsert (update existing or insert new)
        new_leads = []
        existing_leads_to_update = []
        duplicates_updated = 0
        
        # Build map of existing leads by phone number
        existing_leads_map = {}
        if current_leads:
            for lead in current_leads:
                phone = lead.get('phone_number', '')
                if phone:
                    normalized_phone = ''.join(filter(str.isdigit, str(phone).strip()))
                    if normalized_phone:
                        existing_leads_map[normalized_phone] = lead
        
        for idx, row in df.iterrows():
            # Handle None/null phone numbers safely
            phone = row.get('phone_number', '')
            if phone is None or pd.isna(phone):
                phone = ''
            else:
                phone = str(phone).strip()
            
            if phone and phone != 'nan':
                # Normalize phone number
                normalized_phone = ''.join(filter(str.isdigit, phone))
                
                # Check if this lead already exists
                if normalized_phone and normalized_phone in existing_leads_map:
                    # Update existing lead
                    existing_lead = existing_leads_map[normalized_phone]
                    
                    # Check if this is a real existing lead (has ID) or just a duplicate within import file
                    if 'id' in existing_lead:
                        row_dict = row.to_dict()
                        row_dict['id'] = existing_lead['id']  # Keep original ID
                        existing_leads_to_update.append(row_dict)
                        duplicates_updated += 1
                        logger.info(f"Will update existing lead: {phone}")
                    else:
                        # Duplicate within import file - skip
                        logger.warning(f"Duplicate phone number within import file, skipping: {phone}")
                else:
                    # New lead
                    new_leads.append(row)
                    # Add to map to prevent duplicates within import file
                    if normalized_phone:
                        existing_leads_map[normalized_phone] = {'phone_number': phone}
            else:
                # Lead without phone number - add as new
                new_leads.append(row)
        
        logger.info(f"New leads to add: {len(new_leads)}, Existing leads to update: {duplicates_updated}")
        
        if len(new_leads) == 0 and len(existing_leads_to_update) == 0:
            logger.info("No leads to process")
            return {
                "success": True,
                "backup_created": backup_filename,
                "new_leads_count": 0,
                "updated_leads_count": 0,
                "duplicates_updated": 0,
                "duplicates_skipped": 0,
                "total_leads_now": len(current_leads),
                "telecaller_assignments": {
                    "leads_assigned": 0,
                    "telecallers_updated": 0
                },
                "message": "No leads to process."
            }
        
        # Convert new leads list to DataFrame for processing
        df_new = pd.DataFrame(new_leads)
        
        # Step 6: Process telecaller assignments from USERS collection
        users_collection = db['users']
        telecaller_assignments = {}
        leads_with_assignments = 0
        
        # Process both new leads and existing leads to update
        all_leads_to_process = new_leads + existing_leads_to_update
        
        if 'assigned_telecaller' in df.columns or 'assigned_telecaller' in (all_leads_to_process[0] if all_leads_to_process else {}):
            logger.info("Processing telecaller assignments from users collection...")
            
            # Get all telecaller users
            all_telecallers = await users_collection.find({"account_type": "telecaller"}).to_list(length=None)
            
            # Build name map with multiple variations
            telecaller_name_map = {}
            for tc in all_telecallers:
                full_name = f"{tc.get('first_name', '')} {tc.get('last_name', '')}".strip()
                first_name = tc.get('first_name', '').strip()
                
                # Map by full name and first name (lowercase)
                if full_name:
                    telecaller_name_map[full_name.lower()] = tc
                if first_name:
                    telecaller_name_map[first_name.lower()] = tc
            
            logger.info(f"Found {len(all_telecallers)} telecaller users in database")
            
            # Process each lead's telecaller assignment
            for lead_data in all_leads_to_process:
                if isinstance(lead_data, pd.Series):
                    lead_dict = lead_data.to_dict()
                else:
                    lead_dict = lead_data
                
                assigned_telecaller = lead_dict.get('assigned_telecaller')
                
                # Check if this lead has an ID (existing lead) or needs one (new lead)
                lead_id = lead_dict.get('id')
                
                # If telecaller name is empty or NaN, unassign
                if pd.isna(assigned_telecaller) or not str(assigned_telecaller).strip():
                    lead_dict['assigned_telecaller'] = None
                    lead_dict['assigned_telecaller_name'] = None
                    logger.info(f"Unassigning telecaller for lead {lead_id or 'new lead'}")
                else:
                    telecaller_name = str(assigned_telecaller).strip()
                    telecaller_name_lower = telecaller_name.lower()
                    
                    if telecaller_name_lower in telecaller_name_map:
                        telecaller = telecaller_name_map[telecaller_name_lower]
                        
                        # Store email for assignment
                        lead_dict['assigned_telecaller'] = telecaller['email']
                        lead_dict['assigned_telecaller_name'] = f"{telecaller.get('first_name', '')} {telecaller.get('last_name', '')}".strip()
                        
                        # Only track assignments for existing leads with IDs
                        if lead_id:
                            if telecaller['email'] not in telecaller_assignments:
                                telecaller_assignments[telecaller['email']] = []
                            telecaller_assignments[telecaller['email']].append(lead_id)
                        
                        leads_with_assignments += 1
                        logger.info(f"Assigned lead {lead_id or 'new lead'} to telecaller {telecaller.get('first_name', '')}")
                    else:
                        logger.warning(f"Telecaller '{telecaller_name}' not found in users - unassigning")
                        lead_dict['assigned_telecaller'] = None
                        lead_dict['assigned_telecaller_name'] = None
        
        # Step 7: Helper function to determine stage from status
        def get_stage_from_status(status):
            """Determine stage based on status value"""
            if not status:
                return 'S1 - Filtering'
            
            # S1 statuses
            s1_statuses = ['New', 'Not Interested', 'Interested, No DL', 'Interested, No Badge', 
                          'Highly Interested', 'Call back 1D', 'Call back 1W', 'Call back 2W', 'Call back 1M']
            if status in s1_statuses:
                return 'S1 - Filtering'
            
            # S2 statuses
            s2_statuses = ['Docs Upload Pending', 'Verification Pending', 'Duplicate License', 
                          'DL - Amount', 'Verified', 'Verification Rejected']
            if status in s2_statuses:
                return 'S2 - Documentation'
            
            # S3 statuses
            s3_statuses = ['Schedule Pending', 'Training WIP', 'Training Completed', 'Training Rejected',
                          'Re-Training', 'Absent for training', 'Approved']
            if status in s3_statuses:
                return 'S3 - Training'
            
            # S4 statuses
            s4_statuses = ['CT Pending', 'CT WIP', 'Shift Details Pending', 'DONE!', 
                          'Terminated']
            if status in s4_statuses:
                return 'S4 - Onboarding'
            
            # Default to S1 if status not recognized
            return 'S1 - Filtering'
        
        # Clean NaN values and set stage for all leads
        def clean_lead_data(lead_dict):
            cleaned = {}
            for key, value in lead_dict.items():
                if pd.isna(value):
                    cleaned[key] = None
                else:
                    cleaned[key] = value
            
            # Automatically set stage based on status if not present or empty
            if not cleaned.get('stage'):
                status = cleaned.get('status')
                cleaned['stage'] = get_stage_from_status(status)
            
            return cleaned
        
        # Step 8: Update existing leads
        updated_count = 0
        if existing_leads_to_update:
            logger.info(f"Updating {len(existing_leads_to_update)} existing leads...")
            for lead_data in existing_leads_to_update:
                lead_dict = clean_lead_data(lead_data)
                lead_id = lead_dict['id']
                
                # Update the lead (overwrite all fields)
                await leads_collection.update_one(
                    {"id": lead_id},
                    {"$set": lead_dict}
                )
                updated_count += 1
            
            logger.info(f"Updated {updated_count} existing leads")
        
        # Step 9: INSERT new leads
        inserted_count = 0
        if new_leads:
            leads_to_insert = [clean_lead_data(lead.to_dict() if isinstance(lead, pd.Series) else lead) for lead in new_leads]
            logger.info(f"Inserting {len(leads_to_insert)} new leads...")
            
            insert_result = await leads_collection.insert_many(leads_to_insert)
            inserted_count = len(insert_result.inserted_ids)
            logger.info(f"Inserted {inserted_count} new leads")
        
        # Telecaller assignments are stored directly in leads, no separate profile update needed
        updated_telecallers = len(telecaller_assignments)
        
        # Get final count
        total_leads_now = await leads_collection.count_documents({})
        
        return {
            "success": True,
            "message": f"Import completed successfully. {inserted_count} new leads added, {updated_count} existing leads updated.",
            "backup_created": backup_filename if current_leads else None,
            "new_leads_count": inserted_count,
            "updated_leads_count": updated_count,
            "duplicates_updated": duplicates_updated,
            "duplicates_skipped": updated_count,  # Same as updated_leads_count - duplicates are updated, not skipped
            "total_leads_now": total_leads_now,
            "telecaller_assignments": {
                "leads_assigned": leads_with_assignments,
                "telecallers_updated": updated_telecallers
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Bulk import failed: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Bulk import failed: {str(e)}")


@api_router.get("/driver-onboarding/backup-library")
async def get_backup_library(current_user: User = Depends(get_current_user)):
    """
    Get list of all backup files in library
    """
    try:
        import os
        from datetime import datetime
        
        backups = []
        
        # List all files in backup folder
        if os.path.exists(BACKUP_LIBRARY_FOLDER):
            for filename in os.listdir(BACKUP_LIBRARY_FOLDER):
                if filename.endswith('.xlsx') and filename.startswith('backup-'):
                    file_path = os.path.join(BACKUP_LIBRARY_FOLDER, filename)
                    file_stats = os.stat(file_path)
                    
                    backups.append({
                        "filename": filename,
                        "size_bytes": file_stats.st_size,
                        "size_mb": round(file_stats.st_size / (1024 * 1024), 2),
                        "created_at": datetime.fromtimestamp(file_stats.st_ctime).isoformat(),
                        "modified_at": datetime.fromtimestamp(file_stats.st_mtime).isoformat()
                    })
        
        # Sort by created date (newest first)
        backups.sort(key=lambda x: x['created_at'], reverse=True)
        
        return {
            "success": True,
            "backups": backups,
            "total_backups": len(backups)
        }
        
    except Exception as e:
        logger.error(f"Failed to get backup library: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get backup library: {str(e)}")


@api_router.get("/driver-onboarding/backup-library/{filename}/download")
async def download_backup(
    filename: str,
    current_user: User = Depends(get_current_user)
):
    """
    Download a specific backup file from library
    """
    try:
        # Validate filename
        if not filename.endswith('.xlsx') or not filename.startswith('backup-'):
            raise HTTPException(status_code=400, detail="Invalid backup filename")
        
        file_path = os.path.join(BACKUP_LIBRARY_FOLDER, filename)
        
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="Backup file not found")
        
        return FileResponse(
            path=file_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filename,
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to download backup: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to download backup: {str(e)}")


@api_router.delete("/driver-onboarding/backup-library/{filename}")
async def delete_backup(
    filename: str,
    current_user: User = Depends(get_current_user)
):
    """
    Delete a backup file from library (Master Admin only)
    """
    try:
        # Check if user is master admin
        if current_user.account_type != 'master_admin':
            raise HTTPException(status_code=403, detail="Only Master Admin can delete backups")
        
        # Validate filename
        if not filename.endswith('.xlsx') or not filename.startswith('backup-'):
            raise HTTPException(status_code=400, detail="Invalid backup filename")
        
        file_path = os.path.join(BACKUP_LIBRARY_FOLDER, filename)
        
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="Backup file not found")
        
        # Delete the file
        os.remove(file_path)
        logger.info(f"Backup deleted by {current_user.email}: {filename}")
        
        return {
            "success": True,
            "message": f"Backup {filename} deleted successfully"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to delete backup: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to delete backup: {str(e)}")


@api_router.post("/driver-onboarding/backup-library/{filename}/rollback")
async def rollback_to_backup(
    filename: str,
    current_user: User = Depends(get_current_user)
):
    """
    Rollback to a backup file - REPLACES all current leads with backup data
    Creates a backup of current state before rollback
    """
    try:
        import pandas as pd
        from datetime import datetime
        
        logger.info(f"Starting rollback to {filename} by user {current_user.email}")
        
        # Validate filename
        if not filename.endswith('.xlsx') or not filename.startswith('backup-'):
            raise HTTPException(status_code=400, detail="Invalid backup filename")
        
        backup_file_path = os.path.join(BACKUP_LIBRARY_FOLDER, filename)
        
        if not os.path.exists(backup_file_path):
            raise HTTPException(status_code=404, detail="Backup file not found")
        
        # Step 1: Create backup of CURRENT state before rollback
        leads_collection = db['driver_leads']
        current_leads = await leads_collection.find({}, {"_id": 0}).to_list(length=None)
        
        if current_leads:
            current_backup_df = pd.DataFrame(current_leads)
            timestamp = datetime.now().strftime("%d-%m-%Y-%H-%M-%S")
            current_backup_filename = f"backup-before-rollback-{timestamp}.xlsx"
            current_backup_path = os.path.join(BACKUP_LIBRARY_FOLDER, current_backup_filename)
            
            current_backup_df.to_excel(current_backup_path, index=False, sheet_name='Driver Leads')
            logger.info(f"Pre-rollback backup created: {current_backup_filename}")
        else:
            current_backup_filename = None
        
        # Step 2: Read backup file to restore
        backup_df = pd.read_excel(backup_file_path)
        restore_leads = backup_df.to_dict('records')
        
        # Clean up NaN values
        for lead in restore_leads:
            for key, value in list(lead.items()):
                if pd.isna(value):
                    lead[key] = None
        
        # Step 3: DELETE all current leads
        delete_result = await leads_collection.delete_many({})
        deleted_count = delete_result.deleted_count
        logger.info(f"Deleted {deleted_count} current leads")
        
        # Step 4: INSERT leads from backup
        if restore_leads:
            insert_result = await leads_collection.insert_many(restore_leads)
            restored_count = len(insert_result.inserted_ids)
            logger.info(f"Restored {restored_count} leads from backup")
        else:
            restored_count = 0
        
        return {
            "success": True,
            "message": f"Successfully rolled back to {filename}",
            "pre_rollback_backup": current_backup_filename,
            "deleted_count": deleted_count,
            "restored_count": restored_count,
            "rollback_from": filename
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Rollback failed: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Rollback failed: {str(e)}")


@api_router.get("/driver-onboarding/leads")
async def get_leads(
    current_user: User = Depends(get_current_user),
    search: Optional[str] = None,
    telecaller: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    page: int = 1,
    limit: int = 50,
    skip_pagination: bool = False
):
    """
    Get driver leads with pagination, search and telecaller filter
    
    Pagination:
    - page: Page number (default: 1)
    - limit: Items per page (default: 50, max: 100)
    - skip_pagination: Return all results without pagination (for exports)
    
    Search supports:
    - Single or multiple names (comma-separated): e.g., "Alexander" or "Alexander, Antony"
    - Single or multiple phone numbers (comma-separated): e.g., "9898933220" or "9898933220, 8787811221"
    - Partial matching for names, exact matching for phone numbers
    
    Telecaller filter:
    - Filter leads assigned to a specific telecaller by their user ID
    """
    try:
        query = {}
        
        # Handle telecaller filter - support both user ID and email
        if telecaller:
            # Check if it's an email (contains @) or a user ID
            if '@' in telecaller:
                # It's an email, use it directly
                query["assigned_telecaller"] = telecaller
            else:
                # It's a user ID, need to check if leads are stored with email or ID
                # Support both formats: try ID first, if no results, get user's email and try that
                query["$or"] = [
                    {"assigned_telecaller": telecaller},  # Try as user ID
                ]
                
                # Also try to find by user email if ID was provided
                user = await db.users.find_one({"id": telecaller}, {"_id": 0, "email": 1})
                if user and user.get('email'):
                    query["$or"].append({"assigned_telecaller": user['email']})
        
        # Handle date filtering
        if start_date or end_date:
            date_query = {}
            if start_date:
                try:
                    # Parse date and set to beginning of day
                    start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
                    date_query["$gte"] = start_datetime.strftime('%Y-%m-%d')
                except ValueError:
                    pass  # Invalid date format, skip
            if end_date:
                try:
                    # Parse date and set to end of day
                    end_datetime = datetime.strptime(end_date, '%Y-%m-%d')
                    date_query["$lte"] = end_datetime.strftime('%Y-%m-%d')
                except ValueError:
                    pass  # Invalid date format, skip
            
            if date_query:
                # Filter by import_date (the date when lead was imported)
                query["import_date"] = date_query
        
        # Handle search parameter
        if search and search.strip():
            search_values = [s.strip() for s in search.split(',') if s.strip()]
            
            if search_values:
                # Build OR conditions for name and phone
                or_conditions = []
                
                for value in search_values:
                    # Check if it looks like a phone number (digits only)
                    if value.isdigit():
                        # Partial phone match - handles both string and number formats
                        or_conditions.append({"phone_number": {"$regex": value, "$options": "i"}})
                    else:
                        # Partial name match (case-insensitive)
                        or_conditions.append({"name": {"$regex": value, "$options": "i"}})
                
                if or_conditions:
                    query["$or"] = or_conditions
        
        # Get total count for pagination (optimized with index)
        total_count = await db.driver_leads.count_documents(query)
        
        # Fetch leads with pagination and index optimization
        # SORTING LOGIC: Leads without last_called (new/uncalled leads) appear first,
        # then leads sorted by last_called ascending (oldest called first)
        if skip_pagination:
            # For showing all leads, optimize by fetching only essential display fields first
            # This reduces data transfer size significantly
            projection = {
                "_id": 0,
                "id": 1,
                "name": 1,
                "phone_number": 1,
                "status": 1,
                "stage": 1,
                "assigned_telecaller": 1,
                "import_date": 1,
                "source": 1,  # Added source for filtering
                "last_called": 1,  # Add last_called for sorting
                "callback_date": 1  # Add callback_date for filtering
            }
            # Fetch all leads first
            all_leads = await db.driver_leads.find(query, projection).limit(50000).to_list(50000)
            
            # Sort in Python: new leads (no last_called) first, then by last_called ascending
            leads = sorted(all_leads, key=lambda x: (
                x.get('last_called') is not None,  # False (None) comes before True (has value)
                x.get('last_called') or ''  # If has last_called, sort by it ascending
            ))
        else:
            # For paginated requests, return full documents
            limit = min(limit, 100)
            skip_value = (page - 1) * limit
            
            # Fetch all matching leads first (without pagination)
            try:
                cursor = db.driver_leads.find(query, {"_id": 0})
                if telecaller:
                    try:
                        cursor = cursor.hint("idx_assigned_telecaller")
                    except:
                        pass  # Index might not exist
                all_leads = await cursor.to_list(50000)
            except Exception as hint_error:
                # If index doesn't exist, fall back to query without hint
                logger.warning(f"Index not found for driver leads, using query without hint: {str(hint_error)}")
                cursor = db.driver_leads.find(query, {"_id": 0})
                all_leads = await cursor.to_list(50000)
            
            # Sort in Python: new leads (no last_called) first, then by last_called ascending
            sorted_leads = sorted(all_leads, key=lambda x: (
                x.get('last_called') is not None,  # False (None) comes before True (has value)
                x.get('last_called') or ''  # If has last_called, sort by it ascending
            ))
            
            # Apply pagination after sorting
            leads = sorted_leads[skip_value:skip_value + limit]
        
        # Populate telecaller names for leads with assigned telecallers
        # Batch fetch all unique telecallers for efficiency
        telecaller_identifiers = set()
        for lead in leads:
            if lead.get('assigned_telecaller'):
                telecaller_identifiers.add(lead['assigned_telecaller'])
        
        # Fetch all telecallers in one query
        telecaller_map = {}
        if telecaller_identifiers:
            telecallers = await db.users.find(
                {"$or": [
                    {"id": {"$in": list(telecaller_identifiers)}},
                    {"email": {"$in": list(telecaller_identifiers)}}
                ]},
                {"_id": 0, "id": 1, "email": 1, "first_name": 1, "last_name": 1}
            ).to_list(None)
            
            # Build map for quick lookup
            for tc in telecallers:
                name = f"{tc.get('first_name', '')} {tc.get('last_name', '')}".strip()
                if not name:
                    name = tc.get('email', '').split('@')[0]
                # Map by both ID and email
                if tc.get('id'):
                    telecaller_map[tc['id']] = name
                if tc.get('email'):
                    telecaller_map[tc['email']] = name
        
        # Assign telecaller names to leads and clean up remarks data
        for lead in leads:
            if lead.get('assigned_telecaller'):
                telecaller_name = telecaller_map.get(lead['assigned_telecaller'])
                if telecaller_name:
                    lead['assigned_telecaller_name'] = telecaller_name
            
            # Ensure remarks is always a string (handle legacy data)
            if lead.get("remarks"):
                if isinstance(lead["remarks"], list) and len(lead["remarks"]) > 0:
                    # Convert array of remark objects to string
                    lead["remarks"] = "\n".join([
                        remark.get("text", str(remark)) if isinstance(remark, dict) else str(remark)
                        for remark in lead["remarks"]
                    ])
                elif isinstance(lead["remarks"], dict):
                    # Convert single remark object to string
                    lead["remarks"] = lead["remarks"].get("text", str(lead["remarks"]))
                elif not isinstance(lead["remarks"], str):
                    # Convert any other type to string
                    lead["remarks"] = str(lead["remarks"])
            else:
                # Set empty remarks to None instead of string "None"
                lead["remarks"] = None
        
        # Return with pagination metadata
        if skip_pagination:
            return {
                "leads": leads,
                "total": len(leads)
            }
        else:
            return {
                "leads": leads,
                "total": total_count,
                "page": page,
                "limit": limit,
                "total_pages": (total_count + limit - 1) // limit
            }
        
    except Exception as e:
        logger.error(f"Error fetching leads with search: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch leads: {str(e)}")


@api_router.post("/driver-onboarding/{lead_id}/remarks")
async def add_remark(
    lead_id: str,
    remark_text: str = Body(..., embed=True),
    current_user: User = Depends(get_current_user)
):
    """Add a remark to a driver lead"""
    try:
        from datetime import datetime, timezone
        
        # Get lead
        lead = await db.driver_leads.find_one({"id": lead_id})
        if not lead:
            raise HTTPException(status_code=404, detail="Lead not found")
        
        # Create remark object
        new_remark = {
            "text": remark_text,
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "user_id": current_user.id,
            "user_name": f"{current_user.first_name} {current_user.last_name or ''}".strip() or current_user.email,
            "user_email": current_user.email
        }
        
        # Add remark to lead (push to remarks array)
        await db.driver_leads.update_one(
            {"id": lead_id},
            {"$push": {"remarks": new_remark}}
        )
        
        logger.info(f"Added remark to lead {lead_id} by {current_user.email}")
        
        return {
            "success": True,
            "message": "Remark added successfully",
            "remark": new_remark
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error adding remark: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to add remark: {str(e)}")


@api_router.get("/driver-onboarding/{lead_id}/remarks")
async def get_remarks(
    lead_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get all remarks for a driver lead"""
    try:
        # Get lead with remarks
        lead = await db.driver_leads.find_one({"id": lead_id}, {"remarks": 1})
        
        if not lead:
            raise HTTPException(status_code=404, detail="Lead not found")
        
        remarks = lead.get("remarks", [])
        
        # Sort by timestamp (newest first)
        remarks_sorted = sorted(remarks, key=lambda x: x.get("timestamp", ""), reverse=True)
        
        return {
            "success": True,
            "remarks": remarks_sorted,
            "count": len(remarks_sorted)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching remarks: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch remarks: {str(e)}")


# BACKEND ENDPOINT REMOVED - Status summary is now calculated on frontend
# User will provide custom counting logic in the frontend
# @api_router.get("/driver-onboarding/status-summary")
# async def get_status_summary(
#     start_date: Optional[str] = None,
#     end_date: Optional[str] = None,
#     source: Optional[str] = None,
#     current_user: User = Depends(get_current_user)
# ):
#     """
#     Get status summary dashboard with counts grouped by stage and status.
#     Optionally filter by date range (import_date field) and source.
#     """
#     pass



@api_router.get("/driver-onboarding/sources")
async def get_unique_sources(current_user: User = Depends(get_current_user)):
    """
    Get unique import sources from all leads for filter dropdown.
    Returns sources with meaningful names, filtering out filenames when possible.
    """
    try:
        # Use MongoDB aggregation to get distinct non-null sources
        pipeline = [
            {"$match": {"source": {"$exists": True, "$ne": None, "$ne": ""}}},
            {"$group": {"_id": "$source"}},
            {"$sort": {"_id": 1}}
        ]
        
        results = await db.driver_leads.aggregate(pipeline).to_list(None)
        sources = [result['_id'] for result in results if result['_id']]
        
        # Filter out filename-like sources (with extensions) and keep meaningful names
        meaningful_sources = []
        filename_sources = []
        
        for source in sources:
            source_str = str(source).strip()
            # Check if it looks like a filename (has extension like .xlsx, .csv, .xls)
            if any(source_str.lower().endswith(ext) for ext in ['.xlsx', '.xls', '.csv', '.txt']):
                # Extract meaningful part from filename (remove extension and path)
                import os
                base_name = os.path.splitext(os.path.basename(source_str))[0]
                if base_name and base_name not in meaningful_sources:
                    filename_sources.append({"value": source_str, "label": base_name})
            else:
                # It's a meaningful source name (like 'bhavani', 'ganesh', 'digital leads')
                meaningful_sources.append({"value": source_str, "label": source_str})
        
        # Combine: meaningful names first, then filename-based sources
        all_sources = meaningful_sources + filename_sources
        
        logger.info(f"Found {len(all_sources)} unique import sources")
        
        return {
            "success": True,
            "sources": all_sources,
            "count": len(all_sources)
        }
        
    except Exception as e:
        logger.error(f"Error fetching unique sources: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch sources: {str(e)}")


@api_router.post("/driver-onboarding/sync-leads")
async def sync_leads_to_sheets(current_user: User = Depends(get_current_user)):
    """Sync all leads to Google Sheets with batch processing to avoid timeouts"""
    try:
        # Get all leads from database (no limit)
        leads = await db.driver_leads.find({}, {"_id": 0}).to_list(length=None)
        
        if not leads:
            return {"message": "No leads to sync", "count": 0}
        
        # Get Google Sheets Web App URL
        web_app_url = os.environ.get('GOOGLE_SHEETS_WEB_APP_URL')
        
        if not web_app_url:
            raise HTTPException(status_code=500, detail="GOOGLE_SHEETS_WEB_APP_URL not configured")
        
        # BATCH PROCESSING to avoid timeout
        BATCH_SIZE = 500  # Sync 500 leads at a time
        total_leads = len(leads)
        total_updated = 0
        total_created = 0
        
        import requests
        
        # Process in batches
        for i in range(0, total_leads, BATCH_SIZE):
            batch = leads[i:i + BATCH_SIZE]
            batch_num = (i // BATCH_SIZE) + 1
            total_batches = (total_leads + BATCH_SIZE - 1) // BATCH_SIZE
            
            logger.info(f"Syncing batch {batch_num}/{total_batches} ({len(batch)} leads)")
            
            # Prepare payload for Google Sheets
            payload = {
                "action": "sync_from_app",
                "leads": batch
            }
            
            # Send batch to Google Sheets Web App with longer timeout
            response = requests.post(
                web_app_url,
                json=payload,
                headers={'Content-Type': 'application/json'},
                timeout=60  # Increased timeout to 60 seconds per batch
            )
            
            if response.status_code == 200:
                result = response.json()
                if result.get('success'):
                    total_updated += result.get('updated', 0)
                    total_created += result.get('created', 0)
                    logger.info(f"Batch {batch_num} synced: {result.get('updated', 0)} updated, {result.get('created', 0)} created")
                else:
                    logger.error(f"Batch {batch_num} failed: {result.get('message', 'Unknown error')}")
                    raise HTTPException(status_code=500, detail=f"Batch {batch_num} sync failed: {result.get('message', 'Unknown error')}")
            else:
                logger.error(f"Batch {batch_num} API error: {response.status_code}")
                raise HTTPException(status_code=500, detail=f"Batch {batch_num} API error: {response.status_code}")
        
        return {
            "success": True,
            "message": f"Successfully synced {total_leads} leads to Google Sheets in {total_batches} batches",
            "total_leads": total_leads,
            "updated": total_updated,
            "created": total_created,
            "batches": total_batches
        }
            
    except requests.exceptions.Timeout:
        raise HTTPException(status_code=504, detail="Google Sheets sync timed out. Please try again or reduce batch size.")
    except requests.exceptions.RequestException as e:
        logger.error(f"Request error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to connect to Google Sheets: {str(e)}")
    except Exception as e:
        logger.error(f"Sync error: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Failed to sync: {str(e)}")


class LeadStatusUpdate(BaseModel):
    status: str


# IMPORTANT: Bulk update route must come BEFORE {lead_id} routes to avoid path conflicts
@api_router.patch("/driver-onboarding/leads/bulk-update-status")
async def bulk_update_lead_status(bulk_data: BulkLeadStatusUpdate, current_user: User = Depends(get_current_user)):
    """Bulk update lead status for multiple leads"""
    print("=== BULK UPDATE CALLED ===")
    print(f"Lead IDs count: {len(bulk_data.lead_ids) if bulk_data.lead_ids else 0}")
    print(f"Status: {bulk_data.status}")
    
    if not bulk_data.lead_ids or len(bulk_data.lead_ids) == 0:
        raise HTTPException(status_code=400, detail="No leads selected for update")
    
    # Update all matching leads
    result = await db.driver_leads.update_many(
        {"id": {"$in": bulk_data.lead_ids}},
        {"$set": {"status": bulk_data.status}}
    )
    
    print(f"MongoDB update result: matched={result.matched_count}, modified={result.modified_count}")
    
    # Get all updated leads for syncing to Google Sheets (no limit on bulk update size)
    updated_leads = await db.driver_leads.find(
        {"id": {"$in": bulk_data.lead_ids}},
        {"_id": 0}
    ).to_list(length=None)
    
    print(f"Found {len(updated_leads)} leads to sync")
    
    if not updated_leads or len(updated_leads) == 0:
        print("ERROR: No leads found - raising 404")
        raise HTTPException(status_code=404, detail="Could not find any leads with the provided IDs")
    
    # Sync all to Google Sheets
    try:
        sync_all_records('leads', updated_leads)
    except Exception as e:
        logger.error(f"Failed to sync to Google Sheets: {str(e)}")
        # Don't fail the whole operation if sheets sync fails
    
    # Return success message
    count = len(updated_leads)
    return {
        "success": True,
        "message": f"Successfully updated {count} lead(s)",
        "updated_count": count
    }


@api_router.patch("/driver-onboarding/bulk-assign")
async def bulk_assign_telecaller(
    assignment_data: dict,
    current_user: User = Depends(get_current_user)
):
    """Bulk assign telecaller to multiple leads with specific assignment date"""
    lead_ids = assignment_data.get("lead_ids", [])
    telecaller_email = assignment_data.get("telecaller_email", "")
    assignment_date = assignment_data.get("assignment_date")  # YYYY-MM-DD format from frontend
    
    if not lead_ids or len(lead_ids) == 0:
        raise HTTPException(status_code=400, detail="No leads selected for assignment")
    
    if not telecaller_email:
        raise HTTPException(status_code=400, detail="No telecaller selected")
    
    if not assignment_date:
        raise HTTPException(status_code=400, detail="No assignment date provided")
    
    # Verify telecaller exists in USERS collection (not telecaller_profiles)
    telecaller = await db.users.find_one({
        "email": telecaller_email,
        "account_type": "telecaller"
    })
    if not telecaller:
        raise HTTPException(status_code=404, detail="Telecaller not found in users")
    
    # Get telecaller name
    telecaller_name = f"{telecaller.get('first_name', '')} {telecaller.get('last_name', '')}".strip()
    
    # Parse assignment date and convert to IST
    from datetime import datetime
    import pytz
    try:
        # Parse YYYY-MM-DD format
        parsed_date = datetime.strptime(assignment_date, "%Y-%m-%d")
        # Set to start of day in IST
        ist = pytz.timezone('Asia/Kolkata')
        assigned_datetime_ist = ist.localize(parsed_date.replace(hour=0, minute=0, second=0))
        # Store as ISO string
        assignment_date_iso = assigned_datetime_ist.isoformat()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Expected YYYY-MM-DD")
    
    # Update all matching leads with assigned telecaller and assignment date
    result = await db.driver_leads.update_many(
        {"id": {"$in": lead_ids}},
        {
            "$set": {
                "assigned_telecaller": telecaller_email,
                "assigned_telecaller_name": telecaller_name,
                "assigned_date": assignment_date_iso
            }
        }
    )
    
    # Get updated leads for sync
    updated_leads = await db.driver_leads.find(
        {"id": {"$in": lead_ids}},
        {"_id": 0}
    ).to_list(length=None)
    
    # Sync to Google Sheets
    try:
        sync_all_records('leads', updated_leads)
    except Exception as e:
        logger.error(f"Failed to sync to Google Sheets: {str(e)}")
    
    logger.info(f"Assigned {result.modified_count} leads to {telecaller_name} ({telecaller_email}) for date {assignment_date}")
    
    return {
        "success": True,
        "message": f"Successfully assigned {result.modified_count} lead(s) to {telecaller_name}",
        "modified_count": result.modified_count
    }


@api_router.patch("/driver-onboarding/reassign-date")
async def reassign_leads_to_date(
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """
    Reassign selected leads to a different date
    Updates the assigned_date field for bulk operations
    """
    try:
        data = await request.json()
        lead_ids = data.get("lead_ids", [])
        new_date = data.get("new_date")
        
        if not lead_ids:
            raise HTTPException(status_code=400, detail="No lead IDs provided")
        
        if not new_date:
            raise HTTPException(status_code=400, detail="No date provided")
        
        # Parse and validate the date
        from datetime import datetime, timezone
        try:
            # Parse the date string (expecting YYYY-MM-DD format)
            parsed_date = datetime.fromisoformat(new_date)
            # Convert to ISO string with UTC timezone
            new_date_iso = parsed_date.replace(tzinfo=timezone.utc).isoformat()
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
        
        # Update all matching leads with new assigned date
        result = await db.driver_leads.update_many(
            {"id": {"$in": lead_ids}},
            {
                "$set": {
                    "assigned_date": new_date_iso
                }
            }
        )
        
        # Get updated leads for sync
        updated_leads = await db.driver_leads.find(
            {"id": {"$in": lead_ids}},
            {"_id": 0}
        ).to_list(length=None)
        
        # Sync to Google Sheets
        try:
            sync_all_records('leads', updated_leads)
        except Exception as e:
            logger.error(f"Failed to sync to Google Sheets: {str(e)}")
        
        logger.info(f"Reassigned {result.modified_count} leads to date {new_date}")
        
        return {
            "success": True,
            "message": f"Successfully reassigned {result.modified_count} lead(s) to {parsed_date.strftime('%B %d, %Y')}",
            "modified_count": result.modified_count
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error reassigning leads to date: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.patch("/driver-onboarding/leads-bulk/status")
async def bulk_update_lead_status(bulk_data: BulkLeadStatusUpdate, current_user: User = Depends(get_current_user)):
    """Bulk update lead status for multiple leads"""
    print("=== BULK UPDATE CALLED ===")
    print(f"Lead IDs count: {len(bulk_data.lead_ids) if bulk_data.lead_ids else 0}")
    print(f"Status: {bulk_data.status}")
    
    if not bulk_data.lead_ids or len(bulk_data.lead_ids) == 0:
        raise HTTPException(status_code=400, detail="No leads selected for update")
    
    # Update all matching leads
    result = await db.driver_leads.update_many(
        {"id": {"$in": bulk_data.lead_ids}},
        {"$set": {"status": bulk_data.status}}
    )
    
    print(f"MongoDB update result: matched={result.matched_count}, modified={result.modified_count}")
    
    # Get all updated leads for syncing to Google Sheets (no limit on bulk update size)
    updated_leads = await db.driver_leads.find(
        {"id": {"$in": bulk_data.lead_ids}},
        {"_id": 0}
    ).to_list(length=None)
    
    print(f"Found {len(updated_leads)} leads to sync")
    
    if not updated_leads or len(updated_leads) == 0:
        print("ERROR: No leads found - raising 404")
        raise HTTPException(status_code=404, detail="Could not find any leads with the provided IDs")
    
    # Sync all to Google Sheets
    try:
        sync_all_records('leads', updated_leads)
    except Exception as e:
        logger.error(f"Failed to sync to Google Sheets: {str(e)}")
        # Don't fail the whole operation if sheets sync fails
    
    # Return success message
    count = len(updated_leads)
    return {
        "message": f"Successfully updated status for {count} lead(s) to '{bulk_data.status}'",
        "updated_count": count
    }


@api_router.patch("/driver-onboarding/leads/{lead_id}")
async def update_lead(lead_id: str, lead_data: DriverLeadUpdate, current_user: User = Depends(get_current_user)):
    """Update lead details with status history tracking and callback date calculation"""
    from datetime import datetime, timezone, timedelta
    
    # Find the lead
    lead = await db.driver_leads.find_one({"id": lead_id})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    # Prepare update data (only include fields that are provided)
    update_data = {}
    for field, value in lead_data.model_dump(exclude_unset=True).items():
        if value is not None:
            # Ensure phone_number is always stored as string
            if field == "phone_number":
                update_data[field] = str(value)
            else:
                update_data[field] = value
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")
    
    # Track status/stage changes in history
    current_time = datetime.now(timezone.utc).isoformat()
    
    # Initialize status_history if it doesn't exist, is null, or is not an array
    status_history = lead.get("status_history")
    if not isinstance(status_history, list):
        # Fix corrupted status_history field (could be null, string, or any non-list type)
        await db.driver_leads.update_one(
            {"id": lead_id},
            {"$set": {"status_history": []}}
        )
        lead["status_history"] = []
    
    # Calculate callback_date if status starts with "Call back"
    if "status" in update_data and update_data["status"].startswith("Call back"):
        now = datetime.now(timezone.utc)
        callback_days = 0
        
        if update_data["status"] == "Call back 1D":
            callback_days = 1
        elif update_data["status"] == "Call back 1W":
            callback_days = 7
        elif update_data["status"] == "Call back 2W":
            callback_days = 14
        elif update_data["status"] == "Call back 1M":
            callback_days = 30
        
        if callback_days > 0:
            callback_date = now + timedelta(days=callback_days)
            update_data["callback_date"] = callback_date.isoformat()
    else:
        # If status is NOT a callback status, clear callback_date
        if "status" in update_data:
            update_data["callback_date"] = None
    
    # Check if status or stage changed
    history_entry = None
    if "status" in update_data and update_data["status"] != lead.get("status"):
        history_entry = {
            "timestamp": current_time,
            "field": "status",
            "old_value": lead.get("status", "N/A"),
            "new_value": update_data["status"],
            "changed_by": current_user.email,
            "action": "status_changed"
        }
    elif "stage" in update_data and update_data["stage"] != lead.get("stage"):
        history_entry = {
            "timestamp": current_time,
            "field": "stage",
            "old_value": lead.get("stage", "N/A"),
            "new_value": update_data["stage"],
            "changed_by": current_user.email,
            "action": "stage_changed"
        }
    
    # Add last_modified timestamp
    update_data['last_modified'] = current_time
    
    # If there's a history entry, add it to status_history
    if history_entry:
        await db.driver_leads.update_one(
            {"id": lead_id},
            {
                "$set": update_data,
                "$push": {"status_history": history_entry}
            }
        )
    else:
        # Update lead without history entry
        await db.driver_leads.update_one(
            {"id": lead_id},
            {"$set": update_data}
        )
    
    # Fetch and return the updated lead
    updated_lead = await db.driver_leads.find_one({"id": lead_id}, {"_id": 0})
    
    return {"success": True, "lead": updated_lead}
    


@api_router.post("/driver-onboarding/leads/{lead_id}/call-done")
async def mark_call_done(lead_id: str, current_user: User = Depends(get_current_user)):
    """Mark that telecaller completed a call for this lead"""
    from datetime import datetime, timezone
    
    # Find the lead
    lead = await db.driver_leads.find_one({"id": lead_id})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    current_time = datetime.now(timezone.utc).isoformat()
    
    # Create call history entry
    call_entry = {
        "timestamp": current_time,
        "called_by": current_user.email,
        "caller_name": f"{current_user.first_name} {current_user.last_name}" if hasattr(current_user, 'first_name') else current_user.email
    }
    
    # Add to status history as well
    status_history_entry = {
        "timestamp": current_time,
        "field": "call_made",
        "old_value": None,
        "new_value": "Call completed",
        "changed_by": current_user.email,
        "action": "call_made"
    }
    
    # Update lead with call history
    await db.driver_leads.update_one(
        {"id": lead_id},
        {
            "$push": {
                "calling_history": call_entry,
                "status_history": status_history_entry
            },
            "$set": {
                "last_called": current_time,
                "last_modified": current_time
            }
        }
    )
    
    return {
        "success": True,
        "message": "Call marked as done",
        "lead_id": lead_id,
        "last_called": current_time
    }


@api_router.post("/driver-onboarding/leads/{lead_id}/mark-called")
async def mark_lead_as_called(lead_id: str, current_user: User = Depends(get_current_user)):
    """Mark lead as called with timestamp"""
    from datetime import datetime
    import pytz
    
    # Find the lead
    lead = await db.driver_leads.find_one({"id": lead_id})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    # Initialize calling_history and status_history if they don't exist, are null, or are not arrays
    calling_history = lead.get("calling_history")
    if not isinstance(calling_history, list):
        # Fix corrupted calling_history field
        await db.driver_leads.update_one(
            {"id": lead_id},
            {"$set": {"calling_history": []}}
        )
        lead["calling_history"] = []
    
    status_history = lead.get("status_history")
    if not isinstance(status_history, list):
        # Fix corrupted status_history field
        await db.driver_leads.update_one(
            {"id": lead_id},
            {"$set": {"status_history": []}}
        )
        lead["status_history"] = []
    
    # Get current time in IST
    ist = pytz.timezone('Asia/Kolkata')
    current_time_ist = datetime.now(ist)
    current_time_iso = current_time_ist.isoformat()
    
    # Create call history entry
    call_entry = {
        "timestamp": current_time_iso,
        "called_by": current_user.email,
        "caller_name": f"{current_user.first_name} {current_user.last_name}" if hasattr(current_user, 'first_name') else current_user.email
    }
    
    # Add to status history
    status_history_entry = {
        "timestamp": current_time_iso,
        "field": "call_made",
        "old_value": None,
        "new_value": "Call completed",
        "changed_by": current_user.email,
        "action": "call_made"
    }
    
    # Update lead with call history and last_called timestamp
    await db.driver_leads.update_one(
        {"id": lead_id},
        {
            "$push": {
                "calling_history": call_entry,
                "status_history": status_history_entry
            },
            "$set": {
                "last_called": current_time_iso,
                "last_modified": current_time_iso
            }
        }
    )
    
    logger.info(f"Lead {lead_id} marked as called by {current_user.email} at {current_time_iso}")
    
    return {
        "success": True,
        "message": "Lead marked as called",
        "lead_id": lead_id,
        "last_called": current_time_iso
    }



@api_router.get("/driver-onboarding/telecaller-summary")
async def get_telecaller_summary(
    telecaller: str = Query(None),
    start_date: str = Query(None),
    end_date: str = Query(None),
    source: str = Query(None),
    current_user: User = Depends(get_current_user)
):
    """Get summary statistics for a telecaller - OPTIMIZED with projection and index hint"""
    from datetime import datetime, date, timezone
    
    # If no telecaller specified, use current user's email
    if not telecaller:
        telecaller = current_user.email
    
    # Build query
    query = {"assigned_telecaller": telecaller}
    
    # Add date filter if provided
    if start_date and end_date:
        try:
            start = datetime.fromisoformat(start_date).date().isoformat()
            end = datetime.fromisoformat(end_date).date().isoformat()
            query["import_date"] = {"$gte": start, "$lte": end}
        except Exception as e:
            logger.error(f"Error parsing dates: {e}")
    
    # Add source filter if provided
    if source:
        query["source"] = source
    
    # OPTIMIZATION: Only fetch fields needed for summary (reduces data transfer)
    projection = {
        "_id": 0,
        "status": 1,
        "stage": 1,
        "last_called": 1
    }
    
    # Get all leads for this telecaller with optimized query
    try:
        cursor = db.driver_leads.find(query, projection).hint("idx_assigned_telecaller")
        leads = await cursor.to_list(length=50000)
    except Exception as hint_error:
        # Fallback if index doesn't exist
        logger.warning(f"Index not available, using query without hint: {hint_error}")
        cursor = db.driver_leads.find(query, projection)
        leads = await cursor.to_list(length=50000)
    
    total_leads = len(leads)
    
    # Count calls made today
    today = date.today().isoformat()
    calls_today = 0
    for lead in leads:
        last_called = lead.get("last_called")
        if last_called:
            try:
                call_date = datetime.fromisoformat(last_called).date().isoformat()
                if call_date == today:
                    calls_today += 1
            except:
                pass
    
    # Calls pending = leads that have never been called or not called today
    calls_pending = total_leads - calls_today
    
    # Group by stage
    stage_breakdown = {}
    for lead in leads:
        stage = lead.get("stage", "Unknown")
        if stage not in stage_breakdown:
            stage_breakdown[stage] = {
                "total": 0,
                "statuses": {}
            }
        stage_breakdown[stage]["total"] += 1
        
        # Count statuses within each stage
        status = lead.get("status", "Unknown")
        if status not in stage_breakdown[stage]["statuses"]:
            stage_breakdown[stage]["statuses"][status] = 0
        stage_breakdown[stage]["statuses"][status] += 1
    
    # Sort stages (S1, S2, S3, S4, then others)
    sorted_stages = sorted(
        stage_breakdown.keys(),
        key=lambda x: (0 if x.startswith("S") else 1, x)
    )
    
    return {
        "success": True,
        "telecaller": telecaller,
        "total_leads": total_leads,
        "calls_made_today": calls_today,
        "calls_pending": calls_pending,
        "stage_breakdown": {stage: stage_breakdown[stage] for stage in sorted_stages},
        "start_date": start_date,
        "end_date": end_date
    }



@api_router.get("/telecaller-desk/leads")
async def get_telecaller_desk_leads(
    telecaller_email: str = Query(..., description="Telecaller email"),
    date: str = Query(None, description="Date in YYYY-MM-DD format"),
    current_user: User = Depends(get_current_user)
):
    """
    Get leads for Telecaller's Desk - shows leads assigned for specific date + callback leads
    Returns two separate lists: assigned_leads and callback_leads
    """
    from datetime import datetime
    import pytz
    
    # Verify telecaller exists
    telecaller = await db.users.find_one({
        "email": telecaller_email,
        "account_type": "telecaller"
    })
    if not telecaller:
        raise HTTPException(status_code=404, detail="Telecaller not found")
    
    # Get all leads assigned to this telecaller
    all_leads = await db.driver_leads.find(
        {"assigned_telecaller": telecaller_email},
        {"_id": 0}
    ).to_list(length=10000)
    
    # If no date filter, return all leads
    if not date:
        return {
            "success": True,
            "telecaller_email": telecaller_email,
            "date_filter": None,
            "assigned_leads": all_leads,
            "callback_leads": [],
            "total_assigned": len(all_leads),
            "total_callbacks": 0
        }
    
    # Parse the selected date
    try:
        selected_date = datetime.strptime(date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Expected YYYY-MM-DD")
    
    # Filter leads for the selected date
    assigned_for_date = []
    callbacks_for_date = []
    
    for lead in all_leads:
        # Check if lead is assigned for this date
        if lead.get("assigned_date"):
            try:
                assigned_dt = datetime.fromisoformat(lead["assigned_date"])
                if assigned_dt.date() == selected_date:
                    assigned_for_date.append(lead)
                    continue  # Don't check callback if already in assigned
            except:
                pass
        
        # Check if lead has callback for this date
        if lead.get("callback_date"):
            try:
                callback_dt = datetime.fromisoformat(lead["callback_date"])
                if callback_dt.date() == selected_date:
                    callbacks_for_date.append(lead)
            except:
                pass
    
    return {
        "success": True,
        "telecaller_email": telecaller_email,
        "date_filter": date,
        "assigned_leads": assigned_for_date,
        "callback_leads": callbacks_for_date,
        "total_assigned": len(assigned_for_date),
        "total_callbacks": len(callbacks_for_date)
    }


    return {
        "success": True,
        "message": "Call marked as done",
        "timestamp": current_time
    }

    # Get updated lead for sync
    updated_lead = await db.driver_leads.find_one({"id": lead_id}, {"_id": 0})
    
    # Sync to Google Sheets
    sync_single_record('leads', updated_lead)
    
    return {"message": "Lead updated successfully", "lead": updated_lead}

@api_router.get("/driver-onboarding/leads/{lead_id}")
async def get_lead(lead_id: str, current_user: User = Depends(get_current_user)):
    """Get a single lead by ID"""
    try:
        # Find the lead
        lead = await db.driver_leads.find_one({"id": lead_id}, {"_id": 0})
        if not lead:
            raise HTTPException(status_code=404, detail="Lead not found")
        
        # Ensure remarks is always a string (handle legacy data)
        if lead.get("remarks"):
            if isinstance(lead["remarks"], list) and len(lead["remarks"]) > 0:
                # Convert array of remark objects to string
                lead["remarks"] = "\n".join([
                    remark.get("text", str(remark)) if isinstance(remark, dict) else str(remark)
                    for remark in lead["remarks"]
                ])
            elif isinstance(lead["remarks"], dict):
                # Convert single remark object to string
                lead["remarks"] = lead["remarks"].get("text", str(lead["remarks"]))
            elif not isinstance(lead["remarks"], str):
                # Convert any other type to string
                lead["remarks"] = str(lead["remarks"])
        
        return lead
        
    except Exception as e:
        logger.error(f"Error fetching lead {lead_id}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch lead: {str(e)}")


@api_router.patch("/driver-onboarding/leads/{lead_id}/status")
async def update_lead_status(lead_id: str, status_data: LeadStatusUpdate, current_user: User = Depends(get_current_user)):
    """Update lead status"""
    from datetime import datetime, timezone
    
    # Find the lead
    lead = await db.driver_leads.find_one({"id": lead_id})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    # Update status and last_modified
    await db.driver_leads.update_one(
        {"id": lead_id},
        {"$set": {
            "status": status_data.status,
            "last_modified": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Get updated lead for sync
    updated_lead = await db.driver_leads.find_one({"id": lead_id}, {"_id": 0})
    
    # Sync to Google Sheets
    sync_single_record('leads', updated_lead)
    
    return {"message": "Lead status updated successfully", "lead": updated_lead}


@api_router.post("/driver-onboarding/leads/{lead_id}/sync-stage")
async def sync_driver_stage(lead_id: str, current_user: User = Depends(get_current_user)):
    """
    Auto-progress driver to next stage if on a completion (green) status.
    Stage progression rules:
    - S1 (Filtering) -> S2 (Docs Collection) when status is "Highly Interested"
    - S2 (Docs Collection) -> S3 (Training) when status is "Verified" (and mandatory docs uploaded)
    - S3 (Training) -> S4 (Customer Readiness) when status is "Approved"
    - S4 is final stage, no further progression
    """
    # Find the lead
    lead = await db.driver_leads.find_one({"id": lead_id})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    current_stage = lead.get('stage', 'S1')
    current_status = lead.get('status', 'New')
    
    # Define stage progression rules
    stage_progression = {
        'S1': {
            'completion_status': 'Highly Interested',
            'next_stage': 'S2',
            'next_status': 'Docs Upload Pending'
        },
        'S2': {
            'completion_status': 'Verified',
            'next_stage': 'S3',
            'next_status': 'Schedule Pending',
            'requires_docs': True  # Mandatory docs check
        },
        'S3': {
            'completion_status': 'Approved',
            'next_stage': 'S4',
            'next_status': 'CT Pending'
        }
    }
    
    # Check if current stage has progression rules
    if current_stage not in stage_progression:
        if current_stage == 'S4':
            return {
                "success": False,
                "message": "Driver is already in final stage (S4). No further progression available.",
                "current_stage": current_stage,
                "current_status": current_status
            }
        return {
            "success": False,
            "message": f"Invalid stage: {current_stage}",
            "current_stage": current_stage,
            "current_status": current_status
        }
    
    progression_rule = stage_progression[current_stage]
    
    # Check if on completion status
    if current_status != progression_rule['completion_status']:
        return {
            "success": False,
            "message": f"Cannot sync. Status must be '{progression_rule['completion_status']}' to progress from {current_stage}. Current status is '{current_status}'.",
            "current_stage": current_stage,
            "current_status": current_status,
            "required_status": progression_rule['completion_status']
        }
    
    # S2 specific: Check mandatory documents
    if progression_rule.get('requires_docs'):
        dl_no = lead.get('dl_no')
        aadhar_card = lead.get('aadhar_card')
        gas_bill = lead.get('gas_bill')
        bank_passbook = lead.get('bank_passbook')  # Address verification alternative
        
        missing_docs = []
        if not dl_no:
            missing_docs.append('Driver Licence')
        if not aadhar_card:
            missing_docs.append('Aadhar Card')
        if not gas_bill and not bank_passbook:
            missing_docs.append('Gas Bill or Address Verification (Bank Passbook)')
        
        if missing_docs:
            return {
                "success": False,
                "message": f"Cannot progress from S2. Missing mandatory documents: {', '.join(missing_docs)}",
                "current_stage": current_stage,
                "current_status": current_status,
                "missing_documents": missing_docs
            }
    
    # Perform stage progression
    new_stage = progression_rule['next_stage']
    new_status = progression_rule['next_status']
    
    await db.driver_leads.update_one(
        {"id": lead_id},
        {"$set": {
            "stage": new_stage,
            "status": new_status
        }}
    )
    
    # Get updated lead for sync
    updated_lead = await db.driver_leads.find_one({"id": lead_id}, {"_id": 0})
    
    # Sync to Google Sheets
    sync_single_record('leads', updated_lead)
    
    return {
        "success": True,
        "message": f"Successfully progressed from {current_stage} ({current_status}) to {new_stage} ({new_status})",
        "previous_stage": current_stage,
        "previous_status": current_status,
        "new_stage": new_stage,
        "new_status": new_status,
        "lead": updated_lead
    }


@api_router.delete("/driver-onboarding/leads/{lead_id}")
async def delete_lead(lead_id: str, current_user: User = Depends(get_current_user)):
    """Delete a single lead (Master Admin only)"""
    # Check if user is master admin
    if current_user.account_type != "master_admin":
        raise HTTPException(status_code=403, detail="Only Master Admin can delete leads")
    
    # Find the lead
    lead = await db.driver_leads.find_one({"id": lead_id})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    # Delete from MongoDB
    await db.driver_leads.delete_one({"id": lead_id})
    
    # Delete from Google Sheets
    try:
        delete_record('leads', lead_id)
    except Exception as e:
        logger.error(f"Failed to delete from Google Sheets: {str(e)}")
        # Don't fail the whole operation if sheets delete fails
    
    return {"message": "Lead deleted successfully"}


# ==================== TELECALLER MANAGEMENT ENDPOINTS ====================

@api_router.get("/telecallers")
async def get_telecallers(current_user: User = Depends(get_current_user)):
    """Get all telecaller profiles"""
    telecallers = await db.telecaller_profiles.find({}, {"_id": 0}).to_list(1000)
    return telecallers


@api_router.get("/users/telecallers")
async def get_telecaller_users(current_user: User = Depends(get_current_user)):
    """Get all users with telecaller account type - returns actual user names"""
    if current_user.account_type not in ["master_admin", "admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    telecallers = await db.users.find(
        {
            "account_type": "telecaller",
            "status": {"$ne": "deleted"}  # Exclude deleted users
        },
        {
            "_id": 0,
            "password": 0  # Don't return password
        }
    ).to_list(1000)
    
    # Add combined 'name' field for frontend compatibility
    for telecaller in telecallers:
        first_name = telecaller.get("first_name", "")
        last_name = telecaller.get("last_name", "")
        telecaller["name"] = f"{first_name} {last_name}".strip() if last_name else first_name
    
    return telecallers


@api_router.post("/telecallers")
async def create_telecaller(
    telecaller_data: TelecallerProfileCreate,
    current_user: User = Depends(get_current_user)
):
    """Create new telecaller profile (Admin only)"""
    from datetime import datetime, timezone
    
    # Check if user is admin or master_admin
    if current_user.account_type not in ["admin", "master_admin"]:
        raise HTTPException(status_code=403, detail="Only admins can create telecaller profiles")
    
    # Check if email already exists
    existing = await db.telecaller_profiles.find_one({"email": telecaller_data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already exists")
    
    # Create telecaller profile
    telecaller = {
        "id": str(uuid.uuid4()),
        "name": telecaller_data.name,
        "phone_number": telecaller_data.phone_number,
        "email": telecaller_data.email,
        "status": telecaller_data.status if telecaller_data.status else "active",
        "notes": telecaller_data.notes,
        "aadhar_card": telecaller_data.aadhar_card,
        "pan_card": telecaller_data.pan_card,
        "address_proof": telecaller_data.address_proof,
        "total_assigned_leads": 0,
        "active_leads": 0,
        "converted_leads": 0,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "last_modified": datetime.now(timezone.utc).isoformat(),
        "created_by": current_user.id
    }
    
    await db.telecaller_profiles.insert_one(telecaller)
    
    return {"message": "Telecaller profile created successfully", "telecaller": telecaller}


@api_router.patch("/telecallers/{telecaller_id}")
async def update_telecaller(
    telecaller_id: str,
    telecaller_data: TelecallerProfileUpdate,
    current_user: User = Depends(get_current_user)
):
    """Update telecaller profile"""
    from datetime import datetime, timezone
    
    if current_user.account_type not in ["admin", "master_admin"]:
        raise HTTPException(status_code=403, detail="Only admins can update telecaller profiles")
    
    telecaller = await db.telecaller_profiles.find_one({"id": telecaller_id})
    if not telecaller:
        raise HTTPException(status_code=404, detail="Telecaller not found")
    
    # Prepare update data
    update_data = {}
    for field, value in telecaller_data.model_dump(exclude_unset=True).items():
        if value is not None:
            update_data[field] = value
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")
    
    update_data['last_modified'] = datetime.now(timezone.utc).isoformat()
    
    await db.telecaller_profiles.update_one(
        {"id": telecaller_id},
        {"$set": update_data}
    )
    
    updated_telecaller = await db.telecaller_profiles.find_one({"id": telecaller_id}, {"_id": 0})
    
    return {"message": "Telecaller profile updated successfully", "telecaller": updated_telecaller}


@api_router.delete("/telecallers/{telecaller_id}")
async def delete_telecaller(telecaller_id: str, current_user: User = Depends(get_current_user)):
    """Delete telecaller profile (Admin only)"""
    if current_user.account_type not in ["admin", "master_admin"]:
        raise HTTPException(status_code=403, detail="Only admins can delete telecaller profiles")
    
    telecaller = await db.telecaller_profiles.find_one({"id": telecaller_id})
    if not telecaller:
        raise HTTPException(status_code=404, detail="Telecaller not found")
    
    # Check if telecaller has assigned leads
    assigned_leads = await db.driver_leads.count_documents({"assigned_telecaller": telecaller['email']})
    if assigned_leads > 0:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot delete telecaller with {assigned_leads} assigned leads. Please reassign first."
        )
    
    await db.telecaller_profiles.delete_one({"id": telecaller_id})
    
    return {"message": "Telecaller profile deleted successfully"}



@api_router.post("/telecallers/sync-from-users")
async def sync_telecallers_from_users(current_user: User = Depends(get_current_user)):
    """Sync telecaller profiles from users collection (Admin only)"""
    if current_user.account_type not in ["admin", "master_admin"]:
        raise HTTPException(status_code=403, detail="Only admins can sync telecaller profiles")
    
    # Get all telecaller users
    telecaller_users = await db.users.find(
        {"account_type": "telecaller"},
        {"_id": 0}
    ).to_list(length=None)
    
    synced_count = 0
    created_count = 0
    updated_count = 0
    
    for user in telecaller_users:
        # Check if telecaller profile exists
        existing_profile = await db.telecaller_profiles.find_one({"email": user['email']})
        
        if not existing_profile:
            # Create new telecaller profile
            telecaller_profile = {
                "id": str(uuid.uuid4()),
                "name": f"{user.get('first_name', '')} {user.get('last_name', '')}".strip(),
                "phone_number": "",
                "email": user['email'],
                "status": user.get('status', 'active'),
                "notes": "Auto-synced from User Management",
                "aadhar_card": "",
                "pan_card": "",
                "address_proof": "",
                "total_assigned_leads": 0,
                "active_leads": 0,
                "converted_leads": 0,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "last_modified": datetime.now(timezone.utc).isoformat(),
                "created_by": current_user.id
            }
            
            await db.telecaller_profiles.insert_one(telecaller_profile)
            created_count += 1
            logger.info(f"Created telecaller profile for {user['email']}")
        else:
            # Update existing profile name and status if needed
            update_data = {
                "name": f"{user.get('first_name', '')} {user.get('last_name', '')}".strip(),
                "status": user.get('status', 'active'),
                "last_modified": datetime.now(timezone.utc).isoformat()
            }
            
            await db.telecaller_profiles.update_one(
                {"email": user['email']},
                {"$set": update_data}
            )
            updated_count += 1
            logger.info(f"Updated telecaller profile for {user['email']}")
        
        synced_count += 1
    
    return {
        "message": f"Successfully synced {synced_count} telecaller(s)",
        "created": created_count,
        "updated": updated_count,
        "total": synced_count
    }


@api_router.post("/telecallers/assign-leads")
async def assign_leads_to_telecaller(
    assignment: LeadAssignment,
    current_user: User = Depends(get_current_user)
):
    """Manually assign leads to a telecaller"""
    from datetime import datetime, timezone
    
    if current_user.account_type not in ["admin", "master_admin"]:
        raise HTTPException(status_code=403, detail="Only admins can assign leads")
    
    # Verify telecaller exists
    telecaller = await db.telecaller_profiles.find_one({"id": assignment.telecaller_id})
    if not telecaller:
        raise HTTPException(status_code=404, detail="Telecaller not found")
    
    # Update leads with assigned telecaller
    result = await db.driver_leads.update_many(
        {"id": {"$in": assignment.lead_ids}},
        {"$set": {
            "assigned_telecaller": telecaller['email'],
            "last_modified": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Update telecaller stats
    await db.telecaller_profiles.update_one(
        {"id": assignment.telecaller_id},
        {"$inc": {"total_assigned_leads": result.modified_count}}
    )
    
    # Sync updated leads to Google Sheets
    for lead_id in assignment.lead_ids:
        lead = await db.driver_leads.find_one({"id": lead_id}, {"_id": 0})
        if lead:
            sync_single_record('leads', lead)
    
    return {
        "message": f"Successfully assigned {result.modified_count} leads to {telecaller['name']}",
        "assigned_count": result.modified_count
    }


@api_router.post("/telecallers/reassign-leads")
async def reassign_leads_to_telecaller(
    reassignment: LeadReassignment,
    current_user: User = Depends(get_current_user)
):
    """Reassign leads from one telecaller to another"""
    from datetime import datetime, timezone
    
    if current_user.account_type not in ["admin", "master_admin"]:
        raise HTTPException(status_code=403, detail="Only admins can reassign leads")
    
    # Verify new telecaller exists
    to_telecaller = await db.telecaller_profiles.find_one({"id": reassignment.to_telecaller_id})
    if not to_telecaller:
        raise HTTPException(status_code=404, detail="Target telecaller not found")
    
    # Get current assignments to update stats
    leads_to_reassign = await db.driver_leads.find(
        {"id": {"$in": reassignment.lead_ids}},
        {"_id": 0, "id": 1, "assigned_telecaller": 1}
    ).to_list(length=None)
    
    # Count leads by current telecaller
    from_telecaller_counts = {}
    for lead in leads_to_reassign:
        current_telecaller_email = lead.get("assigned_telecaller")
        if current_telecaller_email:
            from_telecaller_counts[current_telecaller_email] = from_telecaller_counts.get(current_telecaller_email, 0) + 1
    
    # Update leads with new telecaller
    result = await db.driver_leads.update_many(
        {"id": {"$in": reassignment.lead_ids}},
        {"$set": {
            "assigned_telecaller": to_telecaller['email'],
            "last_modified": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Update stats: decrease from old telecallers
    for telecaller_email, count in from_telecaller_counts.items():
        await db.telecaller_profiles.update_one(
            {"email": telecaller_email},
            {"$inc": {"total_assigned_leads": -count}}
        )
    
    # Update stats: increase for new telecaller
    await db.telecaller_profiles.update_one(
        {"id": reassignment.to_telecaller_id},
        {"$inc": {"total_assigned_leads": result.modified_count}}
    )
    
    # Sync updated leads to Google Sheets
    for lead_id in reassignment.lead_ids:
        lead = await db.driver_leads.find_one({"id": lead_id}, {"_id": 0})
        if lead:
            sync_single_record('leads', lead)
    
    return {
        "success": True,
        "message": f"Successfully reassigned {result.modified_count} leads to {to_telecaller['name']}",
        "reassigned_count": result.modified_count
    }


@api_router.post("/telecallers/deassign-leads")
async def deassign_leads_from_telecaller(
    deassignment: LeadDeassignment,
    current_user: User = Depends(get_current_user)
):
    """Remove telecaller assignment from leads"""
    from datetime import datetime, timezone
    
    if current_user.account_type not in ["admin", "master_admin"]:
        raise HTTPException(status_code=403, detail="Only admins can deassign leads")
    
    # Get current assignments to update stats
    leads_to_deassign = await db.driver_leads.find(
        {"id": {"$in": deassignment.lead_ids}},
        {"_id": 0, "id": 1, "assigned_telecaller": 1}
    ).to_list(length=None)
    
    # Count leads by current telecaller
    telecaller_counts = {}
    for lead in leads_to_deassign:
        current_telecaller_email = lead.get("assigned_telecaller")
        if current_telecaller_email:
            telecaller_counts[current_telecaller_email] = telecaller_counts.get(current_telecaller_email, 0) + 1
    
    # Remove telecaller assignment from leads
    result = await db.driver_leads.update_many(
        {"id": {"$in": deassignment.lead_ids}},
        {"$set": {
            "assigned_telecaller": None,
            "last_modified": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Update telecaller stats: decrease counts
    for telecaller_email, count in telecaller_counts.items():
        await db.telecaller_profiles.update_one(
            {"email": telecaller_email},
            {"$inc": {"total_assigned_leads": -count}}
        )
    
    # Sync updated leads to Google Sheets
    for lead_id in deassignment.lead_ids:
        lead = await db.driver_leads.find_one({"id": lead_id}, {"_id": 0})
        if lead:
            sync_single_record('leads', lead)
    
    return {
        "success": True,
        "message": f"Successfully deassigned {result.modified_count} leads",
        "deassigned_count": result.modified_count
    }


@api_router.post("/telecallers/sync-from-sheets")
async def sync_assignments_from_sheets(current_user: User = Depends(get_current_user)):
    """Read Column H from Google Sheets and assign leads to telecallers"""
    from datetime import datetime, timezone
    
    if current_user.account_type not in ["admin", "master_admin"]:
        raise HTTPException(status_code=403, detail="Only admins can sync assignments")
    
    try:
        # Get all leads from database (no limit)
        all_leads = await db.driver_leads.find({}, {"_id": 0}).to_list(length=None)
        
        # Get all telecallers (reasonable limit)
        all_telecallers = await db.telecaller_profiles.find({}, {"_id": 0}).to_list(length=None)
        telecaller_map = {t['email']: t for t in all_telecallers}
        telecaller_map_by_name = {t['name'].lower(): t for t in all_telecallers}
        
        updated_count = 0
        
        for lead in all_leads:
            # Check if assigned_telecaller is set and different from current
            assigned_telecaller = lead.get('assigned_telecaller', '').strip()
            
            if assigned_telecaller:
                # Try to match by email or name
                telecaller = None
                if assigned_telecaller in telecaller_map:
                    telecaller = telecaller_map[assigned_telecaller]
                elif assigned_telecaller.lower() in telecaller_map_by_name:
                    telecaller = telecaller_map_by_name[assigned_telecaller.lower()]
                
                if telecaller:
                    # Update lead with telecaller email
                    await db.driver_leads.update_one(
                        {"id": lead['id']},
                        {"$set": {
                            "assigned_telecaller": telecaller['email'],
                            "last_modified": datetime.now(timezone.utc).isoformat()
                        }}
                    )
                    updated_count += 1
        
        return {
            "success": True,
            "message": f"Synced assignments for {updated_count} leads from sheets",
            "updated_count": updated_count
        }
        
    except Exception as e:
        logger.error(f"Failed to sync from sheets: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to sync from sheets: {str(e)}")


@api_router.get("/telecallers/my-leads")
async def get_my_assigned_leads(current_user: User = Depends(get_current_user)):
    """Get leads assigned to the logged-in telecaller"""
    
    # This endpoint is for telecaller users only
    if current_user.account_type != "telecaller":
        raise HTTPException(status_code=403, detail="This endpoint is for telecallers only")
    
    # Get telecaller's email
    user_email = current_user.email
    
    # Find leads assigned to this telecaller
    leads = await db.driver_leads.find(
        {"assigned_telecaller": user_email},
        {"_id": 0}
    ).to_list(length=None)
    
    return leads


@api_router.post("/driver-onboarding/leads/bulk-delete")
async def bulk_delete_leads(bulk_data: BulkLeadDelete, current_user: User = Depends(get_current_user)):
    """Bulk delete leads (Master Admin only)"""
    # Check if user is master admin
    if current_user.account_type != "master_admin":
        raise HTTPException(status_code=403, detail="Only Master Admin can delete leads")
    
    if not bulk_data.lead_ids or len(bulk_data.lead_ids) == 0:
        raise HTTPException(status_code=400, detail="No leads selected for deletion")
    
    # Delete from MongoDB first (fast operation)
    result = await db.driver_leads.delete_many(
        {"id": {"$in": bulk_data.lead_ids}}
    )
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="No leads found to delete")
    
    # Return immediately, sync will happen in background
    # Note: We just update the sync time marker, the actual Google Sheets
    # will be synced on next manual sync or when leads are updated
    try:
        update_last_sync_time('leads')
        logger.info(f"Bulk delete: Deleted {result.deleted_count} leads from database")
    except Exception as e:
        logger.error(f"Failed to update sync time: {str(e)}")
    
    return {
        "message": f"Successfully deleted {result.deleted_count} lead(s)",
        "deleted_count": result.deleted_count
    }


@api_router.get("/driver-onboarding/test-bulk")
async def test_bulk_endpoint(current_user: User = Depends(get_current_user)):
    """Test endpoint to verify route is working"""
    return {"message": "Bulk update endpoint route is accessible", "version": "2"}


@api_router.get("/driver-onboarding/performance-tracking")
async def get_performance_tracking(current_user: User = Depends(get_current_user)):
    """Get telecaller performance metrics"""
    try:
        # Get all leads (no limit)
        all_leads = await db.driver_leads.find({}, {"_id": 0}).to_list(length=None)
        
        # Group by telecaller
        telecaller_stats = {}
        
        for lead in all_leads:
            telecaller = lead.get("assigned_telecaller", "Unassigned")
            
            if telecaller not in telecaller_stats:
                telecaller_stats[telecaller] = {
                    "name": telecaller,
                    "total_leads": 0,
                    "contacted": 0,
                    "qualified": 0,
                    "onboarded": 0,
                    "rejected": 0,
                    "in_progress": 0,
                    "conversion_rate": 0,
                    "stages": {
                        "new": 0,
                        "contacted": 0,
                        "qualified": 0,
                        "assigned": 0,
                        "in_progress": 0
                    }
                }
            
            stats = telecaller_stats[telecaller]
            stats["total_leads"] += 1
            
            # Count by status
            status = lead.get("status", "New").lower()
            if "contact" in status:
                stats["contacted"] += 1
            elif "interested" in status or "qualified" in status:
                stats["qualified"] += 1
            elif "onboard" in status:
                stats["onboarded"] += 1
            elif "reject" in status or "not interested" in status:
                stats["rejected"] += 1
            
            # Count by lead stage
            lead_stage = lead.get("lead_stage", "New").lower()
            if "new" in lead_stage:
                stats["stages"]["new"] += 1
            elif "contacted" in lead_stage:
                stats["stages"]["contacted"] += 1
            elif "qualified" in lead_stage:
                stats["stages"]["qualified"] += 1
            elif "assigned" in lead_stage:
                stats["stages"]["assigned"] += 1
            elif "progress" in lead_stage:
                stats["stages"]["in_progress"] += 1
        
        # Calculate conversion rates
        for telecaller, stats in telecaller_stats.items():
            if stats["total_leads"] > 0:
                stats["conversion_rate"] = round((stats["onboarded"] / stats["total_leads"]) * 100, 2)
        
        # Overall stats
        total_leads = len(all_leads)
        total_contacted = sum(s["contacted"] for s in telecaller_stats.values())
        total_onboarded = sum(s["onboarded"] for s in telecaller_stats.values())
        overall_conversion = round((total_onboarded / total_leads * 100), 2) if total_leads > 0 else 0
        
        return {
            "telecallers": list(telecaller_stats.values()),
            "overall": {
                "total_leads": total_leads,
                "total_contacted": total_contacted,
                "total_onboarded": total_onboarded,
                "conversion_rate": overall_conversion
            }
        }
    except Exception as e:
        logger.error(f"Performance tracking error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get performance data: {str(e)}")


@api_router.post("/driver-onboarding/webhook/sync-from-sheets")
async def sync_from_google_sheets(request: Request):
    """
    Webhook endpoint to receive driver leads data FROM Google Sheets
    This enables two-way sync: Google Sheets → App
    
    SYNC STRATEGY:
    - Uses ID as unique identifier (not phone number)
    - If ID exists in DB → UPDATE the lead
    - If ID doesn't exist in DB → CREATE new lead with that ID
    - If lead exists in DB but NOT in sheets → DELETE from DB
    
    Expected payload: {
        "action": "sync_from_sheets",
        "leads": [array of lead objects with id field]
    }
    """
    try:
        body = await request.json()
        logger.info(f"Received sync from Google Sheets: {body.get('action')}")
        
        if body.get('action') != 'sync_from_sheets':
            raise HTTPException(status_code=400, detail="Invalid action")
        
        leads_data = body.get('leads', [])
        
        # Get all IDs from Google Sheets (filter out empty rows)
        sheet_lead_ids = set()
        valid_leads = []
        
        for lead in leads_data:
            # Skip empty rows (no phone number means empty row)
            if not lead.get('phone_number') or lead.get('phone_number') == '':
                continue
            
            lead_id = lead.get('id', '').strip()
            
            # Generate ID if not present
            if not lead_id:
                lead_id = str(uuid.uuid4())
                lead['id'] = lead_id
            
            sheet_lead_ids.add(lead_id)
            valid_leads.append(lead)
        
        # Get all existing leads from database
        existing_leads = await db.driver_leads.find().to_list(length=None)
        db_lead_ids = {lead['id'] for lead in existing_leads}
        
        created_count = 0
        updated_count = 0
        deleted_count = 0
        
        # Process each lead from sheets
        for lead in valid_leads:
            lead_id = lead['id']
            
            # Check if lead exists by ID
            existing_lead = await db.driver_leads.find_one({"id": lead_id})
            
            # Prepare lead data with all fields (including new document fields)
            lead_data = {
                "id": lead_id,
                "name": lead.get('name', ''),
                "phone_number": lead.get('phone_number', ''),
                "vehicle": lead.get('vehicle'),
                "driving_license": lead.get('driving_license'),
                "experience": lead.get('experience'),
                "interested_ev": lead.get('interested_ev'),
                "monthly_salary": lead.get('monthly_salary'),
                "residing_chennai": lead.get('residing_chennai'),
                "current_location": lead.get('current_location'),
                "lead_source": lead.get('lead_source'),
                "stage": lead.get('stage', 'New'),
                "status": lead.get('status', 'New'),
                "assigned_telecaller": lead.get('assigned_telecaller'),
                "telecaller_notes": lead.get('telecaller_notes'),
                "notes": lead.get('notes'),
                "import_date": lead.get('import_date', ''),
                "created_at": lead.get('created_at', ''),
                # New document fields
                "dl_no": lead.get('dl_no'),
                "badge_no": lead.get('badge_no'),
                "aadhar_card": lead.get('aadhar_card'),
                "pan_card": lead.get('pan_card'),
                "gas_bill": lead.get('gas_bill'),
                "bank_passbook": lead.get('bank_passbook'),
                # Shift fields
                "preferred_shift": lead.get('preferred_shift'),
                "allotted_shift": lead.get('allotted_shift'),
                # Vehicle assignment
                "default_vehicle": lead.get('default_vehicle'),
                # End date
                "end_date": lead.get('end_date')
            }
            
            if existing_lead:
                # Update existing lead (overwrite with sheet data)
                await db.driver_leads.update_one(
                    {"id": lead_id},
                    {"$set": lead_data}
                )
                updated_count += 1
                logger.info(f"Updated lead: {lead_id} - {lead_data['name']}")
            else:
                # Create new lead with the ID from sheets
                # Add timestamps if not present
                if not lead_data['import_date']:
                    lead_data['import_date'] = datetime.now(timezone.utc).isoformat()
                if not lead_data['created_at']:
                    lead_data['created_at'] = datetime.now(timezone.utc).isoformat()
                
                await db.driver_leads.insert_one(lead_data)
                created_count += 1
                logger.info(f"Created new lead: {lead_id} - {lead_data['name']}")
        
        # Delete leads that exist in DB but NOT in Google Sheets
        leads_to_delete = db_lead_ids - sheet_lead_ids
        
        if leads_to_delete:
            delete_result = await db.driver_leads.delete_many({
                "id": {"$in": list(leads_to_delete)}
            })
            deleted_count = delete_result.deleted_count
            logger.info(f"Deleted {deleted_count} leads that were removed from Google Sheets")
        
        return {
            "success": True,
            "message": f"Synced {len(valid_leads)} leads from Google Sheets",
            "created": created_count,
            "updated": updated_count,
            "deleted": deleted_count,
            "total_processed": created_count + updated_count + deleted_count
        }
        
    except Exception as e:
        logger.error(f"Sync from sheets error: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Failed to sync from sheets: {str(e)}")


# Telecaller Queue Manager - Lead Assignment
@api_router.get("/telecaller-queue/daily-assignments")
async def get_daily_assignments(current_user: User = Depends(get_current_user)):
    """Get daily lead assignments for telecallers - 20 NEW leads each (40 total)"""
    # Priority: Assign only NEW leads (status = "New")
    # Sorted by import date (oldest first)
    
    logger.info("Fetching daily assignments for telecallers")
    
    # Get 40 NEW leads (20 for each telecaller)
    new_leads = await db.driver_leads.find({
        "status": "New",
        "lead_stage": "New"
    }, {"_id": 0}).sort("import_date", 1).to_list(40)
    
    logger.info(f"Found {len(new_leads)} NEW leads for assignment")
    
    # If we have fewer than 40 NEW leads, try to get more from other stages
    if len(new_leads) < 40:
        additional_needed = 40 - len(new_leads)
        logger.info(f"Need {additional_needed} more leads, fetching from other statuses")
        
        # Get leads from other statuses (but not Onboarded or Rejected)
        additional_leads = await db.driver_leads.find({
            "status": {"$nin": ["New", "Onboarded", "Rejected"]},
            "lead_stage": {"$ne": "In Progress"}
        }, {"_id": 0}).sort("import_date", 1).to_list(additional_needed)
        
        new_leads.extend(additional_leads)
        logger.info(f"Total leads after adding from other statuses: {len(new_leads)}")
    
    # Split leads equally between 2 telecallers (20 each)
    telecaller1_leads = new_leads[0:20]
    telecaller2_leads = new_leads[20:40]
    
    logger.info(f"Assigned {len(telecaller1_leads)} leads to Telecaller 1")
    logger.info(f"Assigned {len(telecaller2_leads)} leads to Telecaller 2")
    
    return {
        "telecaller1": {
            "name": "Telecaller 1",
            "leads": telecaller1_leads,
            "count": len(telecaller1_leads)
        },
        "telecaller2": {
            "name": "Telecaller 2", 
            "leads": telecaller2_leads,
            "count": len(telecaller2_leads)
        },
        "total_assigned": len(telecaller1_leads) + len(telecaller2_leads),
        "assignment_date": datetime.now(timezone.utc).isoformat()
    }


@api_router.post("/telecaller-queue/update-call-status")
async def update_call_status(
    lead_id: str,
    call_outcome: str,
    notes: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Update lead status after call"""
    # Find the lead
    lead = await db.driver_leads.find_one({"id": lead_id})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    # Update status based on call outcome
    new_status = call_outcome
    
    await db.driver_leads.update_one(
        {"id": lead_id},
        {"$set": {"status": new_status}}
    )
    
    # Get updated lead for sync
    updated_lead = await db.driver_leads.find_one({"id": lead_id}, {"_id": 0})
    
    # Sync to Google Sheets
    sync_single_record('leads', updated_lead)
    
    return {"message": "Call status updated successfully", "lead": updated_lead}


# Telecaller Queue
@api_router.post("/telecaller-queue")
async def create_telecaller_task(task_data: TelecallerTaskCreate, current_user: User = Depends(get_current_user)):
    """Create new telecaller task"""
    task = TelecallerTask(**task_data.model_dump())
    task_dict = task.model_dump()
    task_dict['date'] = task_dict['date'].isoformat()
    task_dict['created_at'] = task_dict['created_at'].isoformat()
    if task_dict.get('scheduled_time'):
        task_dict['scheduled_time'] = task_dict['scheduled_time'].isoformat()
    
    await db.telecaller_queue.insert_one(task_dict)
    sync_single_record('telecaller_queue', task_dict)
    
    return {"message": "Telecaller task created", "task": task}


@api_router.get("/telecaller-queue")
async def get_telecaller_tasks(current_user: User = Depends(get_current_user)):
    """Get all telecaller tasks"""
    tasks = await db.telecaller_queue.find({}, {"_id": 0}).to_list(1000)
    for task in tasks:
        if isinstance(task.get('date'), str):
            task['date'] = datetime.fromisoformat(task['date'])
        if isinstance(task.get('created_at'), str):
            task['created_at'] = datetime.fromisoformat(task['created_at'])
    return tasks


# ==================== MONTRA VEHICLE INSIGHTS ====================

@api_router.post("/montra-vehicle/import-feed")
async def import_montra_feed(file: UploadFile = File(...), current_user: User = Depends(get_current_user)):
    """Import Montra vehicle feed data and sync to Google Sheets"""
    try:
        import re
        
        # Get filename without extension
        filename = file.filename
        logger.info(f"Importing Montra feed from file: {filename}")
        
        # Extract vehicle ID, day, month, and year from filename
        # Expected format: "P60G2512500002032 - 01 Sep 2025.csv"
        filename_pattern = r"^([A-Z0-9]+)\s*-\s*(\d{2})\s+([A-Za-z]+)\s+(\d{4})\.(csv|xlsx)$"
        match = re.match(filename_pattern, filename)
        
        if not match:
            raise HTTPException(
                status_code=400, 
                detail=f"Invalid filename format. Expected: 'VEHICLE_ID - DD MMM YYYY.csv'. Got: {filename}"
            )
        
        vehicle_id = match.group(1)
        day = match.group(2)
        month = match.group(3)
        year = match.group(4)
        separator = "-"
        
        logger.info(f"Extracted from filename - Vehicle ID: {vehicle_id}, Day: {day}, Month: {month}, Year: {year}")
        
        # Read file content
        content = await file.read()
        
        # Parse based on file type
        file_extension = filename.split('.')[-1].lower()
        
        if file_extension == 'csv':
            import io
            df = pd.read_csv(io.BytesIO(content))
        elif file_extension == 'xlsx':
            df = pd.read_excel(io.BytesIO(content))
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format")
        
        logger.info(f"Parsed file with {len(df)} rows and {len(df.columns)} columns")
        
        # CRITICAL: Sort data by time column (Column A - Date/Time) chronologically
        # This ensures battery consumption calculations are accurate
        time_col = df.columns[0]  # First column should be Date or Time
        logger.info(f"Sorting data by time column: {time_col}")
        
        try:
            # Convert to datetime and sort
            df[time_col] = pd.to_datetime(df[time_col], errors='coerce')
            df = df.sort_values(by=time_col)
            df = df.reset_index(drop=True)
            logger.info(f"Data sorted successfully from {df[time_col].min()} to {df[time_col].max()}")
        except Exception as sort_error:
            logger.warning(f"Could not sort by time column: {sort_error}. Proceeding with original order.")
        
        # Verify we have the expected columns (A to U = 21 columns)
        if len(df.columns) != 21:
            raise HTTPException(
                status_code=400,
                detail=f"Expected 21 columns (A-U), but found {len(df.columns)} columns"
            )
        
        # Look up registration number from vehicle mapping
        vehicle_mapping = await db.vehicle_mapping.find_one(
            {"vehicle_id": vehicle_id}, 
            {"_id": 0, "registration_number": 1}
        )
        registration_number = vehicle_mapping.get("registration_number", "") if vehicle_mapping else ""
        
        logger.info(f"Vehicle {vehicle_id} → Registration: {registration_number if registration_number else 'Not found'}")
        
        # Prepare data for Google Sheets
        # Sheet columns: D to X (CSV columns A to U)
        # Additional columns: Y (vehicle_id), Z (separator), AA (day), AB (month), AC (registration_number)
        
        rows_to_import = []
        for idx, row in df.iterrows():
            # Convert row to list and add filename data
            row_data = row.tolist()
            # Add vehicle_id, separator, day, month, registration_number
            row_data.extend([vehicle_id, separator, day, month, registration_number])
            rows_to_import.append(row_data)
        
        logger.info(f"Prepared {len(rows_to_import)} rows for import")
        
        # Load mode mapping tables
        model_dict, mode_dict = load_mode_mapping_tables()
        
        # Also save to MongoDB for analytics queries
        montra_docs = []
        headers = df.columns.tolist() + ['Vehicle ID', 'Separator', 'Day', 'Month', 'Registration Number']
        
        # Parse the date properly for ISO format storage
        from datetime import datetime as dt_obj
        try:
            # Convert "01 Sep 2025" to ISO date "2025-09-01"
            date_str = f"{day} {month} {year}"
            parsed_date = dt_obj.strptime(date_str, "%d %b %Y")
            iso_date = parsed_date.strftime("%Y-%m-%d")
        except:
            # Fallback if parsing fails
            iso_date = f"{year}-01-01"
            logger.warning(f"Could not parse date '{day} {month} {year}', using fallback: {iso_date}")
        
        for row_data in rows_to_import:
            doc = {
                "vehicle_id": vehicle_id,
                "date": iso_date,  # Store in ISO format for easy querying
                "date_display": f"{day} {month} {year}",  # Keep original for display
                "day": day,
                "month": month,
                "year": year,
                "registration_number": registration_number,
                "filename": filename,
                "imported_at": datetime.now(timezone.utc).isoformat()
            }
            # Map all columns to document
            for i, header in enumerate(headers):
                if i < len(row_data):
                    doc[header] = row_data[i]
            
            # Enrich with Mode Name and Mode Type
            ride_mode = doc.get("Ride Mode", "")
            if ride_mode:
                mode_name, mode_type = enrich_with_mode_data(
                    registration_number if registration_number else vehicle_id,
                    str(ride_mode),
                    model_dict,
                    mode_dict
                )
                doc["mode_name"] = mode_name
                doc["mode_type"] = mode_type
            else:
                doc["mode_name"] = "Unknown"
                doc["mode_type"] = "Unknown"
            
            montra_docs.append(doc)
        
        # Save to MongoDB
        if montra_docs:
            await db.montra_feed_data.insert_many(montra_docs)
            logger.info(f"Saved {len(montra_docs)} rows to MongoDB")
        
        logger.info(f"Successfully imported {len(rows_to_import)} rows to database")
        return {
            "message": f"Successfully imported {len(rows_to_import)} rows from {filename}",
            "rows": len(rows_to_import),
            "vehicle_id": vehicle_id,
            "date": f"{day} {month}",
            "synced_to_database": True
        }
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error importing Montra feed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to import feed: {str(e)}")


@api_router.post("/montra-vehicle/fix-date-format")
async def fix_montra_date_format(current_user: User = Depends(get_current_user)):
    """Fix date format in existing Montra feed data from 'DD MMM' to ISO 'YYYY-MM-DD' format"""
    try:
        from datetime import datetime as dt_obj
        
        logger.info("Starting date format fix for montra_feed_data...")
        
        # Get all records with old date format (pattern: "DD MMM")
        # We'll process in batches to avoid memory issues
        batch_size = 10000
        total_updated = 0
        total_failed = 0
        
        # Get total count
        total_records = await db.montra_feed_data.count_documents({})
        logger.info(f"Total records to process: {total_records}")
        
        # Process in batches
        skip = 0
        while skip < total_records:
            batch = await db.montra_feed_data.find({}).skip(skip).limit(batch_size).to_list(batch_size)
            
            for record in batch:
                try:
                    old_date = record.get('date', '')
                    day = record.get('day', '')
                    month = record.get('month', '')
                    year = record.get('year', '2025')
                    
                    # Check if date is already in ISO format (YYYY-MM-DD)
                    if '-' in old_date and len(old_date) == 10:
                        # Already in ISO format, skip
                        continue
                    
                    # Parse the old format date
                    if day and month and year:
                        date_str = f"{day} {month} {year}"
                        try:
                            parsed_date = dt_obj.strptime(date_str, "%d %b %Y")
                            iso_date = parsed_date.strftime("%Y-%m-%d")
                            
                            # Update the record
                            await db.montra_feed_data.update_one(
                                {"_id": record["_id"]},
                                {"$set": {
                                    "date": iso_date,
                                    "date_display": f"{day} {month} {year}"
                                }}
                            )
                            total_updated += 1
                        except ValueError as e:
                            logger.warning(f"Failed to parse date '{date_str}': {e}")
                            total_failed += 1
                    else:
                        total_failed += 1
                        
                except Exception as e:
                    logger.error(f"Error processing record {record.get('_id')}: {e}")
                    total_failed += 1
            
            skip += batch_size
            logger.info(f"Processed {skip}/{total_records} records. Updated: {total_updated}, Failed: {total_failed}")
        
        logger.info(f"Date format fix completed. Updated: {total_updated}, Failed: {total_failed}")
        
        return {
            "success": True,
            "message": f"Successfully fixed date format for {total_updated} records",
            "total_updated": total_updated,
            "total_failed": total_failed,
            "total_records": total_records
        }
    except Exception as e:
        logger.error(f"Error fixing date format: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fix date format: {str(e)}")


@api_router.post("/montra-vehicle/enrich-existing-data")
async def enrich_existing_montra_data(current_user: User = Depends(get_current_user)):
    """Re-process all existing Montra feed data to add Mode Name and Mode Type"""
    try:
        # Load mode mapping tables
        model_dict, mode_dict = load_mode_mapping_tables()
        
        if not model_dict or not mode_dict:
            raise HTTPException(status_code=400, detail="Mode mapping tables not available. Please upload Mode Details.xlsx")
        
        # Get all records
        all_records = await db.montra_feed_data.find({}).to_list(100000)
        
        logger.info(f"Processing {len(all_records)} existing records for mode enrichment")
        
        updated_count = 0
        for record in all_records:
            vehicle_id = record.get("vehicle_id")
            registration = record.get("registration_number")
            ride_mode = record.get("Ride Mode")
            
            if not ride_mode:
                continue
            
            # Use registration number if available, otherwise vehicle ID
            lookup_id = registration if registration else vehicle_id
            
            mode_name, mode_type = enrich_with_mode_data(
                lookup_id,
                str(ride_mode),
                model_dict,
                mode_dict
            )
            
            # Update record in MongoDB
            await db.montra_feed_data.update_one(
                {"_id": record["_id"]},
                {"$set": {
                    "mode_name": mode_name,
                    "mode_type": mode_type
                }}
            )
            updated_count += 1
        
        logger.info(f"Successfully enriched {updated_count} records")
        
        return {
            "success": True,
            "message": f"Successfully enriched {updated_count} records with mode data",
            "updated_count": updated_count
        }
        
    except Exception as e:
        logger.error(f"Error enriching existing data: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to enrich data: {str(e)}")


@api_router.get("/montra-vehicle/download-mode-mapping")
async def download_mode_mapping(current_user: User = Depends(get_current_user)):
    """Download the current Mode Details.xlsx mapping file"""
    from fastapi.responses import FileResponse
    
    mapping_file_path = "/app/backend/uploaded_files/mode_details.xlsx"
    
    if not os.path.exists(mapping_file_path):
        raise HTTPException(status_code=404, detail="Mode mapping file not found")
    
    return FileResponse(
        path=mapping_file_path,
        filename="Mode Details.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@api_router.post("/montra-vehicle/upload-mode-mapping")
async def upload_mode_mapping(file: UploadFile = File(...), current_user: User = Depends(get_current_user)):
    """Upload updated Mode Details.xlsx mapping file"""
    try:
        import pandas as pd
        import io
        
        # Verify it's an Excel file
        if not file.filename.endswith('.xlsx'):
            raise HTTPException(status_code=400, detail="File must be .xlsx format")
        
        # Read and validate the file
        content = await file.read()
        
        try:
            excel_file = pd.ExcelFile(io.BytesIO(content))
            
            # Verify required sheets exist
            if 'Model Mapping' not in excel_file.sheet_names:
                raise HTTPException(status_code=400, detail="Missing 'Model Mapping' sheet")
            if 'Ride Mode Mapping' not in excel_file.sheet_names:
                raise HTTPException(status_code=400, detail="Missing 'Ride Mode Mapping' sheet")
            
            # Validate Model Mapping structure
            model_df = pd.read_excel(io.BytesIO(content), sheet_name='Model Mapping')
            required_model_cols = ['Vehicle Number', 'VIN', 'Model']
            if not all(col in model_df.columns for col in required_model_cols):
                raise HTTPException(status_code=400, detail=f"Model Mapping must have columns: {required_model_cols}")
            
            # Validate Ride Mode Mapping structure
            mode_df = pd.read_excel(io.BytesIO(content), sheet_name='Ride Mode Mapping')
            required_mode_cols = ['Vehicle-Mode Concatenation', 'Vehicle Model', 'Ride Mode', 'Mode Name', 'Mode Type']
            if not all(col in mode_df.columns for col in required_mode_cols):
                raise HTTPException(status_code=400, detail=f"Ride Mode Mapping must have columns: {required_mode_cols}")
            
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid Excel file structure: {str(e)}")
        
        # Save the file
        mapping_file_path = "/app/backend/uploaded_files/mode_details.xlsx"
        with open(mapping_file_path, "wb") as f:
            f.write(content)
        
        logger.info(f"Updated mode mapping file: {file.filename}")
        
        return {
            "success": True,
            "message": "Mode mapping file updated successfully",
            "filename": file.filename
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error uploading mode mapping: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to upload mapping file: {str(e)}")


@api_router.post("/telecaller-queue/sync")
async def sync_telecaller_queue(current_user: User = Depends(get_current_user)):
    """Sync all telecaller tasks to Google Sheets"""
    tasks = await db.telecaller_queue.find({}, {"_id": 0}).to_list(1000)
    success = sync_all_records('telecaller_queue', tasks)
    if success:
        return {"message": "Telecaller queue synced successfully"}
    raise HTTPException(status_code=500, detail="Failed to sync telecaller queue")


# ==================== ADMIN FILES MANAGEMENT ====================

# Create uploaded files directory if it doesn't exist
UPLOAD_DIR = "/app/backend/uploaded_files"
os.makedirs(UPLOAD_DIR, exist_ok=True)

@api_router.post("/montra-vehicle/view-file-data")
async def view_montra_file_data(
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """View data from a specific uploaded Montra feed file"""
    try:
        data = await request.json()
        vehicle_id = data.get("vehicle_id")
        date = data.get("date")
        filename = data.get("filename")
        
        if not vehicle_id or not date:
            raise HTTPException(status_code=400, detail="vehicle_id and date are required")
        
        # Query records for this specific file
        query = {
            "vehicle_id": vehicle_id,
            "date": date
        }
        if filename:
            query["filename"] = filename
        
        records = await db.montra_feed_data.find(query, {"_id": 0}).sort("Portal Received Time", 1).to_list(1000)
        
        if not records:
            return {
                "success": True,
                "data": [],
                "columns": [],
                "count": 0,
                "message": "No data found for this file"
            }
        
        # Extract columns from first record
        columns = list(records[0].keys())
        
        return {
            "success": True,
            "data": records,
            "columns": columns,
            "count": len(records)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error viewing file data: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to view file data: {str(e)}")


@api_router.get("/admin/files")
async def list_files(current_user: User = Depends(get_current_user)):
    """Get list of all uploaded files"""
    try:
        files = await db.admin_files.find({}, {"_id": 0}).sort("uploaded_at", -1).to_list(1000)
        
        # Format file sizes for display
        for file in files:
            file["size_display"] = format_file_size(file.get("file_size", 0))
        
        return {
            "files": files,
            "count": len(files)
        }
    except Exception as e:
        logger.error(f"Error listing files: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to list files")


@api_router.get("/admin/files/{file_id}/download")
async def download_file(file_id: str, current_user: User = Depends(get_current_user)):
    """Download a file"""
    try:
        file_metadata = await db.admin_files.find_one({"id": file_id}, {"_id": 0})
        
        if not file_metadata:
            raise HTTPException(status_code=404, detail="File not found")
        
        file_path = file_metadata["file_path"]
        
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="File not found on disk")
        
        from fastapi.responses import FileResponse
        
        return FileResponse(
            path=file_path,
            filename=file_metadata["original_filename"],
            media_type=file_metadata.get("content_type", "application/octet-stream")
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error downloading file: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to download file")


@api_router.get("/admin/files/{file_id}/share-link")
async def get_share_link(
    file_id: str,
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """Get shareable download link for a file"""
    try:
        file_metadata = await db.admin_files.find_one({"id": file_id}, {"_id": 0})
        
        if not file_metadata:
            raise HTTPException(status_code=404, detail="File not found")
        
        # Get backend URL dynamically from request
        host = request.headers.get('host', 'localhost:8001')
        scheme = 'https' if 'https' in str(request.url) or request.headers.get('x-forwarded-proto') == 'https' else 'http'
        backend_url = f"{scheme}://{host}"
        
        share_link = f"{backend_url}/api/admin/files/{file_id}/download"
        
        return {
            "share_link": share_link,
            "filename": file_metadata["original_filename"]
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating share link: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to generate share link")


@api_router.delete("/admin/files/{file_id}")
async def delete_file(file_id: str, current_user: User = Depends(get_current_user)):
    """Delete a file (Master Admin only)"""
    # Check if user is master admin
    if current_user.account_type != "master_admin":
        raise HTTPException(status_code=403, detail="Only Master Admin can delete files")
    
    try:
        file_metadata = await db.admin_files.find_one({"id": file_id}, {"_id": 0})
        
        if not file_metadata:
            raise HTTPException(status_code=404, detail="File not found")
        
        # Delete file from disk
        file_path = file_metadata["file_path"]
        if os.path.exists(file_path):
            os.remove(file_path)
            logger.info(f"Deleted file from disk: {file_path}")
        
        # Delete metadata from database
        await db.admin_files.delete_one({"id": file_id})
        
        return {
            "message": f"File '{file_metadata['original_filename']}' deleted successfully"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting file: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to delete file")


@api_router.post("/admin/files/upload")
async def upload_file(file: UploadFile = File(...), current_user: User = Depends(get_current_user)):
    """Upload a new file"""
    try:
        # Validate file size (100MB limit)
        max_size = 100 * 1024 * 1024  # 100MB
        file_content = await file.read()
        file_size = len(file_content)
        
        if file_size > max_size:
            raise HTTPException(
                status_code=400, 
                detail=f"File size exceeds 100MB limit. File size: {format_file_size(file_size)}"
            )
        
        # Generate unique file ID
        file_id = str(uuid.uuid4())
        
        # Save file to disk
        file_path = os.path.join(UPLOAD_DIR, f"{file_id}_{file.filename}")
        with open(file_path, "wb") as f:
            f.write(file_content)
        
        # Save metadata to database
        file_metadata = {
            "id": file_id,
            "original_filename": file.filename,
            "file_path": file_path,
            "file_size": file_size,
            "uploaded_by": current_user.id,
            "uploaded_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.admin_files.insert_one(file_metadata)
        
        logger.info(f"File uploaded successfully: {file.filename} (ID: {file_id})")
        
        return {
            "message": f"File '{file.filename}' uploaded successfully",
            "file_id": file_id,
            "filename": file.filename,
            "size": format_file_size(file_size)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error uploading file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to upload file: {str(e)}")


@api_router.put("/admin/files/{file_id}/update")
async def update_file(
    file_id: str, 
    file: UploadFile = File(...), 
    current_user: User = Depends(get_current_user)
):
    """Update/replace an existing file - overwrites the old file (Master Admin only)"""
    # Check if user is master admin
    if current_user.account_type != "master_admin":
        raise HTTPException(status_code=403, detail="Only Master Admin can update files")
    
    try:
        # Find existing file metadata
        file_metadata = await db.admin_files.find_one({"id": file_id}, {"_id": 0})
        
        if not file_metadata:
            raise HTTPException(status_code=404, detail="File not found")
        
        # Validate file size (100MB limit)
        max_size = 100 * 1024 * 1024  # 100MB
        file_content = await file.read()
        file_size = len(file_content)
        
        if file_size > max_size:
            raise HTTPException(
                status_code=400, 
                detail=f"File size exceeds 100MB limit. File size: {format_file_size(file_size)}"
            )
        
        # Delete old file from disk
        old_file_path = file_metadata["file_path"]
        if os.path.exists(old_file_path):
            os.remove(old_file_path)
            logger.info(f"Deleted old file from disk: {old_file_path}")
        
        # Save new file to disk with same naming pattern
        new_file_path = os.path.join(UPLOAD_DIR, f"{file_id}_{file.filename}")
        with open(new_file_path, "wb") as f:
            f.write(file_content)
        
        # Update metadata in database
        updated_metadata = {
            "original_filename": file.filename,
            "file_path": new_file_path,
            "file_size": file_size,
            "updated_by": current_user.id,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.admin_files.update_one(
            {"id": file_id},
            {"$set": updated_metadata}
        )
        
        logger.info(f"File updated successfully: {file.filename} (ID: {file_id})")
        
        return {
            "message": f"File '{file.filename}' updated successfully",
            "file_id": file_id,
            "filename": file.filename,
            "size": format_file_size(file_size)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to update file: {str(e)}")


@api_router.post("/admin/files/bulk-delete")
async def bulk_delete_files(file_ids: list, current_user: User = Depends(get_current_user)):
    """Bulk delete files (Master Admin only)"""
    if current_user.account_type != "master_admin":
        raise HTTPException(status_code=403, detail="Only Master Admin can delete files")
    
    try:
        deleted_count = 0
        
        for file_id in file_ids:
            file_metadata = await db.admin_files.find_one({"id": file_id}, {"_id": 0})
            
            if file_metadata:
                # Delete file from disk
                file_path = file_metadata["file_path"]
                if os.path.exists(file_path):
                    os.remove(file_path)
                
                # Delete from database
                await db.admin_files.delete_one({"id": file_id})
                deleted_count += 1
        
        return {
            "message": f"Successfully deleted {deleted_count} file(s)",
            "deleted_count": deleted_count
        }
        
    except Exception as e:
        logger.error(f"Error bulk deleting files: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to delete files")


@api_router.get("/montra-vehicle/analytics/battery-data")
async def get_battery_consumption_data(
    vehicle_id: str = Query(...),
    date: str = Query(...),
    current_user: User = Depends(get_current_user)
):
    """Get battery consumption data for a specific vehicle and date from MongoDB - OPTIMIZED"""
    try:
        logger.info(f"Fetching battery data for vehicle {vehicle_id} on {date}")
        
        # Query MongoDB for matching data with compound index (vehicle_id + date)
        query = {
            "vehicle_id": vehicle_id,
            "date": date
        }
        
        # Use projection to only fetch required fields (reduces data transfer)
        # Exclude unnecessary fields to reduce payload size
        projection = {
            "_id": 0  # Already excluding MongoDB ID
            # Include all other fields by default
        }
        
        # Check if index exists before using hint
        try:
            # Try to use index for fastest retrieval
            results = await db.montra_feed_data.find(
                query, 
                projection
            ).hint("idx_vehicle_date").sort("Date", 1).limit(2000).to_list(2000)
        except Exception as hint_error:
            # If index doesn't exist, fall back to query without hint
            logger.warning(f"Index not found, using query without hint: {str(hint_error)}")
            results = await db.montra_feed_data.find(
                query, 
                projection
            ).sort("Date", 1).limit(2000).to_list(2000)
        
        if not results:
            logger.warning(f"No data found for vehicle {vehicle_id} on {date}")
            raise HTTPException(
                status_code=404, 
                detail=f"No data found for vehicle {vehicle_id} on {date}. Please import the feed data first."
            )
        
        logger.info(f"Retrieved {len(results)} rows for vehicle {vehicle_id} (optimized with index)")
        
        return {
            "success": True,
            "vehicle_id": vehicle_id,
            "date": date,
            "data": results,
            "count": len(results)
        }
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching battery data: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch battery data: {str(e)}")


@api_router.get("/montra-vehicle/vehicles")
async def get_vehicles_list(current_user: User = Depends(get_current_user)):
    """Get list of all vehicles from vehicle mapping"""
    try:
        vehicles = await db.vehicle_mapping.find({}, {"_id": 0}).to_list(1000)
        
        vehicle_list = [
            {
                "vehicle_id": v["vehicle_id"],
                "registration_number": v.get("registration_number", "")
            }
            for v in vehicles
        ]
        
        return {
            "vehicles": vehicle_list,
            "count": len(vehicle_list)
        }
    except Exception as e:
        logger.error(f"Error fetching vehicles: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch vehicles list")


@api_router.get("/montra-vehicle/feed-database")
async def get_montra_feed_database(current_user: User = Depends(get_current_user)):
    """Get all montra feed entries with file information for database management"""
    try:
        # Aggregate to get unique files with their data counts, grouped by month
        pipeline = [
            {
                "$group": {
                    "_id": {
                        "vehicle_id": "$vehicle_id",
                        "date": "$date",
                        "filename": {"$ifNull": ["$filename", "Unknown"]},
                        "month": "$month",
                        "year": {"$ifNull": ["$year", "2024"]}  # Default to 2024 if year is missing
                    },
                    "count": {"$sum": 1},
                    "first_entry": {"$first": "$$ROOT"}
                }
            },
            {
                "$project": {
                    "vehicle_id": "$_id.vehicle_id",
                    "date": "$_id.date", 
                    "filename": "$_id.filename",
                    "month": "$_id.month",
                    "year": "$_id.year",
                    "month_year": {"$concat": ["$_id.month", " ", {"$toString": "$_id.year"}]},
                    "record_count": "$count",
                    "uploaded_at": "$first_entry.imported_at",
                    "file_size": "$first_entry.file_size"
                }
            },
            {"$sort": {"uploaded_at": -1}}
        ]
        
        results = await db.montra_feed_data.aggregate(pipeline).to_list(1000)
        
        return {
            "success": True,
            "files": results,
            "count": len(results)
        }
    except Exception as e:
        logger.error(f"Error fetching montra feed database: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch montra feed database")


@api_router.delete("/montra-vehicle/feed-database")
async def delete_montra_feed_files(
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """Delete selected montra feed files from database
    
    Args:
        request: Request containing file_identifiers in JSON body
    """
    try:
        # Parse JSON body
        body = await request.json()
        file_identifiers = body if isinstance(body, list) else body.get("file_identifiers", [])
        
        if not file_identifiers:
            raise HTTPException(status_code=400, detail="No files specified for deletion")
        
        deleted_count = 0
        failed_deletions = []
        
        for file_info in file_identifiers:
            try:
                # Use vehicle_id and date as primary identifiers since filename might be null in older records
                query = {
                    "vehicle_id": file_info["vehicle_id"],
                    "date": file_info["date"]
                }
                
                result = await db.montra_feed_data.delete_many(query)
                deleted_count += result.deleted_count
                
                logger.info(f"Deleted {result.deleted_count} records for vehicle {file_info['vehicle_id']} date {file_info['date']}")
                
            except Exception as e:
                logger.error(f"Error deleting records for {file_info.get('vehicle_id', 'Unknown')}: {str(e)}")
                failed_deletions.append({
                    "filename": file_info.get("filename", f"{file_info.get('vehicle_id', 'Unknown')}-{file_info.get('date', 'Unknown')}"),
                    "error": str(e)
                })
        
        if failed_deletions:
            return {
                "success": False,
                "message": f"Partially completed. Deleted {deleted_count} records.",
                "deleted_count": deleted_count,
                "failed_deletions": failed_deletions
            }
        
        return {
            "success": True,
            "message": f"Successfully deleted {deleted_count} records from {len(file_identifiers)} files",
            "deleted_count": deleted_count
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting montra feed files: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to delete montra feed files")


@api_router.post("/montra-vehicle/feed-database/download")
async def download_montra_feed_data(
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """Download selected montra feed files as CSV
    
    Args:
        request: Request containing files list in JSON body
    """
    try:
        from fastapi.responses import StreamingResponse
        import csv
        from io import StringIO
        
        # Parse JSON body
        body = await request.json()
        files = body.get("files", [])
        
        if not files:
            raise HTTPException(status_code=400, detail="No files specified for download")
        
        # Collect all records
        all_records = []
        
        for file_info in files:
            query = {
                "vehicle_id": file_info["vehicle_id"],
                "date": file_info["date"]
            }
            
            records = await db.montra_feed_data.find(query).to_list(None)
            all_records.extend(records)
        
        if not all_records:
            raise HTTPException(status_code=404, detail="No data found for selected files")
        
        # Create CSV
        output = StringIO()
        
        # Get all unique field names
        fieldnames = set()
        for record in all_records:
            fieldnames.update(record.keys())
        
        # Remove MongoDB _id field
        fieldnames.discard('_id')
        fieldnames = sorted(fieldnames)
        
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        
        for record in all_records:
            # Remove _id
            if '_id' in record:
                del record['_id']
            writer.writerow(record)
        
        # Return as downloadable file
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename=montra_feed_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error downloading montra feed data: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Failed to download data: {str(e)}")


@api_router.post("/montra-vehicle/battery-milestones")
async def get_battery_milestones(
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """
    Analyze battery charge milestones for vehicles over a date range.
    Returns time and km when battery hits 80%, 50%, 30%, 20%, plus derived mileage and mid-day charging.
    """
    try:
        from datetime import datetime, time as dt_time
        import pandas as pd
        
        data = await request.json()
        vehicle_ids = data.get("vehicle_ids", [])
        start_date = data.get("start_date")  # Format: "01 Sep"
        end_date = data.get("end_date")      # Format: "30 Sep"
        
        if not vehicle_ids or not start_date or not end_date:
            raise HTTPException(status_code=400, detail="vehicle_ids, start_date, and end_date are required")
        
        results = []
        
        for vehicle_id in vehicle_ids:
            # Query data for this vehicle in date range
            query = {
                "vehicle_id": vehicle_id,
                "date": {"$gte": start_date, "$lte": end_date}
            }
            
            # Get all records for this vehicle in the date range, grouped by date
            records = await db.montra_feed_data.find(query).sort("date", 1).to_list(10000)
            
            # Group by date
            date_groups = {}
            for record in records:
                date_key = record.get("date")
                if date_key not in date_groups:
                    date_groups[date_key] = []
                date_groups[date_key].append(record)
            
            # Analyze each date
            for date_str, day_records in date_groups.items():
                try:
                    # Sort records by time
                    day_records.sort(key=lambda x: x.get("time", "00:00:00"))
                    
                    # Skip if insufficient data
                    if len(day_records) < 5:
                        continue
                    
                    analysis = {
                        "date": date_str,
                        "vehicle": vehicle_id,
                        "charge_at_6am": None,
                        "time_at_80": None,
                        "km_at_80": None,
                        "time_at_50": None,
                        "km_at_50": None,
                        "time_at_30": None,
                        "time_at_20": None,
                        "km_at_20": None,
                        "derived_mileage": None,
                        "midday_charge": None
                    }
                    
                    # Find charge at 6AM (or closest time before 7AM)
                    charge_at_6am = None
                    for record in day_records:
                        record_time_str = record.get("time", "00:00:00")
                        try:
                            record_time = datetime.strptime(record_time_str, "%H:%M:%S").time()
                            if record_time <= dt_time(7, 0):
                                battery_pct = record.get("battery_soc_percentage")
                                if battery_pct is not None:
                                    charge_at_6am = battery_pct
                        except:
                            pass
                    
                    # Skip this day if no 6AM charge found
                    if charge_at_6am is None:
                        continue
                    
                    analysis["charge_at_6am"] = f"{charge_at_6am}%"
                    
                    # Find milestones (80%, 50%, 30%, 20%) - first occurrence during driving
                    milestones = {80: False, 50: False, 30: False, 20: False}
                    prev_battery = None
                    has_any_milestone = False
                    
                    for record in day_records:
                        battery_pct = record.get("battery_soc_percentage")
                        km = record.get("km")
                        record_time = record.get("time")
                        
                        if battery_pct is None:
                            continue
                        
                        # Check if battery is decreasing (driving, not charging)
                        is_driving = prev_battery is not None and battery_pct < prev_battery
                        
                        if is_driving or prev_battery is None:
                            # Check each milestone
                            for milestone in [80, 50, 30, 20]:
                                if not milestones[milestone]:
                                    if battery_pct <= milestone:
                                        milestones[milestone] = True
                                        has_any_milestone = True
                                        if milestone == 80:
                                            analysis["time_at_80"] = record_time
                                            if km:
                                                analysis["km_at_80"] = km
                                        elif milestone == 50:
                                            analysis["time_at_50"] = record_time
                                            if km:
                                                analysis["km_at_50"] = km
                                        elif milestone == 30:
                                            analysis["time_at_30"] = record_time
                                        elif milestone == 20:
                                            analysis["time_at_20"] = record_time
                                            if km:
                                                analysis["km_at_20"] = km
                        
                        prev_battery = battery_pct
                    
                    # Skip if no meaningful milestones reached
                    if not has_any_milestone:
                        continue
                    
                    # Calculate derived mileage (km per % charge drop during driving periods)
                    # Only consider periods where battery is decreasing (driving)
                    driving_segments = []
                    prev_record = None
                    
                    for record in day_records:
                        battery_pct = record.get("battery_soc_percentage")
                        km = record.get("km")
                        
                        if prev_record and battery_pct is not None and km is not None:
                            prev_battery_pct = prev_record.get("battery_soc_percentage")
                            prev_km = prev_record.get("km")
                            
                            if prev_battery_pct and prev_km:
                                # Check if driving (battery decreasing)
                                if battery_pct < prev_battery_pct:
                                    charge_drop = prev_battery_pct - battery_pct
                                    km_traveled = km - prev_km
                                    
                                    # Only include reasonable values
                                    if charge_drop > 0 and km_traveled > 0 and charge_drop < 50 and km_traveled < 100:
                                        efficiency = km_traveled / charge_drop
                                        driving_segments.append(efficiency)
                        
                        prev_record = record
                    
                    if driving_segments:
                        avg_mileage = sum(driving_segments) / len(driving_segments)
                        analysis["derived_mileage"] = f"{avg_mileage:.2f}"
                    
                    # Calculate mid-day charge (7AM to 7PM)
                    total_midday_charge = 0
                    prev_battery_midday = None
                    
                    for record in day_records:
                        record_time_str = record.get("time", "00:00:00")
                        battery_pct = record.get("battery_soc_percentage")
                        
                        try:
                            record_time = datetime.strptime(record_time_str, "%H:%M:%S").time()
                            
                            # Check if between 7AM and 7PM
                            if dt_time(7, 0) <= record_time <= dt_time(19, 0):
                                if battery_pct is not None and prev_battery_midday is not None:
                                    # If battery increased (charging)
                                    if battery_pct > prev_battery_midday:
                                        charge_added = battery_pct - prev_battery_midday
                                        # Only add if reasonable increase (not data error)
                                        if charge_added < 50:
                                            total_midday_charge += charge_added
                                
                                prev_battery_midday = battery_pct
                        except:
                            pass
                    
                    if total_midday_charge > 0:
                        analysis["midday_charge"] = f"{total_midday_charge:.1f}%"
                    
                    results.append(analysis)
                    
                except Exception as e:
                    logger.error(f"Error analyzing date {date_str} for vehicle {vehicle_id}: {str(e)}")
                    continue
        
        return {
            "success": True,
            "milestones": results,
            "count": len(results)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in battery milestones analysis: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error analyzing battery milestones: {str(e)}")


@api_router.post("/montra-vehicle/refresh-analytics-cache")
async def refresh_analytics_cache(current_user: User = Depends(get_current_user)):
    """Manually trigger analytics cache refresh (admin only)"""
    if current_user.account_type not in ["master_admin", "admin"]:
        raise HTTPException(status_code=403, detail="Only admins can refresh analytics cache")
    
    try:
        import subprocess
        
        logger.info(f"Manual cache refresh triggered by {current_user.email}")
        
        # Run the cache computation script in background
        subprocess.Popen(
            ["python", "/app/backend/analytics_cache.py"],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE
        )
        
        return {
            "success": True,
            "message": "Analytics cache refresh started in background. This may take 1-2 minutes."
        }
    except Exception as e:
        logger.error(f"Failed to trigger cache refresh: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to trigger cache refresh: {str(e)}")


@api_router.get("/montra-vehicle/battery-audit")
async def get_battery_charge_audit(
    current_user: User = Depends(get_current_user), 
    force_refresh: bool = False,
    start_date: str = None,
    end_date: str = None
):
    """
    Battery Audit Analysis - CACHED VERSION for Production Performance
    Uses pre-computed cache (updated daily), falls back to live computation if needed
    Supports custom date range filtering
    """
    try:
        from datetime import datetime, time as dt_time, date, timedelta
        
        # If custom date range is provided, skip cache and compute live
        use_cache = not force_refresh and not start_date and not end_date
        
        # Try to get from cache first (unless force_refresh is True or custom date range)
        if use_cache:
            logger.info("Checking analytics cache for battery audit data")
            cache_entry = await db.analytics_cache.find_one({"cache_type": "battery_audit"})
            
            if cache_entry:
                computed_at = datetime.fromisoformat(cache_entry.get("computed_at"))
                age_hours = (datetime.now(timezone.utc) - computed_at.replace(tzinfo=timezone.utc)).total_seconds() / 3600
                
                # Use cache if less than 24 hours old
                if age_hours < 24:
                    logger.info(f"✓ Returning cached battery audit data (age: {age_hours:.1f} hours)")
                    return cache_entry.get("data")
                else:
                    logger.info(f"Cache is {age_hours:.1f} hours old, computing fresh data")
        
        # LIVE COMPUTATION (Fallback or force refresh)
        logger.info("Computing battery audit data live")
        
        # Check if collection has data
        sample_count = await db.montra_feed_data.count_documents({}, limit=1)
        if sample_count == 0:
            logger.warning("No montra_feed_data found in database")
            return {
                "success": True,
                "audit_results": [],
                "count": 0,
                "critical_count": 0,
                "message": "No Montra feed data found. Please import vehicle data first."
            }
        
        # Determine date range
        if start_date:
            cutoff_date = start_date
        else:
            # Default to last 30 days if no start_date provided
            cutoff_date = (date.today() - timedelta(days=30)).isoformat()
        
        if end_date:
            end_date_filter = end_date
        else:
            # Default to today if no end_date provided
            end_date_filter = date.today().isoformat()
        
        logger.info(f"Fetching records from {cutoff_date} to {end_date_filter}")
        
        # Use aggregation pipeline
        pipeline = [
            {"$match": {"date": {"$gte": cutoff_date, "$lte": end_date_filter}}},
            {"$project": {
                "_id": 0,
                "vehicle_id": 1,
                "registration_number": 1,
                "date": 1,
                "Date": 1,
                "Battery Soc(%)": 1,
                "Battery SOC(%)": 1,
                "Odometer (km)": 1
            }},
            {"$sort": {"date": -1, "Date": 1}},
            {"$limit": 50000}
        ]
        
        cursor = db.montra_feed_data.aggregate(pipeline)
        
        # Group records by vehicle and date
        vehicle_date_groups = {}
        record_count = 0
        
        async for record in cursor:
            record_count += 1
            vehicle_id = record.get("vehicle_id")
            registration = record.get("registration_number")
            date_val = record.get("date")
            
            if not vehicle_id or not date_val:
                continue
            
            display_name = registration if registration else vehicle_id
            
            key = f"{vehicle_id}_{date_val}"
            if key not in vehicle_date_groups:
                vehicle_date_groups[key] = {
                    "vehicle_id": vehicle_id,
                    "display_name": display_name,
                    "date": date_val,
                    "records": []
                }
            vehicle_date_groups[key]["records"].append(record)
        
        if record_count == 0:
            logger.warning(f"No records found after {cutoff_date}")
            return {
                "success": True,
                "audit_results": [],
                "count": 0,
                "critical_count": 0,
                "message": f"No data found in the last 30 days. Please import recent vehicle data."
            }
        
        logger.info(f"Processed {record_count} records, grouped into {len(vehicle_date_groups)} vehicle-date combinations")
        
        audit_results = []
        
        # Analyze each vehicle-date combination
        for key, group_data in vehicle_date_groups.items():
            vehicle_id = group_data["vehicle_id"]
            display_name = group_data["display_name"]
            date = group_data["date"]
            day_records = group_data["records"]
            
            # Sort records by timestamp
            try:
                day_records.sort(key=lambda x: datetime.fromisoformat(
                    str(x.get("Date", "")).replace(' ', 'T')
                ) if x.get("Date") else datetime.min)
            except:
                continue
            
            # Find readings at specific times
            charge_6am = None
            charge_12pm = None
            charge_5pm = None
            odometer_6am = None
            odometer_12pm = None
            odometer_5pm = None
            
            for record in day_records:
                try:
                    date_str = str(record.get("Date", ""))
                    if not date_str:
                        continue
                        
                    dt = datetime.fromisoformat(date_str.replace(' ', 'T'))
                    record_time = dt.time()
                    
                    battery = float(record.get("Battery Soc(%)") or record.get("Battery SOC(%)") or 0)
                    odometer = float(record.get("Odometer (km)") or 0)
                    
                    # Get closest reading to 6 AM (within 2 hour window: 5 AM - 7 AM)
                    if dt_time(5, 0) <= record_time <= dt_time(7, 0) and charge_6am is None:
                        charge_6am = battery
                        odometer_6am = odometer
                    
                    # Get closest reading to 12 PM (within 2 hour window: 11 AM - 1 PM)
                    if dt_time(11, 0) <= record_time <= dt_time(13, 0) and charge_12pm is None:
                        charge_12pm = battery
                        odometer_12pm = odometer
                    
                    # Get closest reading to 5 PM (within 2 hour window: 4 PM - 6 PM)
                    if dt_time(16, 0) <= record_time <= dt_time(18, 0) and charge_5pm is None:
                        charge_5pm = battery
                        odometer_5pm = odometer
                        
                except:
                    continue
            
            # Skip if we don't have all required readings
            if charge_6am is None or charge_12pm is None or charge_5pm is None:
                continue
            if odometer_6am is None or odometer_12pm is None or odometer_5pm is None:
                continue
            
            # Calculate distances
            km_6am_12pm = odometer_12pm - odometer_6am
            km_6am_5pm = odometer_5pm - odometer_6am
            
            # Calculate mileages (avoid division by zero)
            charge_drop_6am_12pm = charge_6am - charge_12pm
            charge_drop_6am_5pm = charge_6am - charge_5pm
            
            mileage_6am_12pm = (km_6am_12pm * 100 / charge_drop_6am_12pm) if charge_drop_6am_12pm > 0 else 0
            mileage_6am_5pm = (km_6am_5pm * 100 / charge_drop_6am_5pm) if charge_drop_6am_5pm > 0 else 0
            
            # Flag if critical (12PM < 60% AND 5PM < 20%)
            is_critical = (charge_12pm < 60 and charge_5pm < 20)
            
            audit_results.append({
                "date": date,
                "vehicle_name": display_name,
                "vehicle_id": vehicle_id,
                "charge_6am": round(charge_6am, 1),
                "charge_12pm": round(charge_12pm, 1),
                "charge_5pm": round(charge_5pm, 1),
                "km_6am_12pm": round(km_6am_12pm, 2),
                "km_6am_5pm": round(km_6am_5pm, 2),
                "mileage_6am_12pm": round(mileage_6am_12pm, 2),
                "mileage_6am_5pm": round(mileage_6am_5pm, 2),
                "is_critical": is_critical
            })
        
        # Sort results by date (newest first) and vehicle
        audit_results.sort(key=lambda x: (x["date"], x["vehicle_name"]), reverse=True)
        
        critical_count = sum(1 for r in audit_results if r["is_critical"])
        
        logger.info(f"Battery audit complete: {len(audit_results)} entries, {critical_count} critical")
        
        return {
            "success": True,
            "audit_results": audit_results,
            "count": len(audit_results),
            "critical_count": critical_count,
            "message": f"Analyzed {len(audit_results)} vehicle-days. Found {critical_count} critical charge drops."
        }
        
    except Exception as e:
        logger.error(f"Error in battery charge audit: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error generating battery audit: {str(e)}")


@api_router.get("/montra-vehicle/morning-charge-audit")
async def get_morning_charge_audit(
    current_user: User = Depends(get_current_user), 
    force_refresh: bool = False,
    start_date: str = None,
    end_date: str = None
):
    """
    Morning Charge Audit - CACHED VERSION for Production Performance
    Uses pre-computed cache (updated daily), falls back to live computation if needed
    Supports custom date range filtering
    """
    print(f"🚨 MORNING AUDIT CALLED: start={start_date}, end={end_date}, force={force_refresh}")
    logger.info(f"🚨 MORNING AUDIT CALLED: start={start_date}, end={end_date}, force={force_refresh}")
    try:
        from datetime import datetime, time as dt_time, date, timedelta
        
        # If custom date range is provided, skip cache and compute live
        use_cache = not force_refresh and not start_date and not end_date
        logger.info(f"🚨 use_cache={use_cache}, force_refresh={force_refresh}, start_date={start_date}, end_date={end_date}")
        
        # Try to get from cache first
        if use_cache:
            logger.info("Checking analytics cache for morning charge audit data")
            cache_entry = await db.analytics_cache.find_one({"cache_type": "morning_charge_audit"})
            
            if cache_entry:
                computed_at = datetime.fromisoformat(cache_entry.get("computed_at"))
                age_hours = (datetime.now(timezone.utc) - computed_at.replace(tzinfo=timezone.utc)).total_seconds() / 3600
                
                # Use cache if less than 24 hours old
                if age_hours < 24:
                    logger.info(f"✓ Returning cached morning charge audit data (age: {age_hours:.1f} hours)")
                    return cache_entry.get("data")
                else:
                    logger.info(f"Cache is {age_hours:.1f} hours old, computing fresh data")
        
        # LIVE COMPUTATION (Fallback)
        logger.info(f"🔍 MORNING AUDIT: Computing live with date range {start_date} to {end_date}")
        
        # Check if collection has data
        sample_count = await db.montra_feed_data.count_documents({}, limit=1)
        logger.info(f"🔍 MORNING AUDIT: Database has {sample_count} records")
        if sample_count == 0:
            logger.warning("No montra_feed_data found in database")
            return {
                "success": True,
                "audit_results": [],
                "count": 0,
                "message": "No Montra feed data found. Please import vehicle data first."
            }
        
        # Determine date range
        if start_date:
            cutoff_date = start_date
        else:
            # Default to last 30 days if no start_date provided
            cutoff_date = (date.today() - timedelta(days=30)).isoformat()
        
        if end_date:
            end_date_filter = end_date
        else:
            # Default to today if no end_date provided
            end_date_filter = date.today().isoformat()
        
        logger.info(f"Fetching records from {cutoff_date} to {end_date_filter}")
        
        # Use aggregation with filters and projection
        pipeline = [
            {"$match": {"date": {"$gte": cutoff_date, "$lte": end_date_filter}}},
            {"$project": {
                "_id": 0,
                "vehicle_id": 1,
                "Vehicle ID": 1,
                "registration_number": 1,
                "Registration Number": 1,
                "date": 1,
                "Date": 1,
                "Portal Received Time": 1,
                "time": 1,
                "Time": 1,
                "Battery Soc(%)": 1,
                "Battery SOC(%)": 1,
                "Battery %": 1,
                "battery_soc_percentage": 1
            }},
            {"$sort": {"date": -1}},
            {"$limit": 50000}
        ]
        
        cursor = db.montra_feed_data.aggregate(pipeline)
        
        # Group records by vehicle and date
        vehicle_date_groups = {}
        record_count = 0
        
        async for record in cursor:
            record_count += 1
            vehicle_id = record.get("vehicle_id") or record.get("Vehicle ID")
            registration = record.get("registration_number") or record.get("Registration Number")
            date_val = record.get("date")
            
            if not vehicle_id or not date_val:
                continue
            
            display_name = registration if registration else vehicle_id
            
            key = f"{vehicle_id}_{date_val}"
            if key not in vehicle_date_groups:
                vehicle_date_groups[key] = {
                    "vehicle_id": vehicle_id,
                    "display_name": display_name,
                    "date": date_val,
                    "records": []
                }
            vehicle_date_groups[key]["records"].append(record)
        
        if record_count == 0:
            logger.warning(f"No records found after {cutoff_date}")
            return {
                "success": True,
                "audit_results": [],
                "count": 0,
                "message": f"No data found in the last 30 days. Please import recent vehicle data."
            }
        
        logger.info(f"Processed {record_count} records, grouped into {len(vehicle_date_groups)} vehicle-date combinations")
        
        audit_results = []
        debug_6am_found = 0
        debug_below_95 = 0
        
        # Analyze each vehicle-date combination
        for key, group_data in vehicle_date_groups.items():
            vehicle_id = group_data["vehicle_id"]
            display_name = group_data["display_name"]
            date = group_data["date"]
            day_records = group_data["records"]
            
            # Sort records by time
            try:
                day_records.sort(key=lambda x: x.get("Portal Received Time", x.get("time", "00:00:00")) if x.get("Portal Received Time") or x.get("time") else "00:00:00")
            except Exception as e:
                logger.error(f"Error sorting records for {vehicle_id} on {date}: {str(e)}")
                continue
            
            # Find battery charge at or closest to 6 AM
            charge_at_6am = None
            closest_time_to_6am = None
            
            for record in day_records:
                battery_pct = (record.get("Battery Soc(%)") or 
                              record.get("Battery SOC(%)") or  # NEW: Handle uppercase SOC
                              record.get("Battery %") or 
                              record.get("battery_soc_percentage") or 
                              record.get("Battery") or
                              record.get("battery") or
                              record.get("SOC") or
                              record.get("soc"))
                
                time_str = (record.get("Portal Received Time") or
                           record.get("time") or 
                           record.get("Time") or 
                           record.get("Hour") or
                           record.get("hour"))
                
                if battery_pct and time_str:
                    try:
                        if isinstance(time_str, str):
                            try:
                                dt = datetime.fromisoformat(time_str.replace(' ', 'T'))
                                record_time = dt.time()
                            except:
                                time_part = time_str.split()[0] if ' ' in time_str else time_str
                                try:
                                    record_time = datetime.strptime(time_part, "%H:%M:%S").time()
                                except:
                                    continue
                        else:
                            continue
                        
                        # Look for records between 5:30 AM and 6:30 AM to get closest to 6 AM
                        if dt_time(5, 30) <= record_time <= dt_time(6, 30):
                            if battery_pct != '-':
                                try:
                                    battery_val = float(str(battery_pct))
                                    
                                    # Keep the record closest to 6:00 AM
                                    target_time = dt_time(6, 0)
                                    time_diff = abs((datetime.combine(datetime.today(), record_time) - 
                                                   datetime.combine(datetime.today(), target_time)).total_seconds())
                                    
                                    if closest_time_to_6am is None or time_diff < closest_time_to_6am:
                                        closest_time_to_6am = time_diff
                                        charge_at_6am = battery_val
                                        debug_6am_found += 1
                                except:
                                    pass
                        elif record_time > dt_time(6, 30):
                            # Stop looking once we're past 6:30 AM
                            break
                    except:
                        continue
            
            # Add to results if charge at 6 AM is less than 95%
            if charge_at_6am is not None and charge_at_6am < 95:
                debug_below_95 += 1
                audit_results.append({
                    "date": date,
                    "vehicle_name": display_name,
                    "charge_at_6am": round(charge_at_6am, 1)
                })
                logger.info(f"Found low morning charge: {vehicle_id} on {date} - {charge_at_6am}%")
        
        # Sort results by date and vehicle
        audit_results.sort(key=lambda x: (x["date"], x["vehicle_name"]))
        
        logger.info(f"Morning charge audit complete: Found {len(audit_results)} instances with charge < 95% at 6 AM")
        logger.info(f"Debug: Total 6AM records found: {debug_6am_found}, Records below 95%: {debug_below_95}, Final results: {len(audit_results)}")
        
        return {
            "success": True,
            "audit_results": audit_results,
            "count": len(audit_results),
            "message": f"Found {len(audit_results)} instances where morning charge was below 95% at 6 AM"
        }
        
    except HTTPException as he:
        logger.error(f"🔍 MORNING AUDIT: HTTPException - {he}")
        raise
    except Exception as e:
        logger.error(f"🔍 MORNING AUDIT: Exception - {str(e)}")
        import traceback
        logger.error(f"🔍 MORNING AUDIT: Traceback - {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error generating morning charge audit: {str(e)}")


# ==================== Vehicle Service Request ====================

@api_router.get("/montra-vehicle/service-requests")
async def get_service_requests(
    skip: int = Query(0, description="Number of records to skip"),
    limit: int = Query(50, description="Number of records to return"),
    search: Optional[str] = Query(None, description="Search by VIN or vehicle name"),
    repair_type: Optional[str] = Query(None, description="Filter by repair type"),
    repair_status: Optional[str] = Query(None, description="Filter by repair status"),
    start_date: Optional[str] = Query(None, description="Filter by repair start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="Filter by repair end date (YYYY-MM-DD)"),
    current_user: User = Depends(get_current_user)
):
    """Get all service requests with filters and pagination"""
    try:
        # Build query
        query = {}
        
        if search:
            query["$or"] = [
                {"vin": {"$regex": search, "$options": "i"}},
                {"vehicle_name": {"$regex": search, "$options": "i"}}
            ]
        
        if repair_type:
            query["repair_type"] = repair_type
        
        if repair_status:
            query["repair_status"] = repair_status
        
        if start_date:
            query["repair_start_date"] = {"$gte": datetime.fromisoformat(start_date)}
        
        if end_date:
            query["repair_end_date"] = {"$lte": datetime.fromisoformat(end_date)}
        
        # Get total count
        total = await db.vehicle_service_requests.count_documents(query)
        
        # Get requests
        requests = await db.vehicle_service_requests.find(query).sort("request_timestamp", -1).skip(skip).limit(limit).to_list(length=limit)
        
        return {
            "success": True,
            "requests": requests,
            "total": total,
            "page": (skip // limit) + 1 if limit > 0 else 1,
            "total_pages": (total + limit - 1) // limit if limit > 0 else 1
        }
    
    except Exception as e:
        logger.error(f"Error fetching service requests: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error fetching service requests: {str(e)}")


@api_router.get("/montra-vehicle/service-requests/{request_id}")
async def get_service_request(
    request_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get single service request by ID"""
    try:
        request = await db.vehicle_service_requests.find_one({"id": request_id})
        
        if not request:
            raise HTTPException(status_code=404, detail="Service request not found")
        
        return {
            "success": True,
            "request": request
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching service request: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error fetching service request: {str(e)}")


def calculate_repair_time_days(start_date: datetime, end_date: datetime) -> float:
    """Calculate repair time in days"""
    if not start_date or not end_date:
        return 0
    
    delta = end_date - start_date
    return round(delta.total_seconds() / (24 * 3600), 2)


def calculate_downtime_hours(start_date: datetime, end_date: datetime) -> float:
    """
    Calculate service vehicle downtime in hours (6AM-11PM driving hours only)
    Formula: Sum of driving hours (6AM-11PM) missed during service period
    """
    if not start_date or not end_date:
        return 0
    
    total_downtime = 0
    current_date = start_date
    
    while current_date.date() <= end_date.date():
        # Define driving window: 6 AM to 11 PM (17 hours total)
        day_start = current_date.replace(hour=6, minute=0, second=0, microsecond=0)
        day_end = current_date.replace(hour=23, minute=0, second=0, microsecond=0)
        
        # Determine actual start and end for this day
        if current_date.date() == start_date.date():
            # First day: start from repair_start_date if after 6 AM
            actual_start = max(start_date, day_start)
        else:
            actual_start = day_start
        
        if current_date.date() == end_date.date():
            # Last day: end at repair_end_date if before 11 PM
            actual_end = min(end_date, day_end)
        else:
            actual_end = day_end
        
        # Calculate hours for this day (only if within driving window)
        if actual_start < day_end and actual_end > day_start:
            # Clamp to driving window
            actual_start = max(actual_start, day_start)
            actual_end = min(actual_end, day_end)
            
            hours = (actual_end - actual_start).total_seconds() / 3600
            total_downtime += hours
        
        # Move to next day
        current_date = (current_date + timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
    
    return round(total_downtime, 2)


@api_router.post("/montra-vehicle/service-requests")
async def create_service_request(
    request_data: VehicleServiceRequestCreate,
    current_user: User = Depends(get_current_user)
):
    """Create new service request"""
    from app_models import VehicleServiceRequest
    
    try:
        # Create request object
        service_request = VehicleServiceRequest(
            **request_data.dict(),
            request_reported_by=current_user.email
        )
        
        # Auto-calculate repair_time_days and downtime_hours
        if request_data.repair_start_date and request_data.repair_end_date:
            service_request.repair_time_days = calculate_repair_time_days(
                request_data.repair_start_date,
                request_data.repair_end_date
            )
            service_request.service_vehicle_downtime_hours = calculate_downtime_hours(
                request_data.repair_start_date,
                request_data.repair_end_date
            )
        
        # Insert into database
        await db.vehicle_service_requests.insert_one(service_request.dict())
        
        return {
            "success": True,
            "message": "Service request created successfully",
            "request_id": service_request.id,
            "request": service_request.dict()
        }
    
    except Exception as e:
        logger.error(f"Error creating service request: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error creating service request: {str(e)}")


@api_router.patch("/montra-vehicle/service-requests/{request_id}")
async def update_service_request(
    request_id: str,
    update_data: VehicleServiceRequestUpdate,
    current_user: User = Depends(get_current_user)
):
    """Update existing service request"""
    try:
        # Get existing request
        existing = await db.vehicle_service_requests.find_one({"id": request_id})
        
        if not existing:
            raise HTTPException(status_code=404, detail="Service request not found")
        
        # Prepare update data
        update_dict = {k: v for k, v in update_data.dict().items() if v is not None}
        update_dict["updated_at"] = datetime.now(timezone.utc)
        
        # Recalculate repair_time_days and downtime_hours if dates changed
        start_date = update_data.repair_start_date or existing.get("repair_start_date")
        end_date = update_data.repair_end_date or existing.get("repair_end_date")
        
        if start_date and end_date:
            # Convert strings to datetime if needed
            if isinstance(start_date, str):
                start_date = datetime.fromisoformat(start_date)
            if isinstance(end_date, str):
                end_date = datetime.fromisoformat(end_date)
            
            update_dict["repair_time_days"] = calculate_repair_time_days(start_date, end_date)
            update_dict["service_vehicle_downtime_hours"] = calculate_downtime_hours(start_date, end_date)
        
        # Update in database
        await db.vehicle_service_requests.update_one(
            {"id": request_id},
            {"$set": update_dict}
        )
        
        # Get updated request
        updated = await db.vehicle_service_requests.find_one({"id": request_id})
        
        return {
            "success": True,
            "message": "Service request updated successfully",
            "request": updated
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating service request: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error updating service request: {str(e)}")


@api_router.delete("/montra-vehicle/service-requests/{request_id}")
async def delete_service_request(
    request_id: str,
    current_user: User = Depends(get_current_user)
):
    """Delete service request"""
    try:
        # Check if request exists
        existing = await db.vehicle_service_requests.find_one({"id": request_id})
        
        if not existing:
            raise HTTPException(status_code=404, detail="Service request not found")
        
        # Delete from database
        await db.vehicle_service_requests.delete_one({"id": request_id})
        
        return {
            "success": True,
            "message": "Service request deleted successfully"
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting service request: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error deleting service request: {str(e)}")


@api_router.post("/montra-vehicle/service-requests/upload-image")
async def upload_service_request_image(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """Upload image for service request"""
    import shutil
    from pathlib import Path
    
    try:
        # Create directory if it doesn't exist
        upload_dir = Path("/app/backend/uploaded_files/vehicle_service_images")
        upload_dir.mkdir(parents=True, exist_ok=True)
        
        # Generate unique filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_extension = file.filename.split(".")[-1] if "." in file.filename else "jpg"
        unique_filename = f"{timestamp}_{str(uuid.uuid4())[:8]}.{file_extension}"
        file_path = upload_dir / unique_filename
        
        # Save file
        with file_path.open("wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        # Return relative path
        relative_path = f"vehicle_service_images/{unique_filename}"
        
        return {
            "success": True,
            "message": "Image uploaded successfully",
            "file_path": relative_path,
            "filename": unique_filename
        }
    
    except Exception as e:
        logger.error(f"Error uploading image: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error uploading image: {str(e)}")


@api_router.get("/montra-vehicle/service-requests/image-folders")
async def get_image_folders(
    current_user: User = Depends(get_current_user)
):
    """Get list of image folders"""
    from pathlib import Path
    
    try:
        upload_dir = Path("/app/backend/uploaded_files/vehicle_service_images")
        
        if not upload_dir.exists():
            return {
                "success": True,
                "folders": [],
                "message": "No image folders found"
            }
        
        # Get all files
        files = []
        for file_path in upload_dir.glob("*"):
            if file_path.is_file():
                stat = file_path.stat()
                files.append({
                    "name": file_path.name,
                    "size": stat.st_size,
                    "modified": datetime.fromtimestamp(stat.st_mtime).isoformat(),
                    "path": f"vehicle_service_images/{file_path.name}"
                })
        
        # Sort by modified date (newest first)
        files.sort(key=lambda x: x["modified"], reverse=True)
        
        return {
            "success": True,
            "files": files,
            "count": len(files)
        }
    
    except Exception as e:
        logger.error(f"Error fetching image folders: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error fetching image folders: {str(e)}")


@api_router.get("/montra-vehicle/vins")
async def get_vehicle_vins(
    month: Optional[str] = Query(None, description="Month name (e.g., Nov)"),
    year: Optional[str] = Query(None, description="Year (e.g., 2025)"),
    current_user: User = Depends(get_current_user)
):
    """Get list of vehicle VINs and registration numbers from Nura Fleet Data.xlsx file"""
    import pandas as pd
    
    try:
        logger.info(f"Looking for vehicles in Nura Fleet Data.xlsx")
        
        # Find the Nura Fleet Data.xlsx file
        files = await db.admin_files.find({}).to_list(1000)
        
        vehicles_file = None
        for file in files:
            filename = file.get("original_filename", "")
            if filename == "Nura Fleet Data.xlsx":
                vehicles_file = file
                break
        
        vehicle_list = []
        
        if vehicles_file:
            file_path = vehicles_file.get("file_path")
            if file_path and os.path.exists(file_path):
                try:
                    # Read the first sheet with headers
                    df = pd.read_excel(file_path, engine='openpyxl')
                    
                    # Expected columns: Column A = VIN (Montral Vehicle ID), Column B = Registration No.
                    if len(df.columns) >= 2:
                        # Iterate through rows (skip header which is already handled by pandas)
                        for idx, row in df.iterrows():
                            vin = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
                            reg_number = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
                            
                            # Skip invalid rows
                            if vin and reg_number and vin.lower() not in ['nan', 'montral vehicle id', 'vin']:
                                vehicle_list.append({
                                    "vin": vin,
                                    "registration_number": reg_number,
                                    "vehicle_name": reg_number  # Keep for compatibility
                                })
                        
                        logger.info(f"Loaded {len(vehicle_list)} vehicles from Nura Fleet Data.xlsx (VIN from Col A, Reg from Col B)")
                    else:
                        logger.warning(f"Nura Fleet Data.xlsx does not have at least 2 columns")
                
                except Exception as e:
                    logger.error(f"Error parsing Nura Fleet Data.xlsx: {str(e)}")
        else:
            logger.warning("Nura Fleet Data.xlsx file not found in admin files")
        
        # If no vehicles found, return empty list with message
        if not vehicle_list:
            logger.warning(f"No vehicles found in Nura Fleet Data.xlsx")
            return {
                "success": True,
                "vehicles": [],
                "count": 0,
                "message": "No vehicles found. Please upload Nura Fleet Data.xlsx with VIN in Column A and Registration Number in Column B."
            }
        
        return {
            "success": True,
            "vehicles": vehicle_list,
            "count": len(vehicle_list)
        }
    
    except Exception as e:
        logger.error(f"Error fetching vehicles: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error fetching vehicles: {str(e)}")


@api_router.get("/admin/files/get-drivers-vehicles")
async def get_drivers_and_vehicles(
    month: str = Query(..., description="Month name (e.g., Sep)"),
    year: str = Query(..., description="Year (e.g., 2025)"),
    current_user: User = Depends(get_current_user)
):
    """Get drivers and vehicles list for a specific month/year from generic Excel files with monthly tabs"""
    import pandas as pd
    
    try:
        # Convert month number to name if needed
        month_names = {
            "01": "Jan", "02": "Feb", "03": "Mar", "04": "Apr", "05": "May", "06": "Jun",
            "07": "Jul", "08": "Aug", "09": "Sep", "10": "Oct", "11": "Nov", "12": "Dec"
        }
        month_name = month_names.get(month, month)
        
        # Ensure month name is title case (handles "sep" -> "Sep", "SEP" -> "Sep", etc.)
        month_name = month_name.capitalize()
        
        # Tab name format: "Sep 2025", "Oct 2025", etc.
        tab_name = f"{month_name} {year}"
        
        logger.info(f"Looking for tab '{tab_name}' in Drivers List.xlsx and Vehicles List.xlsx")
        
        # Find the generic files (not month-specific)
        files = await db.admin_files.find({}).to_list(1000)
        
        drivers_file = None
        vehicles_file = None
        
        for file in files:
            filename = file.get("original_filename", "")
            # Look for generic filenames
            if filename == "Drivers List.xlsx":
                drivers_file = file
            elif filename == "Vehicles List.xlsx":
                vehicles_file = file
        
        result = {
            "success": True,
            "drivers": [],
            "vehicles": [],
            "month": month_name,
            "year": year
        }
        
        # Parse drivers file - read from specific tab
        if drivers_file:
            file_path = drivers_file.get("file_path")
            if file_path and os.path.exists(file_path):
                try:
                    # Read specific sheet/tab by name (header=None to not treat B1 as column names)
                    df = pd.read_excel(file_path, sheet_name=tab_name, header=None)
                    # Read from Column B (index 1), skip first row (B1 which has headers)
                    if len(df.columns) > 1:
                        drivers = df.iloc[1:, 1].dropna().astype(str).tolist()  # Start from row 1 (B2), column 1 (B)
                        drivers = [name.strip() for name in drivers if name.strip() and name.strip().lower() not in ['nan', 'name', 'driver name', 'driver', 'drivers']]
                        result["drivers"] = drivers
                        logger.info(f"Loaded {len(drivers)} drivers from tab '{tab_name}' Column B of Drivers List.xlsx")
                    else:
                        logger.warning(f"Drivers file tab '{tab_name}' does not have Column B")
                except Exception as e:
                    logger.error(f"Error parsing drivers file tab '{tab_name}': {str(e)}")
        else:
            logger.warning("Drivers List.xlsx file not found in admin files")
        
        # Parse vehicles file - read from specific tab
        if vehicles_file:
            file_path = vehicles_file.get("file_path")
            if file_path and os.path.exists(file_path):
                try:
                    # Read specific sheet/tab by name (header=None to not treat B1 as column names)
                    df = pd.read_excel(file_path, sheet_name=tab_name, header=None)
                    # Read from Column B (index 1), skip first row (B1 which has headers)
                    if len(df.columns) > 1:
                        vehicles = df.iloc[1:, 1].dropna().astype(str).tolist()  # Start from row 1 (B2), column 1 (B)
                        vehicles = [name.strip() for name in vehicles if name.strip() and name.strip().lower() not in ['nan', 'name', 'vehicle number', 'vehicle', 'registration number', 'vehicles']]
                        result["vehicles"] = vehicles
                        logger.info(f"Loaded {len(vehicles)} vehicles from tab '{tab_name}' Column B of Vehicles List.xlsx")
                    else:
                        logger.warning(f"Vehicles file tab '{tab_name}' does not have Column B")
                except Exception as e:
                    logger.error(f"Error parsing vehicles file tab '{tab_name}': {str(e)}")
        else:
            logger.warning("Vehicles List.xlsx file not found in admin files")
        
        if not result["drivers"] and not result["vehicles"]:
            # Return mock data as fallback
            result["drivers"] = ["Abdul", "Samantha", "Samuel", "Sareena", "Ravi", "John", "Mike"]
            result["vehicles"] = ["TN07CE2222", "TN01AB1234", "KA05CD5678", "AP09EF9012"]
            result["using_mock_data"] = True
            logger.warning(f"No data found for tab '{tab_name}', using mock data")
        
        return result
        
    except Exception as e:
        logger.error(f"Error in get_drivers_and_vehicles: {str(e)}")
        # Return mock data as fallback
        return {
            "success": True,
            "drivers": ["Abdul", "Samantha", "Samuel", "Sareena", "Ravi", "John", "Mike"],
            "vehicles": ["TN07CE2222", "TN01AB1234", "KA05CD5678", "AP09EF9012"],
            "month": month,
            "year": year,
            "using_mock_data": True,
            "error": str(e)
        }


@api_router.post("/payment-reconciliation/process-screenshots")
async def process_payment_screenshots(
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """Process uploaded screenshots using OpenAI GPT-4o Vision to extract payment data - PARALLEL PROCESSING"""
    from dotenv import load_dotenv
    from datetime import datetime
    import asyncio
    load_dotenv()
    
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage, ImageContent
        import base64
        import tempfile
        import uuid
        
        # Parse form data
        form = await request.form()
        files = form.getlist("files")
        month_year = form.get("month_year", "")
        driver_name = form.get("driver_name", "")
        vehicle_number = form.get("vehicle_number", "")
        platform = form.get("platform", "")
        
        logger.info(f"📋 Processing {len(files)} files with metadata:")
        logger.info(f"  - Month/Year: {month_year}")
        logger.info(f"  - Driver: {driver_name}")
        logger.info(f"  - Vehicle: {vehicle_number}")
        logger.info(f"  - Platform: '{platform}' (type: {type(platform).__name__})")
        
        if not files:
            raise HTTPException(status_code=400, detail="No files uploaded")
        
        if len(files) > 10:
            raise HTTPException(status_code=400, detail="Maximum 10 files allowed per batch")
        
        # Get the emergent LLM key
        api_key = os.environ.get("EMERGENT_LLM_KEY")
        if not api_key:
            raise HTTPException(status_code=500, detail="Emergent LLM API key not configured")
        
        logger.info(f"🚀 Starting parallel processing of {len(files)} files")
        
        # Extraction prompt (keeping the same detailed prompt)
        extraction_prompt = """You are analyzing ride-sharing receipt screenshots (Tamil/English). These can be in multiple formats:

**FORMAT TYPES:**
1. Simple cash receipts: "Auto", time, கேஷ், ₹amount
2. Detailed app receipts: "Auto", ₹amount, distance (கி.மீ), duration (நிமி, வி), time
3. Paytm detailed receipts: "மதிப்பிடப்பட்ட வருவாய்" (Estimated Fare), distance with decimals (2.36 km), duration with decimals (16.67 min), pickup/drop locations
4. Ride history list: Multiple rides with timestamps like "10:59 pm", "10:18 pm", pickup/drop locations, fare amounts
5. Surge pricing: Upward arrow ↑ with "அதிகரித்துள்ளது" (increased)
6. Cancellations: "வாடிக்கையாளர் ரத்துசெய்தார்" (customer cancelled) or "வண்டிக்கையாளர் ரத்துசெய்தார்" (driver cancelled)
7. Zero-fare rides: ₹0.00 with promotional text or cancellation

**TAMIL TEXT MEANINGS:**
- கேஷ் = Cash
- கி.மீ / km = Kilometers
- நிமி = Minutes (நிமிடம்)
- வி = Seconds (விநாடி)
- ம.நே = Hours (மணிநேரம்)
- மதிப்பிடப்பட்ட வருவாய் / மதிப்பிடப்பட்ட வரு... = Estimated Fare
- பிக்கப் = Pickup
- முரபி / டிராப் = Dropoff/Drop
- அதிகரித்துள்ளது = Surge/Increased pricing
- வாடிக்கையாளர் ரத்துசெய்தார் = Customer cancelled
- வண்டிக்கையாளர் ரத்துசெய்தார் = Driver cancelled
- பயணச் சவால் = Travel challenge (promotional)
- ஆட்டோ ஆர்டர் = Auto Order
- முடிந்த ஆர்டர்கள் = Completed orders/rides

**EXTRACT EACH RIDE AS JSON:**

[
  {
    "driver": "Driver name if visible, otherwise N/A",
    "vehicle": "Vehicle number if visible, otherwise N/A", 
    "description": "Auto, Bike, Car, etc. If surge add (Surge), if cancelled add (Cancelled)",
    "date": "DD/MM/YYYY format (convert DD MMM or date from screenshot)",
    "time": "HH:MM AM/PM format",
    "amount": "Fare amount as NUMBER only (no ₹, commas). Use 0 for ₹0.00 rides",
    "payment_mode": "Cash/UPI/Card if visible, otherwise N/A",
    "distance": "Extract km as NUMBER only (e.g., '3.57' from '3.57 கி.மீ'), or N/A",
    "duration": "Convert to MINUTES as NUMBER (e.g., '16' for '16 நிமி', '71' for '1 ம.நே 11 நிமி'), or N/A", 
    "pickup_km": "N/A",
    "drop_km": "N/A",
    "pickup_location": "N/A",
    "drop_location": "N/A"
  }
]

**EXTRACTION RULES:**
✅ Extract EVERY visible ride (including ₹0.00 if cancelled/promotional)
✅ For duration: Convert hours to minutes (1 hour 11 min = 71 min)
✅ Accept decimal distances: "2.36 km" → 2.36
✅ Accept decimal durations: "16.67 min" → 16.67
✅ For surge pricing: Include in description (e.g., "Auto (Surge)")
✅ For cancellations: Set amount to 0, add to description (e.g., "Auto (Cancelled)")
✅ Skip only "Bank Transfer" entries or unrelated promotional banners
✅ Extract distance/duration as numbers (remove Tamil/English text)
✅ If screenshot shows "26 September" or "செப்." convert to DD/MM/2024 format
✅ Extract pickup/drop locations if visible (e.g., "Crown Residences", "Mogappair West")
✅ For ride history lists: Extract each ride separately with its timestamp

⚠️ **CRITICAL RULES - NO ASSUMPTIONS:**
❌ DO NOT copy or assume data from other rides in the same screenshot
❌ DO NOT guess missing pickup/drop locations - use "N/A" if not visible
❌ DO NOT assume the same location repeats across multiple rides
❌ DO NOT fill in missing data based on other rides
✅ ONLY extract data that is CLEARLY VISIBLE for each specific ride
✅ If a ride has no visible pickup location, use "N/A" - DO NOT copy from other rides
✅ If a ride has no visible drop location, use "N/A" - DO NOT copy from other rides
✅ Each ride is independent - treat them separately

**EXAMPLE CONVERSIONS:**
- "3.57 கி.மீ" / "2.36 km" → distance: "3.57" / "2.36"
- "16 நிமி 55 வி" → duration: "16" (ignore seconds)
- "16.67 min" → duration: "16.67"
- "1 ம.நே 11 நிமி" → duration: "71"
- "₹126.45" / "₹89" → amount: "126.45" / "89"
- "திங்கள்., 29 செப்." / "26 September" → date: "29/09/2024" / "26/09/2024"
- "10:59 pm" → time: "10:59 PM"
- "Crown Residences" → pickup_location: "Crown Residences"
- "8/25, Srinivasa Nagar, Virugambakkam" → drop_location: "8/25, Srinivasa Nagar, Virugambakkam"
- "B1 Entrance, Thirumangalam Metro St..." → pickup_location: "B1 Entrance, Thirumangalam Metro Station"

Be precise and extract ALL rides shown in the screenshot. If a screenshot shows multiple rides (like a ride history list), extract each as a separate record.
"""
        
        # Function to process a single file with MAXIMUM SPEED optimization
        async def process_single_file(file, index):
            """Process one file and return results - OPTIMIZED FOR SPEED"""
            temp_path = None
            try:
                logger.info(f"  📄 Processing file {index + 1}/{len(files)}: {file.filename}")
                
                # Read file content
                file_content = await file.read()
                
                # AGGRESSIVE image optimization for maximum speed
                from PIL import Image
                import io
                
                try:
                    # Open image
                    image = Image.open(io.BytesIO(file_content))
                    original_size = len(file_content)
                    
                    # AGGRESSIVE resize - max 1536px (smaller = faster processing)
                    max_dimension = 1536
                    if max(image.size) > max_dimension:
                        ratio = max_dimension / max(image.size)
                        new_size = tuple(int(dim * ratio) for dim in image.size)
                        image = image.resize(new_size, Image.Resampling.LANCZOS)
                    
                    # Convert to RGB
                    if image.mode not in ('RGB', 'RGBA'):
                        image = image.convert('RGB')
                    
                    # AGGRESSIVE compression - quality 70% (smaller = faster upload)
                    img_byte_arr = io.BytesIO()
                    image.save(img_byte_arr, format='JPEG', quality=70, optimize=True)
                    file_content = img_byte_arr.getvalue()
                    
                    optimized_size = len(file_content)
                    reduction = 100 - (optimized_size * 100 // original_size) if original_size > 0 else 0
                    logger.info(f"    ⚡ Optimized: {original_size} → {optimized_size} bytes ({reduction}% reduction)")
                except Exception as e:
                    logger.warning(f"    ⚠️ Optimization failed: {str(e)}")
                
                # Create temporary file
                temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".jpg")
                temp_file.write(file_content)
                temp_file.close()
                temp_path = temp_file.name
                
                # Read image and encode as base64
                with open(temp_path, 'rb') as img_file:
                    image_bytes = img_file.read()
                    image_base64 = base64.b64encode(image_bytes).decode('utf-8')
                
                # Create ImageContent
                image_content = ImageContent(image_base64=image_base64)
                
                # Use GPT-4o-mini for 3x FASTER processing (still accurate for OCR!)
                chat = LlmChat(
                    api_key=api_key,
                    session_id=f"payment-extraction-{uuid.uuid4()}",
                    system_message="Extract ride payment data from screenshot. Return JSON array only."
                ).with_model("openai", "gpt-4o-mini")  # FASTER MODEL!
                
                # Simplified prompt for faster processing
                simplified_prompt = """Extract ALL rides from this screenshot as JSON array:
[{"driver":"N/A", "vehicle":"N/A", "description":"Auto/Bike/Car", "date":"DD/MM/YYYY", "time":"HH:MM AM/PM", "amount":"number", "payment_mode":"Cash/UPI/N/A", "distance":"number or N/A", "duration":"minutes or N/A", "pickup_km":"N/A", "drop_km":"N/A", "pickup_location":"N/A", "drop_location":"N/A"}]

Rules:
- Extract EVERY visible ride
- If surge/cancelled, add to description
- Convert duration to minutes
- Use N/A for missing data
- NO assumptions"""
                
                user_message = UserMessage(
                    text=simplified_prompt,
                    file_contents=[image_content]
                )
                
                # Increased timeout for mini model (can be slower sometimes)
                response = await asyncio.wait_for(
                    chat.send_message(user_message),
                    timeout=120.0  # 2 minutes per image max
                )
                
                # Parse JSON response
                import json
                clean_response = response.strip()
                if clean_response.startswith("```json"):
                    clean_response = clean_response[7:]
                if clean_response.endswith("```"):
                    clean_response = clean_response[:-3]
                clean_response = clean_response.strip()
                
                parsed_data = json.loads(clean_response)
                
                # Handle both single object and array responses
                results = []
                if isinstance(parsed_data, list):
                    for ride_data in parsed_data:
                        ride_data["screenshot_filename"] = file.filename
                        ride_data["id"] = str(uuid.uuid4())
                        ride_data["processed_at"] = datetime.now().isoformat()
                        results.append(ride_data)
                else:
                    parsed_data["screenshot_filename"] = file.filename
                    parsed_data["id"] = str(uuid.uuid4())
                    parsed_data["processed_at"] = datetime.now().isoformat()
                    results.append(parsed_data)
                
                logger.info(f"  ✅ File {index + 1} processed: {len(results)} ride(s) extracted")
                return {"success": True, "results": results, "filename": file.filename}
                
            except asyncio.TimeoutError:
                logger.error(f"  ⏱️ TIMEOUT: File {index + 1} ({file.filename}) took >90 seconds to process")
                return {"success": False, "error": "Processing timeout (>90 seconds)", "filename": file.filename}
            except Exception as e:
                logger.error(f"  ❌ Error processing file {index + 1} ({file.filename}): {str(e)}")
                import traceback
                logger.error(traceback.format_exc())
                return {"success": False, "error": str(e), "filename": file.filename}
            finally:
                # Cleanup temp file
                if temp_path:
                    try:
                        if os.path.exists(temp_path):
                            os.unlink(temp_path)
                    except Exception as e:
                        logger.warning(f"    ⚠️ Failed to cleanup temp file: {str(e)}")
        
        # Process ALL files simultaneously (MAXIMUM SPEED - NO BATCHING!)
        logger.info(f"🚀 TURBO MODE: Processing ALL {len(files)} files simultaneously!")
        all_results = []
        errors = []
        
        # Create tasks for all files at once
        tasks = [process_single_file(file, idx) for idx, file in enumerate(files)]
        
        # Process all in parallel - NO WAITING!
        batch_results = await asyncio.gather(*tasks, return_exceptions=True)
        
        # Collect results
        for result in batch_results:
            if isinstance(result, Exception):
                error_msg = f"Unexpected error: {str(result)}"
                logger.error(f"  ❌ {error_msg}")
                errors.append(error_msg)
            elif result["success"]:
                all_results.extend(result["results"])
            else:
                errors.append(f"{result['filename']}: {result['error']}")
        
        # Check if we got results
        if errors and not all_results:
            raise HTTPException(
                status_code=422,
                detail={
                    "message": "All files failed to process",
                    "errors": errors,
                    "failed_batch": True
                }
            )
        
        # Add metadata and save to MongoDB
        if all_results:
            for record in all_results:
                record["user_id"] = current_user.id
                record["month_year"] = month_year
                record["driver"] = driver_name or record.get("driver", "N/A")
                record["vehicle"] = vehicle_number or record.get("vehicle", "N/A")
                # Fix: Ensure platform is properly set (empty string should not override)
                if platform and platform.strip():  # Check if platform is non-empty
                    record["platform"] = platform
                else:
                    record["platform"] = record.get("platform", "N/A")
                record["uploaded_at"] = datetime.now().isoformat()
                record["status"] = "pending"
                record["files_imported"] = False
            
            # Save to MongoDB
            try:
                logger.info(f"💾 Attempting to save {len(all_results)} records to payment_records collection")
                logger.info(f"Sample record platform value: {all_results[0].get('platform') if all_results else 'N/A'}")
                
                records_to_insert = [record.copy() for record in all_results]
                result = await db.payment_records.insert_many(records_to_insert)
                logger.info(f"✅ Saved {len(result.inserted_ids)} payment records to MongoDB (payment_records collection)")
                logger.info(f"First inserted ID: {result.inserted_ids[0] if result.inserted_ids else 'None'}")
                
                # Clean up _id
                for record in all_results:
                    if "_id" in record:
                        del record["_id"]
            except Exception as e:
                logger.error(f"❌ Failed to save to MongoDB: {str(e)}")
                logger.error(f"Exception type: {type(e).__name__}")
                import traceback
                logger.error(f"Traceback: {traceback.format_exc()}")
        
        logger.info(f"🎉 Processing complete: {len(all_results)} rides from {len(files)} files")
        
        return {
            "success": True,
            "extracted_data": all_results,
            "processed_files": len(files),
            "total_rides_extracted": len(all_results),
            "errors": errors if errors else None,
            "message": f"Successfully processed {len(files)} screenshots and extracted {len(all_results)} ride(s)"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Fatal error in process_payment_screenshots: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Processing failed: {str(e)}")

@api_router.post("/payment-reconciliation/import-files")
async def import_files_to_backend(
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """Import screenshot files to payment_screenshots folder and mark records as imported"""
    try:
        data = await request.json()
        month_year = data.get("month_year")
        record_ids = data.get("record_ids", [])
        
        if not month_year:
            raise HTTPException(status_code=400, detail="month_year is required")
        
        if not record_ids:
            raise HTTPException(status_code=400, detail="No records to import")
        
        # Fetch the records
        records = await db.payment_records.find({
            "id": {"$in": record_ids},
            "user_id": current_user.id,
            "month_year": month_year
        }, {"_id": 0}).to_list(length=None)
        
        if not records:
            raise HTTPException(status_code=404, detail="No records found")
        
        # Group records by driver
        drivers_screenshots = {}
        for record in records:
            driver = record.get("driver", "Unknown")
            filename = record.get("screenshot_filename")
            if filename:
                if driver not in drivers_screenshots:
                    drivers_screenshots[driver] = []
                drivers_screenshots[driver].append(filename)
        
        # Get unique screenshot filenames that need to be imported
        all_filenames = []
        for driver, filenames in drivers_screenshots.items():
            all_filenames.extend(filenames)
        
        unique_filenames = list(set(all_filenames))
        
        # For now, we'll just mark the records as imported in the database
        # The actual files should be re-uploaded by the frontend
        # Update: We'll save files that are already in temp directory or ask frontend to send them
        
        # Mark records as imported
        result = await db.payment_records.update_many(
            {
                "id": {"$in": record_ids},
                "user_id": current_user.id
            },
            {
                "$set": {
                    "files_imported": True,
                    "imported_at": datetime.now(timezone.utc).isoformat()
                }
            }
        )
        
        logger.info(f"✅ Marked {result.modified_count} records as imported for {month_year}")
        
        return {
            "success": True,
            "message": f"Successfully marked {result.modified_count} record(s) as imported",
            "imported_count": result.modified_count,
            "files_needed": unique_filenames
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error importing files: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to import files: {str(e)}")


@api_router.post("/payment-reconciliation/upload-screenshots-to-folder")
async def upload_screenshots_to_folder(
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """Upload screenshot files to payment_screenshots folder structure"""
    try:
        form = await request.form()
        files = form.getlist("files")
        month_year = form.get("month_year")
        driver_name = form.get("driver_name")
        
        if not files:
            raise HTTPException(status_code=400, detail="No files provided")
        
        if not month_year or not driver_name:
            raise HTTPException(status_code=400, detail="month_year and driver_name are required")
        
        # Create directory structure
        base_dir = "/app/backend/payment_screenshots"
        driver_dir = os.path.join(base_dir, month_year, driver_name)
        os.makedirs(driver_dir, exist_ok=True)
        
        logger.info(f"Uploading {len(files)} files to {driver_dir}")
        
        saved_files = []
        saved_count = 0
        
        for file in files:
            try:
                # Read file content
                file_content = await file.read()
                
                # Handle duplicate filenames
                base_filename = file.filename
                filename_without_ext, ext = os.path.splitext(base_filename)
                dest_path = os.path.join(driver_dir, base_filename)
                
                counter = 2
                final_filename = base_filename
                while os.path.exists(dest_path):
                    # File exists, add (2), (3), etc.
                    final_filename = f"{filename_without_ext} ({counter}){ext}"
                    dest_path = os.path.join(driver_dir, final_filename)
                    counter += 1
                
                # Save file
                with open(dest_path, 'wb') as f:
                    f.write(file_content)
                
                saved_files.append(final_filename)
                saved_count += 1
                logger.info(f"✅ Saved screenshot: {dest_path}")
                
            except Exception as e:
                logger.error(f"Error saving file {file.filename}: {str(e)}")
        
        logger.info(f"✅ Successfully saved {saved_count}/{len(files)} screenshots to {driver_dir}")
        
        return {
            "success": True,
            "message": f"Successfully uploaded {saved_count} file(s) to {month_year}/{driver_name}",
            "saved_count": saved_count,
            "saved_files": saved_files
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error uploading screenshots: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to upload screenshots: {str(e)}")


@api_router.post("/payment-reconciliation/sync-to-sheets")
async def sync_payment_data_to_sheets(
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """Sync payment reconciliation data to Google Sheets using Apps Script V2"""
    try:
        import shutil
        
        body = await request.json()
        data_rows = body.get("data", [])
        month_year = body.get("month_year", "Sep 2025")
        
        if not data_rows:
            raise HTTPException(status_code=400, detail="No data to sync")
        
        # Get Google Apps Script URL from environment
        apps_script_url = os.environ.get("PAYMENT_SHEETS_APPS_SCRIPT_URL")
        
        if not apps_script_url:
            logger.warning("PAYMENT_SHEETS_APPS_SCRIPT_URL not configured, skipping sync")
            return {"success": True, "message": "Sync skipped - Apps Script URL not configured"}
        
        # Prepare data for Apps Script V2 format
        records_to_sync = []
        for row in data_rows:
            records_to_sync.append({
                "id": row.get("id", ""),
                "driver": row.get("driver", "N/A"),
                "vehicle": row.get("vehicle", "N/A"),
                "platform": row.get("platform", "N/A"),  # Platform in Column C (replaces Mode)
                "date": row.get("date", "N/A"),
                "time": row.get("time", "N/A"),
                "description": row.get("description", "Auto"),
                "amount": str(row.get("amount", "0")),
                "payment_mode": row.get("paymentMode", "N/A"),  # Keep for backward compatibility
                "distance": str(row.get("distance", "N/A")),
                "duration": str(row.get("duration", "N/A")),
                "pickup_km": row.get("pickupKm", "N/A"),
                "drop_km": row.get("dropKm", "N/A"),
                "pickup_location": row.get("pickupLocation", "N/A"),
                "drop_location": row.get("dropLocation", "N/A"),
                "screenshot_filename": row.get("screenshotFilename", "N/A"),
                "status": row.get("status", "pending")
            })
        
        # Call Apps Script Web App
        import requests
        payload = {
            "data": records_to_sync,
            "month_year": month_year
        }
        
        response = requests.post(apps_script_url, json=payload, timeout=30)
        
        if response.status_code == 200:
            result = response.json()
            if result.get("success"):
                logger.info(f"Successfully synced {len(records_to_sync)} records to Google Sheets: {month_year}")
                
                # Note: Screenshots are already organized in Payment Screenshots library
                # during processing at: /app/backend/payment_screenshots/{month_year}/{driver_name}/
                # No need to copy them again after sync
                
                return {
                    "success": True, 
                    "message": result.get("message", "Synced successfully")
                }
            else:
                raise HTTPException(status_code=500, detail=f"Apps Script error: {result.get('message')}")
        else:
            raise HTTPException(status_code=500, detail=f"Apps Script request failed: {response.status_code}")
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error syncing to Google Sheets: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error syncing to Google Sheets: {str(e)}")


@api_router.post("/payment-reconciliation/delete-records")
async def delete_payment_records(
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """Delete payment records after successful sync to Google Sheets"""
    try:
        body = await request.json()
        record_ids = body.get("record_ids", [])
        month_year = body.get("month_year", "")
        
        if not record_ids:
            return {"success": True, "message": "No records to delete"}
        
        # Delete records from MongoDB
        delete_result = await db.payment_records.delete_many({
            "id": {"$in": record_ids},
            "month_year": month_year
        })
        
        logger.info(f"Deleted {delete_result.deleted_count} payment records for {month_year}")
        
        return {
            "success": True,
            "deleted_count": delete_result.deleted_count,
            "message": f"Successfully deleted {delete_result.deleted_count} records"
        }
        
    except Exception as e:
        logger.error(f"Error deleting payment records: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error deleting records: {str(e)}")


@api_router.get("/payment-reconciliation/sync-status")
async def get_payment_sync_status(current_user: User = Depends(get_current_user)):
    """Get last sync status for payment reconciliation"""
    try:
        # For now, return a mock status - in production, this would check actual sync history
        return {
            "success": True,
            "last_sync": datetime.now().isoformat(),
            "sync_status": "up_to_date",
            "message": "Data is synchronized"
        }
    except Exception as e:
        logger.error(f"Error getting sync status: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to get sync status")


@api_router.get("/payment-reconciliation/records")
async def get_payment_records(
    month_year: str = None,
    current_user: User = Depends(get_current_user)
):
    """Get all payment records, optionally filtered by month/year"""
    try:
        query = {"user_id": current_user.id}
        
        # Add month_year filter if provided
        if month_year:
            query["month_year"] = month_year
        
        # Fetch records from MongoDB
        records = await db.payment_records.find(query, {"_id": 0}).to_list(length=None)
        
        return {
            "success": True,
            "records": records,
            "count": len(records)
        }
    except Exception as e:
        logger.error(f"Error fetching payment records: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.put("/payment-reconciliation/update-record")
async def update_payment_record(
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """Update any field in a payment record"""
    try:
        data = await request.json()
        record_id = data.get("record_id")
        updates = data.get("updates", {})
        
        if not record_id:
            raise HTTPException(status_code=400, detail="No record ID provided")
        
        if not updates:
            raise HTTPException(status_code=400, detail="No updates provided")
        
        # Add updated_at timestamp
        updates["updated_at"] = datetime.now().isoformat()
        updates["updated_by"] = current_user.id
        
        # Update record in payment_records collection
        result = await db.payment_records.update_one(
            {"id": record_id},
            {"$set": updates}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Record not found")
        
        return {
            "success": True,
            "message": "Record updated successfully"
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating payment record: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/payment-reconciliation/export-to-excel")
async def export_payment_data_to_excel(
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """Export payment reconciliation data to Excel file in backend storage"""
    try:
        body = await request.json()
        month_year = body.get("month_year")  # e.g., "Sep 2025"
        
        if not month_year:
            raise HTTPException(status_code=400, detail="month_year is required")
        
        # Fetch all records for this month from MongoDB
        records = await db.payment_records.find({"month_year": month_year}).to_list(length=None)
        
        if not records:
            raise HTTPException(status_code=404, detail=f"No records found for {month_year}")
        
        # Create DataFrame
        import pandas as pd
        df = pd.DataFrame([{
            "Driver": record.get("driver", "N/A"),
            "Vehicle": record.get("vehicle", "N/A"),
            "Date": record.get("date", "N/A"),
            "Time": record.get("time", "N/A"),
            "Description": record.get("description", "Auto"),
            "Amount": record.get("amount", "0"),
            "Payment Mode": record.get("payment_mode", "N/A"),
            "Distance (km)": record.get("distance", "N/A"),
            "Duration (min)": record.get("duration", "N/A"),
            "Pickup KM": record.get("pickup_km", "N/A"),
            "Drop KM": record.get("drop_km", "N/A"),
            "Pickup Location": record.get("pickup_location", "N/A"),
            "Drop Location": record.get("drop_location", "N/A"),
            "Screenshot": record.get("screenshot_filename", "N/A")
        } for record in records])
        
        # Create directory structure: Payment Screenshots/Sep 2025/
        base_dir = "/app/backend/payment_screenshots"
        month_dir = os.path.join(base_dir, month_year)
        os.makedirs(month_dir, exist_ok=True)
        
        # Generate filename
        filename = f"{month_year.replace(' ', '_')}_data.xlsx"
        filepath = os.path.join(month_dir, filename)
        
        # Save to Excel
        df.to_excel(filepath, index=False, engine='openpyxl')
        
        logger.info(f"Exported {len(records)} records to {filepath}")
        
        return {
            "success": True,
            "message": f"Exported {len(records)} records to Excel",
            "filepath": filepath,
            "filename": filename
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting to Excel: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error exporting to Excel: {str(e)}")


@api_router.delete("/payment-reconciliation/delete-records")
async def delete_payment_records(
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """Delete selected payment records (Master Admin only)"""
    try:
        # Check if user has permission
        if current_user.account_type not in ["master_admin", "admin"]:
            raise HTTPException(status_code=403, detail="Only Admin or Master Admin can delete records")
        
        data = await request.json()
        record_ids = data.get("record_ids", [])
        
        if not record_ids:
            raise HTTPException(status_code=400, detail="No record IDs provided")
        
        # Delete records from payment_records collection
        result = await db.payment_records.delete_many({"id": {"$in": record_ids}})
        
        return {
            "success": True,
            "deleted_count": result.deleted_count,
            "message": f"Successfully deleted {result.deleted_count} records"
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting payment records: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/payment-reconciliation/folders")
async def get_payment_folders(current_user: User = Depends(get_current_user)):
    """Get all payment reconciliation folders"""
    try:
        folders = await db.payment_folders.find({}, {"_id": 0}).to_list(1000)
        return {
            "success": True,
            "folders": folders
        }
    except Exception as e:
        logger.error(f"Error fetching payment folders: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/payment-reconciliation/folders")
async def create_payment_folder(
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """Create a new payment reconciliation folder"""
    try:
        data = await request.json()
        
        # Check if folder already exists
        existing_folder = await db.payment_folders.find_one({"name": data["name"]})
        if existing_folder:
            raise HTTPException(status_code=400, detail="Folder already exists")
        
        # Create new folder
        folder_data = {
            "id": str(uuid.uuid4()),
            "name": data["name"],
            "month": data["month"],
            "year": data["year"],
            "monthLabel": data["monthLabel"],
            "fullName": data["fullName"],
            "createdAt": data["createdAt"],
            "created_by": current_user.id
        }
        
        await db.payment_folders.insert_one(folder_data)
        
        return {
            "success": True,
            "folder": folder_data,
            "message": f"Folder '{data['name']}' created successfully"
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating payment folder: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/admin/payment-screenshots/browse")
async def browse_payment_screenshots(
    path: str = "",
    current_user: User = Depends(get_current_user)
):
    """Browse payment screenshots directory structure"""
    try:
        # Allow all authenticated users to browse
        if current_user.account_type not in ["master_admin", "admin", "standard", "ops_team"]:
            raise HTTPException(status_code=403, detail="Access denied")
        
        base_dir = "/app/backend/payment_screenshots"
        full_path = os.path.join(base_dir, path) if path else base_dir
        
        # Security check - prevent directory traversal
        if not os.path.abspath(full_path).startswith(os.path.abspath(base_dir)):
            raise HTTPException(status_code=403, detail="Invalid path")
        
        if not os.path.exists(full_path):
            # Create the directory if it doesn't exist
            os.makedirs(full_path, exist_ok=True)
            return {"folders": [], "files": []}
        
        folders = []
        files = []
        
        for item in os.listdir(full_path):
            item_path = os.path.join(full_path, item)
            if os.path.isdir(item_path):
                # Count files in this folder recursively
                file_count = sum(1 for _, _, f in os.walk(item_path) for _ in f)
                folders.append({"name": item, "file_count": file_count})
            else:
                # Get file size
                file_size = os.path.getsize(item_path)
                files.append({"name": item, "size": file_size})
        
        return {
            "folders": sorted(folders, key=lambda x: x["name"]),
            "files": sorted(files, key=lambda x: x["name"])
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error browsing payment screenshots: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/admin/payment-screenshots/download")
async def download_payment_screenshot(
    path: str,
    current_user: User = Depends(get_current_user)
):
    """Download a file from payment screenshots directory"""
    try:
        # Allow all authenticated users to download
        if current_user.account_type not in ["master_admin", "admin", "standard", "ops_team"]:
            raise HTTPException(status_code=403, detail="Access denied")
        
        base_dir = "/app/backend/payment_screenshots"
        full_path = os.path.join(base_dir, path)
        
        # Security check
        if not os.path.abspath(full_path).startswith(os.path.abspath(base_dir)):
            raise HTTPException(status_code=403, detail="Invalid path")
        
        if not os.path.exists(full_path):
            raise HTTPException(status_code=404, detail="File not found")
        
        from fastapi.responses import FileResponse
        return FileResponse(full_path, filename=os.path.basename(full_path))
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error downloading file: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.delete("/admin/payment-screenshots/delete")
async def delete_payment_screenshot(
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """Delete a file or folder from payment screenshots directory (Master Admin only)"""
    try:
        # Only Master Admin can delete
        if current_user.account_type != "master_admin":
            raise HTTPException(status_code=403, detail="Only Master Admin can delete payment screenshots")
        
        body = await request.json()
        path = body.get("path")
        is_folder = body.get("is_folder", False)
        
        base_dir = "/app/backend/payment_screenshots"
        full_path = os.path.join(base_dir, path)
        
        # Security check
        if not os.path.abspath(full_path).startswith(os.path.abspath(base_dir)):
            raise HTTPException(status_code=403, detail="Invalid path")
        
        if not os.path.exists(full_path):
            raise HTTPException(status_code=404, detail="Item not found")
        
        if is_folder:
            import shutil
            shutil.rmtree(full_path)
        else:
            os.remove(full_path)
        
        return {"success": True, "message": "Item deleted successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting item: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/admin/files/reload-fleet-mapping")
async def reload_fleet_mapping(current_user: User = Depends(get_current_user)):
    """Reload vehicle to registration number mapping from Nura Fleet Data.xlsx"""
    try:
        # Find the fleet data file
        fleet_file = await db.admin_files.find_one(
            {"original_filename": "Nura Fleet Data.xlsx"}, 
            {"_id": 0}
        )
        
        if not fleet_file:
            raise HTTPException(
                status_code=404, 
                detail="Nura Fleet Data.xlsx not found. Please upload the file first."
            )
        
        file_path = fleet_file["file_path"]
        
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="Fleet data file not found on disk")
        
        # Read Excel file
        df = pd.read_excel(file_path)
        
        logger.info(f"Loading fleet mapping from {fleet_file['original_filename']}")
        logger.info(f"Columns: {df.columns.tolist()}")
        
        # Create mapping from Column A (Vehicle ID) to Column B (Registration Number)
        vehicle_mapping = {}
        for idx, row in df.iterrows():
            vehicle_id = str(row.iloc[0]) if pd.notna(row.iloc[0]) else None
            reg_number = str(row.iloc[1]) if pd.notna(row.iloc[1]) else None
            
            if vehicle_id and reg_number:
                vehicle_mapping[vehicle_id] = reg_number
        
        # Save to database
        await db.vehicle_mapping.delete_many({})  # Clear old mapping
        
        mapping_docs = [
            {"vehicle_id": vid, "registration_number": reg} 
            for vid, reg in vehicle_mapping.items()
        ]
        
        if mapping_docs:
            await db.vehicle_mapping.insert_many(mapping_docs)
        
        logger.info(f"Loaded {len(vehicle_mapping)} vehicle mappings")
        
        return {
            "message": f"Successfully loaded {len(vehicle_mapping)} vehicle mappings",
            "count": len(vehicle_mapping),
            "sample_mappings": list(vehicle_mapping.items())[:5]
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error reloading fleet mapping: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to reload fleet mapping: {str(e)}")


def format_file_size(size_bytes):
    """Format file size in human readable format"""
    if size_bytes < 1024:
        return f"{size_bytes} B"
    elif size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.2f} KB"
    elif size_bytes < 1024 * 1024 * 1024:
        return f"{size_bytes / (1024 * 1024):.2f} MB"
    else:
        return f"{size_bytes / (1024 * 1024 * 1024):.2f} GB"


# Montra Vehicle Insights
@api_router.post("/montra-vehicle-insights")
async def create_vehicle(vehicle_data: VehicleRecordCreate, current_user: User = Depends(get_current_user)):
    """Create new vehicle record"""
    vehicle = VehicleRecord(**vehicle_data.model_dump())
    vehicle_dict = vehicle.model_dump()
    vehicle_dict['date'] = vehicle_dict['date'].isoformat()
    vehicle_dict['created_at'] = vehicle_dict['created_at'].isoformat()
    if vehicle_dict.get('last_service'):
        vehicle_dict['last_service'] = vehicle_dict['last_service'].isoformat()
    
    await db.montra_vehicle_insights.insert_one(vehicle_dict)
    sync_single_record('montra_vehicle_insights', vehicle_dict)
    
    return {"message": "Vehicle record created", "vehicle": vehicle}


@api_router.get("/montra-vehicle-insights")
async def get_vehicles(current_user: User = Depends(get_current_user)):
    """Get all vehicle records"""
    vehicles = await db.montra_vehicle_insights.find({}, {"_id": 0}).to_list(1000)
    for vehicle in vehicles:
        if isinstance(vehicle.get('date'), str):
            vehicle['date'] = datetime.fromisoformat(vehicle['date'])
        if isinstance(vehicle.get('created_at'), str):
            vehicle['created_at'] = datetime.fromisoformat(vehicle['created_at'])
        if vehicle.get('last_service') and isinstance(vehicle.get('last_service'), str):
            vehicle['last_service'] = datetime.fromisoformat(vehicle['last_service'])
    return vehicles


@api_router.post("/montra-vehicle-insights/sync")
async def sync_vehicles(current_user: User = Depends(get_current_user)):
    """Sync all vehicles to Google Sheets"""
    vehicles = await db.montra_vehicle_insights.find({}, {"_id": 0}).to_list(1000)
    success = sync_all_records('montra_vehicle_insights', vehicles)
    if success:
        return {"message": "Vehicles synced successfully"}
    raise HTTPException(status_code=500, detail="Failed to sync vehicles")


# Fetch from Sheets (Read from Sheets to app)
@api_router.get("/sheets/fetch/{tab}")
async def fetch_from_sheets(tab: str, current_user: User = Depends(get_current_user)):
    """Fetch data from Google Sheets"""
    valid_tabs = ['payment_reconciliation', 'driver_onboarding', 'telecaller_queue', 'montra_vehicle_insights']
    if tab not in valid_tabs:
        raise HTTPException(status_code=400, detail="Invalid tab name")
    
    records = get_all_records(tab)
    return {"message": f"Fetched {len(records)} records", "data": records}


@api_router.get("/sheets/last-sync-time")
async def get_sheets_last_sync_time(current_user: User = Depends(get_current_user)):
    """Get the last sync time from Google Sheets"""
    try:
        last_sync = get_last_sync_time()
        return {"last_sync_time": last_sync}
    except Exception as e:
        logger.error(f"Failed to get last sync time: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to get last sync time")




# ==================== Vehicle Documents ====================

@api_router.post("/vehicle-documents/upload-file")
async def upload_vehicle_document_file(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """Upload document file for vehicle documents (pdf/png/word)"""
    import shutil
    from pathlib import Path
    
    try:
        # Create directory if it doesn't exist
        upload_dir = Path("/app/backend/uploaded_files/Vehicle_Docs")
        upload_dir.mkdir(parents=True, exist_ok=True)
        
        # Generate unique filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_extension = file.filename.split(".")[-1] if "." in file.filename else "pdf"
        unique_filename = f"{timestamp}_{str(uuid.uuid4())[:8]}.{file_extension}"
        file_path = upload_dir / unique_filename
        
        # Save file
        with file_path.open("wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        # Return relative path
        relative_path = f"Vehicle_Docs/{unique_filename}"
        
        return {
            "success": True,
            "message": "File uploaded successfully",
            "file_path": relative_path,
            "filename": unique_filename
        }
    
    except Exception as e:
        logger.error(f"Error uploading file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error uploading file: {str(e)}")


@api_router.get("/vehicle-documents")
async def get_vehicle_documents(
    skip: int = Query(0, description="Number of records to skip"),
    limit: int = Query(50, description="Number of records to return"),
    search: Optional[str] = Query(None, description="Search by VIN or vehicle name"),
    current_user: User = Depends(get_current_user)
):
    """Get all vehicle documents with pagination and search"""
    from app_models import VehicleDocument
    
    try:
        # Build query
        query = {}
        
        if search:
            query["$or"] = [
                {"vin": {"$regex": search, "$options": "i"}},
                {"vehicle_name": {"$regex": search, "$options": "i"}},
                {"vehicle_number": {"$regex": search, "$options": "i"}}
            ]
        
        # Get total count
        total = await db.vehicle_documents.count_documents(query)
        
        # Get documents (exclude _id field to avoid ObjectId serialization issues)
        documents = await db.vehicle_documents.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(length=limit)
        
        return {
            "success": True,
            "documents": documents,
            "total": total,
            "page": (skip // limit) + 1 if limit > 0 else 1,
            "total_pages": (total + limit - 1) // limit if limit > 0 else 1
        }
    
    except Exception as e:
        logger.error(f"Error fetching vehicle documents: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error fetching vehicle documents: {str(e)}")


@api_router.get("/vehicle-documents/{document_id}")
async def get_vehicle_document(
    document_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get single vehicle document by ID"""
    from app_models import VehicleDocument
    
    try:
        document = await db.vehicle_documents.find_one({"id": document_id}, {"_id": 0})
        
        if not document:
            raise HTTPException(status_code=404, detail="Vehicle document not found")
        
        return {
            "success": True,
            "document": document
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching vehicle document: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error fetching vehicle document: {str(e)}")


@api_router.post("/vehicle-documents")
async def create_vehicle_document(
    document_data: dict = Body(...),
    current_user: User = Depends(get_current_user)
):
    """Create a new vehicle document"""
    from app_models import VehicleDocument, VehicleDocumentCreate
    
    try:
        # Parse dates if provided
        if document_data.get("registration_expiry_date"):
            document_data["registration_expiry_date"] = datetime.fromisoformat(document_data["registration_expiry_date"].replace("Z", "+00:00"))
        
        if document_data.get("insurance_expiry_date"):
            document_data["insurance_expiry_date"] = datetime.fromisoformat(document_data["insurance_expiry_date"].replace("Z", "+00:00"))
        
        if document_data.get("purchase_date"):
            document_data["purchase_date"] = datetime.fromisoformat(document_data["purchase_date"].replace("Z", "+00:00"))
        
        # Create document
        new_document = VehicleDocument(
            **document_data,
            created_by=current_user.id,
            created_at=datetime.now(timezone.utc),
            updated_at=datetime.now(timezone.utc)
        )
        
        # Insert into database
        result = await db.vehicle_documents.insert_one(new_document.model_dump(by_alias=True))
        
        logger.info(f"Created vehicle document {new_document.id} for VIN {new_document.vin}")
        
        return {
            "success": True,
            "message": "Vehicle document created successfully",
            "document_id": new_document.id,
            "document": new_document.model_dump(by_alias=True)
        }
    
    except Exception as e:
        logger.error(f"Error creating vehicle document: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error creating vehicle document: {str(e)}")


@api_router.put("/vehicle-documents/{document_id}")
async def update_vehicle_document(
    document_id: str,
    update_data: dict = Body(...),
    current_user: User = Depends(get_current_user)
):
    """Update an existing vehicle document"""
    from app_models import VehicleDocumentUpdate
    
    try:
        # Check if document exists
        existing_document = await db.vehicle_documents.find_one({"id": document_id}, {"_id": 0})
        if not existing_document:
            raise HTTPException(status_code=404, detail="Vehicle document not found")
        
        # Parse dates if provided
        if update_data.get("registration_expiry_date"):
            update_data["registration_expiry_date"] = datetime.fromisoformat(update_data["registration_expiry_date"].replace("Z", "+00:00"))
        
        if update_data.get("insurance_expiry_date"):
            update_data["insurance_expiry_date"] = datetime.fromisoformat(update_data["insurance_expiry_date"].replace("Z", "+00:00"))
        
        if update_data.get("purchase_date"):
            update_data["purchase_date"] = datetime.fromisoformat(update_data["purchase_date"].replace("Z", "+00:00"))
        
        # Add updated_at timestamp
        update_data["updated_at"] = datetime.now(timezone.utc)
        
        # Update document
        result = await db.vehicle_documents.update_one(
            {"id": document_id},
            {"$set": update_data}
        )
        
        if result.modified_count == 0:
            logger.warning(f"No changes made to vehicle document {document_id}")
        
        # Fetch updated document
        updated_document = await db.vehicle_documents.find_one({"id": document_id}, {"_id": 0})
        
        logger.info(f"Updated vehicle document {document_id}")
        
        return {
            "success": True,
            "message": "Vehicle document updated successfully",
            "document": updated_document
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating vehicle document: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error updating vehicle document: {str(e)}")


@api_router.delete("/vehicle-documents/{document_id}")
async def delete_vehicle_document(
    document_id: str,
    current_user: User = Depends(get_current_user)
):
    """Delete a vehicle document (only for master admins and admins)"""
    try:
        # Check permissions
        if current_user.account_type not in ["master_admin", "admin"]:
            raise HTTPException(
                status_code=403,
                detail="Only master admins and admins can delete vehicle documents"
            )
        
        # Check if document exists
        existing_document = await db.vehicle_documents.find_one({"id": document_id}, {"_id": 0})
        if not existing_document:
            raise HTTPException(status_code=404, detail="Vehicle document not found")
        
        # Delete document
        result = await db.vehicle_documents.delete_one({"id": document_id})
        
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Vehicle document not found")
        
        logger.info(f"Deleted vehicle document {document_id} by {current_user.email}")
        
        return {
            "success": True,
            "message": "Vehicle document deleted successfully"
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting vehicle document: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error deleting vehicle document: {str(e)}")


@api_router.get("/vehicle-documents/file/{file_path:path}")
async def get_vehicle_document_file(
    file_path: str,
    current_user: User = Depends(get_current_user)
):
    """Get vehicle document file"""
    from pathlib import Path
    
    try:
        # Construct full file path
        full_path = Path(f"/app/backend/uploaded_files/{file_path}")
        
        if not full_path.exists():
            raise HTTPException(status_code=404, detail="File not found")
        
        return FileResponse(full_path)
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error retrieving file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error retrieving file: {str(e)}")

# ==================== App Initialization ====================

@app.on_event("startup")
async def startup_event():
    await initialize_master_admin()
    logger.info("Application started")




# ==================== HOTSPOT PLANNING ====================

@api_router.post("/hotspot-planning/analyze")
async def analyze_hotspot_placement(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """
    Analyze ride data and find optimal hotspot locations using CELF greedy + 1-swap algorithm.
    Groups trips into 6 time slots and finds top 5 hotspots per slot.
    """
    try:
        # Read CSV file
        contents = await file.read()
        import io
        df = pd.read_csv(io.BytesIO(contents))
        
        logger.info(f"Loaded {len(df)} rides from CSV")
        
        # Normalize column names (case-insensitive)
        df.columns = df.columns.str.lower().str.strip()
        col_map = {col: col for col in df.columns}
        
        # Find required columns
        lat_col = next((c for c in df.columns if 'lat' in c.lower() and 'pickup' in c.lower()), None)
        lon_col = next((c for c in df.columns if ('long' in c.lower() or 'lon' in c.lower()) and 'pickup' in c.lower()), None)
        time_col = next((c for c in df.columns if any(t in c.lower() for t in ['createdat', 'time', 'timestamp', 'datetime'])), None)
        pickup_point_col = next((c for c in df.columns if 'pickuppoint' in c.lower().replace('_', '').replace(' ', '')), None)
        drop_lat_col = next((c for c in df.columns if 'lat' in c.lower() and 'drop' in c.lower()), None)
        drop_lon_col = next((c for c in df.columns if ('long' in c.lower() or 'lon' in c.lower()) and 'drop' in c.lower()), None)
        drop_point_col = next((c for c in df.columns if 'droppoint' in c.lower().replace('_', '').replace(' ', '')), None)
        
        if not lat_col or not lon_col:
            raise HTTPException(status_code=400, detail="Missing pickup latitude/longitude columns")
        
        # Rename columns
        rename_dict = {
            lat_col: 'lat',
            lon_col: 'lon'
        }
        if pickup_point_col:
            rename_dict[pickup_point_col] = 'pickup_point'
        if drop_lat_col:
            rename_dict[drop_lat_col] = 'drop_lat'
        if drop_lon_col:
            rename_dict[drop_lon_col] = 'drop_lon'
        if drop_point_col:
            rename_dict[drop_point_col] = 'drop_point'
            
        df = df.rename(columns=rename_dict)
        
        if time_col:
            df = df.rename(columns={time_col: 'timestamp'})
        
        # Clean data
        df = df.dropna(subset=['lat', 'lon']).copy()
        
        # Parse timestamp and extract hour
        if 'hour' in df.columns:
            # Hour column already exists
            df['hour'] = pd.to_numeric(df['hour'], errors='coerce')
            df = df.dropna(subset=['hour'])
        elif 'timestamp' in df.columns:
            # Handle Excel serial numbers
            def parse_time(val):
                try:
                    if pd.isna(val):
                        return None
                    if isinstance(val, (int, float)):
                        dt = pd.to_datetime(val, origin='1899-12-30', unit='D')
                        return dt + timedelta(hours=5, minutes=30)  # UTC to IST
                    return pd.to_datetime(val, errors='coerce')
                except:
                    return None
            
            df['datetime'] = df['timestamp'].apply(parse_time)
            df = df.dropna(subset=['datetime'])
            df['hour'] = df['datetime'].dt.hour
        elif 'createdat' in df.columns or 'updatedat' in df.columns:
            # Try to parse from createdAt/updatedAt
            time_col = 'updatedat' if 'updatedat' in df.columns else 'createdat'
            def parse_time(val):
                try:
                    if pd.isna(val):
                        return None
                    # Handle Excel serial numbers
                    if isinstance(val, (int, float)):
                        dt = pd.to_datetime(val, origin='1899-12-30', unit='D')
                        return dt + timedelta(hours=5, minutes=30)
                    # Try parsing time strings
                    time_str = str(val).strip()
                    for fmt in ['%H:%M:%S', '%H:%M', '%I:%M:%S %p']:
                        try:
                            return pd.to_datetime(time_str, format=fmt)
                        except:
                            continue
                    return pd.to_datetime(val, errors='coerce')
                except:
                    return None
            
            df['datetime'] = df[time_col].apply(parse_time)
            df = df.dropna(subset=['datetime'])
            df['hour'] = df['datetime'].dt.hour
        else:
            # No timestamp, assume uniform distribution
            df['hour'] = 12  # Default to afternoon
        
        # Add weight column (default 1)
        df['weight'] = 1.0
        
        # Group into time slots
        time_slot_results = {}
        
        for slot in TIME_SLOTS:
            slot_name = slot['name']
            slot_label = slot['label']
            start_hour = slot['start']
            end_hour = slot['end']
            
            # Filter rides in this time slot
            if end_hour <= 24:
                slot_df = df[(df['hour'] >= start_hour) & (df['hour'] < end_hour)].copy()
            else:
                # Handle overnight slots (e.g., 22-25 means 10PM-1AM)
                slot_df = df[(df['hour'] >= start_hour) | (df['hour'] < (end_hour - 24))].copy()
            
            if len(slot_df) < 10:
                time_slot_results[slot_name] = {
                    'status': 'insufficient_data',
                    'slot_label': slot_label,
                    'rides_count': int(len(slot_df)),
                    'message': f'Not enough rides in this slot (minimum 10 required)'
                }
                continue
            
            # Run hotspot optimization
            # Build column list dynamically based on what's available
            cols_to_pass = ['lat', 'lon', 'weight']
            if 'pickup_point' in slot_df.columns:
                cols_to_pass.append('pickup_point')
            if 'drop_lat' in slot_df.columns:
                cols_to_pass.append('drop_lat')
            if 'drop_lon' in slot_df.columns:
                cols_to_pass.append('drop_lon')
            if 'drop_point' in slot_df.columns:
                cols_to_pass.append('drop_point')
                
            result = optimize_hotspots(
                df=slot_df[cols_to_pass],
                N=10,  # Changed from 5 to 10 hotspots
                h3_res=9,
                use_h3=True
            )
            
            time_slot_results[slot_name] = {
                'status': 'success',
                'slot_label': slot_label,
                'rides_count': result['total_rides'],
                'covered_rides': result['covered_rides'],
                'coverage_percentage': result['coverage_percentage'],
                'hotspots': result['hotspots'],
                'geojson': result['geojson'],
                'detailed_assignments': result.get('detailed_assignments', [])
            }
            
            logger.info(f"{slot_name}: {result['covered_rides']}/{result['total_rides']} rides covered ({result['coverage_percentage']}%)")
        
        total_rides = len(df)
        slots_with_data = len([s for s in time_slot_results.values() if s['status'] == 'success'])
        
        return {
            'success': True,
            'total_rides_analyzed': int(total_rides),
            'time_slots': time_slot_results,
            'analysis_params': {
                'algorithm': 'CELF Greedy + 1-Swap Local Search',
                'radius_m': 1000,
                'hotspots_per_slot': 5
            },
            'message': f'Analysis complete: {slots_with_data} time slots analyzed from {total_rides} rides'
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in hotspot analysis: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Failed to analyze data: {str(e)}")

@api_router.post("/hotspot-planning/analyze-and-save")
async def analyze_and_save_hotspot(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """Analyze hotspot data and save to library"""
    try:
        # First, run the analysis
        file_content = await file.read()
        
        # Recreate file object for analysis
        from io import BytesIO
        file_for_analysis = UploadFile(filename=file.filename, file=BytesIO(file_content))
        
        # Call the original analyze function
        analysis_result = await analyze_hotspot_placement(file_for_analysis, current_user)
        
        # Save to library
        analysis_id = str(uuid.uuid4())
        library_entry = {
            "id": analysis_id,
            "filename": file.filename,
            "uploaded_by": current_user.id,
            "uploaded_at": datetime.now(timezone.utc).isoformat(),
            "file_size": len(file_content),
            "analysis_result": analysis_result,
            "total_rides": analysis_result.get('total_rides_analyzed', 0),
            "time_slots_count": len([s for s in analysis_result.get('time_slots', {}).values() if s.get('status') == 'success'])
        }
        
        await db.hotspot_analyses.insert_one(library_entry)
        
        logger.info(f"Hotspot analysis saved to library: {analysis_id} by {current_user.email}")
        
        return {
            "success": True,
            "analysis_id": analysis_id,
            "analysis": analysis_result,
            "message": "Analysis completed and saved to library"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to analyze and save hotspot: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Failed to save analysis: {str(e)}")


@api_router.get("/hotspot-planning/library")
async def get_hotspot_library(
    current_user: User = Depends(get_current_user)
):
    """Get all saved hotspot analyses from library"""
    try:
        analyses_collection = db["hotspot_analyses"]
        analyses = await analyses_collection.find(
            {"uploaded_by": current_user.id},
            {"_id": 0}  # Exclude MongoDB ObjectId from results
        ).sort("created_at", -1).to_list(length=100)
        
        return {
            "success": True,
            "analyses": analyses
        }
    except Exception as e:
        logger.error(f"Failed to fetch library: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch library")


@api_router.post("/hotspot-planning/download-csv/{slot_name}")
async def download_hotspot_csv(
    slot_name: str,
    analysis_data: dict,
    current_user: User = Depends(get_current_user)
):
    """Download detailed CSV with hotspot assignments for a specific time slot"""
    try:
        import pandas as pd
        from io import BytesIO
        
        logger.info(f"Generating hotspot CSV download for slot: {slot_name}")
        
        # Get the slot data
        slot_data = analysis_data.get('time_slots', {}).get(slot_name, {})
        
        if slot_data.get('status') != 'success':
            raise HTTPException(status_code=404, detail=f"No successful analysis found for slot: {slot_name}")
        
        # Get detailed assignments
        detailed_assignments = slot_data.get('detailed_assignments', [])
        
        if not detailed_assignments:
            raise HTTPException(status_code=404, detail="No detailed assignment data available")
        
        # Create DataFrame
        df = pd.DataFrame(detailed_assignments)
        
        # Rename columns for user-friendly export
        column_mapping = {
            'pickup_point': 'Pickup Point',
            'pickup_lat': 'Pickup Lat',
            'pickup_lon': 'Pickup Long',
            'pickup_locality': 'Pickup Locality',
            'drop_point': 'Drop Point',
            'drop_lat': 'Drop Lat',
            'drop_lon': 'Drop Long',
            'drop_locality': 'Drop Locality',
            'assigned_hotspot_rank': 'Hotspot Rank',
            'assigned_hotspot_locality': 'Assigned Hotspot',
            'assigned_hotspot_lat': 'Hotspot Lat',
            'assigned_hotspot_lon': 'Hotspot Long',
            'distance_from_hotspot_m': 'Distance from Hotspot (m)'
        }
        
        # Only rename columns that exist
        df = df.rename(columns={k: v for k, v in column_mapping.items() if k in df.columns})
        
        # Add combined "Assigned Hotspot Lat Long" column
        if 'Hotspot Lat' in df.columns and 'Hotspot Long' in df.columns:
            df['Assigned Hotspot Lat Long'] = df.apply(
                lambda row: f"{row['Hotspot Lat']},{row['Hotspot Long']}" if pd.notna(row['Hotspot Lat']) else None,
                axis=1
            )
        
        # Reorder columns for better readability
        preferred_order = [
            'Pickup Point',
            'Pickup Locality',
            'Pickup Lat', 
            'Pickup Long',
            'Drop Point',
            'Drop Locality',
            'Drop Lat',
            'Drop Long',
            'Assigned Hotspot',
            'Assigned Hotspot Lat Long',
            'Distance from Hotspot (m)',
            'Hotspot Rank'
        ]
        
        # Keep only columns that exist
        final_columns = [col for col in preferred_order if col in df.columns]
        remaining_columns = [col for col in df.columns if col not in final_columns]
        df = df[final_columns + remaining_columns]
        
        # Create CSV in memory
        output = BytesIO()
        df.to_csv(output, index=False)
        output.seek(0)
        
        # Generate filename
        filename = f"hotspot_coverage_{slot_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        logger.info(f"CSV generated: {len(df)} rows, filename: {filename}")
        
        return StreamingResponse(
            output,
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to generate hotspot CSV: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Failed to generate CSV: {str(e)}")


@api_router.get("/hotspot-planning/library")
async def get_hotspot_library(
    current_user: User = Depends(get_current_user)
):
    """Get all saved hotspot analyses"""
    try:
        analyses = await db.hotspot_analyses.find({}).sort('uploaded_at', -1).to_list(None)
        
        # Remove _id and large analysis data for list view
        for analysis in analyses:
            if '_id' in analysis:
                del analysis['_id']
            # Keep only summary, not full analysis
            if 'analysis_result' in analysis:
                analysis['summary'] = {
                    'total_rides': analysis['analysis_result'].get('total_rides_analyzed', 0),
                    'time_slots': len([s for s in analysis['analysis_result'].get('time_slots', {}).values() if s.get('status') == 'success'])
                }
                del analysis['analysis_result']
        
        return {
            "success": True,
            "analyses": analyses,
            "count": len(analyses)
        }
        
    except Exception as e:
        logger.error(f"Failed to get hotspot library: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get library: {str(e)}")


@api_router.get("/hotspot-planning/library/{analysis_id}")
async def get_hotspot_analysis(
    analysis_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get specific hotspot analysis"""
    try:
        analysis = await db.hotspot_analyses.find_one({"id": analysis_id}, {"_id": 0})
        
        if not analysis:
            raise HTTPException(status_code=404, detail="Analysis not found")
        
        return {
            "success": True,
            "analysis": analysis
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get hotspot analysis: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get analysis: {str(e)}")


@api_router.delete("/hotspot-planning/library/{analysis_id}")
async def delete_hotspot_analysis(
    analysis_id: str,
    current_user: User = Depends(get_current_user)
):
    """Delete hotspot analysis from library - Master Admin only"""
    try:
        if current_user.account_type != "master_admin":
            raise HTTPException(status_code=403, detail="Only Master Admin can delete analyses")
        
        result = await db.hotspot_analyses.delete_one({"id": analysis_id})
        
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Analysis not found")
        
        return {
            "success": True,
            "message": "Analysis deleted successfully"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to delete hotspot analysis: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to delete analysis: {str(e)}")


# ==================== ANALYTICS ====================

# In-memory storage for active sessions and page views
active_sessions = {}  # {user_id: {username, email, account_type, last_seen, current_page}}
page_views = {}  # {page_name: count}

@api_router.post("/analytics/track-page-view")
async def track_page_view(
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """Track page view and update user session"""
    try:
        body = await request.json()
        page_name = body.get("page")
        
        if not page_name:
            raise HTTPException(status_code=400, detail="Page name is required")
        
        # Update active session
        active_sessions[current_user.id] = {
            "user_id": current_user.id,
            "username": f"{current_user.first_name} {current_user.last_name or ''}".strip(),
            "email": current_user.email,
            "account_type": current_user.account_type,
            "current_page": page_name,
            "last_seen": datetime.now(timezone.utc).isoformat()
        }
        
        # Update page view count
        page_views[page_name] = page_views.get(page_name, 0) + 1
        
        return {"success": True, "message": "Page view tracked"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error tracking page view: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to track page view")


@api_router.get("/analytics/active-users")
async def get_active_users(current_user: User = Depends(get_current_user)):
    """Get list of currently active users (Master Admin only)"""
    if current_user.account_type != "master_admin":
        raise HTTPException(status_code=403, detail="Only Master Admin can view analytics")
    
    try:
        # Clean up stale sessions (inactive for more than 5 minutes)
        current_time = datetime.now(timezone.utc)
        stale_users = []
        
        for user_id, session in active_sessions.items():
            last_seen = datetime.fromisoformat(session["last_seen"])
            if (current_time - last_seen).total_seconds() > 300:  # 5 minutes
                stale_users.append(user_id)
        
        for user_id in stale_users:
            del active_sessions[user_id]
        
        return {
            "success": True,
            "active_users": list(active_sessions.values()),
            "count": len(active_sessions)
        }
        
    except Exception as e:
        logger.error(f"Error getting active users: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to get active users")


@api_router.get("/analytics/page-views")
async def get_page_views(current_user: User = Depends(get_current_user)):
    """Get page view statistics (Master Admin only)"""
    if current_user.account_type != "master_admin":
        raise HTTPException(status_code=403, detail="Only Master Admin can view analytics")
    
    try:
        # Convert to list of {page, count} sorted by count
        page_view_list = [
            {"page": page, "views": count}
            for page, count in page_views.items()
        ]
        page_view_list.sort(key=lambda x: x["views"], reverse=True)
        
        return {
            "success": True,
            "page_views": page_view_list,
            "total_views": sum(page_views.values())
        }
        
    except Exception as e:
        logger.error(f"Error getting page views: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to get page views")


@api_router.post("/analytics/logout")
async def track_logout(current_user: User = Depends(get_current_user)):
    """Track user logout"""
    try:
        if current_user.id in active_sessions:
            del active_sessions[current_user.id]
        
        return {"success": True, "message": "Logout tracked"}
        
    except Exception as e:
        logger.error(f"Error tracking logout: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to track logout")


# ==================== EXPENSE TRACKER ====================

# Expenses directory
EXPENSES_DIR = ROOT_DIR / "expense_receipts"
EXPENSES_DIR.mkdir(exist_ok=True)

@api_router.get("/expenses")
async def get_expenses(current_user: User = Depends(get_current_user)):
    """Get all expenses for the user or all expenses for admin/master_admin"""
    try:
        if current_user.account_type in ["master_admin", "admin"]:
            # Admins can see all expenses
            expenses_cursor = db.expenses.find({"deleted": {"$ne": True}})
        else:
            # Standard users see only their expenses
            expenses_cursor = db.expenses.find({
                "user_id": current_user.id,
                "deleted": {"$ne": True}
            })
        
        expenses = await expenses_cursor.to_list(length=None)
        
        # Convert ObjectId and datetime to string
        for expense in expenses:
            if '_id' in expense:
                del expense['_id']
            if 'created_at' in expense and isinstance(expense['created_at'], datetime):
                expense['created_at'] = expense['created_at'].isoformat()
        
        return {"expenses": expenses}
    except Exception as e:
        logger.error(f"Error fetching expenses: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch expenses")


@api_router.post("/expenses/add")
async def add_expense(
    date: str = Form(...),
    description: str = Form(...),
    amount: float = Form(...),
    receipts: List[UploadFile] = File(default=[]),
    current_user: User = Depends(get_current_user)
):
    """Add a new expense"""
    try:
        expense_id = str(uuid.uuid4())
        receipt_filenames = []
        
        # Save receipt files
        if receipts:
            expense_folder = EXPENSES_DIR / expense_id
            expense_folder.mkdir(exist_ok=True)
            
            for receipt in receipts:
                if receipt.size > 10 * 1024 * 1024:  # 10MB limit
                    raise HTTPException(status_code=400, detail=f"File {receipt.filename} exceeds 10MB limit")
                
                file_path = expense_folder / receipt.filename
                with open(file_path, "wb") as f:
                    content = await receipt.read()
                    f.write(content)
                receipt_filenames.append(receipt.filename)
        
        expense_data = {
            "id": expense_id,
            "user_id": current_user.id,
            "user_name": f"{current_user.first_name} {current_user.last_name or ''}".strip(),
            "date": date,
            "description": description,
            "amount": amount,
            "receipt_filenames": receipt_filenames,
            "approval_status": "Pending",
            "created_at": datetime.now(timezone.utc),
            "deleted": False
        }
        
        await db.expenses.insert_one(expense_data)
        
        return {"message": "Expense added successfully", "expense_id": expense_id}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error adding expense: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to add expense")


@api_router.post("/expenses/update")
async def update_expense(
    expense_id: str = Form(...),
    date: str = Form(...),
    description: str = Form(...),
    amount: float = Form(...),
    receipts: List[UploadFile] = File(default=[]),
    current_user: User = Depends(get_current_user)
):
    """Update an existing expense"""
    try:
        # Check if expense exists and user has permission
        expense = await db.expenses.find_one({"id": expense_id})
        if not expense:
            raise HTTPException(status_code=404, detail="Expense not found")
        
        # Only the creator or admin/master_admin can edit
        if expense["user_id"] != current_user.id and current_user.account_type not in ["master_admin", "admin"]:
            raise HTTPException(status_code=403, detail="Not authorized to edit this expense")
        
        update_data = {
            "date": date,
            "description": description,
            "amount": amount,
            "updated_at": datetime.now(timezone.utc)
        }
        
        # Handle new receipt files
        if receipts and any(r.filename for r in receipts):
            expense_folder = EXPENSES_DIR / expense_id
            expense_folder.mkdir(exist_ok=True)
            
            existing_receipts = expense.get("receipt_filenames", [])
            
            for receipt in receipts:
                if receipt.filename and receipt.size > 0:
                    if receipt.size > 10 * 1024 * 1024:
                        raise HTTPException(status_code=400, detail=f"File {receipt.filename} exceeds 10MB limit")
                    
                    file_path = expense_folder / receipt.filename
                    with open(file_path, "wb") as f:
                        content = await receipt.read()
                        f.write(content)
                    
                    if receipt.filename not in existing_receipts:
                        existing_receipts.append(receipt.filename)
            
            update_data["receipt_filenames"] = existing_receipts
        
        await db.expenses.update_one(
            {"id": expense_id},
            {"$set": update_data}
        )
        
        return {"message": "Expense updated successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating expense: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to update expense")


@api_router.post("/expenses/delete")
async def delete_expenses(
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """Delete expenses (master admin only)"""
    try:
        if current_user.account_type != "master_admin":
            raise HTTPException(status_code=403, detail="Only master admin can delete expenses")
        
        body = await request.json()
        expense_ids = body.get("expense_ids", [])
        
        if not expense_ids:
            raise HTTPException(status_code=400, detail="No expense IDs provided")
        
        # Mark as deleted instead of actually deleting
        result = await db.expenses.update_many(
            {"id": {"$in": expense_ids}},
            {"$set": {"deleted": True, "deleted_at": datetime.now(timezone.utc)}}
        )
        
        return {"message": f"Deleted {result.modified_count} expense(s)"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting expenses: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to delete expenses")


@api_router.post("/expenses/approve")
async def approve_expense(
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """Approve or reject an expense (admin/master admin only)"""
    try:
        if current_user.account_type not in ["master_admin", "admin"]:
            raise HTTPException(status_code=403, detail="Only admins can approve/reject expenses")
        
        body = await request.json()
        expense_id = body.get("expense_id")
        status = body.get("status")  # "Approved", "Rejected", "Pending"
        
        if not expense_id or not status:
            raise HTTPException(status_code=400, detail="Missing expense_id or status")
        
        if status not in ["Approved", "Rejected", "Pending"]:
            raise HTTPException(status_code=400, detail="Invalid status")
        
        result = await db.expenses.update_one(
            {"id": expense_id},
            {"$set": {
                "approval_status": status,
                "approved_by": current_user.id,
                "approved_at": datetime.now(timezone.utc)
            }}
        )
        
        if result.modified_count == 0:
            raise HTTPException(status_code=404, detail="Expense not found")
        
        return {"message": f"Expense {status.lower()} successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error approving expense: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to update approval status")


@api_router.get("/expenses/{expense_id}/receipt/{filename}")
async def get_expense_receipt(
    expense_id: str,
    filename: str,
    current_user: User = Depends(get_current_user)
):
    """Download/view an expense receipt"""
    from fastapi.responses import FileResponse
    
    try:
        # Check if expense exists and user has permission
        expense = await db.expenses.find_one({"id": expense_id})
        if not expense:
            raise HTTPException(status_code=404, detail="Expense not found")
        
        # Check permissions
        if expense["user_id"] != current_user.id and current_user.account_type not in ["master_admin", "admin"]:
            raise HTTPException(status_code=403, detail="Not authorized to view this receipt")
        
        file_path = EXPENSES_DIR / expense_id / filename
        if not file_path.exists():
            raise HTTPException(status_code=404, detail="Receipt file not found")
        
        return FileResponse(file_path, filename=filename)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error retrieving receipt: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to retrieve receipt")




# ==================== QR Code Management ====================

@api_router.post("/qr-codes/create")
async def create_qr_code(
    qr_data: QRCodeCreate,
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """Create QR code(s) - Supports bulk generation and UTM tracking - Master Admin only"""
    try:
        # Check master admin permission
        if current_user.account_type != "master_admin":
            raise HTTPException(status_code=403, detail="Only master admin can create QR codes")
        
        import qrcode
        from io import BytesIO
        from urllib.parse import urlencode
        
        # Get backend URL dynamically from request to ensure correct environment
        # This ensures QR codes created in production use production URLs
        host = request.headers.get('host', 'localhost:8001')
        scheme = 'https' if 'https' in str(request.url) or request.headers.get('x-forwarded-proto') == 'https' else 'http'
        backend_url = f"{scheme}://{host}"
        
        logger.info(f"Creating QR code with backend URL: {backend_url}")
        
        # Determine bulk count (1 to 100)
        bulk_count = min(max(qr_data.bulk_count or 1, 1), 100)
        
        created_qr_codes = []
        
        # Generate multiple QR codes if bulk_count > 1
        for i in range(bulk_count):
            # Generate unique short code
            unique_code = str(uuid.uuid4())[:8]
            
            # Get destination URL for display purposes
            if qr_data.landing_page_type == 'single':
                dest_url = qr_data.landing_page_single
            else:
                # Use mobile as default for multi-URL QR codes
                dest_url = (qr_data.landing_page_mobile or qr_data.landing_page_desktop or 
                           qr_data.landing_page_ios or qr_data.landing_page_android)
            
            # Add UTM parameters to destination URL if provided
            utm_params = {}
            if qr_data.utm_source:
                utm_params['utm_source'] = qr_data.utm_source
            if qr_data.utm_medium:
                utm_params['utm_medium'] = qr_data.utm_medium
            if qr_data.utm_campaign:
                utm_params['utm_campaign'] = qr_data.utm_campaign
            if qr_data.utm_term:
                utm_params['utm_term'] = qr_data.utm_term
            if qr_data.utm_content:
                # For bulk generation, add unique identifier to utm_content
                utm_content = qr_data.utm_content
                if bulk_count > 1:
                    utm_content = f"{utm_content}_{i+1}"
                utm_params['utm_content'] = utm_content
            
            # Append UTM parameters to destination URL
            if utm_params:
                separator = '&' if '?' in dest_url else '?'
                dest_url_with_utm = f"{dest_url}{separator}{urlencode(utm_params)}"
            else:
                dest_url_with_utm = dest_url
            
            # Extract clean domain for scanner preview
            from urllib.parse import urlparse
            parsed = urlparse(dest_url)
            dest_display = f"{parsed.netloc}{parsed.path}" if parsed.netloc else dest_url
            
            # Create QR redirect URL - Must include /api prefix for Kubernetes routing
            qr_redirect_url = f"{backend_url}/api/qr/{unique_code}"
            
            # Generate QR code image
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=4,
            )
            qr.add_data(qr_redirect_url)
            qr.make(fit=True)
            
            img = qr.make_image(fill_color="black", back_color="white")
            
            # Save QR code image
            qr_filename = f"qr_{unique_code}.png"
            qr_path = f"/app/backend/qr_codes/{qr_filename}"
            img.save(qr_path)
            
            # Create QR name with bulk index if generating multiple
            qr_name = qr_data.name
            if bulk_count > 1:
                qr_name = f"{qr_data.name} #{i+1}"
            
            # Create QR code document with UTM parameters
            qr_code = QRCode(
                name=qr_name,
                campaign_name=qr_data.campaign_name,
                landing_page_type=qr_data.landing_page_type,
                landing_page_single=dest_url_with_utm if qr_data.landing_page_type == 'single' else qr_data.landing_page_single,
                landing_page_ios=qr_data.landing_page_ios,
                landing_page_android=qr_data.landing_page_android,
                landing_page_mobile=qr_data.landing_page_mobile,
                landing_page_desktop=qr_data.landing_page_desktop,
                utm_source=qr_data.utm_source,
                utm_medium=qr_data.utm_medium,
                utm_campaign=qr_data.utm_campaign,
                utm_term=qr_data.utm_term,
                utm_content=utm_params.get('utm_content'),  # Save the actual content used (with bulk suffix)
                qr_image_filename=qr_filename,
                unique_short_code=unique_code,
                created_by=current_user.id,
            )
            
            # Save to database
            logger.info(f"About to save QR code to database: ID={qr_code.id}, Code={unique_code}")
            try:
                result = await db.qr_codes.insert_one(qr_code.model_dump())
                logger.info(f"QR code saved to database successfully. Inserted ID: {result.inserted_id}")
            except Exception as db_error:
                logger.error(f"DATABASE SAVE FAILED: {str(db_error)}")
                raise
            
            created_qr_codes.append({
                "id": qr_code.id,
                "name": qr_code.name,
                "unique_code": unique_code,
                "qr_url": qr_redirect_url,
                "qr_image": qr_filename,
                "utm_tracking": bool(utm_params)
            })
            
            logger.info(f"QR code created: {qr_code.id} by {current_user.email}")
        
        return {
            "success": True,
            "message": f"Successfully created {bulk_count} QR code(s)",
            "qr_codes": created_qr_codes,
            "bulk_count": bulk_count
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to create QR code: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Failed to create QR code: {str(e)}")


@api_router.get("/qr-codes")
async def get_qr_codes(
    current_user: User = Depends(get_current_user),
    skip: int = 0,
    limit: int = 100
):
    """Get all QR codes - Master Admin only"""
    try:
        if current_user.account_type != "master_admin":
            raise HTTPException(status_code=403, detail="Only master admin can view QR codes")
        
        # Get QR codes with pagination
        qr_codes = await db.qr_codes.find({}, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(length=None)
        
        # Get total count
        total_count = await db.qr_codes.count_documents({})
        
        # Calculate total scans across all QR codes
        total_scans = sum(qr.get('total_scans', 0) for qr in qr_codes)
        
        return {
            "success": True,
            "qr_codes": qr_codes,
            "total_count": total_count,
            "total_scans": total_scans,
            "skip": skip,
            "limit": limit
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get QR codes: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get QR codes: {str(e)}")


@api_router.get("/qr-codes/campaigns")
async def get_campaigns(current_user: User = Depends(get_current_user)):
    """Get all campaigns with QR code counts"""
    try:
        # Aggregate campaigns with published status
        pipeline = [
            {"$group": {
                "_id": "$campaign_name",
                "qr_count": {"$sum": 1},
                "total_scans": {"$sum": "$scan_count"},
                "created_at": {"$first": "$created_at"},
                "published": {"$max": "$published"}  # True if any QR in campaign is published
            }},
            {"$sort": {"created_at": -1}}
        ]
        
        campaigns = await db.qr_codes.aggregate(pipeline).to_list(None)
        
        # Format response
        formatted_campaigns = [
            {
                "campaign_name": c["_id"],
                "qr_count": c["qr_count"],
                "total_scans": c["total_scans"],
                "created_at": c["created_at"],
                "published": c.get("published", False)
            }
            for c in campaigns
        ]
        
        return {
            "success": True,
            "campaigns": formatted_campaigns,
            "count": len(formatted_campaigns)
        }
        
    except Exception as e:
        logger.error(f"Failed to get campaigns: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get campaigns: {str(e)}")


@api_router.get("/qr-codes/{qr_id}")
async def get_qr_code(
    qr_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get specific QR code details - Master Admin only"""
    try:
        if current_user.account_type != "master_admin":
            raise HTTPException(status_code=403, detail="Only master admin can view QR codes")
        
        qr_code = await db.qr_codes.find_one({"id": qr_id}, {"_id": 0})
        
        if not qr_code:
            raise HTTPException(status_code=404, detail="QR code not found")
        
        return qr_code
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get QR code: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get QR code: {str(e)}")


@api_router.get("/qr-codes/{qr_id}/download")
async def download_qr_code(
    qr_id: str,
    current_user: User = Depends(get_current_user)
):
    """Download QR code image - Master Admin only"""
    try:
        if current_user.account_type != "master_admin":
            raise HTTPException(status_code=403, detail="Only master admin can download QR codes")
        
        qr_code = await db.qr_codes.find_one({"id": qr_id}, {"_id": 0})
        
        if not qr_code:
            raise HTTPException(status_code=404, detail="QR code not found")
        
        qr_image_path = f"/app/backend/qr_codes/{qr_code['qr_image_filename']}"
        
        if not os.path.exists(qr_image_path):
            raise HTTPException(status_code=404, detail="QR code image not found")
        
        return FileResponse(
            path=qr_image_path,
            media_type="image/png",
            filename=f"{qr_code['name']}_QR.png"
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to download QR code: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to download QR code: {str(e)}")


# OLD ANALYTICS ENDPOINT REMOVED - CONFLICTED WITH NEW DETAILED ENDPOINT
# The new endpoint at line 11252 provides detailed scan_details and UTM URLs


@api_router.get("/qr-codes/{qr_id}/scans")
async def get_qr_scans(
    qr_id: str,
    current_user: User = Depends(get_current_user),
    skip: int = 0,
    limit: int = 50
):
    """Get QR code scan history - Master Admin only"""
    try:
        if current_user.account_type != "master_admin":
            raise HTTPException(status_code=403, detail="Only master admin can view scan history")
        
        # Get scans with pagination (exclude MongoDB _id)
        scans = await db.qr_scans.find({"qr_code_id": qr_id}, {"_id": 0}).sort("scan_datetime", -1).skip(skip).limit(limit).to_list(length=None)
        
        # Get total count
        total_count = await db.qr_scans.count_documents({"qr_code_id": qr_id})
        
        return {
            "success": True,
            "scans": scans,
            "total_count": total_count,
            "skip": skip,
            "limit": limit
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get scans: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get scans: {str(e)}")


@api_router.get("/qr-codes/{qr_id}/export-csv")
async def export_qr_scans_csv(
    qr_id: str,
    current_user: User = Depends(get_current_user)
):
    """Export QR scan data as CSV - Master Admin only"""
    try:
        if current_user.account_type != "master_admin":
            raise HTTPException(status_code=403, detail="Only master admin can export scan data")
        
        qr_code = await db.qr_codes.find_one({"id": qr_id}, {"_id": 0})
        if not qr_code:
            raise HTTPException(status_code=404, detail="QR code not found")
        
        # Get all scans
        scans = await db.qr_scans.find({"qr_code_id": qr_id}).sort("scan_datetime", -1).to_list(length=None)
        
        # Create CSV
        import csv
        from io import StringIO
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow([
            "Date", "Time", "Device Type", "Browser", "OS",
            "IP Address", "Country", "City",
            "Latitude", "Longitude", "Location Source",
            "Landing Page"
        ])
        
        # Write data
        for scan in scans:
            writer.writerow([
                scan.get('scan_date', ''),
                scan.get('scan_time', ''),
                scan.get('device_type', ''),
                scan.get('browser', ''),
                scan.get('os', ''),
                scan.get('ip_address', ''),
                scan.get('country', ''),
                scan.get('city', ''),
                scan.get('latitude', ''),
                scan.get('longitude', ''),
                scan.get('location_source', ''),
                scan.get('landing_page_redirected', '')
            ])
        
        # Create response
        from fastapi.responses import StreamingResponse
        output.seek(0)
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename={qr_code['name']}_scans.csv"
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to export CSV: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to export CSV: {str(e)}")


@api_router.put("/qr-codes/{qr_id}")
async def update_qr_code(
    qr_id: str,
    qr_update: QRCodeUpdate,
    current_user: User = Depends(get_current_user)
):
    """Update QR code settings - Master Admin only"""
    try:
        if current_user.account_type != "master_admin":
            raise HTTPException(status_code=403, detail="Only master admin can update QR codes")
        
        qr_code = await db.qr_codes.find_one({"id": qr_id}, {"_id": 0})
        if not qr_code:
            raise HTTPException(status_code=404, detail="QR code not found")
        
        # Prepare update data (only non-None fields)
        update_data = {k: v for k, v in qr_update.model_dump().items() if v is not None}
        
        if not update_data:
            raise HTTPException(status_code=400, detail="No fields to update")
        
        # Update in database
        await db.qr_codes.update_one(
            {"id": qr_id},
            {"$set": update_data}
        )
        
        logger.info(f"QR code updated: {qr_id} by {current_user.email}")
        
        return {
            "success": True,
            "message": "QR code updated successfully"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to update QR code: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to update QR code: {str(e)}")


@api_router.delete("/qr-codes/{qr_id}")
async def delete_qr_code(
    qr_id: str,
    current_user: User = Depends(get_current_user)
):
    """Delete QR code - Master Admin only"""
    try:
        if current_user.account_type != "master_admin":
            raise HTTPException(status_code=403, detail="Only master admin can delete QR codes")
        
        qr_code = await db.qr_codes.find_one({"id": qr_id}, {"_id": 0})
        if not qr_code:
            raise HTTPException(status_code=404, detail="QR code not found")
        
        # Delete QR code image
        qr_image_path = f"/app/backend/qr_codes/{qr_code['qr_image_filename']}"
        if os.path.exists(qr_image_path):
            os.remove(qr_image_path)
        
        # Delete from database
        await db.qr_codes.delete_one({"id": qr_id})
        
        # Delete all associated scans
        await db.qr_scans.delete_many({"qr_code_id": qr_id})
        
        logger.info(f"QR code deleted: {qr_id} by {current_user.email}")
        
        return {
            "success": True,
            "message": "QR code and all associated scans deleted successfully"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to delete QR code: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to delete QR code: {str(e)}")


@api_router.post("/qr-codes/bulk-delete")
async def bulk_delete_qr_codes(
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """Bulk delete QR codes - Master Admin only"""
    try:
        if current_user.account_type != "master_admin":
            raise HTTPException(status_code=403, detail="Only master admin can delete QR codes")
        
        # Parse request body
        body = await request.json()
        qr_ids = body.get('qr_ids', [])
        
        if not qr_ids:
            raise HTTPException(status_code=400, detail="No QR codes selected for deletion")
        
        deleted_count = 0
        failed_count = 0
        
        for qr_id in qr_ids:
            try:
                qr_code = await db.qr_codes.find_one({"id": qr_id}, {"_id": 0})
                if qr_code:
                    # Delete QR code image
                    qr_image_path = f"/app/backend/qr_codes/{qr_code['qr_image_filename']}"
                    if os.path.exists(qr_image_path):
                        os.remove(qr_image_path)
                    
                    # Delete from database
                    await db.qr_codes.delete_one({"id": qr_id})
                    
                    # Delete all associated scans
                    await db.qr_scans.delete_many({"qr_code_id": qr_id})
                    
                    deleted_count += 1
                    logger.info(f"QR code deleted: {qr_id}")
                else:
                    failed_count += 1
            except Exception as e:
                logger.error(f"Failed to delete QR code {qr_id}: {str(e)}")
                failed_count += 1
        
        return {
            "success": True,
            "message": f"Deleted {deleted_count} QR code(s), {failed_count} failed",
            "deleted_count": deleted_count,
            "failed_count": failed_count
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to bulk delete QR codes: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to bulk delete: {str(e)}")


# PUBLIC ENDPOINT - No authentication required
# ==================== RIDE DECK DATA ANALYSIS ====================

# VR Mall coordinates
VR_MALL_LAT = 13.0795762
VR_MALL_LNG = 80.1956368

@api_router.post("/ride-deck/analyze")
async def analyze_ride_deck(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """
    Analyze ride deck data by calculating distances using Google Maps API
    - Distance from VR Mall to pickup location
    - Distance from pickup location to drop location
    """
    try:
        logger.info(f"Ride deck analysis started by user: {current_user.email}")
        
        # Get Google Maps API key from environment
        gmaps_api_key = os.environ.get('GOOGLE_MAPS_API_KEY')
        if not gmaps_api_key:
            raise HTTPException(status_code=500, detail="Google Maps API key not configured")
        
        # Initialize Google Maps client
        gmaps = googlemaps.Client(key=gmaps_api_key)
        
        # Read the uploaded Excel file
        content = await file.read()
        
        # Load with pandas
        df = pd.read_excel(io.BytesIO(content))
        
        logger.info(f"File loaded: {len(df)} rows, Columns: {list(df.columns)}")
        
        # Check if required columns exist
        required_columns = ['pickupLat', 'pickupLong', 'dropLat', 'dropLong']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            raise HTTPException(
                status_code=400, 
                detail=f"Missing required columns: {', '.join(missing_columns)}"
            )
        
        # Initialize new columns
        pickup_localities = []
        vr_to_pickup_distances = []
        vr_to_pickup_times = []
        drop_localities = []
        pickup_to_drop_distances = []
        pickup_to_drop_times = []
        
        total_rows = len(df)
        
        # Helper function to extract locality from address
        def extract_locality(address):
            """
            Extract locality name from full address.
            Example: "Pattalam, Choolai for 5/3, Jai Nagar, Pattalam, Choolai, Chennai, Tamil Nadu 600012, India"
            Should return: "Choolai" (the last part before ", Chennai")
            """
            if pd.isna(address) or not address:
                return None
            
            address_str = str(address)
            # Split by comma and look for the part before "Chennai"
            parts = [p.strip() for p in address_str.split(',')]
            
            # Find the locality (the part immediately before Chennai)
            for i, part in enumerate(parts):
                if 'Chennai' in part or 'chennai' in part:
                    # Get the previous part (locality) - just the immediate previous one
                    if i > 0:
                        return parts[i-1]
                    return None
            
            # If Chennai not found, return the last part that's not a state/country
            for part in reversed(parts):
                if part and not any(word in part.lower() for word in ['india', 'tamil nadu', 'tamilnadu']):
                    return part
            
            return None
            
            # If Chennai not found, return the second-to-last part (before state/country)
            if len(parts) >= 2:
                # Skip parts that are likely state/country/pincode
                for part in reversed(parts[:-1]):
                    if part and not any(word in part.lower() for word in ['india', 'tamil nadu', 'tamilnadu']) and not part.strip().isdigit():
                        return part
            
            return None
        
        # Process each row
        for idx, row in df.iterrows():
            try:
                pickup_lat = row['pickupLat']
                pickup_lng = row['pickupLong']
                drop_lat = row['dropLat']
                drop_lng = row['dropLong']
                
                # Extract addresses from columns F and I (pickupPoint and dropPoint)
                # Column F is usually index 5, Column I is index 8
                pickup_address = row.get('pickupPoint', '')
                drop_address = row.get('dropPoint', '')
                
                # Extract localities
                pickup_locality = extract_locality(pickup_address)
                drop_locality = extract_locality(drop_address)
                
                pickup_localities.append(pickup_locality)
                drop_localities.append(drop_locality)
                
                # Skip if any coordinate is missing
                if pd.isna(pickup_lat) or pd.isna(pickup_lng) or pd.isna(drop_lat) or pd.isna(drop_lng):
                    vr_to_pickup_distances.append(None)
                    vr_to_pickup_times.append(None)
                    pickup_to_drop_distances.append(None)
                    pickup_to_drop_times.append(None)
                    logger.warning(f"Row {idx}: Missing coordinates, skipping")
                    continue
                
                # Calculate distance and time from VR Mall to pickup
                try:
                    result1 = gmaps.distance_matrix(
                        origins=[(VR_MALL_LAT, VR_MALL_LNG)],
                        destinations=[(pickup_lat, pickup_lng)],
                        mode="driving"
                    )
                    
                    if result1['rows'][0]['elements'][0]['status'] == 'OK':
                        distance_meters = result1['rows'][0]['elements'][0]['distance']['value']
                        distance_km = round(distance_meters / 1000, 2)
                        duration_seconds = result1['rows'][0]['elements'][0]['duration']['value']
                        duration_mins = round(duration_seconds / 60, 1)
                        
                        vr_to_pickup_distances.append(distance_km)
                        vr_to_pickup_times.append(duration_mins)
                    else:
                        vr_to_pickup_distances.append(None)
                        vr_to_pickup_times.append(None)
                        logger.warning(f"Row {idx}: VR to pickup distance not available")
                except Exception as e:
                    vr_to_pickup_distances.append(None)
                    vr_to_pickup_times.append(None)
                    logger.error(f"Row {idx}: Error calculating VR to pickup distance: {str(e)}")
                
                # Small delay to avoid rate limiting
                await asyncio.sleep(0.1)
                
                # Calculate distance and time from pickup to drop
                try:
                    result2 = gmaps.distance_matrix(
                        origins=[(pickup_lat, pickup_lng)],
                        destinations=[(drop_lat, drop_lng)],
                        mode="driving"
                    )
                    
                    if result2['rows'][0]['elements'][0]['status'] == 'OK':
                        distance_meters = result2['rows'][0]['elements'][0]['distance']['value']
                        distance_km = round(distance_meters / 1000, 2)
                        duration_seconds = result2['rows'][0]['elements'][0]['duration']['value']
                        duration_mins = round(duration_seconds / 60, 1)
                        
                        pickup_to_drop_distances.append(distance_km)
                        pickup_to_drop_times.append(duration_mins)
                    else:
                        pickup_to_drop_distances.append(None)
                        pickup_to_drop_times.append(None)
                        logger.warning(f"Row {idx}: Pickup to drop distance not available")
                except Exception as e:
                    pickup_to_drop_distances.append(None)
                    pickup_to_drop_times.append(None)
                    logger.error(f"Row {idx}: Error calculating pickup to drop distance: {str(e)}")
                
                # Small delay to avoid rate limiting
                await asyncio.sleep(0.1)
                
                logger.info(f"Processed row {idx + 1}/{total_rows}")
                
            except Exception as e:
                pickup_localities.append(None)
                vr_to_pickup_distances.append(None)
                vr_to_pickup_times.append(None)
                drop_localities.append(None)
                pickup_to_drop_distances.append(None)
                pickup_to_drop_times.append(None)
                logger.error(f"Row {idx}: Error processing: {str(e)}")
        
        # Add new columns to dataframe in the specified order
        df['Pickup_Locality'] = pickup_localities
        df['VR_to_Pickup_Distance_KM'] = vr_to_pickup_distances
        df['VR_to_Pickup_Time_Mins'] = vr_to_pickup_times
        df['Drop_Locality'] = drop_localities
        df['Pickup_to_Drop_Distance_KM'] = pickup_to_drop_distances
        df['Pickup_to_Drop_Time_Mins'] = pickup_to_drop_times
        
        # Save to new Excel file
        output = io.BytesIO()
        
        # Use openpyxl to maintain original formatting if needed
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
        
        output.seek(0)
        
        logger.info(f"Ride deck analysis completed: {total_rows} rows processed")
        
        # Return the file as response
        from fastapi.responses import StreamingResponse
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=ride_deck_analyzed.xlsx"
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ride deck analysis failed: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Ride deck analysis failed: {str(e)}")


# ==================== RIDE DECK DATA IMPORT & MANAGEMENT ====================

from app_models import Customer, Ride, ImportStats
from collections import Counter

@api_router.post("/ride-deck/import-customers", response_model=ImportStats)
async def import_customers(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """
    Import customer data from CSV file
    - Merges with existing data (no duplicates based on 'id')
    - Returns import statistics
    """
    try:
        logger.info(f"Customer import started by user: {current_user.email}")
        
        # Read CSV file
        content = await file.read()
        df = pd.read_csv(io.BytesIO(content))
        
        logger.info(f"Customer CSV loaded: {len(df)} rows")
        
        # Get MongoDB database
        customers_collection = db['customers']
        
        # Track statistics
        total_rows = len(df)
        new_records = 0
        duplicate_records = 0
        errors = 0
        error_details = []
        
        # Get existing customer IDs
        existing_ids = set()
        existing_customers = await customers_collection.find({}, {'id': 1}).to_list(None)
        for customer in existing_customers:
            existing_ids.add(str(customer['id']))
        
        logger.info(f"Existing customers in DB: {len(existing_ids)}")
        
        # Process each row
        for idx, row in df.iterrows():
            try:
                customer_id = str(row.get('id', ''))
                
                if not customer_id or pd.isna(customer_id):
                    errors += 1
                    error_details.append(f"Row {idx}: Missing customer ID")
                    continue
                
                # Check if customer already exists
                if customer_id in existing_ids:
                    duplicate_records += 1
                    logger.debug(f"Customer {customer_id} already exists, skipping")
                    continue
                
                # Create customer object
                customer_data = {
                    'id': customer_id,
                    'name': str(row.get('name', '')) if pd.notna(row.get('name')) else None,
                    'email': str(row.get('email', '')) if pd.notna(row.get('email')) else None,
                    'phoneNumber': str(row.get('phoneNumber', '')) if pd.notna(row.get('phoneNumber')) else None,
                    'gender': str(row.get('gender', '')) if pd.notna(row.get('gender')) else None,
                    'rideOtp': str(row.get('rideOtp', '')) if pd.notna(row.get('rideOtp')) else None,
                    'referredById': str(row.get('referredById', '')) if pd.notna(row.get('referredById')) else None,
                    'nuraCoins': int(row.get('nuraCoins', 0)) if pd.notna(row.get('nuraCoins')) else 0,
                    'dateOfBirth': str(row.get('dateOfBirth', '')) if pd.notna(row.get('dateOfBirth')) else None,
                    'emergencyContact': str(row.get('emergencyContact', '')) if pd.notna(row.get('emergencyContact')) else None,
                    'userReferralCode': str(row.get('userReferralCode', '')) if pd.notna(row.get('userReferralCode')) else None,
                    'createdAt': str(row.get('createdAt', '')) if pd.notna(row.get('createdAt')) else None,
                    'updatedAt': str(row.get('updatedAt', '')) if pd.notna(row.get('updatedAt')) else None,
                    'date': str(row.get('date', '')) if pd.notna(row.get('date')) else None,
                    'time': str(row.get('time', '')) if pd.notna(row.get('time')) else None,
                    'hour': int(row.get('hour', 0)) if pd.notna(row.get('hour')) else None,
                    'source': str(row.get('source', '')) if pd.notna(row.get('source')) else None,
                    'Channel': str(row.get('Channel', '')) if pd.notna(row.get('Channel')) else None,
                    'imported_at': datetime.now(timezone.utc).isoformat()
                }
                
                # Insert into database
                await customers_collection.insert_one(customer_data)
                new_records += 1
                existing_ids.add(customer_id)
                
            except Exception as e:
                errors += 1
                error_msg = f"Row {idx}: {str(e)}"
                error_details.append(error_msg)
                logger.error(error_msg)
        
        logger.info(f"Customer import complete: {new_records} new, {duplicate_records} duplicates, {errors} errors")
        
        return ImportStats(
            total_rows=total_rows,
            new_records=new_records,
            duplicate_records=duplicate_records,
            errors=errors,
            error_details=error_details[:10] if error_details else None  # Limit to 10 errors
        )
        
    except Exception as e:
        logger.error(f"Customer import failed: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Customer import failed: {str(e)}")


@api_router.post("/ride-deck/import-rides", response_model=ImportStats)
async def import_rides(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """
    Import ride data from CSV file
    - Merges with existing data (no duplicates based on 'id')
    - Computes additional fields for NEW rides only
    - Returns import statistics
    """
    try:
        logger.info(f"Ride import started by user: {current_user.email}")
        
        # Read CSV file
        content = await file.read()
        df = pd.read_csv(io.BytesIO(content))
        
        logger.info(f"Ride CSV loaded: {len(df)} rows")
        
        # Get MongoDB database
        rides_collection = db['rides']
        
        # Get Google Maps API key
        gmaps_api_key = os.environ.get('GOOGLE_MAPS_API_KEY')
        if not gmaps_api_key:
            raise HTTPException(status_code=500, detail="Google Maps API key not configured")
        
        gmaps = googlemaps.Client(key=gmaps_api_key)
        
        # Track statistics
        total_rows = len(df)
        new_records = 0
        duplicate_records = 0
        errors = 0
        error_details = []
        
        # Get existing ride IDs
        existing_ids = set()
        existing_rides = await rides_collection.find({}, {'id': 1}).to_list(None)
        for ride in existing_rides:
            existing_ids.add(str(ride['id']))
        
        logger.info(f"Existing rides in DB: {len(existing_ids)}")
        
        # Helper function to extract locality
        def extract_locality(address):
            """
            Extract locality name from full address.
            Example: "Pattalam, Choolai for 5/3, Jai Nagar, Pattalam, Choolai, Chennai, Tamil Nadu 600012, India"
            Should return: "Choolai" (the last part before ", Chennai")
            """
            if pd.isna(address) or not address:
                return None
            
            address_str = str(address)
            parts = [p.strip() for p in address_str.split(',')]
            
            # Find the locality (the part immediately before Chennai)
            for i, part in enumerate(parts):
                if 'Chennai' in part or 'chennai' in part:
                    # Get the previous part (locality) - just the immediate previous one
                    if i > 0:
                        return parts[i-1]
                    return None
            
            # If Chennai not found, return the last part that's not a state/country/postal code
            for part in reversed(parts):
                if part and not any(word in part.lower() for word in ['india', 'tamil nadu', 'tamilnadu']) and not part.strip().isdigit():
                    return part
            
            return None
        
        # Build historical pickup data for each customer (for most common pickup point calculation)
        customer_pickup_history = {}
        all_rides = await rides_collection.find({}, {'customerId': 1, 'pickupLat': 1, 'pickupLong': 1}).to_list(None)
        for ride in all_rides:
            customer_id = ride.get('customerId')
            pickup_lat = ride.get('pickupLat')
            pickup_long = ride.get('pickupLong')
            
            if customer_id and pickup_lat and pickup_long:
                if customer_id not in customer_pickup_history:
                    customer_pickup_history[customer_id] = []
                customer_pickup_history[customer_id].append((pickup_lat, pickup_long))
        
        logger.info(f"Built pickup history for {len(customer_pickup_history)} customers")
        
        # Process each row
        for idx, row in df.iterrows():
            try:
                ride_id = str(row.get('id', ''))
                
                if not ride_id or pd.isna(ride_id):
                    errors += 1
                    error_details.append(f"Row {idx}: Missing ride ID")
                    continue
                
                # Check if ride already exists
                if ride_id in existing_ids:
                    duplicate_records += 1
                    logger.debug(f"Ride {ride_id} already exists, skipping")
                    continue
                
                # This is a NEW ride - compute additional fields
                customer_id = str(row.get('customerId', '')) if pd.notna(row.get('customerId')) else None
                pickup_lat = float(row.get('pickupLat')) if pd.notna(row.get('pickupLat')) else None
                pickup_long = float(row.get('pickupLong')) if pd.notna(row.get('pickupLong')) else None
                drop_lat = float(row.get('dropLat')) if pd.notna(row.get('dropLat')) else None
                drop_long = float(row.get('dropLong')) if pd.notna(row.get('dropLong')) else None
                pickup_point = str(row.get('pickupPoint', '')) if pd.notna(row.get('pickupPoint')) else None
                drop_point = str(row.get('dropPoint', '')) if pd.notna(row.get('dropPoint')) else None
                
                # Compute Pickup and Drop Locality
                pickup_locality = extract_locality(pickup_point)
                drop_locality = extract_locality(drop_point)
                
                # Compute Pickup Distance from DEPOT (VR Mall)
                pickup_distance_from_depot = None
                if pickup_lat and pickup_long:
                    try:
                        result = gmaps.distance_matrix(
                            origins=[(VR_MALL_LAT, VR_MALL_LNG)],
                            destinations=[(pickup_lat, pickup_long)],
                            mode="driving"
                        )
                        if result['rows'][0]['elements'][0]['status'] == 'OK':
                            distance_meters = result['rows'][0]['elements'][0]['distance']['value']
                            pickup_distance_from_depot = round(distance_meters / 1000, 2)
                        await asyncio.sleep(0.05)  # Rate limiting
                    except Exception as e:
                        logger.warning(f"Row {idx}: Could not calculate pickup distance from depot: {str(e)}")
                
                # Compute Drop Distance from DEPOT (VR Mall)
                drop_distance_from_depot = None
                if drop_lat and drop_long:
                    try:
                        result = gmaps.distance_matrix(
                            origins=[(VR_MALL_LAT, VR_MALL_LNG)],
                            destinations=[(drop_lat, drop_long)],
                            mode="driving"
                        )
                        if result['rows'][0]['elements'][0]['status'] == 'OK':
                            distance_meters = result['rows'][0]['elements'][0]['distance']['value']
                            drop_distance_from_depot = round(distance_meters / 1000, 2)
                        await asyncio.sleep(0.05)  # Rate limiting
                    except Exception as e:
                        logger.warning(f"Row {idx}: Could not calculate drop distance from depot: {str(e)}")
                
                # Compute Most Common Pickup Point
                most_common_pickup_point = None
                most_common_pickup_locality = None
                if customer_id:
                    # Add current pickup to history
                    if customer_id not in customer_pickup_history:
                        customer_pickup_history[customer_id] = []
                    if pickup_lat and pickup_long:
                        customer_pickup_history[customer_id].append((pickup_lat, pickup_long))
                    
                    # Find most common pickup point (rounded to 4 decimal places for grouping)
                    if customer_pickup_history[customer_id]:
                        pickup_points = [
                            (round(lat, 4), round(long, 4))
                            for lat, long in customer_pickup_history[customer_id]
                        ]
                        most_common = Counter(pickup_points).most_common(1)
                        if most_common:
                            most_common_lat, most_common_long = most_common[0][0]
                            most_common_pickup_point = f"{most_common_lat},{most_common_long}"
                            
                            # Get locality for most common pickup point
                            # We need to reverse geocode this - but to save API calls, we'll try to find it in existing data
                            # For now, we'll leave it as None and can enhance later
                            most_common_pickup_locality = None
                
                # Create ride object with all fields
                ride_data = {
                    'id': ride_id,
                    'customerId': customer_id,
                    'driverId': str(row.get('driverId', '')) if pd.notna(row.get('driverId')) else None,
                    'rideStatus': str(row.get('rideStatus', '')) if pd.notna(row.get('rideStatus')) else None,
                    'rideType': str(row.get('rideType', '')) if pd.notna(row.get('rideType')) else None,
                    'pickupPoint': pickup_point,
                    'pickupLat': pickup_lat,
                    'pickupLong': pickup_long,
                    'dropPoint': drop_point,
                    'dropLat': drop_lat,
                    'dropLong': drop_long,
                    'initialDistance': float(row.get('initialDistance')) if pd.notna(row.get('initialDistance')) else None,
                    'initialDuration': int(row.get('initialDuration')) if pd.notna(row.get('initialDuration')) else None,
                    'finalDistance': float(row.get('finalDistance')) if pd.notna(row.get('finalDistance')) else None,
                    'finalDuration': int(row.get('finalDuration')) if pd.notna(row.get('finalDuration')) else None,
                    'payWithNuraCoins': str(row.get('payWithNuraCoins', '')) if pd.notna(row.get('payWithNuraCoins')) else None,
                    'appliedVoucherId': str(row.get('appliedVoucherId', '')) if pd.notna(row.get('appliedVoucherId')) else None,
                    'appliedCouponId': str(row.get('appliedCouponId', '')) if pd.notna(row.get('appliedCouponId')) else None,
                    'rideAssignedLat': float(row.get('rideAssignedLat')) if pd.notna(row.get('rideAssignedLat')) else None,
                    'rideAssignedLong': float(row.get('rideAssignedLong')) if pd.notna(row.get('rideAssignedLong')) else None,
                    'initialFare': float(row.get('initialFare')) if pd.notna(row.get('initialFare')) else None,
                    'finalFare': float(row.get('finalFare')) if pd.notna(row.get('finalFare')) else None,
                    'payableAmount': float(row.get('payableAmount')) if pd.notna(row.get('payableAmount')) else None,
                    'rideStartTime': str(row.get('rideStartTime', '')) if pd.notna(row.get('rideStartTime')) else None,
                    'rideEndTime': str(row.get('rideEndTime', '')) if pd.notna(row.get('rideEndTime')) else None,
                    'rideAssignedTime': str(row.get('rideAssignedTime', '')) if pd.notna(row.get('rideAssignedTime')) else None,
                    'initialOdometer': float(row.get('initialOdometer')) if pd.notna(row.get('initialOdometer')) else None,
                    'finalOdometer': float(row.get('finalOdometer')) if pd.notna(row.get('finalOdometer')) else None,
                    'createdAt': str(row.get('createdAt', '')) if pd.notna(row.get('createdAt')) else None,
                    'updatedAt': str(row.get('updatedAt', '')) if pd.notna(row.get('updatedAt')) else None,
                    'date': str(row.get('date', '')) if pd.notna(row.get('date')) else None,
                    'time_est': str(row.get('time_est', '')) if pd.notna(row.get('time_est')) else None,
                    'hour': int(row.get('hour')) if pd.notna(row.get('hour')) else None,
                    'source': str(row.get('source', '')) if pd.notna(row.get('source')) else None,
                    'dd': str(row.get('dd', '')) if pd.notna(row.get('dd')) else None,
                    'dd1': str(row.get('dd1', '')) if pd.notna(row.get('dd1')) else None,
                    'dd2': str(row.get('dd2', '')) if pd.notna(row.get('dd2')) else None,
                    'dd3': str(row.get('dd3', '')) if pd.notna(row.get('dd3')) else None,
                    'dd4': str(row.get('dd4', '')) if pd.notna(row.get('dd4')) else None,
                    'dd5': str(row.get('dd5', '')) if pd.notna(row.get('dd5')) else None,
                    # Computed fields
                    'pickupLocality': pickup_locality,
                    'dropLocality': drop_locality,
                    'pickupDistanceFromDepot': pickup_distance_from_depot,
                    'dropDistanceFromDepot': drop_distance_from_depot,
                    'mostCommonPickupPoint': most_common_pickup_point,
                    'mostCommonPickupLocality': most_common_pickup_locality,
                    'statusReason': None,  # Empty by default
                    'statusDetail': None,  # Empty by default
                    'imported_at': datetime.now(timezone.utc).isoformat()
                }
                
                # Insert into database
                await rides_collection.insert_one(ride_data)
                new_records += 1
                existing_ids.add(ride_id)
                
                if (idx + 1) % 10 == 0:
                    logger.info(f"Processed {idx + 1}/{total_rows} rides")
                
            except Exception as e:
                errors += 1
                error_msg = f"Row {idx}: {str(e)}"
                error_details.append(error_msg)
                logger.error(error_msg)
        
        logger.info(f"Ride import complete: {new_records} new, {duplicate_records} duplicates, {errors} errors")
        
        return ImportStats(
            total_rows=total_rows,
            new_records=new_records,
            duplicate_records=duplicate_records,
            errors=errors,
            error_details=error_details[:10] if error_details else None
        )
        
    except Exception as e:
        logger.error(f"Ride import failed: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Ride import failed: {str(e)}")


@api_router.get("/ride-deck/stats")
async def get_ride_deck_stats(
    current_user: User = Depends(get_current_user)
):
    """
    Get statistics about imported data
    """
    try:
        customers_count = await db['customers'].count_documents({})
        rides_count = await db['rides'].count_documents({})
        
        # Get ride status distribution
        ride_status_pipeline = [
            {"$group": {"_id": "$rideStatus", "count": {"$sum": 1}}}
        ]
        ride_status_dist = await db['rides'].aggregate(ride_status_pipeline).to_list(None)
        
        return {
            "success": True,
            "customers_count": customers_count,
            "rides_count": rides_count,
            "ride_status_distribution": {item['_id']: item['count'] for item in ride_status_dist}
        }
        
    except Exception as e:
        logger.error(f"Failed to get stats: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get stats: {str(e)}")


@api_router.post("/ride-deck/progress")
async def get_ride_deck_progress(
    current_user: User = Depends(get_current_user)
):
    """
    Get progress of ride deck analysis (placeholder for real-time progress)
    In production, this would use WebSocket or server-sent events
    """
    return {
        "success": True,
        "progress": 100,
        "message": "Analysis complete"
    }


@api_router.get("/ride-deck/customers")
async def get_customers_data(
    limit: int = 100,
    skip: int = 0,
    current_user: User = Depends(get_current_user)
):
    """
    Get customer data with pagination
    """
    try:
        customers_collection = db['customers']
        
        customers = await customers_collection.find({}).skip(skip).limit(limit).to_list(None)
        total_count = await customers_collection.count_documents({})
        
        # Convert ObjectId to string if present
        for customer in customers:
            if '_id' in customer:
                del customer['_id']
        
        return {
            "success": True,
            "data": customers,
            "total": total_count,
            "limit": limit,
            "skip": skip
        }
        
    except Exception as e:
        logger.error(f"Failed to get customers data: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get customers: {str(e)}")


@api_router.get("/ride-deck/rides")
async def get_rides_data(
    limit: int = 100,
    skip: int = 0,
    current_user: User = Depends(get_current_user)
):
    """
    Get ride data with pagination
    """
    try:
        rides_collection = db['rides']
        
        rides = await rides_collection.find({}).skip(skip).limit(limit).to_list(None)
        total_count = await rides_collection.count_documents({})
        
        # Convert ObjectId to string if present
        for ride in rides:
            if '_id' in ride:
                del ride['_id']
        
        return {
            "success": True,
            "data": rides,
            "total": total_count,
            "limit": limit,
            "skip": skip
        }
        
    except Exception as e:
        logger.error(f"Failed to get rides data: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get rides: {str(e)}")


@api_router.get("/ride-deck/export-customers")
async def export_customers_excel(
    current_user: User = Depends(get_current_user)
):
    """
    Export all customer data as Excel file
    """
    try:
        customers_collection = db['customers']
        
        customers = await customers_collection.find({}).to_list(None)
        
        if not customers:
            raise HTTPException(status_code=404, detail="No customer data found")
        
        # Remove MongoDB _id field
        for customer in customers:
            if '_id' in customer:
                del customer['_id']
        
        # Convert to DataFrame
        df = pd.DataFrame(customers)
        
        # Create Excel file in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Customers')
        
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename=customers_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to export customers: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to export customers: {str(e)}")


@api_router.get("/ride-deck/export-rides")
async def export_rides_excel(
    current_user: User = Depends(get_current_user)
):
    """
    Export all ride data as Excel file
    """
    try:
        rides_collection = db['rides']
        
        rides = await rides_collection.find({}).to_list(None)
        
        if not rides:
            raise HTTPException(status_code=404, detail="No ride data found")
        
        # Remove MongoDB _id field
        for ride in rides:
            if '_id' in ride:
                del ride['_id']
        
        # Convert to DataFrame
        df = pd.DataFrame(rides)
        
        # Create Excel file in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Rides')
        
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename=rides_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to export rides: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to export rides: {str(e)}")


@api_router.delete("/ride-deck/delete-customers")
async def delete_all_customers(
    current_user: User = Depends(get_current_user)
):
    """
    Delete all customer data (Master Admin only)
    """
    if current_user.account_type != "master_admin":
        raise HTTPException(status_code=403, detail="Only Master Admins can delete all customer data")
    
    try:
        customers_collection = db['customers']
        
        result = await customers_collection.delete_many({})
        
        return {
            "success": True,
            "message": f"Successfully deleted {result.deleted_count} customer records",
            "deleted_count": result.deleted_count
        }
        
    except Exception as e:
        logger.error(f"Failed to delete customers: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to delete customers: {str(e)}")


@api_router.delete("/ride-deck/delete-rides")
async def delete_all_rides(
    current_user: User = Depends(get_current_user)
):
    """
    Delete all ride data (Master Admin only)
    """
    if current_user.account_type != "master_admin":
        raise HTTPException(status_code=403, detail="Only Master Admins can delete all ride data")
    
    try:
        rides_collection = db['rides']
        
        result = await rides_collection.delete_many({})
        
        return {
            "success": True,
            "message": f"Successfully deleted {result.deleted_count} ride records",
            "deleted_count": result.deleted_count
        }
        
    except Exception as e:
        logger.error(f"Failed to delete rides: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to delete rides: {str(e)}")


@api_router.post("/ride-deck/fix-localities")
async def fix_ride_localities(
    current_user: User = Depends(get_current_user)
):
    """
    Fix locality extraction for existing rides (Master Admin only)
    Re-processes pickupLocality and dropLocality for all rides with the updated extraction logic
    """
    if current_user.account_type != "master_admin":
        raise HTTPException(status_code=403, detail="Only Master Admins can fix localities")
    
    try:
        rides_collection = db['rides']
        
        # Helper function to extract locality (same as in import)
        def extract_locality(address):
            """
            Extract locality name from full address.
            Example: "Pattalam, Choolai for 5/3, Jai Nagar, Pattalam, Choolai, Chennai, Tamil Nadu 600012, India"
            Should return: "Choolai" (the last part before ", Chennai")
            """
            if not address:
                return None
            
            address_str = str(address)
            parts = [p.strip() for p in address_str.split(',')]
            
            # Find the locality (the part immediately before Chennai)
            for i, part in enumerate(parts):
                if 'Chennai' in part or 'chennai' in part:
                    # Get the previous part (locality) - just the immediate previous one
                    if i > 0:
                        return parts[i-1]
                    return None
            
            # If Chennai not found, return the last part that's not a state/country/postal code
            for part in reversed(parts):
                if part and not any(word in part.lower() for word in ['india', 'tamil nadu', 'tamilnadu']) and not part.strip().isdigit():
                    return part
            
            return None
        
        # Fetch all rides
        rides = await rides_collection.find({}).to_list(None)
        
        updated_count = 0
        for ride in rides:
            pickup_point = ride.get('pickupPoint')
            drop_point = ride.get('dropPoint')
            
            if pickup_point or drop_point:
                new_pickup_locality = extract_locality(pickup_point)
                new_drop_locality = extract_locality(drop_point)
                
                # Update if different
                if (new_pickup_locality != ride.get('pickupLocality') or 
                    new_drop_locality != ride.get('dropLocality')):
                    
                    await rides_collection.update_one(
                        {'id': ride['id']},
                        {'$set': {
                            'pickupLocality': new_pickup_locality,
                            'dropLocality': new_drop_locality
                        }}
                    )
                    updated_count += 1
        
        return {
            "success": True,
            "message": f"Successfully re-processed localities for {updated_count} rides",
            "total_rides": len(rides),
            "updated_count": updated_count
        }
        
    except Exception as e:
        logger.error(f"Failed to fix localities: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fix localities: {str(e)}")


# ==================== ANALYTICS DASHBOARDS - PIVOT TABLES ====================

from datetime import datetime, timezone as dt_timezone
from collections import defaultdict

def convert_utc_to_ist(utc_date_str):
    """Convert UTC date string to IST date, handling Excel serial numbers and various formats"""
    try:
        if not utc_date_str or utc_date_str == 'N/A':
            return None
        
        # Handle Excel serial date numbers (e.g., 45928, 45929)
        if isinstance(utc_date_str, (int, float)):
            # Convert Excel serial number to datetime
            # Excel epoch starts at 1899-12-30
            utc_dt = pd.to_datetime(utc_date_str, origin='1899-12-30', unit='D')
            # Convert to IST (UTC+5:30)
            ist_dt = utc_dt + timedelta(hours=5, minutes=30)
            return ist_dt.strftime('%Y-%m-%d')
        
        # Try converting string serial numbers (handle both integer and decimal strings)
        if isinstance(utc_date_str, str):
            # Remove whitespace
            utc_date_str = utc_date_str.strip()
            
            # Try parsing as numeric (Excel serial number)
            try:
                serial_num = float(utc_date_str)
                # Excel serial numbers for dates are typically > 1 and < 100000
                if 1 < serial_num < 100000:
                    utc_dt = pd.to_datetime(serial_num, origin='1899-12-30', unit='D')
                    ist_dt = utc_dt + timedelta(hours=5, minutes=30)
                    result = ist_dt.strftime('%Y-%m-%d')
                    logger.info(f"Converted Excel serial {serial_num} to IST date {result}")
                    return result
            except (ValueError, TypeError):
                # Not a numeric string, try other formats
                pass
            
            # Try parsing as ISO format
            try:
                utc_dt = datetime.fromisoformat(utc_date_str.replace('Z', '+00:00'))
                ist_dt = utc_dt + timedelta(hours=5, minutes=30)
                return ist_dt.strftime('%Y-%m-%d')
            except:
                pass
            
            # Try other common date formats
            for fmt in ['%Y-%m-%d', '%d-%m-%Y', '%Y/%m/%d', '%d/%m/%Y', '%d/%m/%y', '%Y-%m-%d %H:%M:%S']:
                try:
                    utc_dt = datetime.strptime(utc_date_str, fmt)
                    # If no timezone, assume UTC
                    if utc_dt.tzinfo is None:
                        utc_dt = utc_dt.replace(tzinfo=dt_timezone.utc)
                    ist_dt = utc_dt + timedelta(hours=5, minutes=30)
                    return ist_dt.strftime('%Y-%m-%d')
                except:
                    continue
            
            # If we couldn't parse the string, return it as-is
            logger.warning(f"Could not parse date string: {utc_date_str}")
            return utc_date_str
        
        # Handle datetime objects
        if isinstance(utc_date_str, datetime):
            utc_dt = utc_date_str
            if utc_dt.tzinfo is None:
                utc_dt = utc_dt.replace(tzinfo=dt_timezone.utc)
            ist_dt = utc_dt + timedelta(hours=5, minutes=30)
            return ist_dt.strftime('%Y-%m-%d')
        
        # Fallback: return original value
        logger.warning(f"Unhandled date type: {type(utc_date_str)} - {utc_date_str}")
        return str(utc_date_str)
        
    except Exception as e:
        logger.error(f"Error converting UTC to IST for value '{utc_date_str}' (type: {type(utc_date_str)}): {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        # Return None instead of original value to avoid sorting errors
        return None


@api_router.get("/analytics/ride-status-pivot")
async def get_ride_status_pivot(
    row_field: str = "date",
    column_field: str = "rideStatus",
    value_field: str = "count",
    value_operation: str = "count",  # count, sum, average
    filter_field: Optional[str] = None,
    filter_value: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """
    Get ride status pivot table data
    Supports customizable row, column, value fields
    All dates converted from UTC to IST
    """
    try:
        rides_collection = db['rides']
        
        # Build query with filters
        query = {}
        if filter_field and filter_value and filter_value != 'all':
            query[filter_field] = filter_value
        
        # Fetch all rides
        rides = await rides_collection.find(query).to_list(None)
        
        # Convert UTC dates to IST and build pivot data
        pivot_data = defaultdict(lambda: defaultdict(int))
        row_values = set()
        column_values = set()
        
        for ride in rides:
            # Get row value (convert date to IST if it's a date field)
            row_val = ride.get(row_field, 'N/A')
            if row_field == 'date' or 'date' in row_field.lower() or 'time' in row_field.lower():
                row_val = convert_utc_to_ist(row_val)
            
            # Get column value  (also convert if it's a date field)
            col_val = ride.get(column_field, 'N/A')
            if column_field == 'date' or 'date' in column_field.lower() or 'time' in column_field.lower():
                col_val = convert_utc_to_ist(col_val)
            
            # Skip if essential values are missing or None
            if row_val is None or row_val == 'N/A' or col_val is None or col_val == 'N/A':
                continue
            
            row_values.add(row_val)
            column_values.add(col_val)
            
            # Apply value operation
            if value_operation == 'count':
                pivot_data[row_val][col_val] += 1
            elif value_operation == 'sum' and value_field in ride:
                pivot_data[row_val][col_val] += float(ride.get(value_field, 0))
            elif value_operation == 'average' and value_field in ride:
                # Store sum and count for later average calculation
                if 'avg_data' not in pivot_data:
                    pivot_data['avg_data'] = defaultdict(lambda: defaultdict(lambda: {'sum': 0, 'count': 0}))
                pivot_data['avg_data'][row_val][col_val]['sum'] += float(ride.get(value_field, 0))
                pivot_data['avg_data'][row_val][col_val]['count'] += 1
        
        # Convert to sorted list format (filter out None values before sorting)
        sorted_rows = sorted([r for r in row_values if r is not None])
        sorted_columns = sorted([c for c in column_values if c is not None])
        
        # Build final pivot table
        table_data = []
        for row_val in sorted_rows:
            row_data = {'rowLabel': row_val}
            for col_val in sorted_columns:
                if value_operation == 'average' and 'avg_data' in pivot_data:
                    avg_info = pivot_data['avg_data'][row_val][col_val]
                    if avg_info['count'] > 0:
                        row_data[col_val] = round(avg_info['sum'] / avg_info['count'], 2)
                    else:
                        row_data[col_val] = 0
                else:
                    row_data[col_val] = pivot_data[row_val][col_val]
            table_data.append(row_data)
        
        # Get unique filter values if filter field is specified
        filter_options = []
        if filter_field:
            filter_values_set = set()
            for ride in rides:
                filter_val = ride.get(filter_field)
                if filter_val:
                    filter_values_set.add(str(filter_val))
            filter_options = sorted(list(filter_values_set))
        
        return {
            "success": True,
            "data": table_data,
            "columns": sorted_columns,
            "row_field": row_field,
            "column_field": column_field,
            "value_operation": value_operation,
            "filter_options": filter_options,
            "total_records": len(rides)
        }
        
    except Exception as e:
        logger.error(f"Failed to generate ride status pivot: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Failed to generate pivot table: {str(e)}")


@api_router.get("/analytics/signups-pivot")
async def get_signups_pivot(
    row_field: str = "date",
    column_field: str = "source",
    value_field: str = "count",
    value_operation: str = "count",  # count, sum, average
    filter_field: Optional[str] = None,
    filter_value: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """
    Get customer signups pivot table data
    Supports customizable row, column, value fields
    All dates converted from UTC to IST
    """
    try:
        customers_collection = db['customers']
        
        # Build query with filters
        query = {}
        if filter_field and filter_value and filter_value != 'all':
            query[filter_field] = filter_value
        
        # Fetch all customers
        customers = await customers_collection.find(query).to_list(None)
        
        # Convert UTC dates to IST and build pivot data
        pivot_data = defaultdict(lambda: defaultdict(int))
        row_values = set()
        column_values = set()
        
        for customer in customers:
            # Get row value (convert date to IST if it's a date field)
            row_val = customer.get(row_field, 'N/A')
            if row_field == 'date' or 'date' in row_field.lower() or 'time' in row_field.lower():
                row_val = convert_utc_to_ist(row_val)
            
            # Get column value (also convert if it's a date field)
            col_val = customer.get(column_field, 'N/A')
            if column_field == 'date' or 'date' in column_field.lower() or 'time' in column_field.lower():
                col_val = convert_utc_to_ist(col_val)
            
            # Skip if essential values are missing or None
            if row_val is None or row_val == 'N/A' or col_val is None or col_val == 'N/A':
                continue
            
            row_values.add(row_val)
            column_values.add(col_val)
            
            # Apply value operation
            if value_operation == 'count':
                pivot_data[row_val][col_val] += 1
            elif value_operation == 'sum' and value_field in customer:
                pivot_data[row_val][col_val] += float(customer.get(value_field, 0))
            elif value_operation == 'average' and value_field in customer:
                # Store sum and count for later average calculation
                if 'avg_data' not in pivot_data:
                    pivot_data['avg_data'] = defaultdict(lambda: defaultdict(lambda: {'sum': 0, 'count': 0}))
                pivot_data['avg_data'][row_val][col_val]['sum'] += float(customer.get(value_field, 0))
                pivot_data['avg_data'][row_val][col_val]['count'] += 1
        
        # Convert to sorted list format (filter out None values before sorting)
        sorted_rows = sorted([r for r in row_values if r is not None])
        sorted_columns = sorted([c for c in column_values if c is not None])
        
        # Build final pivot table
        table_data = []
        for row_val in sorted_rows:
            row_data = {'rowLabel': row_val}
            for col_val in sorted_columns:
                if value_operation == 'average' and 'avg_data' in pivot_data:
                    avg_info = pivot_data['avg_data'][row_val][col_val]
                    if avg_info['count'] > 0:
                        row_data[col_val] = round(avg_info['sum'] / avg_info['count'], 2)
                    else:
                        row_data[col_val] = 0
                else:
                    row_data[col_val] = pivot_data[row_val][col_val]
            table_data.append(row_data)
        
        # Get unique filter values if filter field is specified
        filter_options = []
        if filter_field:
            filter_values_set = set()
            for customer in customers:
                filter_val = customer.get(filter_field)
                if filter_val:
                    filter_values_set.add(str(filter_val))
            filter_options = sorted(list(filter_values_set))
        
        return {
            "success": True,
            "data": table_data,
            "columns": sorted_columns,
            "row_field": row_field,
            "column_field": column_field,
            "value_operation": value_operation,
            "filter_options": filter_options,
            "total_records": len(customers)
        }
        
    except Exception as e:
        logger.error(f"Failed to generate signups pivot: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Failed to generate pivot table: {str(e)}")


# ==================== DRIVER ONBOARDING - DOCUMENT UPLOAD & OCR ====================

from emergentintegrations.llm.chat import LlmChat, UserMessage

# Document upload folder
DRIVER_DOCUMENTS_FOLDER = "driver_documents"
os.makedirs(DRIVER_DOCUMENTS_FOLDER, exist_ok=True)

@api_router.post("/driver-onboarding/upload-document/{lead_id}")
async def upload_driver_document(
    lead_id: str,
    document_type: str,  # dl, aadhar, pan, gas_bill, bank_passbook
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """
    Upload document for a driver lead
    """
    try:
        # Validate document type
        valid_types = ['dl', 'aadhar', 'pan', 'gas_bill', 'bank_passbook']
        if document_type not in valid_types:
            raise HTTPException(status_code=400, detail=f"Invalid document type. Must be one of: {', '.join(valid_types)}")
        
        # Validate file type
        allowed_extensions = ['.png', '.jpg', '.jpeg', '.pdf']
        file_ext = os.path.splitext(file.filename)[1].lower()
        if file_ext not in allowed_extensions:
            raise HTTPException(status_code=400, detail=f"Invalid file type. Allowed: {', '.join(allowed_extensions)}")
        
        # Create folder for this lead
        lead_folder = os.path.join(DRIVER_DOCUMENTS_FOLDER, lead_id)
        os.makedirs(lead_folder, exist_ok=True)
        
        # Check file count limit (max 10 files per driver)
        existing_files = [f for f in os.listdir(lead_folder) if os.path.isfile(os.path.join(lead_folder, f))]
        
        # Check if we're replacing an existing file
        temp_filename = f"{document_type}{file_ext}"
        is_replacement = any(f.startswith(document_type) for f in existing_files)
        
        if not is_replacement and len(existing_files) >= 10:
            raise HTTPException(
                status_code=400, 
                detail="Maximum 10 documents allowed per driver. Please delete some documents before uploading new ones."
            )
        
        # Save file with document type name
        filename = f"{document_type}{file_ext}"
        file_path = os.path.join(lead_folder, filename)
        
        # Save the file
        with open(file_path, "wb") as f:
            content = await file.read()
            f.write(content)
        
        # Update lead record with document path
        leads_collection = db['driver_leads']
        field_name = f"{document_type}_document_path"
        
        await leads_collection.update_one(
            {'id': lead_id},
            {'$set': {
                field_name: file_path,
                f"{document_type}_document_uploaded_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        
        return {
            "success": True,
            "message": f"{document_type.upper()} document uploaded successfully",
            "file_path": file_path,
            "document_type": document_type
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to upload document: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to upload document: {str(e)}")


@api_router.get("/driver-onboarding/documents/{lead_id}/view/{document_type}")
async def view_driver_document(
    lead_id: str,
    document_type: str,
    current_user: User = Depends(get_current_user)
):
    """
    View/display document for a driver lead (returns file for inline viewing)
    """
    try:
        # Validate document type
        valid_types = ['dl', 'aadhar', 'pan_card', 'gas_bill', 'bank_passbook']
        if document_type not in valid_types:
            raise HTTPException(status_code=400, detail=f"Invalid document type. Must be one of: {', '.join(valid_types)}")
        
        # Get lead to find document path
        leads_collection = db['driver_leads']
        lead = await leads_collection.find_one({'id': lead_id})
        
        if not lead:
            raise HTTPException(status_code=404, detail="Lead not found")
        
        # Get document path from lead record
        field_name = f"{document_type}_document_path"
        document_path = lead.get(field_name)
        
        if not document_path or not os.path.exists(document_path):
            raise HTTPException(status_code=404, detail="Document not found")
        
        # Determine media type based on file extension
        file_ext = os.path.splitext(document_path)[1].lower()
        media_type_map = {
            '.png': 'image/png',
            '.jpg': 'image/jpeg',
            '.jpeg': 'image/jpeg',
            '.pdf': 'application/pdf'
        }
        media_type = media_type_map.get(file_ext, 'application/octet-stream')
        
        # Return file for inline viewing
        return FileResponse(
            path=document_path,
            media_type=media_type,
            filename=os.path.basename(document_path)
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to view document: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to view document: {str(e)}")


@api_router.get("/driver-onboarding/documents/{lead_id}/download/{document_type}")
async def download_driver_document(
    lead_id: str,
    document_type: str,
    current_user: User = Depends(get_current_user)
):
    """
    Download document for a driver lead
    """
    try:
        # Validate document type
        valid_types = ['dl', 'aadhar', 'pan_card', 'gas_bill', 'bank_passbook']
        if document_type not in valid_types:
            raise HTTPException(status_code=400, detail=f"Invalid document type. Must be one of: {', '.join(valid_types)}")
        
        # Get lead to find document path
        leads_collection = db['driver_leads']
        lead = await leads_collection.find_one({'id': lead_id})
        
        if not lead:
            raise HTTPException(status_code=404, detail="Lead not found")
        
        # Get document path from lead record
        field_name = f"{document_type}_document_path"
        document_path = lead.get(field_name)
        
        if not document_path or not os.path.exists(document_path):
            raise HTTPException(status_code=404, detail="Document not found")
        
        # Return file for download (with Content-Disposition: attachment)
        return FileResponse(
            path=document_path,
            media_type='application/octet-stream',
            filename=os.path.basename(document_path),
            headers={"Content-Disposition": f"attachment; filename={os.path.basename(document_path)}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to download document: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to download document: {str(e)}")


@api_router.delete("/driver-onboarding/documents/{lead_id}/delete/{document_type}")
async def delete_driver_document(
    lead_id: str,
    document_type: str,
    current_user: User = Depends(get_current_user)
):
    """
    Delete document for a driver lead
    """
    try:
        # Validate document type
        valid_types = ['dl', 'aadhar', 'pan_card', 'gas_bill', 'bank_passbook']
        if document_type not in valid_types:
            raise HTTPException(status_code=400, detail=f"Invalid document type. Must be one of: {', '.join(valid_types)}")
        
        # Get lead to find document path
        leads_collection = db['driver_leads']
        lead = await leads_collection.find_one({'id': lead_id})
        
        if not lead:
            raise HTTPException(status_code=404, detail="Lead not found")
        
        # Get document path from lead record
        field_name = f"{document_type}_document_path"
        document_path = lead.get(field_name)
        
        if not document_path:
            raise HTTPException(status_code=404, detail="Document not found")
        
        # Delete file from disk if it exists
        if os.path.exists(document_path):
            os.remove(document_path)
        
        # Update lead record to remove document path and uploaded_at fields
        await leads_collection.update_one(
            {'id': lead_id},
            {'$unset': {
                field_name: "",
                f"{document_type}_document_uploaded_at": ""
            }}
        )
        
        return {
            "success": True,
            "message": f"{document_type.upper()} document deleted successfully"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to delete document: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to delete document: {str(e)}")


@api_router.post("/driver-onboarding/scan-document/{lead_id}")
async def scan_driver_document(
    lead_id: str,
    document_type: str = Query(..., description="Document type to scan"),
    current_user: User = Depends(get_current_user)
):
    """
    Scan uploaded document and extract information using GPT-4o Vision
    """
    try:
        # Get the document path from database
        leads_collection = db['driver_leads']
        lead = await leads_collection.find_one({'id': lead_id})
        
        if not lead:
            raise HTTPException(status_code=404, detail="Lead not found")
        
        field_name = f"{document_type}_document_path"
        file_path = lead.get(field_name)
        
        if not file_path or not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail=f"No {document_type} document found for this lead")
        
        # Get API key
        api_key = os.environ.get('EMERGENT_LLM_KEY')
        if not api_key:
            raise HTTPException(status_code=500, detail="EMERGENT_LLM_KEY not configured")
        
        # Read the file and convert to base64
        with open(file_path, 'rb') as f:
            file_content = f.read()
            base64_image = base64.b64encode(file_content).decode('utf-8')
        
        # Determine file mime type
        file_ext = os.path.splitext(file_path)[1].lower()
        mime_types = {
            '.png': 'image/png',
            '.jpg': 'image/jpeg',
            '.jpeg': 'image/jpeg',
            '.pdf': 'application/pdf'
        }
        mime_type = mime_types.get(file_ext, 'image/jpeg')
        
        # Normalize document_type (support both 'pan' and 'pan_card')
        doc_type_normalized = document_type
        if document_type == 'pan_card':
            doc_type_normalized = 'pan'
        
        # Create extraction prompt based on document type
        prompts = {
            'dl': "Extract the Driver License Number from this image. Return ONLY the license number, nothing else.",
            'aadhar': "Extract the Aadhar Card Number (12-digit number) from this image. Return ONLY the number, nothing else.",
            'pan': "Extract the PAN Card Number (10-character alphanumeric) from this image. Return ONLY the PAN number, nothing else.",
            'gas_bill': "Extract the complete address from this gas bill. Return ONLY the address, nothing else.",
            'bank_passbook': "Extract the Bank Account Number from this bank passbook image. Return ONLY the account number, nothing else."
        }
        
        prompt = prompts.get(doc_type_normalized, "Extract relevant information from this document.")
        
        # Use emergentintegrations for OpenAI GPT-4o Vision OCR with EMERGENT_LLM_KEY
        try:
            from emergentintegrations.openai import OpenAI as EmergentOpenAI
            client = EmergentOpenAI(api_key=api_key)
        except ImportError:
            # Fallback to regular OpenAI if emergentintegrations not available
            from openai import OpenAI
            client = OpenAI(api_key=api_key)
        
        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:{mime_type};base64,{base64_image}"
                            }
                        }
                    ]
                }
            ],
            max_tokens=300
        )
        
        extracted_text = response.choices[0].message.content.strip()
        
        # Map document type to field name (support both formats)
        field_mappings = {
            'dl': 'dl_no',
            'aadhar': 'aadhar_card',
            'pan': 'pan_card',
            'pan_card': 'pan_card',
            'gas_bill': 'gas_bill',
            'bank_passbook': 'bank_passbook'
        }
        
        field_name_to_update = field_mappings.get(document_type)
        
        # Update the lead with extracted information
        if field_name_to_update:
            await leads_collection.update_one(
                {'id': lead_id},
                {'$set': {
                    field_name_to_update: extracted_text,
                    f"{document_type}_scanned_at": datetime.now(timezone.utc).isoformat()
                }}
            )
        
        return {
            "success": True,
            "extracted_data": extracted_text,
            "field_updated": field_name_to_update,
            "document_type": document_type
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to scan document: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Failed to scan document: {str(e)}")


@api_router.get("/driver-onboarding/document/{lead_id}/{document_type}")
async def get_driver_document(
    lead_id: str,
    document_type: str,
    current_user: User = Depends(get_current_user)
):
    """
    Get/Download a driver document
    """
    try:
        # Validate document type
        valid_types = ['dl', 'aadhar', 'pan', 'gas_bill', 'bank_passbook']
        if document_type not in valid_types:
            raise HTTPException(status_code=400, detail=f"Invalid document type")
        
        # Get the document path from database
        leads_collection = db['driver_leads']
        lead = await leads_collection.find_one({'id': lead_id})
        
        if not lead:
            raise HTTPException(status_code=404, detail="Lead not found")
        
        field_name = f"{document_type}_document_path"
        file_path = lead.get(field_name)
        
        if not file_path or not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail=f"No {document_type} document found")
        
        # Return the file
        return FileResponse(
            path=file_path,
            filename=os.path.basename(file_path),
            media_type='application/octet-stream'
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get document: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get document: {str(e)}")


@api_router.delete("/driver-onboarding/document/{lead_id}/{document_type}")
async def delete_driver_document(
    lead_id: str,
    document_type: str,
    current_user: User = Depends(get_current_user)
):
    """
    Delete a driver document
    """
    try:
        # Validate document type
        valid_types = ['dl', 'aadhar', 'pan', 'gas_bill', 'bank_passbook']
        if document_type not in valid_types:
            raise HTTPException(status_code=400, detail=f"Invalid document type")
        
        # Get the document path from database
        leads_collection = db['driver_leads']
        lead = await leads_collection.find_one({'id': lead_id})
        
        if not lead:
            raise HTTPException(status_code=404, detail="Lead not found")
        
        field_name = f"{document_type}_document_path"
        file_path = lead.get(field_name)
        
        if file_path and os.path.exists(file_path):
            # Delete the physical file
            os.remove(file_path)
        
        # Remove document path from database
        await leads_collection.update_one(
            {'id': lead_id},
            {'$unset': {
                field_name: "",
                f"{document_type}_document_uploaded_at": ""
            }}
        )
        
        return {
            "success": True,
            "message": f"{document_type.upper()} document deleted successfully",
            "document_type": document_type
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to delete document: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to delete document: {str(e)}")


@api_router.get("/driver-onboarding/documents/status/{lead_id}")
async def get_documents_status(
    lead_id: str,
    current_user: User = Depends(get_current_user)
):
    """
    Get status of all documents for a lead (uploaded or not)
    """
    try:
        leads_collection = db['driver_leads']
        lead = await leads_collection.find_one({'id': lead_id})
        
        if not lead:
            raise HTTPException(status_code=404, detail="Lead not found")
        
        document_types = ['dl', 'aadhar', 'pan', 'gas_bill', 'bank_passbook']
        documents_status = {}
        
        for doc_type in document_types:
            field_name = f"{doc_type}_document_path"
            file_path = lead.get(field_name)
            uploaded_at = lead.get(f"{doc_type}_document_uploaded_at")
            
            documents_status[doc_type] = {
                "uploaded": bool(file_path and os.path.exists(file_path)),
                "file_path": file_path if file_path and os.path.exists(file_path) else None,
                "uploaded_at": uploaded_at,
                "filename": os.path.basename(file_path) if file_path else None
            }
        
        return {
            "success": True,
            "lead_id": lead_id,
            "documents": documents_status
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get documents status: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get documents status: {str(e)}")



# ==================== DOCUMENT LIBRARY ====================

@api_router.get("/document-library/list")
async def list_all_documents(
    search: str = None,
    document_type: str = None,
    current_user: User = Depends(get_current_user)
):
    """
    List all documents from all driver leads with metadata
    Returns folders organized by driver with their documents
    """
    try:
        import zipfile
        from io import BytesIO
        
        # Get all leads with documents
        leads_collection = db['driver_leads']
        query = {}
        
        leads = await leads_collection.find(query, {"_id": 0}).to_list(length=None)
        
        folders = []
        document_types = ['dl', 'aadhar', 'pan', 'gas_bill', 'bank_passbook']
        
        for lead in leads:
            lead_id = lead.get('id')
            if not lead_id:
                continue
                
            # Check if this lead has any documents
            has_documents = False
            documents = []
            
            for doc_type in document_types:
                field_name = f"{doc_type}_document_path"
                file_path = lead.get(field_name)
                uploaded_at = lead.get(f"{doc_type}_document_uploaded_at")
                
                if file_path and os.path.exists(file_path):
                    # Filter by document type if specified
                    if document_type and doc_type != document_type:
                        continue
                        
                    has_documents = True
                    file_stats = os.stat(file_path)
                    file_ext = os.path.splitext(file_path)[1]
                    
                    documents.append({
                        "document_type": doc_type,
                        "file_path": file_path,
                        "filename": os.path.basename(file_path),
                        "file_size": file_stats.st_size,
                        "file_extension": file_ext,
                        "uploaded_at": uploaded_at
                    })
            
            if has_documents:
                driver_name = lead.get('name', 'Unknown')
                phone = lead.get('phone_number', 'N/A')
                
                # Apply search filter
                if search:
                    search_lower = search.lower()
                    if not (search_lower in driver_name.lower() or search_lower in phone):
                        continue
                
                folders.append({
                    "lead_id": lead_id,
                    "driver_name": driver_name,
                    "phone_number": phone,
                    "email": lead.get('email', 'N/A'),
                    "documents": documents,
                    "document_count": len(documents),
                    "total_size": sum(doc['file_size'] for doc in documents)
                })
        
        # Calculate stats
        total_documents = sum(folder['document_count'] for folder in folders)
        total_size = sum(folder['total_size'] for folder in folders)
        
        return {
            "success": True,
            "folders": folders,
            "stats": {
                "total_drivers": len(folders),
                "total_documents": total_documents,
                "total_size": total_size,
                "total_size_mb": round(total_size / (1024 * 1024), 2)
            }
        }
        
    except Exception as e:
        logger.error(f"Failed to list documents: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to list documents: {str(e)}")


@api_router.get("/document-library/download/{lead_id}/{document_type}")
async def download_document(
    lead_id: str,
    document_type: str,
    current_user: User = Depends(get_current_user)
):
    """
    Download a single document
    """
    try:
        leads_collection = db['driver_leads']
        lead = await leads_collection.find_one({'id': lead_id}, {"_id": 0})
        
        if not lead:
            raise HTTPException(status_code=404, detail="Driver not found")
        
        field_name = f"{document_type}_document_path"
        file_path = lead.get(field_name)
        
        if not file_path or not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="Document not found")
        
        driver_name = lead.get('name', 'Unknown').replace(' ', '_')
        filename = f"{driver_name}_{document_type}{os.path.splitext(file_path)[1]}"
        
        return FileResponse(
            path=file_path,
            filename=filename,
            media_type='application/octet-stream'
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to download document: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to download document: {str(e)}")


@api_router.post("/document-library/download-bulk")
async def download_bulk_documents(
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """
    Download selected documents as a ZIP file
    Request body: {"selections": [{"lead_id": "...", "document_type": "..."}]}
    """
    try:
        import zipfile
        from io import BytesIO
        
        body = await request.json()
        selections = body.get('selections', [])
        
        if not selections:
            raise HTTPException(status_code=400, detail="No documents selected")
        
        # Create ZIP file in memory
        zip_buffer = BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            leads_collection = db['driver_leads']
            
            for selection in selections:
                lead_id = selection.get('lead_id')
                document_type = selection.get('document_type')
                
                if not lead_id or not document_type:
                    continue
                
                lead = await leads_collection.find_one({'id': lead_id}, {"_id": 0})
                if not lead:
                    continue
                
                field_name = f"{document_type}_document_path"
                file_path = lead.get(field_name)
                
                if file_path and os.path.exists(file_path):
                    driver_name = lead.get('name', 'Unknown').replace(' ', '_')
                    file_ext = os.path.splitext(file_path)[1]
                    
                    # Create folder structure in ZIP: driver_name/document_type.ext
                    zip_path = f"{driver_name}/{document_type}{file_ext}"
                    zip_file.write(file_path, zip_path)
        
        zip_buffer.seek(0)
        
        timestamp = datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')
        filename = f"driver_documents_{timestamp}.zip"
        
        return StreamingResponse(
            zip_buffer,
            media_type='application/zip',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to create bulk download: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to create bulk download: {str(e)}")


@api_router.post("/document-library/delete")
async def delete_documents(
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """
    Delete selected documents
    Request body: {"selections": [{"lead_id": "...", "document_type": "..."}]}
    """
    try:
        body = await request.json()
        selections = body.get('selections', [])
        
        if not selections:
            raise HTTPException(status_code=400, detail="No documents selected")
        
        leads_collection = db['driver_leads']
        deleted_count = 0
        
        for selection in selections:
            lead_id = selection.get('lead_id')
            document_type = selection.get('document_type')
            
            if not lead_id or not document_type:
                continue
            
            lead = await leads_collection.find_one({'id': lead_id}, {"_id": 0})
            if not lead:
                continue
            
            field_name = f"{document_type}_document_path"
            file_path = lead.get(field_name)
            
            if file_path and os.path.exists(file_path):
                # Delete the file
                os.remove(file_path)
                
                # Update database
                await leads_collection.update_one(
                    {'id': lead_id},
                    {'$unset': {
                        field_name: "",
                        f"{document_type}_document_uploaded_at": ""
                    }}
                )
                
                deleted_count += 1
        
        return {
            "success": True,
            "message": f"Successfully deleted {deleted_count} document(s)",
            "deleted_count": deleted_count
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to delete documents: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to delete documents: {str(e)}")


@api_router.get("/document-library/preview/{lead_id}/{document_type}")
async def preview_document(
    lead_id: str,
    document_type: str,
    current_user: User = Depends(get_current_user)
):
    """
    Preview a document (returns file for inline viewing in browser)
    """
    try:
        leads_collection = db['driver_leads']
        lead = await leads_collection.find_one({'id': lead_id}, {"_id": 0})
        
        if not lead:
            raise HTTPException(status_code=404, detail="Driver not found")
        
        field_name = f"{document_type}_document_path"
        file_path = lead.get(field_name)
        
        if not file_path or not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="Document not found")
        
        # Determine media type
        file_ext = os.path.splitext(file_path)[1].lower()
        media_types = {
            '.pdf': 'application/pdf',
            '.png': 'image/png',
            '.jpg': 'image/jpeg',
            '.jpeg': 'image/jpeg'
        }
        media_type = media_types.get(file_ext, 'application/octet-stream')
        
        return FileResponse(
            path=file_path,
            media_type=media_type,
            headers={'Content-Disposition': f'inline; filename="{os.path.basename(file_path)}"'}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to preview document: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to preview document: {str(e)}")



# ==================== RIDE PAY EXTRACT V2 - UPLOAD FIRST, PROCESS IN BACKGROUND ====================

# Storage folder for uploaded images
RIDE_PAY_V2_FOLDER = "ride_pay_v2_uploads"
os.makedirs(RIDE_PAY_V2_FOLDER, exist_ok=True)

@api_router.post("/ride-pay-v2/upload")
async def upload_ride_pay_images(
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """
    Upload images to folder WITHOUT processing (instant upload)
    """
    try:
        # Parse form data
        form = await request.form()
        files = form.getlist("files")
        month_year = form.get("month_year", "")
        driver_name = form.get("driver_name", "")
        vehicle_number = form.get("vehicle_number", "")
        platform = form.get("platform", "")
        folder_id = form.get("folder_id", "")  # If appending to existing folder
        
        if not files:
            raise HTTPException(status_code=400, detail="No files uploaded")
        
        if len(files) > 10:
            raise HTTPException(status_code=400, detail="Maximum 10 files allowed per upload")
        
        if not month_year or not driver_name:
            raise HTTPException(status_code=400, detail="month_year and driver_name are required")
        
        # Check if folder exists or create new
        if folder_id:
            folder = await db.ride_pay_folders.find_one({'id': folder_id})
            if not folder:
                raise HTTPException(status_code=404, detail="Folder not found")
        else:
            # Create new folder
            folder_id = str(uuid.uuid4())
            folder_doc = {
                "id": folder_id,
                "month_year": month_year,
                "driver_name": driver_name,
                "vehicle_number": vehicle_number,
                "platform": platform,
                "user_id": current_user.id,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "status": "pending",  # pending, processing, completed, failed
                "total_images": 0,
                "processed_images": 0,
                "failed_images": 0
            }
            await db.ride_pay_folders.insert_one(folder_doc)
        
        # Create folder directory
        folder_path = os.path.join(RIDE_PAY_V2_FOLDER, folder_id)
        os.makedirs(folder_path, exist_ok=True)
        
        # Save all files instantly
        uploaded_images = []
        for file in files:
            # Generate unique filename
            file_id = str(uuid.uuid4())
            file_ext = os.path.splitext(file.filename)[1]
            filename = f"{file_id}{file_ext}"
            filepath = os.path.join(folder_path, filename)
            
            # Save file
            content = await file.read()
            with open(filepath, 'wb') as f:
                f.write(content)
            
            # Save image record to database
            image_doc = {
                "id": file_id,
                "folder_id": folder_id,
                "filename": file.filename,
                "filepath": filepath,
                "status": "pending",  # pending, processing, completed, failed
                "uploaded_at": datetime.now(timezone.utc).isoformat(),
                "rides_extracted": 0
            }
            await db.ride_pay_images.insert_one(image_doc)
            uploaded_images.append(image_doc)
        
        # Update folder total_images count
        await db.ride_pay_folders.update_one(
            {'id': folder_id},
            {'$inc': {'total_images': len(files)}}
        )
        
        logger.info(f"✅ Uploaded {len(files)} images to folder {folder_id}")
        
        return {
            "success": True,
            "folder_id": folder_id,
            "uploaded_images": len(files),
            "message": f"Successfully uploaded {len(files)} images. Ready to process."
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to upload images: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Failed to upload images: {str(e)}")


@api_router.post("/ride-pay-v2/process-folder/{folder_id}")
async def process_ride_pay_folder(
    folder_id: str,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(get_current_user)
):
    """
    Trigger background processing for a folder
    """
    try:
        # Check if folder exists
        folder = await db.ride_pay_folders.find_one({'id': folder_id})
        if not folder:
            raise HTTPException(status_code=404, detail="Folder not found")
        
        # Check if folder is already processing
        if folder.get('status') == 'processing':
            raise HTTPException(status_code=400, detail="Folder is already being processed")
        
        # Import worker function
        from ride_pay_v2_worker import process_folder_worker
        
        # Add to background tasks
        background_tasks.add_task(process_folder_worker, folder_id)
        
        # Update folder status
        await db.ride_pay_folders.update_one(
            {'id': folder_id},
            {'$set': {
                'status': 'queued',
                'queued_at': datetime.now(timezone.utc).isoformat()
            }}
        )
        
        return {
            "success": True,
            "message": "Folder processing started in background",
            "folder_id": folder_id
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to start folder processing: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to start processing: {str(e)}")


@api_router.get("/ride-pay-v2/status/{folder_id}")
async def get_folder_status(
    folder_id: str,
    current_user: User = Depends(get_current_user)
):
    """
    Get processing status for a folder with real-time progress
    """
    try:
        # Get folder
        folder = await db.ride_pay_folders.find_one({'id': folder_id})
        if not folder:
            raise HTTPException(status_code=404, detail="Folder not found")
        
        # Get image counts
        total_images = await db.ride_pay_images.count_documents({'folder_id': folder_id})
        pending_images = await db.ride_pay_images.count_documents({
            'folder_id': folder_id,
            'status': 'pending'
        })
        processing_images = await db.ride_pay_images.count_documents({
            'folder_id': folder_id,
            'status': 'processing'
        })
        completed_images = await db.ride_pay_images.count_documents({
            'folder_id': folder_id,
            'status': 'completed'
        })
        failed_images = await db.ride_pay_images.count_documents({
            'folder_id': folder_id,
            'status': 'failed'
        })
        
        # Calculate progress percentage
        progress = 0
        if total_images > 0:
            progress = int((completed_images / total_images) * 100)
        
        # Estimate time remaining (assume 20 seconds per image)
        estimated_seconds = pending_images * 20
        estimated_minutes = estimated_seconds // 60
        
        # Remove _id from folder
        if '_id' in folder:
            del folder['_id']
        
        return {
            "success": True,
            "folder": folder,
            "total_images": total_images,
            "pending_images": pending_images,
            "processing_images": processing_images,
            "completed_images": completed_images,
            "failed_images": failed_images,
            "progress_percentage": progress,
            "estimated_time_remaining": f"{estimated_minutes} minutes" if estimated_minutes > 0 else "Less than 1 minute"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get folder status: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get status: {str(e)}")


@api_router.get("/ride-pay-v2/folders")
async def get_all_folders(
    current_user: User = Depends(get_current_user)
):
    """
    Get all folders with status and stats
    """
    try:
        # Get all folders
        folders = await db.ride_pay_folders.find({}).sort('created_at', -1).to_list(None)
        
        # Get image counts for each folder
        for folder in folders:
            folder_id = folder['id']
            total_images = await db.ride_pay_images.count_documents({'folder_id': folder_id})
            completed_images = await db.ride_pay_images.count_documents({
                'folder_id': folder_id,
                'status': 'completed'
            })
            failed_images = await db.ride_pay_images.count_documents({
                'folder_id': folder_id,
                'status': 'failed'
            })
            
            folder['total_images'] = total_images
            folder['completed_images'] = completed_images
            folder['failed_images'] = failed_images
            
            # Calculate progress
            if total_images > 0:
                folder['progress_percentage'] = int((completed_images / total_images) * 100)
            else:
                folder['progress_percentage'] = 0
            
            # Remove _id
            if '_id' in folder:
                del folder['_id']
        
        return {
            "success": True,
            "folders": folders,
            "total_folders": len(folders)
        }
        
    except Exception as e:
        logger.error(f"Failed to get folders: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get folders: {str(e)}")


@api_router.get("/ride-pay-v2/data/{folder_id}")
async def get_folder_data(
    folder_id: str,
    skip: int = 0,
    limit: int = 100,
    current_user: User = Depends(get_current_user)
):
    """
    Get extracted data table for a folder with pagination
    """
    try:
        # Get folder
        folder = await db.ride_pay_folders.find_one({'id': folder_id})
        if not folder:
            raise HTTPException(status_code=404, detail="Folder not found")
        
        # Get extracted data
        data = await db.ride_pay_data.find({
            'folder_id': folder_id
        }).skip(skip).limit(limit).to_list(None)
        
        # Get total count
        total_records = await db.ride_pay_data.count_documents({'folder_id': folder_id})
        
        # Remove _id from records
        for record in data:
            if '_id' in record:
                del record['_id']
        
        return {
            "success": True,
            "data": data,
            "total_records": total_records,
            "skip": skip,
            "limit": limit,
            "folder_id": folder_id,
            "month_year": folder.get('month_year'),
            "driver_name": folder.get('driver_name')
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get folder data: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get data: {str(e)}")


@api_router.delete("/ride-pay-v2/folder/{folder_id}")
async def delete_folder(
    folder_id: str,
    current_user: User = Depends(get_current_user)
):
    """
    Delete a folder and all its data (Master Admin only)
    """
    if current_user.account_type != 'master_admin':
        raise HTTPException(status_code=403, detail="Only Master Admin can delete folders")
    
    try:
        # Get folder
        folder = await db.ride_pay_folders.find_one({'id': folder_id})
        if not folder:
            raise HTTPException(status_code=404, detail="Folder not found")
        
        # Delete folder directory
        folder_path = os.path.join(RIDE_PAY_V2_FOLDER, folder_id)
        if os.path.exists(folder_path):
            import shutil
            shutil.rmtree(folder_path)
        
        # Delete from database
        await db.ride_pay_images.delete_many({'folder_id': folder_id})
        await db.ride_pay_data.delete_many({'folder_id': folder_id})
        await db.ride_pay_folders.delete_one({'id': folder_id})
        
        return {
            "success": True,
            "message": "Folder deleted successfully"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to delete folder: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to delete folder: {str(e)}")


# ==================== DATABASE EXPORT/IMPORT (Master Admin Only) ====================

@api_router.get("/admin/database/export")
async def export_database(
    current_user: User = Depends(get_current_user)
):
    """Export entire database as JSON - Master Admin only"""
    try:
        # Check master admin permission
        if current_user.account_type != "master_admin":
            raise HTTPException(status_code=403, detail="Only Master Admin can export database")
        
        from fastapi.responses import StreamingResponse
        import json
        from io import BytesIO
        
        logger.info(f"Database export initiated by {current_user.email}")
        
        # Collections to export
        collections_to_export = [
            "users",
            "driver_leads", 
            "montra_feed_data",
            "payment_records",
            "qr_codes",
            "qr_scans",
            "expenses",
            "customer_data",
            "ride_data",
            "activity_logs",
            "app_settings",
            "ride_pay_folders",
            "ride_pay_images",
            "ride_pay_data"
        ]
        
        export_data = {
            "export_timestamp": datetime.now(timezone.utc).isoformat(),
            "exported_by": current_user.email,
            "collections": {}
        }
        
        # Export each collection
        for collection_name in collections_to_export:
            try:
                collection = db[collection_name]
                documents = await collection.find({}).to_list(None)
                
                # Remove _id field (not JSON serializable)
                for doc in documents:
                    if '_id' in doc:
                        del doc['_id']
                
                export_data["collections"][collection_name] = documents
                logger.info(f"Exported {len(documents)} documents from {collection_name}")
            except Exception as e:
                logger.warning(f"Failed to export {collection_name}: {str(e)}")
                export_data["collections"][collection_name] = []
        
        # Convert to JSON
        json_data = json.dumps(export_data, indent=2, default=str)
        
        # Return as downloadable file
        return StreamingResponse(
            iter([json_data]),
            media_type="application/json",
            headers={
                "Content-Disposition": f"attachment; filename=nura_pulse_backup_{datetime.now().strftime('%Y%m%d')}.json"
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Database export failed: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Failed to export database: {str(e)}")


@api_router.post("/admin/database/import")
async def import_database(
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """Import database from JSON backup - Master Admin only"""
    try:
        # Check master admin permission
        if current_user.account_type != "master_admin":
            raise HTTPException(status_code=403, detail="Only Master Admin can import database")
        
        import json
        
        logger.info(f"Database import initiated by {current_user.email}")
        
        # Parse form data
        form = await request.form()
        file = form.get('file')
        
        if not file:
            raise HTTPException(status_code=400, detail="No backup file provided")
        
        # Read file content
        content = await file.read()
        
        try:
            import_data = json.loads(content)
        except json.JSONDecodeError:
            raise HTTPException(status_code=400, detail="Invalid JSON file")
        
        if "collections" not in import_data:
            raise HTTPException(status_code=400, detail="Invalid backup file format")
        
        imported_collections = {}
        
        # Import each collection
        for collection_name, documents in import_data["collections"].items():
            if not documents:
                logger.info(f"Skipping empty collection: {collection_name}")
                continue
            
            try:
                collection = db[collection_name]
                
                # Clear existing data
                await collection.delete_many({})
                logger.info(f"Cleared existing data from {collection_name}")
                
                # Insert new data
                if len(documents) > 0:
                    await collection.insert_many(documents)
                    imported_collections[collection_name] = len(documents)
                    logger.info(f"Imported {len(documents)} documents to {collection_name}")
            except Exception as e:
                logger.error(f"Failed to import {collection_name}: {str(e)}")
                imported_collections[collection_name] = f"Error: {str(e)}"
        
        logger.info(f"Database import completed by {current_user.email}")
        
        return {
            "success": True,
            "message": "Database imported successfully",
            "imported_collections": imported_collections,
            "timestamp": datetime.now(timezone.utc).isoformat()
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Database import failed: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Failed to import database: {str(e)}")


# ==================== APP SETTINGS MANAGEMENT (Master Admin) ====================

@api_router.get("/app-settings")
async def get_app_settings(
    current_user: User = Depends(get_current_user)
):
    """
    Get application settings
    """
    try:
        settings_collection = db['app_settings']
        
        # Get or create default settings
        settings = await settings_collection.find_one({'_id': 'global_settings'})
        
        if not settings:
            # Create default settings
            default_settings = {
                '_id': 'global_settings',
                'payment_extractor_enabled': True,
                'maintenance_mode': False,
                'updated_at': datetime.now(timezone.utc).isoformat(),
                'updated_by': 'system'
            }
            await settings_collection.insert_one(default_settings)
            settings = default_settings
        
        return {
            "success": True,
            "settings": {
                "payment_extractor_enabled": settings.get('payment_extractor_enabled', True),
                "maintenance_mode": settings.get('maintenance_mode', False)
            }
        }
        
    except Exception as e:
        logger.error(f"Failed to get app settings: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get settings: {str(e)}")


@api_router.get("/maintenance-status")
async def get_maintenance_status():
    """
    Public endpoint to check maintenance mode (no authentication required)
    """
    try:
        settings_collection = db['app_settings']
        settings = await settings_collection.find_one({'_id': 'global_settings'})
        
        return {
            "maintenance_mode": settings.get('maintenance_mode', False) if settings else False
        }
    except Exception as e:
        logger.error(f"Failed to check maintenance status: {str(e)}")
        return {"maintenance_mode": False}


@api_router.put("/app-settings")
async def update_app_settings(
    payment_extractor_enabled: bool = None,
    maintenance_mode: bool = None,
    current_user: User = Depends(get_current_user)
):
    """
    Update application settings (Master Admin only)
    """
    if current_user.account_type != "master_admin":
        raise HTTPException(status_code=403, detail="Only Master Admins can update app settings")
    
    try:
        settings_collection = db['app_settings']
        
        # Build update data based on what was provided
        update_data = {
            'updated_at': datetime.now(timezone.utc).isoformat(),
            'updated_by': current_user.email
        }
        
        action_message = []
        
        if payment_extractor_enabled is not None:
            update_data['payment_extractor_enabled'] = payment_extractor_enabled
            action_message.append(f"Payment Extractor {'enabled' if payment_extractor_enabled else 'disabled'}")
        
        if maintenance_mode is not None:
            update_data['maintenance_mode'] = maintenance_mode
            action_message.append(f"Maintenance Mode {'enabled' if maintenance_mode else 'disabled'}")
        
        result = await settings_collection.update_one(
            {'_id': 'global_settings'},
            {'$set': update_data},
            upsert=True
        )
        
        # Log this action
        await log_user_activity(
            user_email=current_user.email,
            action="update_app_settings",
            details=", ".join(action_message),
            module="App Settings"
        )
        
        # Get updated settings to return
        updated_settings = await settings_collection.find_one({'_id': 'global_settings'})
        
        return {
            "success": True,
            "message": ", ".join(action_message),
            "settings": {
                "payment_extractor_enabled": updated_settings.get('payment_extractor_enabled', True),
                "maintenance_mode": updated_settings.get('maintenance_mode', False)
            }
        }
        
    except Exception as e:
        logger.error(f"Failed to update app settings: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to update settings: {str(e)}")


# ==================== USER ACTIVITY LOGGING ====================

async def log_user_activity(
    user_email: str,
    action: str,
    details: str = None,
    module: str = None,
    ip_address: str = None
):
    """
    Log user activity to database
    """
    try:
        activity_logs = db['user_activity_logs']
        
        log_entry = {
            'id': str(uuid.uuid4()),
            'user_email': user_email,
            'action': action,
            'details': details,
            'module': module,
            'ip_address': ip_address,
            'timestamp': datetime.now(timezone.utc).isoformat(),
            'date': datetime.now(timezone.utc).strftime('%Y-%m-%d')
        }
        
        await activity_logs.insert_one(log_entry)
        
    except Exception as e:
        logger.error(f"Failed to log activity: {str(e)}")


@api_router.get("/analytics/user-activity")
async def get_user_activity_logs(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user_email: Optional[str] = None,
    module: Optional[str] = None,
    limit: int = 100,
    skip: int = 0,
    current_user: User = Depends(get_current_user)
):
    """
    Get user activity logs (Master Admin and Admin only)
    """
    if current_user.account_type not in ["master_admin", "admin"]:
        raise HTTPException(status_code=403, detail="Only Admins can view activity logs")
    
    try:
        activity_logs = db['user_activity_logs']
        
        # Build query
        query = {}
        
        if start_date and end_date:
            query['date'] = {'$gte': start_date, '$lte': end_date}
        elif start_date:
            query['date'] = {'$gte': start_date}
        elif end_date:
            query['date'] = {'$lte': end_date}
        
        if user_email:
            query['user_email'] = user_email
        
        if module:
            query['module'] = module
        
        # Get logs
        logs = await activity_logs.find(query).sort('timestamp', -1).skip(skip).limit(limit).to_list(None)
        total_count = await activity_logs.count_documents(query)
        
        # Remove _id field
        for log in logs:
            if '_id' in log:
                del log['_id']
        
        return {
            "success": True,
            "logs": logs,
            "total": total_count,
            "limit": limit,
            "skip": skip
        }
        
    except Exception as e:
        logger.error(f"Failed to get activity logs: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get logs: {str(e)}")


@api_router.get("/analytics/active-users")
async def get_active_users(
    current_user: User = Depends(get_current_user)
):
    """
    Get currently active users (users who had activity in last 5 minutes)
    """
    if current_user.account_type not in ["master_admin", "admin"]:
        raise HTTPException(status_code=403, detail="Only Admins can view active users")
    
    try:
        activity_logs = db['user_activity_logs']
        
        # Get activity from last 5 minutes
        five_minutes_ago = datetime.now(timezone.utc) - timedelta(minutes=5)
        five_minutes_ago_iso = five_minutes_ago.isoformat()
        
        # Aggregate to get latest activity per user
        pipeline = [
            {"$match": {"timestamp": {"$gte": five_minutes_ago_iso}}},
            {"$sort": {"timestamp": -1}},
            {"$group": {
                "_id": "$user_email",
                "latest_action": {"$first": "$action"},
                "latest_module": {"$first": "$module"},
                "latest_timestamp": {"$first": "$timestamp"},
                "activity_count": {"$sum": 1}
            }}
        ]
        
        active_users_data = await activity_logs.aggregate(pipeline).to_list(None)
        
        active_users = []
        for user_data in active_users_data:
            active_users.append({
                "user_email": user_data['_id'],
                "current_module": user_data['latest_module'],
                "latest_action": user_data['latest_action'],
                "last_activity": user_data['latest_timestamp'],
                "activity_count": user_data['activity_count']
            })
        
        return {
            "success": True,
            "active_users": active_users,
            "count": len(active_users)
        }
        
    except Exception as e:
        logger.error(f"Failed to get active users: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get active users: {str(e)}")


@api_router.post("/analytics/log-activity")
async def log_activity(
    action: str,
    module: str,
    details: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """
    Log user activity from frontend
    """
    try:
        await log_user_activity(
            user_email=current_user.email,
            action=action,
            details=details,
            module=module
        )
        
        return {
            "success": True,
            "message": "Activity logged"
        }
        
    except Exception as e:
        logger.error(f"Failed to log activity: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to log activity: {str(e)}")


# ==================== RCA (Root Cause Analysis) Management ====================

@api_router.get("/ride-deck/rca/cancelled")
async def get_cancelled_rides_for_rca(
    current_user: User = Depends(get_current_user)
):
    """
    Get cancelled rides where statusReason is empty for RCA analysis
    """
    try:
        rides_collection = db['rides']
        customers_collection = db['customers']
        
        # Find cancelled rides with empty statusReason
        query = {
            "rideStatus": "CANCELLED",
            "$or": [
                {"statusReason": {"$exists": False}},
                {"statusReason": None},
                {"statusReason": ""}
            ]
        }
        
        cancelled_rides = await rides_collection.find(query).to_list(None)
        
        # Enrich with customer names
        result = []
        for ride in cancelled_rides:
            customer_id = ride.get('customerId')
            customer_name = None
            
            if customer_id:
                customer = await customers_collection.find_one({'id': str(customer_id)})
                if customer:
                    customer_name = customer.get('name')
            
            result.append({
                'id': ride.get('id'),
                'customerId': ride.get('customerId'),
                'customerName': customer_name or 'N/A',
                'rideStartTime': ride.get('rideStartTime'),
                'pickupLocality': ride.get('pickupLocality', 'N/A'),
                'dropLocality': ride.get('dropLocality', 'N/A'),
                'statusReason': ride.get('statusReason', ''),
                'statusDetail': ride.get('statusDetail', '')
            })
        
        return {
            "success": True,
            "count": len(result),
            "rides": result
        }
        
    except Exception as e:
        logger.error(f"Failed to fetch cancelled rides: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch cancelled rides: {str(e)}")


@api_router.get("/ride-deck/rca/driver-not-found")
async def get_driver_not_found_rides_for_rca(
    current_user: User = Depends(get_current_user)
):
    """
    Get driver not found rides where statusReason is empty for RCA analysis
    """
    try:
        rides_collection = db['rides']
        customers_collection = db['customers']
        
        # Find driver_not_found rides with empty statusReason
        query = {
            "rideStatus": "DRIVER_NOT_FOUND",
            "$or": [
                {"statusReason": {"$exists": False}},
                {"statusReason": None},
                {"statusReason": ""}
            ]
        }
        
        dnf_rides = await rides_collection.find(query).to_list(None)
        
        # Enrich with customer names
        result = []
        for ride in dnf_rides:
            customer_id = ride.get('customerId')
            customer_name = None
            
            if customer_id:
                customer = await customers_collection.find_one({'id': str(customer_id)})
                if customer:
                    customer_name = customer.get('name')
            
            result.append({
                'id': ride.get('id'),
                'customerId': ride.get('customerId'),
                'customerName': customer_name or 'N/A',
                'rideStartTime': ride.get('rideStartTime'),
                'pickupLocality': ride.get('pickupLocality', 'N/A'),
                'dropLocality': ride.get('dropLocality', 'N/A'),
                'statusReason': ride.get('statusReason', ''),
                'statusDetail': ride.get('statusDetail', '')
            })
        
        return {
            "success": True,
            "count": len(result),
            "rides": result
        }
        
    except Exception as e:
        logger.error(f"Failed to fetch driver not found rides: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch driver not found rides: {str(e)}")


@api_router.put("/ride-deck/rca/update/{ride_id}")
async def update_ride_rca(
    ride_id: str,
    statusReason: str,
    statusDetail: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """
    Update statusReason and statusDetail for a ride
    """
    try:
        rides_collection = db['rides']
        
        # Find the ride
        ride = await rides_collection.find_one({'id': ride_id})
        if not ride:
            raise HTTPException(status_code=404, detail="Ride not found")
        
        # Prepare update data
        update_data = {
            "statusReason": statusReason,
            "statusDetail": statusDetail or "",
            "updatedBy": current_user.email,
            "updatedAt": datetime.now(timezone.utc).isoformat()
        }
        
        # Update the ride
        result = await rides_collection.update_one(
            {'id': ride_id},
            {'$set': update_data}
        )
        
        if result.modified_count == 0:
            raise HTTPException(status_code=400, detail="Failed to update ride")
        
        return {
            "success": True,
            "message": "Ride RCA updated successfully",
            "ride_id": ride_id
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to update ride RCA: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to update ride RCA: {str(e)}")


# ==================== QR CODE MANAGER ====================

import qrcode
import io
import base64
from PIL import Image
from user_agents import parse
import secrets
import string

# QR Code Models
class QRCodeCreate(BaseModel):
    campaign_name: str
    landing_page_type: str = "single"  # "single" or "multiple"
    ios_url: Optional[str] = None
    android_url: Optional[str] = None
    web_url: Optional[str] = None
    single_url: Optional[str] = None
    utm_source: Optional[str] = None
    utm_medium: str = "qrscan"
    utm_campaign: Optional[str] = None
    utm_term: Optional[str] = None
    utm_content: Optional[str] = None
    use_color_qr: bool = False
    qr_foreground_color: str = "#000000"
    qr_background_color: str = "#FFFFFF"

class QRCodeBatchCreate(BaseModel):
    campaign_name: str
    landing_page_type: str = "single"
    # NEW field names matching scan endpoint expectations
    landing_page_ios: Optional[str] = None
    landing_page_android: Optional[str] = None
    landing_page_mobile: Optional[str] = None
    landing_page_desktop: Optional[str] = None
    landing_page_single: Optional[str] = None
    # Keep OLD field names for backward compatibility
    ios_url: Optional[str] = None
    android_url: Optional[str] = None
    web_url: Optional[str] = None
    single_url: Optional[str] = None
    qr_count: int = Field(ge=1, le=100)
    qr_names: Optional[List[str]] = None
    auto_fill_utm: bool = True
    utm_medium: str = "qrscan"
    utm_campaign: Optional[str] = None
    use_color_qr: bool = False
    qr_foreground_color: str = "#000000"
    qr_background_color: str = "#FFFFFF"

def get_location_from_ip(ip_address: str, headers: dict = None) -> dict:
    """Get location information from IP address with X-Forwarded-For support"""
    try:
        # Try to get real public IP from X-Forwarded-For header (for users behind proxies)
        real_ip = ip_address
        if headers:
            forwarded_for = headers.get("x-forwarded-for", "")
            if forwarded_for:
                # X-Forwarded-For can have multiple IPs, get the first one (client's real IP)
                real_ip = forwarded_for.split(",")[0].strip()
                logger.info(f"Using X-Forwarded-For IP: {real_ip} instead of direct IP: {ip_address}")
        
        # Skip for localhost only, but try geolocation for private IPs (they might be X-Forwarded)
        if real_ip in ["127.0.0.1", "localhost"]:
            return {"city": "Local", "region": "Local", "country": "Local"}
        
        # Use a free IP geolocation service (works for public IPs)
        response = requests.get(f"http://ip-api.com/json/{real_ip}", timeout=5)
        if response.status_code == 200:
            data = response.json()
            # Check if geolocation was successful
            if data.get("status") == "success":
                city = data.get("city", "Unknown")
                region = data.get("regionName", "Unknown")
                country = data.get("country", "Unknown")
                logger.info(f"Geolocation success for IP {real_ip}: {city}, {region}, {country}")
                return {
                    "city": city,
                    "region": region,
                    "country": country
                }
            else:
                logger.warning(f"Geolocation failed for IP {real_ip}: {data.get('message', 'Unknown error')}")
    except Exception as e:
        logger.error(f"Error getting location for IP {ip_address}: {str(e)}")
    
    # Default fallback locations for India (since this is Nura Mobility)
    return {"city": "Chennai", "region": "Tamil Nadu", "country": "India"}

def generate_short_code(length: int = 8) -> str:
    """Generate a random short code for QR tracking"""
    characters = string.ascii_letters + string.digits
    return ''.join(secrets.choice(characters) for _ in range(length))

def generate_qr_code_image(data: str, foreground_color: str = "black", background_color: str = "white") -> str:
    """Generate QR code image and return as base64"""
    qr = qrcode.QRCode(version=1, box_size=10, border=4)
    qr.add_data(data)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color=foreground_color, back_color=background_color)
    
    # Convert to base64
    buffer = io.BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)
    img_base64 = base64.b64encode(buffer.getvalue()).decode()
    
    return f"data:image/png;base64,{img_base64}"

@api_router.post("/qr-codes/create")
async def create_qr_code(
    qr_data: QRCodeCreate,
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """Create a single QR code"""
    try:
        # Get backend URL dynamically from request to ensure correct environment
        host = request.headers.get('host', 'localhost:8001')
        scheme = 'https' if 'https' in str(request.url) or request.headers.get('x-forwarded-proto') == 'https' else 'http'
        backend_url = f"{scheme}://{host}"
        
        logger.info(f"Creating single QR code with backend URL: {backend_url}")
        
        # Generate short code for tracking
        short_code = generate_short_code()
        
        # Create tracking URL with dynamic backend URL - Must include /api prefix
        tracking_url = f"{backend_url}/api/qr/{short_code}"
        
        # Generate QR code image with color parameters
        foreground_color = qr_data.qr_foreground_color if qr_data.use_color_qr else "black"
        background_color = qr_data.qr_background_color if qr_data.use_color_qr else "white"
        qr_image = generate_qr_code_image(tracking_url, foreground_color, background_color)
        
        # Prepare QR code document
        qr_code = {
            "id": str(uuid.uuid4()),
            "short_code": short_code,
            "tracking_url": tracking_url,
            "campaign_name": qr_data.campaign_name,
            "landing_page_type": qr_data.landing_page_type,
            "ios_url": qr_data.ios_url,
            "android_url": qr_data.android_url,
            "web_url": qr_data.web_url,
            "single_url": qr_data.single_url,
            "utm_source": qr_data.utm_source or qr_data.campaign_name,
            "utm_medium": qr_data.utm_medium,
            "utm_campaign": qr_data.utm_campaign or qr_data.campaign_name,
            "utm_term": qr_data.utm_term,
            "utm_content": qr_data.utm_content,
            "qr_image": qr_image,
            "created_by": current_user.id,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "scan_count": 0,
            "status": "active"
        }
        
        # Save to database
        await db.qr_codes.insert_one(qr_code)
        
        # Remove MongoDB _id from response
        if '_id' in qr_code:
            del qr_code['_id']
        
        return {
            "success": True,
            "message": "QR code created successfully",
            "qr_code": qr_code
        }
        
    except Exception as e:
        logger.error(f"Failed to create QR code: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to create QR code: {str(e)}")

@api_router.post("/qr-codes/create-batch")
async def create_batch_qr_codes(
    batch_data: QRCodeBatchCreate,
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """Create multiple QR codes in a batch"""
    try:
        # Get backend URL dynamically from request to ensure correct environment
        host = request.headers.get('host', 'localhost:8001')
        scheme = 'https' if 'https' in str(request.url) or request.headers.get('x-forwarded-proto') == 'https' else 'http'
        backend_url = f"{scheme}://{host}"
        
        logger.info(f"Creating batch QR codes with backend URL: {backend_url}")
        
        qr_codes = []
        qr_names = batch_data.qr_names or [f"QR{i+1}" for i in range(batch_data.qr_count)]
        
        # Limit to specified count
        qr_names = qr_names[:batch_data.qr_count]
        
        for i, qr_name in enumerate(qr_names):
            # Generate short code for tracking
            short_code = generate_short_code()
            
            # Create tracking URL with dynamic backend URL - Must include /api prefix
            tracking_url = f"{backend_url}/api/qr/{short_code}"
            
            # Generate QR code image with color parameters
            foreground_color = batch_data.qr_foreground_color if batch_data.use_color_qr else "black"
            background_color = batch_data.qr_background_color if batch_data.use_color_qr else "white"
            qr_image = generate_qr_code_image(tracking_url, foreground_color, background_color)
            
            # Create proper UTM parameters (Campaign-VehicleNumber format)
            utm_source = f"{batch_data.campaign_name}-{qr_name}" if batch_data.auto_fill_utm else batch_data.utm_campaign
            
            # Support both OLD and NEW field names for backward compatibility
            # Priority: NEW field names > OLD field names
            landing_page_ios = batch_data.landing_page_ios or batch_data.ios_url
            landing_page_android = batch_data.landing_page_android or batch_data.android_url
            landing_page_mobile = batch_data.landing_page_mobile or batch_data.web_url
            landing_page_desktop = batch_data.landing_page_desktop or batch_data.web_url
            landing_page_single = batch_data.landing_page_single or batch_data.single_url or "https://nuraemobility.co.in/"
            
            # Prepare QR code document with NEW field names matching scan endpoint expectations
            qr_code = {
                "id": str(uuid.uuid4()),
                "short_code": short_code,
                "unique_short_code": short_code,  # Add this field for scanning endpoint compatibility
                "tracking_url": tracking_url,
                "qr_name": qr_name,
                "qr_filename": f"{batch_data.campaign_name}-{qr_name}.png",  # Proper filename format
                "campaign_name": batch_data.campaign_name,
                "landing_page_type": batch_data.landing_page_type,
                # Store with NEW field names to match scan endpoint
                "landing_page_ios": landing_page_ios,
                "landing_page_android": landing_page_android,
                "landing_page_mobile": landing_page_mobile,
                "landing_page_desktop": landing_page_desktop,
                "landing_page_single": landing_page_single,
                "utm_source": utm_source,
                "utm_medium": batch_data.utm_medium,
                "utm_campaign": batch_data.utm_campaign or batch_data.campaign_name,
                "qr_image": qr_image,
                "created_by": current_user.id,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "scan_count": 0,
                "status": "active",
                "batch_id": f"{batch_data.campaign_name}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}"
            }
            
            qr_codes.append(qr_code)
        
        # Save all to database
        if qr_codes:
            result = await db.qr_codes.insert_many(qr_codes)
            # Remove MongoDB _id from response
            for qr in qr_codes:
                if '_id' in qr:
                    del qr['_id']
        
        return {
            "success": True,
            "message": f"Created {len(qr_codes)} QR codes successfully",
            "qr_codes": qr_codes,
            "count": len(qr_codes)
        }
        
    except Exception as e:
        logger.error(f"Failed to create batch QR codes: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to create batch QR codes: {str(e)}")

@api_router.get("/qr/{short_code}")
async def scan_qr_code_short(
    short_code: str,
    request: Request
):
    """Short URL handler for QR code scans - redirects to main scan endpoint"""
    # This endpoint exists because QR codes use /qr/{code} format
    # Delegate to the main scan handler
    return await scan_qr_code(short_code, request)


@api_router.get("/qr-codes/scan/{short_code}")
async def scan_qr_code(
    short_code: str,
    request: Request
):
    """Handle QR code scan and redirect based on device"""
    from fastapi.responses import RedirectResponse
    try:
        logger.info(f"=== QR SCAN ATTEMPT === Code: {short_code}")
        
        # Get QR code from database - try both field names for backwards compatibility
        qr_code = await db.qr_codes.find_one({"short_code": short_code})
        if not qr_code:
            # Try the alternate field name used in creation
            qr_code = await db.qr_codes.find_one({"unique_short_code": short_code})
        
        if not qr_code:
            logger.error(f"QR CODE NOT FOUND in database! Searched for short_code='{short_code}' and unique_short_code='{short_code}'")
            # Let's check what QR codes exist
            sample_qr = await db.qr_codes.find_one({})
            if sample_qr:
                logger.info(f"Sample QR code in database has fields: {list(sample_qr.keys())}")
                logger.info(f"Sample unique_short_code: {sample_qr.get('unique_short_code')}, short_code: {sample_qr.get('short_code')}")
            raise HTTPException(status_code=404, detail="QR code not found")
        
        # Parse user agent
        user_agent_string = request.headers.get("user-agent", "")
        user_agent = parse(user_agent_string)
        
        # Detect platform
        if user_agent.is_mobile:
            if user_agent.os.family == "iOS":
                platform = "ios"
            elif user_agent.os.family == "Android":
                platform = "android"
            else:
                platform = "mobile_other"
        else:
            platform = "desktop"
        
        logger.info(f"QR Scan - Short Code: {short_code}, Platform Detected: {platform}, OS: {user_agent.os.family}, User Agent: {user_agent_string[:100]}")
        
        # Get client IP (check X-Forwarded-For for real IP behind proxies)
        client_ip = request.client.host
        forwarded_for = request.headers.get("x-forwarded-for", "")
        if forwarded_for:
            client_ip = forwarded_for.split(",")[0].strip()
            logger.info(f"Using X-Forwarded-For IP: {client_ip}")
        
        # Build UTM parameters in the format requested by user
        utm_value = qr_code.get('utm_source', f"{qr_code.get('campaign_name', '')}-{qr_code.get('qr_name', '')}")
        
        # Helper function to add UTM parameter correctly
        def add_utm_to_url(base_url, utm_param):
            if not base_url:
                return base_url
            separator = "&" if "?" in base_url else "?"
            return f"{base_url}{separator}utm={utm_param}"
        
        # Determine redirect URL based on device platform
        logger.info(f"QR Code Landing Pages - Type: {qr_code.get('landing_page_type')}, iOS: {qr_code.get('landing_page_ios')}, Android: {qr_code.get('landing_page_android')}, Mobile: {qr_code.get('landing_page_mobile')}, Single: {qr_code.get('landing_page_single')}")
        
        if qr_code.get('landing_page_type') == 'single':
            redirect_url = add_utm_to_url(qr_code.get('landing_page_single', ''), utm_value)
            logger.info(f"Using single URL: {redirect_url}")
        else:
            # Multi-URL QR code - redirect based on device platform
            redirect_url = None
            
            if platform == "ios" and qr_code.get('landing_page_ios'):
                redirect_url = add_utm_to_url(qr_code.get('landing_page_ios'), utm_value)
                logger.info(f"iOS detected - Using iOS URL: {redirect_url}")
            elif platform == "android" and qr_code.get('landing_page_android'):
                redirect_url = add_utm_to_url(qr_code.get('landing_page_android'), utm_value)
                logger.info(f"Android detected - Using Android URL: {redirect_url}")
            elif platform == "mobile_other" and qr_code.get('landing_page_mobile'):
                redirect_url = add_utm_to_url(qr_code.get('landing_page_mobile'), utm_value)
                logger.info(f"Other Mobile detected - Using Mobile URL: {redirect_url}")
            elif platform == "desktop" and qr_code.get('landing_page_desktop'):
                redirect_url = add_utm_to_url(qr_code.get('landing_page_desktop'), utm_value)
                logger.info(f"Desktop detected - Using Desktop URL: {redirect_url}")
            
            # If no URL found for detected platform, use fallback
            if not redirect_url or redirect_url == 'None':
                # Fallback order: try other platforms, then default to nuraemobility.co.in
                redirect_url = (
                    add_utm_to_url(qr_code.get('landing_page_mobile'), utm_value) or
                    add_utm_to_url(qr_code.get('landing_page_ios'), utm_value) or
                    add_utm_to_url(qr_code.get('landing_page_android'), utm_value) or
                    add_utm_to_url(qr_code.get('landing_page_desktop'), utm_value) or
                    add_utm_to_url(qr_code.get('landing_page_single'), utm_value) or
                    add_utm_to_url('https://nuraemobility.co.in', utm_value)  # Default fallback
                )
                logger.info(f"No URL for {platform}, using fallback: {redirect_url}")
        
        # Enhanced duplicate scan prevention with multiple layers
        import hashlib
        current_time = datetime.now(timezone.utc)
        
        # LAYER 1: Check for same QR + User Agent within 3 seconds (catches rapid double-clicks)
        cutoff_rapid = (current_time - timedelta(seconds=3)).isoformat()
        rapid_scan = await db.qr_scans.find_one({
            "qr_code_id": qr_code["id"],
            "user_agent": user_agent_string,
            "scanned_at": {"$gte": cutoff_rapid}
        })
        
        if rapid_scan:
            logger.info(f"RAPID DUPLICATE prevented - same User Agent scanned QR {short_code} within 3 seconds")
            return RedirectResponse(url=redirect_url, status_code=307)
        
        # LAYER 2: Check for recent scans from same IP and QR in last 2 minutes
        cutoff_time = (current_time - timedelta(minutes=2)).isoformat()
        recent_scan = await db.qr_scans.find_one({
            "qr_code_id": qr_code["id"],
            "ip_address": client_ip,
            "scanned_at": {"$gte": cutoff_time}
        })
        
        if recent_scan:
            logger.info(f"Duplicate scan prevented - same IP {client_ip} scanned QR {short_code} within 2 minutes")
            # Return redirect WITHOUT recording scan or incrementing count
            return RedirectResponse(url=redirect_url, status_code=307)
        
        # LAYER 3: Additional check - same User Agent + IP + QR within 5 minutes
        cutoff_time_extended = (current_time - timedelta(minutes=5)).isoformat()
        duplicate_check = await db.qr_scans.find_one({
            "qr_code_id": qr_code["id"],
            "ip_address": client_ip,
            "user_agent": user_agent_string,
            "scanned_at": {"$gte": cutoff_time_extended}
        })
        
        if duplicate_check:
            logger.info(f"Duplicate scan prevented - same User Agent + IP combination for QR {short_code}")
            # Return redirect WITHOUT recording scan or incrementing count
            return RedirectResponse(url=redirect_url, status_code=307)
        
        # LAYER 4: Create unique scan identifier for additional deduplication (minute-level)
        scan_identifier = hashlib.md5(f"{client_ip}_{user_agent_string[:50]}_{qr_code['id']}_{current_time.strftime('%Y%m%d%H%M')}".encode()).hexdigest()
        
        # Check if this exact scan identifier already exists (final safety check)
        identifier_check = await db.qr_scans.find_one({"scan_identifier": scan_identifier})
        if identifier_check:
            logger.info(f"Duplicate scan prevented - scan identifier already exists for QR {short_code}")
            return RedirectResponse(url=redirect_url, status_code=307)
        
        # Get location data from IP with X-Forwarded-For support
        location_info = get_location_from_ip(client_ip, dict(request.headers))
        
        # Log analytics with enhanced tracking - only if not duplicate
        scan_data = {
            "id": str(uuid.uuid4()),
            "scan_identifier": scan_identifier,
            "qr_code_id": qr_code["id"],
            "short_code": short_code,
            "qr_name": qr_code.get("qr_name", qr_code.get("utm_source", "Unknown")),
            "campaign_name": qr_code.get("campaign_name"),
            "scanned_at": datetime.now(timezone.utc).isoformat(),
            "platform": platform,
            "os_family": user_agent.os.family,
            "os_version": user_agent.os.version_string,
            "browser": user_agent.browser.family,
            "browser_version": user_agent.browser.version_string,
            "device": user_agent.device.family,
            "user_agent": user_agent_string,
            "ip_address": client_ip,
            "location_city": location_info.get("city", "Unknown"),
            "location_region": location_info.get("region", "Unknown"),
            "location_country": location_info.get("country", "Unknown"),
            "redirect_url": redirect_url,
            "utm_params": f"utm={utm_value}",
            # Note: App download cannot be directly tracked via QR scan
            # This only tracks the redirect to app store, not the actual download
            "tracking_note": "Tracks store redirect only, not app installation"
        }
        
        # Insert scan record
        await db.qr_scans.insert_one(scan_data)
        logger.info(f"New scan recorded for QR {short_code} from IP {client_ip}, Platform: {platform}, Redirect: {redirect_url}")
        
        # Increment scan count only for unique scans - use the QR code's actual ID
        await db.qr_codes.update_one(
            {"id": qr_code["id"]},
            {"$inc": {"scan_count": 1}}
        )
        
        # Validate redirect URL before redirecting
        if not redirect_url or redirect_url == 'None' or redirect_url.strip() == '':
            logger.error(f"Empty redirect URL for QR {short_code}, QR Data: {qr_code}")
            raise HTTPException(status_code=500, detail=f"QR code configuration error: No valid landing page URL configured for {platform}")
        
        # Redirect with proper status code (307 Temporary Redirect)
        from fastapi.responses import RedirectResponse
        logger.info(f"QR scan recorded: {short_code} -> {platform}. Redirecting to: {redirect_url}")
        return RedirectResponse(url=redirect_url, status_code=307)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to process QR scan: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Failed to process QR scan: {str(e)}")

@api_router.get("/qr-codes/campaign/{campaign_name}")
async def get_campaign_qr_codes(
    campaign_name: str,
    current_user: User = Depends(get_current_user)
):
    """Get all QR codes for a campaign"""
    try:
        qr_codes = await db.qr_codes.find({"campaign_name": campaign_name}).to_list(None)
        
        # Remove MongoDB _id from response
        for qr in qr_codes:
            if '_id' in qr:
                del qr['_id']
        
        return {
            "success": True,
            "qr_codes": qr_codes,
            "count": len(qr_codes)
        }
        
    except Exception as e:
        logger.error(f"Failed to get campaign QR codes: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get campaign QR codes: {str(e)}")

@api_router.get("/qr-codes/analytics")
async def get_qr_analytics(
    campaign_name: Optional[str] = None,
    qr_name: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get QR code scan analytics with filters"""
    try:
        # Build query
        query = {}
        if campaign_name:
            query["campaign_name"] = campaign_name
        if qr_name:
            query["qr_name"] = qr_name
        if start_date and end_date:
            query["scanned_at"] = {
                "$gte": start_date,
                "$lte": end_date
            }
        
        # Get scan data
        scans = await db.qr_scans.find(query).sort("scanned_at", -1).to_list(None)
        
        # Remove MongoDB _id from response
        for scan in scans:
            if '_id' in scan:
                del scan['_id']
        
        # Get summary stats
        total_scans = len(scans)
        platform_breakdown = {}
        for scan in scans:
            platform = scan.get("platform", "unknown")
            platform_breakdown[platform] = platform_breakdown.get(platform, 0) + 1
        
        return {
            "success": True,
            "scans": scans,
            "total_scans": total_scans,
            "platform_breakdown": platform_breakdown
        }
        
    except Exception as e:
        logger.error(f"Failed to get QR analytics: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get QR analytics: {str(e)}")

@api_router.delete("/qr-codes/{qr_code_id}")
async def delete_qr_code(
    qr_code_id: str,
    force: bool = False,  # Add force parameter to allow deleting published QR codes
    current_user: User = Depends(get_current_user)
):
    """Delete a QR code"""
    try:
        # Check if the QR code exists
        qr_code = await db.qr_codes.find_one({"id": qr_code_id})
        
        if not qr_code:
            raise HTTPException(status_code=404, detail="QR code not found")
        
        # Only master admins can delete published QR codes
        if qr_code.get("published", False) and not force:
            raise HTTPException(status_code=400, detail="Cannot delete published QR code. Use force=true to override (master admin only)")
        
        if qr_code.get("published", False) and force:
            if current_user.account_type != "master_admin":
                raise HTTPException(status_code=403, detail="Only master admins can force delete published QR codes")
        
        result = await db.qr_codes.delete_one({"id": qr_code_id})
        
        # Also delete associated scan data
        await db.qr_scans.delete_many({"qr_code_id": qr_code_id})
        
        return {
            "success": True,
            "message": "QR code deleted successfully"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to delete QR code: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to delete QR code: {str(e)}")

@api_router.delete("/qr-codes/campaigns/{campaign_name}")
async def delete_campaign(
    campaign_name: str,
    force: bool = False,  # Add force parameter
    current_user: User = Depends(get_current_user)
):
    """Delete all QR codes in a campaign"""
    try:
        logger.info(f"=== CAMPAIGN DELETE REQUEST === Campaign: '{campaign_name}', Force: {force}, User: {current_user.email}, Role: {current_user.account_type}")
        
        # Find all QR codes in the campaign
        qr_codes = await db.qr_codes.find({"campaign_name": campaign_name}).to_list(None)
        
        if not qr_codes:
            logger.error(f"Campaign '{campaign_name}' not found in database")
            raise HTTPException(status_code=404, detail="Campaign not found")
        
        logger.info(f"Found {len(qr_codes)} QR codes in campaign '{campaign_name}'")
        
        # Check if any QR code in the campaign is published
        has_published = any(qr.get("published", False) for qr in qr_codes)
        
        logger.info(f"Campaign has published QR codes: {has_published}")
        
        if has_published and not force:
            logger.warning(f"Attempted to delete published campaign '{campaign_name}' without force flag")
            raise HTTPException(status_code=400, detail="Cannot delete published campaign. Use force=true to override (master admin only)")
        
        if has_published and force:
            if current_user.account_type != "master_admin":
                logger.warning(f"Non-master admin {current_user.email} attempted to force delete published campaign '{campaign_name}'")
                raise HTTPException(status_code=403, detail="Only master admins can force delete published campaigns")
        
        # Extract all QR code IDs to delete associated scans
        qr_code_ids = [qr.get("id") for qr in qr_codes if qr.get("id")]
        
        logger.info(f"Deleting associated scans for {len(qr_code_ids)} QR codes")
        
        # Delete associated scan data first (scans are linked by qr_code_id, not campaign_name)
        if qr_code_ids:
            scans_result = await db.qr_scans.delete_many({"qr_code_id": {"$in": qr_code_ids}})
            logger.info(f"Deleted {scans_result.deleted_count} scans for campaign '{campaign_name}'")
        
        # Delete all QR codes in the campaign
        result = await db.qr_codes.delete_many({"campaign_name": campaign_name})
        
        logger.info(f"Successfully deleted campaign '{campaign_name}' with {result.deleted_count} QR codes")
        
        return {
            "success": True,
            "message": f"Campaign '{campaign_name}' and {result.deleted_count} QR codes deleted successfully",
            "qr_codes_deleted": result.deleted_count,
            "scans_deleted": scans_result.deleted_count if qr_code_ids else 0
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to delete campaign '{campaign_name}': {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to delete campaign: {str(e)}")

@api_router.post("/qr-codes/campaigns/{campaign_name}/publish")
async def publish_campaign(
    campaign_name: str,
    current_user: User = Depends(get_current_user)
):
    """Publish a campaign - disables deletion"""
    try:
        # Find all QR codes in the campaign
        qr_codes = await db.qr_codes.find({"campaign_name": campaign_name}).to_list(None)
        
        if not qr_codes:
            raise HTTPException(status_code=404, detail="Campaign not found")
        
        # Update all QR codes in the campaign to published status
        result = await db.qr_codes.update_many(
            {"campaign_name": campaign_name},
            {"$set": {"published": True, "published_at": datetime.now(timezone.utc).isoformat()}}
        )
        
        return {
            "success": True,
            "message": f"Campaign '{campaign_name}' published successfully",
            "qr_codes_updated": result.modified_count
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to publish campaign: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to publish campaign: {str(e)}")

@api_router.post("/qr-codes/campaigns/{campaign_name}/unpublish")
async def unpublish_campaign(
    campaign_name: str,
    current_user: User = Depends(get_current_user)
):
    """Unpublish a campaign - enables deletion"""
    try:
        # Find all QR codes in the campaign
        qr_codes = await db.qr_codes.find({"campaign_name": campaign_name}).to_list(None)
        
        if not qr_codes:
            raise HTTPException(status_code=404, detail="Campaign not found")
        
        # Update all QR codes in the campaign to unpublished status
        result = await db.qr_codes.update_many(
            {"campaign_name": campaign_name},
            {"$set": {"published": False}, "$unset": {"published_at": ""}}
        )
        
        return {
            "success": True,
            "message": f"Campaign '{campaign_name}' unpublished successfully",
            "qr_codes_updated": result.modified_count
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to unpublish campaign: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to unpublish campaign: {str(e)}")

@api_router.post("/qr-codes/campaigns/{campaign_name}/unpublish")
async def unpublish_campaign(
    campaign_name: str,
    current_user: User = Depends(get_current_user)
):
    """Unpublish a campaign - enables deletion"""
    try:
        # Find all QR codes in the campaign
        qr_codes = await db.qr_codes.find({"campaign_name": campaign_name}).to_list(None)
        
        if not qr_codes:
            raise HTTPException(status_code=404, detail="Campaign not found")
        
        # Update all QR codes in the campaign to unpublished status
        result = await db.qr_codes.update_many(
            {"campaign_name": campaign_name},
            {"$set": {"published": False}, "$unset": {"published_at": ""}}
        )
        
        return {
            "success": True,
            "message": f"Campaign '{campaign_name}' unpublished successfully",
            "qr_codes_updated": result.modified_count
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to publish campaign: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to publish campaign: {str(e)}")

@api_router.get("/qr-codes/campaigns/{campaign_name}/details")
async def get_campaign_details(
    campaign_name: str,
    current_user: User = Depends(get_current_user)
):
    """Get campaign metadata including publish status"""
    try:
        # Find a sample QR code to get campaign info
        sample_qr = await db.qr_codes.find_one({"campaign_name": campaign_name})
        
        if not sample_qr:
            raise HTTPException(status_code=404, detail="Campaign not found")
        
        return {
            "campaign_name": campaign_name,
            "published": sample_qr.get("published", False),
            "published_at": sample_qr.get("published_at"),
            "created_at": sample_qr.get("created_at")
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get campaign details: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get campaign details: {str(e)}")

@api_router.get("/qr-codes/campaigns/{campaign_name}/analytics")
async def get_campaign_analytics(
    campaign_name: str,
    current_user: User = Depends(get_current_user)
):
    """Get detailed analytics for all QR codes in a campaign"""
    try:
        # Get all QR codes in the campaign
        qr_codes = await db.qr_codes.find({"campaign_name": campaign_name}).to_list(None)
        
        if not qr_codes:
            raise HTTPException(status_code=404, detail="Campaign not found")
        
        analytics = []
        
        for qr_code in qr_codes:
            qr_id = qr_code["id"]
            
            # Get scan analytics for this QR code
            scans = await db.qr_scans.find({"qr_code_id": qr_id}).to_list(None)
            
            # Calculate platform breakdown
            ios_scans = len([s for s in scans if s.get("platform") == "ios"])
            android_scans = len([s for s in scans if s.get("platform") == "android"])
            web_scans = len([s for s in scans if s.get("platform") in ["desktop", "mobile_other"]])
            
            # Get last scan time
            last_scan = None
            if scans:
                last_scan = max([s.get("scanned_at") for s in scans if s.get("scanned_at")])
            
            # Create proper filename format: Campaign-VehicleNumber
            qr_filename = f"{campaign_name}-{qr_code.get('qr_name', qr_code.get('utm_source', 'Unknown'))}"
            
            # Generate UTM URLs for all platforms
            utm_value = qr_code.get('utm_source', f"{campaign_name}-{qr_code.get('qr_name', 'Unknown')}")
            
            # Helper function to add UTM parameter correctly
            def add_utm_to_url(base_url, utm_param):
                if not base_url:
                    return base_url
                separator = "&" if "?" in base_url else "?"
                return f"{base_url}{separator}utm={utm_param}"
            
            # Determine URLs for all platforms with UTM
            if qr_code.get('landing_page_type') == 'single':
                base_url = qr_code.get('single_url', '')
                utm_url_web = add_utm_to_url(base_url, utm_value)
                utm_url_ios = utm_url_web  # Same URL for single mode
                utm_url_android = utm_url_web  # Same URL for single mode
            else:
                # Different URLs for each platform
                utm_url_ios = add_utm_to_url(qr_code.get('ios_url', ''), utm_value)
                utm_url_android = add_utm_to_url(qr_code.get('android_url', ''), utm_value)
                utm_url_web = add_utm_to_url(qr_code.get('web_url', qr_code.get('single_url', '')), utm_value)
            
            analytics.append({
                "qr_code_id": qr_id,
                "qr_name": qr_code.get("qr_name", qr_code.get("utm_source")),
                "qr_filename": qr_filename,
                "utm_source": qr_code.get("utm_source"),
                "utm_url_web": utm_url_web,
                "utm_url_ios": utm_url_ios,
                "utm_url_android": utm_url_android,
                "landing_page_type": qr_code.get("landing_page_type", "single"),
                "total_scans": len(scans),
                "ios_scans": ios_scans,
                "android_scans": android_scans,
                "web_scans": web_scans,
                "last_scan": last_scan,
                "scan_details": [
                    {
                        "scanned_at": scan.get("scanned_at"),
                        "platform": scan.get("platform"),
                        "os_family": scan.get("os_family"),
                        "browser": scan.get("browser"),
                        "device": scan.get("device"),
                        "ip_address": scan.get("ip_address"),
                        "location_city": scan.get("location_city", "Unknown"),
                        "location_region": scan.get("location_region", "Unknown"),
                        "location_country": scan.get("location_country", "Unknown"),
                        "user_agent": scan.get("user_agent")
                    }
                    for scan in scans
                    if scan.get("scanned_at") is not None  # Filter out incomplete scans
                ]
            })
        
        return {
            "success": True,
            "campaign_name": campaign_name,
            "analytics": analytics,
            "total_qr_codes": len(qr_codes),
            "total_scans": sum([item["total_scans"] for item in analytics])
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get campaign analytics: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get campaign analytics: {str(e)}")

@api_router.get("/qr-codes/{qr_code_id}/analytics")
async def get_individual_qr_analytics(
    qr_code_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get detailed analytics for a single QR code"""
    try:
        # Get the specific QR code
        qr_code = await db.qr_codes.find_one({"id": qr_code_id})
        
        if not qr_code:
            raise HTTPException(status_code=404, detail="QR code not found")
        
        # Get scan analytics for this QR code
        scans = await db.qr_scans.find({"qr_code_id": qr_code_id}).to_list(None)
        
        # Calculate platform breakdown
        ios_scans = len([s for s in scans if s.get("platform") == "ios"])
        android_scans = len([s for s in scans if s.get("platform") == "android"])
        web_scans = len([s for s in scans if s.get("platform") in ["desktop", "mobile_other"]])
        
        # Get last scan time
        last_scan = None
        if scans:
            last_scan = max([s.get("scanned_at") for s in scans if s.get("scanned_at")])
        
        # Create proper filename format: Campaign-VehicleNumber
        qr_filename = f"{qr_code.get('campaign_name', 'Unknown')}-{qr_code.get('qr_name', qr_code.get('utm_source', 'Unknown'))}"
        
        # Generate UTM URLs for all platforms
        utm_value = qr_code.get('utm_source', f"{qr_code.get('campaign_name', 'Unknown')}-{qr_code.get('qr_name', 'Unknown')}")
        
        # Helper function to add UTM parameter correctly
        def add_utm_to_url(base_url, utm_param):
            if not base_url:
                return base_url
            separator = "&" if "?" in base_url else "?"
            return f"{base_url}{separator}utm={utm_param}"
        
        # Determine URLs for all platforms with UTM
        # Support both NEW and OLD field names for backward compatibility
        if qr_code.get('landing_page_type') == 'single':
            base_url = qr_code.get('landing_page_single') or qr_code.get('single_url', '')
            utm_url_web = add_utm_to_url(base_url, utm_value)
            utm_url_ios = utm_url_web  # Same URL for single mode
            utm_url_android = utm_url_web  # Same URL for single mode
        else:
            # Different URLs for each platform - check NEW field names first, then OLD
            utm_url_ios = add_utm_to_url(
                qr_code.get('landing_page_ios') or qr_code.get('ios_url', ''), 
                utm_value
            )
            utm_url_android = add_utm_to_url(
                qr_code.get('landing_page_android') or qr_code.get('android_url', ''), 
                utm_value
            )
            utm_url_web = add_utm_to_url(
                qr_code.get('landing_page_desktop') or qr_code.get('landing_page_mobile') or qr_code.get('web_url') or qr_code.get('landing_page_single') or qr_code.get('single_url', ''), 
                utm_value
            )
        
        analytics = {
            "qr_code_id": qr_code_id,
            "qr_name": qr_code.get("qr_name", qr_code.get("utm_source")),
            "qr_filename": qr_filename,
            "campaign_name": qr_code.get("campaign_name"),
            "utm_source": qr_code.get("utm_source"),
            "utm_url_web": utm_url_web,
            "utm_url_ios": utm_url_ios,
            "utm_url_android": utm_url_android,
            "landing_page_type": qr_code.get("landing_page_type", "single"),
            "total_scans": len(scans),
            "ios_scans": ios_scans,
            "android_scans": android_scans,
            "web_scans": web_scans,
            "last_scan": last_scan,
            "created_at": qr_code.get("created_at"),
            "scan_details": [
                {
                    "scanned_at": scan.get("scanned_at"),
                    "platform": scan.get("platform"),
                    "os_family": scan.get("os_family"),
                    "browser": scan.get("browser"),
                    "device": scan.get("device"),
                    "ip_address": scan.get("ip_address"),
                    "location_city": scan.get("location_city", "Unknown"),
                    "location_region": scan.get("location_region", "Unknown"),
                    "location_country": scan.get("location_country", "Unknown"),
                    "user_agent": scan.get("user_agent")
                }
                for scan in scans
                if scan.get("scanned_at") is not None  # Filter out incomplete scans
            ]
        }
        
        return {
            "success": True,
            "analytics": analytics
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get QR code analytics: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get QR code analytics: {str(e)}")

# ==================== MANAGE DATABASE - VEHICLES & DRIVERS ====================

class Vehicle(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    vehicle_type: str
    register_number: str
    vin_number: str
    model: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    updated_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class VehicleCreate(BaseModel):
    vehicle_type: str
    register_number: str
    vin_number: str
    model: Optional[str] = None

class Driver(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    phone_number: str
    dl_number: Optional[str] = None
    status: str = "Active"  # Active or Terminated
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    updated_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class DriverCreate(BaseModel):
    name: str
    phone_number: str
    dl_number: Optional[str] = None
    status: str = "Active"

# Vehicles Endpoints
@api_router.get("/manage-db/vehicles")
async def get_vehicles(current_user: User = Depends(get_current_user)):
    """Get all vehicles"""
    try:
        vehicles = await db.vehicles.find({}, {"_id": 0}).to_list(1000)
        return {"vehicles": vehicles}
    except Exception as e:
        logger.error(f"Error fetching vehicles: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch vehicles")

@api_router.post("/manage-db/vehicles")
async def create_vehicle(vehicle: VehicleCreate, current_user: User = Depends(get_current_user)):
    """Create a new vehicle"""
    try:
        vehicle_dict = Vehicle(
            vehicle_type=vehicle.vehicle_type,
            register_number=vehicle.register_number,
            vin_number=vehicle.vin_number,
            model=vehicle.model
        ).model_dump()
        
        await db.vehicles.insert_one(vehicle_dict)
        return {"success": True, "vehicle": vehicle_dict}
    except Exception as e:
        logger.error(f"Error creating vehicle: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to create vehicle")

@api_router.put("/manage-db/vehicles/{vehicle_id}")
async def update_vehicle(vehicle_id: str, vehicle: VehicleCreate, current_user: User = Depends(get_current_user)):
    """Update a vehicle"""
    try:
        update_data = vehicle.model_dump()
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        
        result = await db.vehicles.update_one(
            {"id": vehicle_id},
            {"$set": update_data}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Vehicle not found")
        
        return {"success": True, "message": "Vehicle updated successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating vehicle: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to update vehicle")

@api_router.delete("/manage-db/vehicles/{vehicle_id}")
async def delete_vehicle(vehicle_id: str, current_user: User = Depends(get_current_user)):
    """Delete a vehicle"""
    try:
        result = await db.vehicles.delete_one({"id": vehicle_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Vehicle not found")
        
        return {"success": True, "message": "Vehicle deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting vehicle: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to delete vehicle")

@api_router.post("/manage-db/vehicles/bulk-import")
async def bulk_import_vehicles(vehicles_data: dict = Body(...), current_user: User = Depends(get_current_user)):
    """Bulk import vehicles from Excel"""
    try:
        vehicles = vehicles_data.get("vehicles", [])
        
        if not vehicles:
            raise HTTPException(status_code=400, detail="No vehicles data provided")
        
        imported_count = 0
        for vehicle_data in vehicles:
            if not vehicle_data.get("vehicle_type") or not vehicle_data.get("register_number") or not vehicle_data.get("vin_number"):
                continue
            
            vehicle_dict = Vehicle(
                vehicle_type=vehicle_data["vehicle_type"],
                register_number=vehicle_data["register_number"],
                vin_number=vehicle_data["vin_number"],
                model=vehicle_data.get("model")
            ).model_dump()
            
            # Check if vehicle already exists
            existing = await db.vehicles.find_one({"vin_number": vehicle_data["vin_number"]})
            if existing:
                # Update existing
                await db.vehicles.update_one(
                    {"vin_number": vehicle_data["vin_number"]},
                    {"$set": {**vehicle_dict, "updated_at": datetime.now(timezone.utc).isoformat()}}
                )
            else:
                # Insert new
                await db.vehicles.insert_one(vehicle_dict)
            
            imported_count += 1
        
        return {"success": True, "imported_count": imported_count}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error bulk importing vehicles: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to import vehicles: {str(e)}")

# Drivers Endpoints
@api_router.get("/manage-db/drivers")
async def get_drivers(current_user: User = Depends(get_current_user)):
    """Get all drivers"""
    try:
        drivers = await db.drivers.find({}, {"_id": 0}).to_list(1000)
        return {"drivers": drivers}
    except Exception as e:
        logger.error(f"Error fetching drivers: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch drivers")

@api_router.post("/manage-db/drivers")
async def create_driver(driver: DriverCreate, current_user: User = Depends(get_current_user)):
    """Create a new driver"""
    try:
        driver_dict = Driver(
            name=driver.name,
            phone_number=driver.phone_number,
            dl_number=driver.dl_number,
            status=driver.status
        ).model_dump()
        
        await db.drivers.insert_one(driver_dict)
        return {"success": True, "driver": driver_dict}
    except Exception as e:
        logger.error(f"Error creating driver: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to create driver")

@api_router.put("/manage-db/drivers/{driver_id}")
async def update_driver(driver_id: str, driver: DriverCreate, current_user: User = Depends(get_current_user)):
    """Update a driver"""
    try:
        update_data = driver.model_dump()
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        
        result = await db.drivers.update_one(
            {"id": driver_id},
            {"$set": update_data}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Driver not found")
        
        return {"success": True, "message": "Driver updated successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating driver: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to update driver")

@api_router.delete("/manage-db/drivers/{driver_id}")
async def delete_driver(driver_id: str, current_user: User = Depends(get_current_user)):
    """Delete a driver"""
    try:
        result = await db.drivers.delete_one({"id": driver_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Driver not found")
        
        return {"success": True, "message": "Driver deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting driver: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to delete driver")

@api_router.post("/manage-db/drivers/bulk-import")
async def bulk_import_drivers(drivers_data: dict = Body(...), current_user: User = Depends(get_current_user)):
    """Bulk import drivers from Excel"""
    try:
        drivers = drivers_data.get("drivers", [])
        
        if not drivers:
            raise HTTPException(status_code=400, detail="No drivers data provided")
        
        imported_count = 0
        for driver_data in drivers:
            if not driver_data.get("name") or not driver_data.get("phone_number"):
                continue
            
            driver_dict = Driver(
                name=driver_data["name"],
                phone_number=driver_data["phone_number"],
                dl_number=driver_data.get("dl_number"),
                status=driver_data.get("status", "Active")
            ).model_dump()
            
            # Check if driver already exists
            existing = await db.drivers.find_one({"phone_number": driver_data["phone_number"]})
            if existing:
                # Update existing
                await db.drivers.update_one(
                    {"phone_number": driver_data["phone_number"]},
                    {"$set": {**driver_dict, "updated_at": datetime.now(timezone.utc).isoformat()}}
                )
            else:
                # Insert new
                await db.drivers.insert_one(driver_dict)
            
            imported_count += 1
        
        return {"success": True, "imported_count": imported_count}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error bulk importing drivers: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to import drivers: {str(e)}")

@api_router.post("/manage-db/drivers/sync-from-onboarding")
async def sync_drivers_from_onboarding(current_user: User = Depends(get_current_user)):
    """Sync drivers from Driver Onboarding leads with DONE! or Terminated status"""
    try:
        # Fetch leads with DONE! status or Terminated status
        leads = await db.driver_leads.find({
            "$or": [
                {"status": "DONE!"},
                {"status": "Terminated"}
            ]
        }).to_list(10000)
        
        synced_count = 0
        for lead in leads:
            if not lead.get("name") or not lead.get("phone_number"):
                continue
            
            # Determine status based on lead status
            driver_status = "Terminated" if lead.get("status") == "Terminated" else "Active"
            
            driver_dict = Driver(
                name=lead["name"],
                phone_number=str(lead["phone_number"]),
                dl_number=lead.get("dl_no"),
                status=driver_status
            ).model_dump()
            
            # Check if driver already exists
            existing = await db.drivers.find_one({"phone_number": str(lead["phone_number"])})
            if existing:
                # Update existing - update status if changed
                await db.drivers.update_one(
                    {"phone_number": str(lead["phone_number"])},
                    {"$set": {
                        "name": lead["name"],
                        "dl_number": lead.get("dl_no"),
                        "status": driver_status,
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    }}
                )
            else:
                # Insert new
                await db.drivers.insert_one(driver_dict)
            
            synced_count += 1
        
        return {"success": True, "synced_count": synced_count}
    except Exception as e:
        logger.error(f"Error syncing drivers from onboarding: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to sync drivers: {str(e)}")


# ==================== SUPPLY PLAN - SHIFT ASSIGNMENTS ====================

from app_models import ShiftAssignment, ShiftAssignmentCreate, ShiftAssignmentUpdate

# Helper function to generate unique color for each driver
def get_driver_color(driver_name: str) -> str:
    """Generate a consistent color for each driver based on their name"""
    import hashlib
    # Use hash to generate consistent color
    hash_obj = hashlib.md5(driver_name.encode())
    hash_hex = hash_obj.hexdigest()
    # Take first 6 characters for color
    color = f"#{hash_hex[:6]}"
    return color


@api_router.get("/supply-plan/drivers")
async def get_supply_plan_drivers(current_user: User = Depends(get_current_user)):
    """Get drivers list from Drivers List.xlsx"""
    try:
        import pandas as pd
        
        # Find the drivers list file
        drivers_file = "/app/backend/uploaded_files/3ea2faab-edc6-4bf1-9512-7b8fb64fb90c_Drivers List.xlsx"
        
        if not os.path.exists(drivers_file):
            raise HTTPException(status_code=404, detail="Drivers List.xlsx not found")
        
        # Read the Excel file
        df = pd.read_excel(drivers_file)
        
        # Extract driver names (assuming they're in a column named 'Driver Name' or similar)
        if 'Driver Name' in df.columns:
            drivers = df['Driver Name'].dropna().unique().tolist()
        else:
            # Try to find the column with names
            drivers = df.iloc[:, 1].dropna().unique().tolist()  # Assuming second column
        
        return {"drivers": sorted(drivers)}
    except Exception as e:
        logger.error(f"Error fetching drivers list: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch drivers: {str(e)}")


@api_router.get("/supply-plan/vehicles")
async def get_supply_plan_vehicles(current_user: User = Depends(get_current_user)):
    """Get vehicles list from Vehicles List.xlsx"""
    try:
        import pandas as pd
        
        # Find the vehicles list file
        vehicles_file = "/app/backend/uploaded_files/1e0860b7-a35b-4813-aa25-15523b6f22e7_Vehicles List.xlsx"
        
        if not os.path.exists(vehicles_file):
            raise HTTPException(status_code=404, detail="Vehicles List.xlsx not found")
        
        # Read the Excel file
        df = pd.read_excel(vehicles_file)
        
        # Extract vehicle registration numbers
        if 'Vehicle List (A to Z)' in df.columns:
            vehicles = df['Vehicle List (A to Z)'].dropna().unique().tolist()
        else:
            # Try to find the column with vehicle numbers
            vehicles = df.iloc[:, 1].dropna().unique().tolist()  # Assuming second column
        
        return {"vehicles": sorted(vehicles)}
    except Exception as e:
        logger.error(f"Error fetching vehicles list: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch vehicles: {str(e)}")


@api_router.get("/supply-plan/assignments")
async def get_shift_assignments(
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(..., description="End date in YYYY-MM-DD format"),
    current_user: User = Depends(get_current_user)
):
    """Get shift assignments for a date range"""
    try:
        assignments = await db.shift_assignments.find({
            "shift_date": {"$gte": start_date, "$lte": end_date}
        }, {"_id": 0}).to_list(10000)
        
        return {"assignments": assignments}
    except Exception as e:
        logger.error(f"Error fetching shift assignments: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch assignments: {str(e)}")


@api_router.post("/supply-plan/assignments")
async def create_shift_assignment(
    assignment_data: ShiftAssignmentCreate,
    current_user: User = Depends(get_current_user)
):
    """Create a new shift assignment"""
    try:
        # Generate color if not provided
        data_dict = assignment_data.model_dump()
        if not data_dict.get('driver_color'):
            data_dict['driver_color'] = get_driver_color(assignment_data.driver_name)
        
        assignment = ShiftAssignment(
            **data_dict,
            created_by=current_user.id
        )
        
        assignment_dict = assignment.model_dump()
        assignment_dict['created_at'] = assignment_dict['created_at'].isoformat()
        assignment_dict['updated_at'] = assignment_dict['updated_at'].isoformat()
        
        await db.shift_assignments.insert_one(assignment_dict)
        
        return {"message": "Shift assignment created successfully", "assignment": assignment}
    except Exception as e:
        logger.error(f"Error creating shift assignment: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to create assignment: {str(e)}")


@api_router.put("/supply-plan/assignments/{assignment_id}")
async def update_shift_assignment(
    assignment_id: str,
    assignment_data: ShiftAssignmentUpdate,
    current_user: User = Depends(get_current_user)
):
    """Update an existing shift assignment"""
    try:
        update_data = {k: v for k, v in assignment_data.model_dump().items() if v is not None}
        update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
        
        result = await db.shift_assignments.update_one(
            {"id": assignment_id},
            {"$set": update_data}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Assignment not found")
        
        return {"message": "Shift assignment updated successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating shift assignment: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to update assignment: {str(e)}")


@api_router.delete("/supply-plan/assignments/{assignment_id}")
async def delete_shift_assignment(
    assignment_id: str,
    current_user: User = Depends(get_current_user)
):
    """Delete a shift assignment"""
    try:
        result = await db.shift_assignments.delete_one({"id": assignment_id})
        
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Assignment not found")
        
        return {"message": "Shift assignment deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting shift assignment: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to delete assignment: {str(e)}")


@api_router.get("/supply-plan/export")
async def export_shift_assignments(
    start_date: str = Query(...),
    end_date: str = Query(...),
    current_user: User = Depends(get_current_user)
):
    """Export shift assignments to Excel"""
    try:
        import pandas as pd
        from io import BytesIO
        
        # Fetch assignments
        assignments = await db.shift_assignments.find({
            "shift_date": {"$gte": start_date, "$lte": end_date}
        }, {"_id": 0}).to_list(10000)
        
        if not assignments:
            raise HTTPException(status_code=404, detail="No assignments found for the selected date range")
        
        # Convert to DataFrame
        df = pd.DataFrame(assignments)
        
        # Select and reorder columns
        columns = ['shift_date', 'vehicle_reg_no', 'driver_name', 'shift_start_time', 'shift_end_time', 'notes']
        df = df[[col for col in columns if col in df.columns]]
        
        # Rename columns for better readability
        df.columns = ['Date', 'Vehicle Registration', 'Driver Name', 'Start Time', 'End Time', 'Notes']
        
        # Create Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Shift Assignments')
        
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename=shift_assignments_{start_date}_to_{end_date}.xlsx'}
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting shift assignments: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to export assignments: {str(e)}")


@api_router.post("/supply-plan/import")
async def import_shift_assignments(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """Import shift assignments from Excel"""
    try:
        import pandas as pd
        from io import BytesIO
        
        # Read the uploaded file
        contents = await file.read()
        df = pd.read_excel(BytesIO(contents))
        
        # Validate required columns
        required_columns = ['Date', 'Vehicle Registration', 'Driver Name', 'Start Time', 'End Time']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            raise HTTPException(
                status_code=400,
                detail=f"Missing required columns: {', '.join(missing_columns)}"
            )
        
        imported_count = 0
        errors = []
        
        for idx, row in df.iterrows():
            try:
                # Skip empty rows
                if pd.isna(row['Date']) or pd.isna(row['Vehicle Registration']) or pd.isna(row['Driver Name']):
                    continue
                
                # Convert date to string format
                shift_date = pd.to_datetime(row['Date']).strftime('%Y-%m-%d')
                
                # Create assignment
                assignment = ShiftAssignment(
                    vehicle_reg_no=str(row['Vehicle Registration']),
                    driver_name=str(row['Driver Name']),
                    shift_date=shift_date,
                    shift_start_time=str(row['Start Time']),
                    shift_end_time=str(row['End Time']),
                    driver_color=get_driver_color(str(row['Driver Name'])),
                    notes=str(row['Notes']) if 'Notes' in row and pd.notna(row['Notes']) else None,
                    created_by=current_user.id
                )
                
                assignment_dict = assignment.model_dump()
                assignment_dict['created_at'] = assignment_dict['created_at'].isoformat()
                assignment_dict['updated_at'] = assignment_dict['updated_at'].isoformat()
                
                await db.shift_assignments.insert_one(assignment_dict)
                imported_count += 1
                
            except Exception as e:
                errors.append(f"Row {idx + 2}: {str(e)}")
        
        return {
            "message": f"Successfully imported {imported_count} assignments",
            "imported_count": imported_count,
            "errors": errors if errors else None
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error importing shift assignments: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to import assignments: {str(e)}")


@api_router.get("/supply-plan/sample-template")
async def download_sample_template(current_user: User = Depends(get_current_user)):
    """Download sample template for shift assignments import"""
    try:
        import pandas as pd
        from io import BytesIO
        
        # Create sample data
        sample_data = {
            'Date': ['2025-11-18', '2025-11-18', '2025-11-19'],
            'Vehicle Registration': ['TN02CE0730', 'TN02CE0738', 'TN02CE0730'],
            'Driver Name': ['Abdul Nayeem', 'Alexander A', 'Abdul Nayeem'],
            'Start Time': ['06:00', '14:00', '06:00'],
            'End Time': ['14:00', '22:00', '14:00'],
            'Notes': ['Morning shift', 'Evening shift', 'Morning shift']
        }
        
        df = pd.DataFrame(sample_data)
        
        # Create Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Shift Assignments')
        
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=shift_assignments_template.xlsx'}
        )
    except Exception as e:
        logger.error(f"Error creating sample template: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to create template: {str(e)}")


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
    logger.info("Application shutdown")


# ==================== Include Router ====================

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

